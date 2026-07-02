# quanly_xnt.py — Quản lý Xuất Nhập Tồn thuốc, vaccine, VTYT (Desktop — Local)
# Dành cho Trung tâm Kiểm soát bệnh tật (CDC)
# - Xuất kho / Cấp phát theo nguyên tắc FEFO
# - Quản lý thuốc, vaccine, vật tư y tế
# - Tích hợp mã vạch, báo cáo XNT, sao lưu tự động

import sqlite3
import tkinter as tk
from tkinter import messagebox, filedialog, simpledialog
import datetime as dt
import os, webbrowser, tempfile
import shutil
import json
import threading
import schedule
import time
from collections import defaultdict

import ttkbootstrap as tb
from ttkbootstrap.widgets import DateEntry
from ttkbootstrap.constants import *

# Import matplotlib for charts
try:
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    from matplotlib.figure import Figure
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

# Import pandas for Excel processing
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

# Import barcode scanner libraries
try:
    import cv2
    from pyzbar import pyzbar
    from PIL import Image, ImageTk
    BARCODE_AVAILABLE = True
except ImportError:
    BARCODE_AVAILABLE = False

# Import PDF export libraries
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
# ==== App info ====
APP_NAME     = "Quản lý XNT thuốc, vaccine và VTYT"
APP_VERSION  = "2.0.0"
AUTHOR_NAME  = "Hồ Sỷ Thoảng"
AUTHOR_EMAIL = "hstptcn5@gmail.com"
AUTHOR_PHONE = "0329381189"
AUTHOR_SITE  = "x/yoshinokuna"
# ==== /App info ====

# ==== App data paths (per-user, có quyền ghi) ====
if os.name == 'nt':
    base = os.environ.get('LOCALAPPDATA') or os.path.expanduser('~')
    APP_DIR = os.path.join(base, 'QuanLyXNT')
else:
    base = os.environ.get('XDG_DATA_HOME', os.path.join(os.path.expanduser('~'), '.local', 'share'))
    APP_DIR = os.path.join(base, 'quanlyxnt')

os.makedirs(APP_DIR, exist_ok=True)

DB_PATH  = os.path.join(APP_DIR, 'pharm.db')
LOG_PATH = os.path.join(APP_DIR, 'app.log')
LIC_PATH = os.path.join(APP_DIR, 'license.lic')  # LicenseManager cũng dùng đường dẫn này
BACKUP_DIR = os.path.join(APP_DIR, 'backups')
os.makedirs(BACKUP_DIR, exist_ok=True)

# Di trú DB cũ (nếu tồn tại cạnh file .py/.exe) sang APP_DIR lần đầu
for _old_name in ('pharm.db',):
    _old_db = os.path.join(os.path.dirname(__file__), _old_name)
    if not os.path.exists(DB_PATH) and os.path.exists(_old_db):
        shutil.copy2(_old_db, DB_PATH)
        break
# Cũng migrate từ thư mục Nhathuoc cũ nếu có
_old_app_dir = os.path.join(os.environ.get('LOCALAPPDATA', ''), 'Nhathuoc')
_old_db_path = os.path.join(_old_app_dir, 'pharm.db')
if not os.path.exists(DB_PATH) and os.path.exists(_old_db_path):
    shutil.copy2(_old_db_path, DB_PATH)
# ==== end paths ====


SCHEMA_SQL = r'''
PRAGMA foreign_keys = ON;

CREATE TABLE IF NOT EXISTS products (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  name TEXT NOT NULL,
  defaultUnit TEXT NOT NULL,
  barcode TEXT,
  productType TEXT DEFAULT 'thuoc',
  registrationNumber TEXT,
  createdAt TEXT DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS product_units (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  productId INTEGER NOT NULL REFERENCES products(id) ON DELETE CASCADE,
  unitCode TEXT NOT NULL,
  toBaseQty REAL NOT NULL,
  price REAL NOT NULL DEFAULT 0,
  UNIQUE(productId, unitCode)
);

CREATE TABLE IF NOT EXISTS batches (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  productId INTEGER NOT NULL REFERENCES products(id) ON DELETE CASCADE,
  lotNo TEXT NOT NULL,
  expiryDate TEXT NOT NULL,
  UNIQUE(productId, lotNo)
);

CREATE TABLE IF NOT EXISTS stock_movements (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  productId INTEGER NOT NULL REFERENCES products(id),
  batchId INTEGER NOT NULL REFERENCES batches(id),
  unitCode TEXT NOT NULL,
  qty REAL NOT NULL,
  type TEXT NOT NULL,
  cost REAL,
  receivingUnit TEXT,
  reason TEXT,
  createdAt TEXT DEFAULT CURRENT_TIMESTAMP
);

-- Bảng đơn vị nhận (cấp phát cho ai)
CREATE TABLE IF NOT EXISTS receiving_units (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  name TEXT NOT NULL UNIQUE,
  address TEXT,
  note TEXT,
  createdAt TEXT DEFAULT CURRENT_TIMESTAMP
);

-- Bảng phiếu xuất kho
CREATE TABLE IF NOT EXISTS dispatch_notes (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  noteNumber TEXT,
  receivingUnit TEXT NOT NULL,
  reason TEXT DEFAULT 'Cấp phát',
  note TEXT,
  createdAt TEXT DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS dispatch_items (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  dispatchId INTEGER NOT NULL REFERENCES dispatch_notes(id) ON DELETE CASCADE,
  productId INTEGER NOT NULL REFERENCES products(id),
  batchId INTEGER,
  unitCode TEXT NOT NULL,
  qty REAL NOT NULL,
  lotNo TEXT,
  expiryDate TEXT
);

CREATE TABLE IF NOT EXISTS sales (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  createdAt TEXT DEFAULT CURRENT_TIMESTAMP,
  total REAL NOT NULL,
  paid REAL NOT NULL,
  change REAL NOT NULL,
  note TEXT
);

CREATE TABLE IF NOT EXISTS sale_items (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  saleId INTEGER NOT NULL REFERENCES sales(id) ON DELETE CASCADE,
  productId INTEGER NOT NULL REFERENCES products(id),
  unitCode TEXT NOT NULL,
  qty REAL NOT NULL,
  price REAL NOT NULL
);

CREATE TABLE IF NOT EXISTS purchase_notes (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  noteNumber TEXT,
  supplier TEXT NOT NULL,
  reason TEXT DEFAULT 'Nhập kho',
  note TEXT,
  createdAt TEXT DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS purchase_items (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  purchaseId INTEGER NOT NULL REFERENCES purchase_notes(id) ON DELETE CASCADE,
  productId INTEGER NOT NULL REFERENCES products(id),
  batchId INTEGER,
  unitCode TEXT NOT NULL,
  qty REAL NOT NULL,
  lotNo TEXT,
  expiryDate TEXT,
  cost REAL NOT NULL
);
'''
# --- Hardware fingerprint (ổn định, không cần quyền admin) ---
def machine_fingerprint() -> str:
    import uuid, platform, hashlib, subprocess, os
    parts = [
        platform.node() or '',
        platform.system() or '',
        platform.machine() or '',
        hex(uuid.getnode()) or ''
    ]
    if os.name == 'nt':
        # cố gắng lấy 1 trong 2: BIOS serial / CSProduct UUID
        for cmd in ('wmic bios get serialnumber', 'wmic csproduct get uuid'):
            try:
                out = subprocess.check_output(cmd, shell=True, stderr=subprocess.DEVNULL)
                lines = [x.strip() for x in out.decode(errors='ignore').splitlines()
                         if x.strip() and 'serial' not in x.lower() and 'uuid' not in x.lower()]
                if lines:
                    parts.append(lines[-1])
                    break
            except Exception:
                pass
    raw = '|'.join(parts)
    return hashlib.sha256(raw.encode('utf-8')).hexdigest()[:16]  # 16 hex là đủ

# ---------------- DB ----------------
class DB:
    def __init__(self, path: str):
        self.path = path
        self.conn = sqlite3.connect(self.path)
        self.conn.row_factory = sqlite3.Row
        self.conn.execute('PRAGMA foreign_keys = ON')
        self.conn.execute("PRAGMA journal_mode=WAL")
        self.conn.execute("PRAGMA synchronous=NORMAL")
        self.conn.executescript(SCHEMA_SQL)

        try: self.conn.execute("ALTER TABLE products ADD COLUMN barcode TEXT")
        except sqlite3.OperationalError: pass

        self.migrate_schema()

        self.conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_products_barcode ON products(barcode) WHERE barcode IS NOT NULL")
        self.conn.execute("CREATE INDEX IF NOT EXISTS idx_sm_product_batch ON stock_movements(productId, batchId)")
        self.conn.execute("CREATE INDEX IF NOT EXISTS idx_batches_product ON batches(productId)")
        self.conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_units_unique ON product_units(productId, unitCode)")
        self.conn.commit()

    def _has_column(self, table, col):
        return any(r[1] == col for r in self.conn.execute(f"PRAGMA table_info({table})"))

    def migrate_schema(self):
        if not self._has_column('stock_movements', 'cost'):
            self.conn.execute("ALTER TABLE stock_movements ADD COLUMN cost REAL")

        # Thêm các trường mới cho stock_movements (v2.0)
        if not self._has_column('stock_movements', 'receivingUnit'):
            self.conn.execute("ALTER TABLE stock_movements ADD COLUMN receivingUnit TEXT")
        if not self._has_column('stock_movements', 'reason'):
            self.conn.execute("ALTER TABLE stock_movements ADD COLUMN reason TEXT")

        # Thêm các trường mới cho products
        if not self._has_column('products', 'productType'):
            self.conn.execute("ALTER TABLE products ADD COLUMN productType TEXT DEFAULT 'thuoc'")
        if not self._has_column('products', 'registrationNumber'):
            self.conn.execute("ALTER TABLE products ADD COLUMN registrationNumber TEXT")

        # Migrate productType cũ: 'general' → 'thuoc', 'medicine' → 'thuoc'
        self.conn.execute("UPDATE products SET productType='thuoc' WHERE productType IN ('general', 'medicine')")

        self.conn.execute("CREATE TABLE IF NOT EXISTS sales (id INTEGER PRIMARY KEY AUTOINCREMENT)")
        for col, ddl in [
            ('createdAt', "ALTER TABLE sales ADD COLUMN createdAt TEXT DEFAULT CURRENT_TIMESTAMP"),
            ('total',     "ALTER TABLE sales ADD COLUMN total REAL DEFAULT 0"),
            ('paid',      "ALTER TABLE sales ADD COLUMN paid REAL DEFAULT 0"),
            ('change',    "ALTER TABLE sales ADD COLUMN change REAL DEFAULT 0"),
            ('note',      "ALTER TABLE sales ADD COLUMN note TEXT"),
        ]:
            if not self._has_column('sales', col): self.conn.execute(ddl)

        self.conn.execute("CREATE TABLE IF NOT EXISTS sale_items (id INTEGER PRIMARY KEY AUTOINCREMENT)")
        for col, ddl in [
            ('saleId',    "ALTER TABLE sale_items ADD COLUMN saleId INTEGER REFERENCES sales(id) ON DELETE CASCADE"),
            ('productId', "ALTER TABLE sale_items ADD COLUMN productId INTEGER REFERENCES products(id)"),
            ('unitCode',  "ALTER TABLE sale_items ADD COLUMN unitCode TEXT"),
            ('qty',       "ALTER TABLE sale_items ADD COLUMN qty REAL DEFAULT 0"),
            ('price',     "ALTER TABLE sale_items ADD COLUMN price REAL DEFAULT 0"),
        ]:
            if not self._has_column('sale_items', col): self.conn.execute(ddl)

        # đảm bảo có dòng đơn vị cơ sở
        for r in self.conn.execute("SELECT id, defaultUnit FROM products"):
            if not self.conn.execute("SELECT 1 FROM product_units WHERE productId=? AND unitCode=?", (r['id'], r['defaultUnit'])).fetchone():
                self.conn.execute("INSERT INTO product_units(productId, unitCode, toBaseQty, price) VALUES(?,?,1,0)", (r['id'], r['defaultUnit']))
        self.conn.commit()

    # utils
    def q(self, sql, params=()):
        return [dict(r) for r in self.conn.execute(sql, params).fetchall()]

    def ex(self, sql, params=()):
        cur = self.conn.execute(sql, params); self.conn.commit(); return cur.lastrowid

    def default_unit_of(self, product_id):
        r = self.q("SELECT defaultUnit FROM products WHERE id=?", (product_id,))
        return r[0]['defaultUnit'] if r else None

    def unit_info(self, product_id, unit_code):
        rs = self.q("SELECT toBaseQty, price FROM product_units WHERE productId=? AND unitCode=?", (product_id, unit_code))
        return (float(rs[0]['toBaseQty']), float(rs[0]['price'])) if rs else (None, None)

    def unit_price(self, product_id, unit_code):
        _, price = self.unit_info(product_id, unit_code); return price or 0.0

    # views
    def stock_view(self):
        sql = '''
        SELECT p.id AS productId, p.name AS productName, sm.batchId, b.lotNo, b.expiryDate,
               ROUND(SUM(sm.qty*1),4) AS qtyBase,
               COALESCE(ROUND((
                    SELECT sm2.cost/1.0 FROM stock_movements sm2
                    WHERE sm2.productId=sm.productId AND sm2.batchId=sm.batchId
                      AND sm2.type='PURCHASE' AND sm2.cost IS NOT NULL
                    ORDER BY sm2.id DESC LIMIT 1
               ),2),0) AS costBase,
               COALESCE(ROUND((
                   (SELECT sm3.cost/1.0 FROM stock_movements sm3
                    WHERE sm3.productId=sm.productId AND sm3.batchId=sm.batchId
                      AND sm3.type='PURCHASE' AND sm3.cost IS NOT NULL
                    ORDER BY sm3.id DESC LIMIT 1) * SUM(sm.qty*1)
               ),2),0) AS valueBase
        FROM stock_movements sm
        JOIN products p ON p.id=sm.productId
        JOIN batches  b ON b.id=sm.batchId
        GROUP BY p.id,p.name,sm.batchId,b.lotNo,b.expiryDate
        HAVING qtyBase<>0
        ORDER BY LOWER(p.name), DATE(b.expiryDate)
        '''
        return self.q(sql)

    def expiring_view(self, days=90):
        sql = '''
        SELECT * FROM (
            SELECT p.id AS productId, p.name AS productName, sm.batchId, b.lotNo, b.expiryDate,
                   ROUND(SUM(sm.qty*1),4) AS qtyBase
            FROM stock_movements sm
            JOIN products p ON p.id=sm.productId
            JOIN batches  b ON b.id=sm.batchId
            GROUP BY sm.productId, sm.batchId, b.lotNo, b.expiryDate
        ) v
        WHERE qtyBase>0 AND DATE(expiryDate) <= DATE('now','+' || ? || ' day')
        ORDER BY LOWER(productName), DATE(expiryDate)
        '''
        return self.q(sql, (days,))

    def stock_summary_by_product(self):
        sql = '''
        SELECT p.id AS productId, p.name AS productName, ROUND(SUM(v.qtyBase),4) AS qtyBaseTotal
        FROM ( SELECT sm.productId, sm.batchId, SUM(sm.qty*1) AS qtyBase
               FROM stock_movements sm GROUP BY sm.productId, sm.batchId ) v
        JOIN products p ON p.id=v.productId
        GROUP BY p.id, p.name HAVING qtyBaseTotal<>0
        ORDER BY LOWER(p.name)
        '''
        return self.q(sql)
    def xnt_report(self, start_date: str, end_date: str):
        """
        Báo cáo Xuất–Nhập–Tồn theo sản phẩm trong khoảng ngày [start_date, end_date].
        - Nhập:  type='PURCHASE'
        - Xuất:  type IN ('SALE','DISCARD','DISPATCH')  (DISCARD và DISPATCH tính như xuất)
        - Đơn vị cơ sở (toBaseQty = 1)
        """
        sql = r'''
        SELECT
          p.id   AS productId,
          p.name AS productName,
          p.defaultUnit AS unit,
          b.lotNo AS lotNo,
          b.expiryDate AS expiryDate,

          COALESCE(ROUND(SUM(CASE
            WHEN DATE(sm.createdAt) < DATE(?) THEN sm.qty * 1
            ELSE 0 END), 4), 0) AS opening,

          COALESCE(ROUND(SUM(CASE
            WHEN DATE(sm.createdAt) BETWEEN DATE(?) AND DATE(?) AND sm.type='PURCHASE'
              THEN sm.qty * 1 ELSE 0 END), 4), 0) AS inbound,

          COALESCE(ROUND(SUM(CASE
            WHEN DATE(sm.createdAt) BETWEEN DATE(?) AND DATE(?) AND sm.type IN ('SALE','DISCARD','DISPATCH')
              THEN -sm.qty * 1 ELSE 0 END), 4), 0) AS outbound,

          COALESCE(ROUND(SUM(CASE
            WHEN DATE(sm.createdAt) <= DATE(?) THEN sm.qty * 1
            ELSE 0 END), 4), 0) AS closing
        FROM products p
        LEFT JOIN stock_movements sm ON sm.productId = p.id
        LEFT JOIN batches b ON sm.batchId = b.id
        GROUP BY p.id, p.name, b.id, b.lotNo, b.expiryDate
        HAVING opening <> 0 OR inbound <> 0 OR outbound <> 0 OR closing <> 0
        ORDER BY LOWER(p.name), b.expiryDate ASC
        '''
        params = (start_date, start_date, end_date, start_date, end_date, end_date)
        return self.q(sql, params)

    def stock_summary_by_product_range(self, start_date: str, end_date: str):
        """
        Tồn theo sản phẩm trong khoảng thời gian (lọc theo createdAt của stock_movements).
        """
        sql = '''
        SELECT p.id AS productId, p.name AS productName,
               ROUND(SUM(v.qtyBase), 4) AS qtyBaseTotal
        FROM (
          SELECT sm.productId, sm.batchId, SUM(sm.qty * 1) AS qtyBase
          FROM stock_movements sm
          WHERE DATE(sm.createdAt) BETWEEN DATE(?) AND DATE(?)
          GROUP BY sm.productId, sm.batchId
        ) v
        JOIN products p ON p.id = v.productId
        GROUP BY p.id, p.name
        HAVING qtyBaseTotal <> 0
        ORDER BY LOWER(p.name)
        '''
        return self.q(sql, (start_date, end_date))

    def ensure_batch(self, product_id, lot_no, expiry_date):
        r = self.q("SELECT id FROM batches WHERE productId=? AND lotNo=?", (product_id, lot_no))
        return r[0]['id'] if r else self.ex("INSERT INTO batches(productId, lotNo, expiryDate) VALUES(?,?,?)", (product_id, lot_no, expiry_date))

    # purchase
    def add_purchase(self, items):
        try:
            self.conn.execute("BEGIN")
            for it in items:
                bid = self.ensure_batch(it['productId'], it['lotNo'], it['expiryDate'])
                self.conn.execute(
                    "INSERT INTO stock_movements(productId, batchId, unitCode, qty, type, cost) VALUES(?,?,?,?, 'PURCHASE', ?)",
                    (it['productId'], bid, it['unitCode'], it['qty'], float(it.get('cost') or 0))
                )
            self.conn.commit()
        except Exception:
            self.conn.rollback(); raise

    # sell (FEFO)
    def sell(self, items):
        total = 0.0
        for it in items:
            to_base, unit_price = self.unit_info(it['productId'], it['unitCode'])
            if to_base is None: raise Exception('Sản phẩm chưa có giá/đơn vị cơ sở')
            need_base = float(it['qty']) * to_base
            lots = self.q('''
              SELECT v.batchId, v.qtyBase, b.expiryDate FROM (
                SELECT sm.batchId, SUM(sm.qty*1) AS qtyBase
                FROM stock_movements sm WHERE sm.productId=? GROUP BY sm.batchId
              ) v JOIN batches b ON b.id=v.batchId
              WHERE v.qtyBase>0 AND DATE(b.expiryDate) >= DATE('now')
              ORDER BY DATE(b.expiryDate)
            ''', (it['productId'],))
            for lot in lots:
                if need_base <= 0: break
                take_base = min(need_base, float(lot['qtyBase']))
                take_in_unit = take_base / to_base
                self.conn.execute("INSERT INTO stock_movements(productId, batchId, unitCode, qty, type) VALUES(?,?,?,?, 'SALE')",
                                  (it['productId'], lot['batchId'], it['unitCode'], -take_in_unit))
                need_base -= take_base
            if need_base > 0: raise Exception('Không đủ tồn kho')
            total += (unit_price or 0.0) * float(it['qty'])
        return round(total, 2)

    def record_sale(self, items, paid: float, note: str = ''):
        finalized = []
        for it in items:
            price = self.unit_price(it['productId'], it['unitCode'])
            finalized.append({'productId': it['productId'], 'productName': it.get('productName') or f"#{it['productId']}",
                              'unitCode': it['unitCode'], 'qty': float(it['qty']), 'price': float(price)})
        total = round(sum(i['qty']*i['price'] for i in finalized), 2)
        paid = float(paid); change = round(paid - total, 2)
        if paid < total: raise Exception('Tiền khách đưa chưa đủ')
        try:
            self.conn.execute("BEGIN")
            self.sell(finalized)
            cur = self.conn.execute("INSERT INTO sales(total, paid, change, note) VALUES(?,?,?,?)", (total, paid, change, note))
            sale_id = cur.lastrowid
            for it in finalized:
                self.conn.execute("INSERT INTO sale_items(saleId, productId, unitCode, qty, price) VALUES(?,?,?,?,?)",
                                  (sale_id, it['productId'], it['unitCode'], it['qty'], it['price']))
            self.conn.commit()
            return sale_id, finalized, total, change
        except Exception:
            self.conn.rollback(); raise

    # dispatch (Xuất kho / Cấp phát — FEFO)
    def dispatch(self, items, receiving_unit: str, reason: str = 'Cấp phát', note: str = '', date_str: str = None):
        """
        Xuất kho / cấp phát hàng theo FEFO.
        items: list of {'productId', 'unitCode', 'qty'}
        receiving_unit: tên đơn vị nhận (VD: TYT Phường X)
        reason: Cấp phát / Hủy / Chuyển kho
        date_str: Ngày xuất tùy chọn (dạng YYYY-MM-DD), nếu None thì lấy thời gian hiện tại
        """
        dispatch_details = []
        try:
            self.conn.execute("BEGIN")

            # Thời gian tạo phiếu xuất
            if date_str:
                created_at = f"{date_str} {dt.datetime.now().strftime('%H:%M:%S')}"
            else:
                created_at = dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # Tạo phiếu xuất kho
            note_number = f"PXK-{dt.datetime.now().strftime('%Y%m%d%H%M%S')}"
            cur = self.conn.execute(
                "INSERT INTO dispatch_notes(noteNumber, receivingUnit, reason, note, createdAt) VALUES(?,?,?,?,?)",
                (note_number, receiving_unit, reason, note, created_at)
            )
            dispatch_id = cur.lastrowid

            for it in items:
                to_base, _ = self.unit_info(it['productId'], it['unitCode'])
                if to_base is None:
                    raise Exception(f"Sản phẩm #{it['productId']} chưa có đơn vị cơ sở")
                need_base = float(it['qty']) * to_base

                # FEFO: lấy lô có hạn gần nhất
                lots = self.q('''
                  SELECT v.batchId, v.qtyBase, b.expiryDate, b.lotNo FROM (
                    SELECT sm.batchId, SUM(sm.qty*1) AS qtyBase
                    FROM stock_movements sm WHERE sm.productId=? GROUP BY sm.batchId
                  ) v JOIN batches b ON b.id=v.batchId
                  WHERE v.qtyBase>0 AND DATE(b.expiryDate) >= DATE('now')
                  ORDER BY DATE(b.expiryDate)
                ''', (it['productId'],))

                for lot in lots:
                    if need_base <= 0:
                        break
                    take_base = min(need_base, float(lot['qtyBase']))
                    take_in_unit = take_base / to_base
                    self.conn.execute(
                        "INSERT INTO stock_movements(productId, batchId, unitCode, qty, type, receivingUnit, reason, createdAt) VALUES(?,?,?,?, 'DISPATCH', ?,?,?)",
                        (it['productId'], lot['batchId'], it['unitCode'], -take_in_unit, receiving_unit, reason, created_at)
                    )
                    # Ghi chi tiết phiếu xuất
                    self.conn.execute(
                        "INSERT INTO dispatch_items(dispatchId, productId, batchId, unitCode, qty, lotNo, expiryDate) VALUES(?,?,?,?,?,?,?)",
                        (dispatch_id, it['productId'], lot['batchId'], it['unitCode'], take_in_unit, lot['lotNo'], lot['expiryDate'])
                    )
                    dispatch_details.append({
                        'productId': it['productId'],
                        'productName': it.get('productName', f"#{it['productId']}"),
                        'unitCode': it['unitCode'],
                        'qty': take_in_unit,
                        'lotNo': lot['lotNo'],
                        'expiryDate': lot['expiryDate'],
                        'batchId': lot['batchId']
                    })
                    need_base -= take_base

                if need_base > 0:
                    raise Exception(f"Không đủ tồn kho cho sản phẩm #{it['productId']}")

            # Lưu đơn vị nhận vào bảng receiving_units (nếu chưa có)
            self._save_receiving_unit(receiving_unit)

            self.conn.commit()
            return dispatch_id, note_number, dispatch_details

        except Exception:
            self.conn.rollback()
            raise

    def _save_receiving_unit(self, name: str):
        """Lưu đơn vị nhận mới (nếu chưa có) để autocomplete lần sau"""
        if not name or not name.strip():
            return
        try:
            self.conn.execute(
                "INSERT OR IGNORE INTO receiving_units(name) VALUES(?)",
                (name.strip(),)
            )
        except Exception:
            pass

    def get_receiving_units(self):
        """Lấy danh sách đơn vị nhận đã lưu"""
        return [r['name'] for r in self.q("SELECT name FROM receiving_units ORDER BY name")]

    def get_dispatch_notes(self, start_date: str = None, end_date: str = None):
        """Lấy danh sách phiếu xuất kho"""
        if start_date and end_date:
            return self.q('''
                SELECT dn.*, COUNT(di.id) as item_count
                FROM dispatch_notes dn
                LEFT JOIN dispatch_items di ON di.dispatchId = dn.id
                WHERE DATE(dn.createdAt) BETWEEN DATE(?) AND DATE(?)
                GROUP BY dn.id
                ORDER BY dn.createdAt DESC
            ''', (start_date, end_date))
        return self.q('''
            SELECT dn.*, COUNT(di.id) as item_count
            FROM dispatch_notes dn
            LEFT JOIN dispatch_items di ON di.dispatchId = dn.id
            GROUP BY dn.id
            ORDER BY dn.createdAt DESC
            LIMIT 50
        ''')

    def get_dispatch_detail(self, dispatch_id: int):
        """Lấy chi tiết phiếu xuất kho"""
        return self.q('''
            SELECT di.*, p.name as productName
            FROM dispatch_items di
            JOIN products p ON p.id = di.productId
            WHERE di.dispatchId = ?
            ORDER BY p.name
        ''', (dispatch_id,))

    # purchase (Nhập kho)
    def record_purchase(self, items, supplier: str, reason: str = 'Nhập kho', note: str = '', date_str: str = None):
        """
        Nhập kho thuốc, vaccine, VTYT và lưu phiếu nhập.
        items: list of {'productId', 'unitCode', 'qty', 'lotNo', 'expiryDate', 'cost'}
        """
        purchase_details = []
        try:
            self.conn.execute("BEGIN")

            # Thời gian tạo phiếu nhập
            if date_str:
                created_at = f"{date_str} {dt.datetime.now().strftime('%H:%M:%S')}"
            else:
                created_at = dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # Tạo số phiếu nhập
            note_number = f"PNK-{dt.datetime.now().strftime('%Y%m%d%H%M%S')}"
            cur = self.conn.execute(
                "INSERT INTO purchase_notes(noteNumber, supplier, reason, note, createdAt) VALUES(?,?,?,?,?)",
                (note_number, supplier, reason, note, created_at)
            )
            purchase_id = cur.lastrowid

            for it in items:
                # Bảo đảm lô hàng tồn tại
                bid = self.ensure_batch(it['productId'], it['lotNo'], it['expiryDate'])
                
                # Ghi chuyển động kho
                self.conn.execute(
                    "INSERT INTO stock_movements(productId, batchId, unitCode, qty, type, cost, receivingUnit, reason, createdAt) VALUES(?,?,?,?, 'PURCHASE', ?,?,?,?)",
                    (it['productId'], bid, it['unitCode'], float(it['qty']), float(it.get('cost') or 0), supplier, reason, created_at)
                )
                
                # Ghi chi tiết phiếu nhập
                self.conn.execute(
                    "INSERT INTO purchase_items(purchaseId, productId, batchId, unitCode, qty, lotNo, expiryDate, cost) VALUES(?,?,?,?,?,?,?,?)",
                    (purchase_id, it['productId'], bid, it['unitCode'], float(it['qty']), it['lotNo'], it['expiryDate'], float(it.get('cost') or 0))
                )
                
                # Đồng bộ giá bán base = giá nhập
                self.conn.execute(
                    "UPDATE product_units SET price=? WHERE productId=? AND unitCode=?",
                    (float(it.get('cost') or 0), it['productId'], it['unitCode'])
                )
                
                purchase_details.append({
                    'productId': it['productId'],
                    'productName': it.get('productName', f"#{it['productId']}"),
                    'unitCode': it['unitCode'],
                    'qty': float(it['qty']),
                    'lotNo': it['lotNo'],
                    'expiryDate': it['expiryDate'],
                    'cost': float(it.get('cost') or 0),
                    'batchId': bid
                })

            self.conn.commit()
            return purchase_id, note_number, purchase_details

        except Exception:
            self.conn.rollback()
            raise

    def get_purchase_notes(self, start_date: str = None, end_date: str = None):
        """Lấy danh sách phiếu nhập kho"""
        if start_date and end_date:
            return self.q('''
                SELECT pn.*, COUNT(pi.id) as item_count
                FROM purchase_notes pn
                LEFT JOIN purchase_items pi ON pi.purchaseId = pn.id
                WHERE DATE(pn.createdAt) BETWEEN DATE(?) AND DATE(?)
                GROUP BY pn.id
                ORDER BY pn.createdAt DESC
            ''', (start_date, end_date))
        return self.q('''
            SELECT pn.*, COUNT(pi.id) as item_count
            FROM purchase_notes pn
            LEFT JOIN purchase_items pi ON pi.purchaseId = pn.id
            GROUP BY pn.id
            ORDER BY pn.createdAt DESC
            LIMIT 50
        ''')

    def get_purchase_detail(self, purchase_id: int):
        """Lấy chi tiết phiếu nhập kho"""
        return self.q('''
            SELECT pi.*, p.name as productName
            FROM purchase_items pi
            JOIN products p ON p.id = pi.productId
            WHERE pi.purchaseId = ?
            ORDER BY p.name
        ''', (purchase_id,))

    def get_suppliers(self):
        """Lấy danh sách nhà cung cấp đã từng nhập hàng"""
        rows = self.q("SELECT DISTINCT supplier FROM purchase_notes WHERE supplier != '' ORDER BY supplier")
        return [r['supplier'] for r in rows]

# ---------------- Backup Manager ----------------
class BackupManager:
    def __init__(self, db_path: str, backup_dir: str):
        self.db_path = db_path
        self.backup_dir = backup_dir
        self.auto_backup_enabled = True
        self.backup_interval_hours = 24  # Mặc định 24 giờ
        self.max_backups = 30  # Giữ tối đa 30 file backup
        
    def create_backup(self, custom_name: str = None) -> str:
        """Tạo backup database"""
        try:
            timestamp = dt.datetime.now().strftime('%Y%m%d_%H%M%S')
            if custom_name:
                backup_name = f"{custom_name}_{timestamp}.db"
            else:
                backup_name = f"backup_{timestamp}.db"
            
            backup_path = os.path.join(self.backup_dir, backup_name)
            shutil.copy2(self.db_path, backup_path)
            
            # Tạo file metadata
            metadata = {
                'created_at': dt.datetime.now().isoformat(),
                'db_size': os.path.getsize(self.db_path),
                'backup_size': os.path.getsize(backup_path),
                'version': APP_VERSION
            }
            
            metadata_path = backup_path.replace('.db', '.json')
            with open(metadata_path, 'w', encoding='utf-8') as f:
                json.dump(metadata, f, indent=2, ensure_ascii=False)
            
            self._cleanup_old_backups()
            return backup_path
            
        except Exception as e:
            raise Exception(f"Lỗi tạo backup: {str(e)}")
    
    def restore_backup(self, backup_path: str) -> bool:
        """Khôi phục từ backup"""
        try:
            if not os.path.exists(backup_path):
                raise Exception("File backup không tồn tại")
            
            # Tạo backup hiện tại trước khi restore
            current_backup = self.create_backup("before_restore")
            
            # Restore database
            shutil.copy2(backup_path, self.db_path)
            
            return True
            
        except Exception as e:
            raise Exception(f"Lỗi khôi phục backup: {str(e)}")
    
    def export_data(self, export_path: str) -> bool:
        """Export toàn bộ dữ liệu ra file JSON"""
        try:
            db = DB(self.db_path)
            
            export_data = {
                'export_info': {
                    'created_at': dt.datetime.now().isoformat(),
                    'version': APP_VERSION,
                    'app_name': APP_NAME
                },
                'products': db.q("SELECT * FROM products"),
                'product_units': db.q("SELECT * FROM product_units"),
                'batches': db.q("SELECT * FROM batches"),
                'stock_movements': db.q("SELECT * FROM stock_movements"),
                'sales': db.q("SELECT * FROM sales"),
                'sale_items': db.q("SELECT * FROM sale_items")
            }
            
            with open(export_path, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, indent=2, ensure_ascii=False)
            
            return True
            
        except Exception as e:
            raise Exception(f"Lỗi export dữ liệu: {str(e)}")
    
    def import_data(self, import_path: str) -> bool:
        """Import dữ liệu từ file JSON"""
        try:
            if not os.path.exists(import_path):
                raise Exception("File import không tồn tại")
            
            with open(import_path, 'r', encoding='utf-8') as f:
                import_data = json.load(f)
            
            # Tạo backup trước khi import
            self.create_backup("before_import")
            
            db = DB(self.db_path)
            
            # Xóa dữ liệu cũ (theo thứ tự để tránh lỗi foreign key)
            db.conn.execute("DELETE FROM sale_items")
            db.conn.execute("DELETE FROM sales")
            db.conn.execute("DELETE FROM stock_movements")
            db.conn.execute("DELETE FROM batches")
            db.conn.execute("DELETE FROM product_units")
            db.conn.execute("DELETE FROM products")
            
            # Import dữ liệu mới
            for table, data in import_data.items():
                if table == 'export_info':
                    continue
                    
                if data:
                    columns = list(data[0].keys())
                    placeholders = ','.join(['?' for _ in columns])
                    sql = f"INSERT INTO {table} ({','.join(columns)}) VALUES ({placeholders})"
                    
                    for row in data:
                        values = [row[col] for col in columns]
                        db.conn.execute(sql, values)
            
            db.conn.commit()
            return True
            
        except Exception as e:
            raise Exception(f"Lỗi import dữ liệu: {str(e)}")
    
    def list_backups(self) -> list:
        """Lấy danh sách các file backup"""
        backups = []
        try:
            for file in os.listdir(self.backup_dir):
                if file.endswith('.db'):
                    backup_path = os.path.join(self.backup_dir, file)
                    metadata_path = backup_path.replace('.db', '.json')
                    
                    backup_info = {
                        'file': file,
                        'path': backup_path,
                        'size': os.path.getsize(backup_path),
                        'created': dt.datetime.fromtimestamp(os.path.getmtime(backup_path))
                    }
                    
                    if os.path.exists(metadata_path):
                        try:
                            with open(metadata_path, 'r', encoding='utf-8') as f:
                                metadata = json.load(f)
                            backup_info.update(metadata)
                        except:
                            pass
                    
                    backups.append(backup_info)
            
            # Sắp xếp theo thời gian tạo (mới nhất trước)
            backups.sort(key=lambda x: x['created'], reverse=True)
            return backups
            
        except Exception as e:
            return []
    
    def _cleanup_old_backups(self):
        """Xóa các backup cũ nếu vượt quá số lượng cho phép"""
        try:
            backups = self.list_backups()
            if len(backups) > self.max_backups:
                for backup in backups[self.max_backups:]:
                    try:
                        os.remove(backup['path'])
                        metadata_path = backup['path'].replace('.db', '.json')
                        if os.path.exists(metadata_path):
                            os.remove(metadata_path)
                    except:
                        pass
        except:
            pass
    
    def start_auto_backup(self):
        """Bắt đầu tự động backup định kỳ"""
        if not self.auto_backup_enabled:
            return
            
        def backup_job():
            try:
                self.create_backup("auto")
                print(f"Auto backup completed at {dt.datetime.now()}")
            except Exception as e:
                print(f"Auto backup failed: {e}")
        
        # Lên lịch backup mỗi ngày lúc 2:00 AM
        schedule.every().day.at("02:00").do(backup_job)
        
        def run_scheduler():
            while True:
                schedule.run_pending()
                time.sleep(60)  # Kiểm tra mỗi phút
        
        # Chạy scheduler trong thread riêng
        scheduler_thread = threading.Thread(target=run_scheduler, daemon=True)
        scheduler_thread.start()

# ---------------- Export Manager ----------------
class ExportManager:
    def __init__(self):
        pass
    
    def export_to_excel(self, data, filename, sheet_name="Báo cáo", headers=None):
        """Xuất dữ liệu ra file Excel"""
        if not PANDAS_AVAILABLE:
            raise Exception("Thư viện pandas chưa được cài đặt. Vui lòng chạy: pip install pandas openpyxl")
        
        try:
            # Tạo DataFrame
            if isinstance(data, list) and len(data) > 0:
                if isinstance(data[0], dict):
                    # Dữ liệu từ database (list of dicts)
                    df = pd.DataFrame(data)
                else:
                    # Dữ liệu từ list thông thường
                    if headers:
                        df = pd.DataFrame(data, columns=headers)
                    else:
                        df = pd.DataFrame(data)
            else:
                raise Exception("Dữ liệu không hợp lệ")
            
            # Tạo file Excel
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Lấy worksheet để format
                worksheet = writer.sheets[sheet_name]
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Format header
                from openpyxl.styles import Font, PatternFill
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
            
            return True
            
        except Exception as e:
            raise Exception(f"Lỗi xuất Excel: {str(e)}")
    
    def export_to_pdf(self, data, filename, title="Báo cáo", headers=None):
        """Xuất dữ liệu ra file PDF"""
        if not PDF_AVAILABLE:
            raise Exception("Thư viện reportlab chưa được cài đặt. Vui lòng chạy: pip install reportlab")
        
        try:
            # Tạo document
            doc = SimpleDocTemplate(filename, pagesize=A4)
            story = []
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Center
            )
            
            # Title
            story.append(Paragraph(title, title_style))
            story.append(Spacer(1, 12))
            
            # Tạo bảng dữ liệu
            if isinstance(data, list) and len(data) > 0:
                if isinstance(data[0], dict):
                    # Dữ liệu từ database
                    if headers:
                        table_data = [headers]
                    else:
                        # Lấy headers từ keys của dict đầu tiên
                        headers = list(data[0].keys())
                        table_data = [headers]
                    
                    # Thêm dữ liệu
                    for row in data:
                        table_data.append([str(row.get(h, '')) for h in headers])
                else:
                    # Dữ liệu từ list thông thường
                    if headers:
                        table_data = [headers] + data
                    else:
                        table_data = data
                
                # Tạo bảng
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                
                story.append(table)
            
            # Build PDF
            doc.build(story)
            return True
            
        except Exception as e:
            raise Exception(f"Lỗi xuất PDF: {str(e)}")

# ---------------- Report Manager ----------------
class ReportManager:
    def __init__(self, db_path: str):
        self.db_path = db_path
        
    def get_revenue_report(self, start_date: str, end_date: str, group_by: str = 'day') -> list:
        """Thống kê phiếu xuất kho theo ngày/tháng/năm (Thay cho báo cáo doanh thu cũ)"""
        try:
            db = DB(self.db_path)
            
            if group_by == 'day':
                date_format = "DATE(createdAt)"
                group_format = "DATE(createdAt)"
            elif group_by == 'month':
                date_format = "strftime('%Y-%m', createdAt)"
                group_format = "strftime('%Y-%m', createdAt)"
            elif group_by == 'year':
                date_format = "strftime('%Y', createdAt)"
                group_format = "strftime('%Y', createdAt)"
            else:
                raise Exception("group_by phải là 'day', 'month', hoặc 'year'")
            
            sql = f'''
            SELECT 
                {date_format} as period,
                COUNT(DISTINCT dn.id) as total_orders, -- Số lượng phiếu xuất
                SUM(di.qty) as total_revenue,          -- Tổng số lượng sản phẩm xuất
                COUNT(di.id) as total_paid,            -- Số lượng danh mục mặt hàng xuất
                AVG(di.qty) as avg_order_value         -- Số lượng TB trên mỗi dòng
            FROM dispatch_notes dn
            LEFT JOIN dispatch_items di ON dn.id = di.dispatchId
            WHERE DATE(dn.createdAt) BETWEEN DATE(?) AND DATE(?)
            GROUP BY {group_format}
            ORDER BY period
            '''
            
            return db.q(sql, (start_date, end_date))
            
        except Exception as e:
            raise Exception(f"Lỗi thống kê xuất kho: {str(e)}")
    
    def get_profit_report(self, start_date: str, end_date: str) -> list:
        """Thống kê cấp phát theo đơn vị nhận (Thay cho báo cáo lợi nhuận cũ)"""
        try:
            db = DB(self.db_path)
            
            sql = '''
            SELECT 
                DATE(dn.createdAt) as sale_date,       -- Ngày xuất
                dn.receivingUnit as product_name,       -- Tên đơn vị nhận (hiển thị ở cột sản phẩm)
                COUNT(DISTINCT dn.id) as qty,          -- Số lượng phiếu nhận
                0 as sell_price,
                0 as cost_price,
                SUM(di.qty) as revenue,                -- Tổng số lượng sản phẩm nhận
                COUNT(di.id) as cost,                  -- Số lượng loại sản phẩm nhận
                SUM(di.qty) as profit                  -- Tổng số lượng
            FROM dispatch_notes dn
            LEFT JOIN dispatch_items di ON dn.id = di.dispatchId
            WHERE DATE(dn.createdAt) BETWEEN DATE(?) AND DATE(?)
            GROUP BY DATE(dn.createdAt), dn.receivingUnit
            ORDER BY dn.createdAt DESC
            '''
            
            return db.q(sql, (start_date, end_date))
            
        except Exception as e:
            raise Exception(f"Lỗi thống kê theo đơn vị nhận: {str(e)}")
    
    def get_top_products(self, start_date: str, end_date: str, limit: int = 10) -> list:
        """Top sản phẩm cấp phát nhiều nhất (Thay cho top sản phẩm bán chạy)"""
        try:
            db = DB(self.db_path)
            
            sql = '''
            SELECT 
                p.id as product_id,
                p.name as product_name,
                SUM(di.qty) as total_qty,              -- Tổng số lượng cấp
                COUNT(DISTINCT dn.id) as total_orders, -- Số lượng phiếu xuất chứa sản phẩm này
                SUM(di.qty) as total_revenue,          -- Tổng số lượng (để hiển thị)
                0 as avg_price
            FROM dispatch_notes dn
            JOIN dispatch_items di ON dn.id = di.dispatchId
            JOIN products p ON di.productId = p.id
            WHERE DATE(dn.createdAt) BETWEEN DATE(?) AND DATE(?)
            GROUP BY p.id, p.name
            ORDER BY total_qty DESC
            LIMIT ?
            '''
            
            return db.q(sql, (start_date, end_date, limit))
            
        except Exception as e:
            raise Exception(f"Lỗi thống kê top sản phẩm: {str(e)}")
    
    def get_daily_sales_summary(self, start_date: str, end_date: str) -> dict:
        """Tóm tắt cấp phát theo ngày (Thay cho tóm tắt bán hàng cũ)"""
        try:
            db = DB(self.db_path)
            
            # Tổng quan
            summary_sql = '''
            SELECT 
                COUNT(DISTINCT dn.id) as total_orders, -- Tổng số phiếu
                SUM(di.qty) as total_revenue,          -- Tổng số lượng xuất
                AVG(di.qty) as avg_order_value,
                MIN(di.qty) as min_order,
                MAX(di.qty) as max_order
            FROM dispatch_notes dn
            LEFT JOIN dispatch_items di ON dn.id = di.dispatchId
            WHERE DATE(dn.createdAt) BETWEEN DATE(?) AND DATE(?)
            '''
            
            summary_result = db.q(summary_sql, (start_date, end_date))
            summary = summary_result[0] if summary_result else {
                'total_orders': 0,
                'total_revenue': 0,
                'avg_order_value': 0,
                'min_order': 0,
                'max_order': 0
            }
            
            # Số phiếu và số lượng theo ngày
            daily_sql = '''
            SELECT 
                DATE(dn.createdAt) as sale_date,
                COUNT(DISTINCT dn.id) as orders,
                SUM(di.qty) as revenue
            FROM dispatch_notes dn
            LEFT JOIN dispatch_items di ON dn.id = di.dispatchId
            WHERE DATE(dn.createdAt) BETWEEN DATE(?) AND DATE(?)
            GROUP BY DATE(dn.createdAt)
            ORDER BY sale_date
            '''
            
            daily_data = db.q(daily_sql, (start_date, end_date))
            
            return {
                'summary': summary,
                'daily_data': daily_data
            }
            
        except Exception as e:
            raise Exception(f"Lỗi tạo tóm tắt cấp phát: {str(e)}")
    
    def get_category_performance(self, start_date: str, end_date: str) -> list:
        """Hiệu suất cấp phát theo phân loại sản phẩm (thuoc, vaccine, vtyt, khac)"""
        try:
            db = DB(self.db_path)
            
            sql = '''
            SELECT 
                p.productType as category,             -- Phân loại sản phẩm
                COUNT(DISTINCT di.productId) as product_count,
                SUM(di.qty) as total_qty,
                SUM(di.qty) as total_revenue,
                0 as avg_price
            FROM dispatch_notes dn
            JOIN dispatch_items di ON dn.id = di.dispatchId
            JOIN products p ON di.productId = p.id
            WHERE DATE(dn.createdAt) BETWEEN DATE(?) AND DATE(?)
            GROUP BY p.productType
            ORDER BY total_qty DESC
            '''
            return db.q(sql, (start_date, end_date))
        except Exception as e:
            raise Exception(f"Lỗi thống kê hiệu suất theo phân loại: {str(e)}")

# ---------------- Medicine Catalog Manager ----------------
class MedicineCatalogManager:
    def __init__(self, db_path: str):
        self.db_path = db_path
        self.catalog_data = None
        self.catalog_file_path = None
        
    def load_catalog_from_excel(self, file_path: str) -> bool:
        """Load danh mục thuốc từ file Excel hoặc CSV"""
        if not PANDAS_AVAILABLE:
            raise Exception("Thư viện pandas chưa được cài đặt. Vui lòng chạy: pip install pandas openpyxl")
        
        try:
            # Đọc file Excel hoặc CSV
            if file_path.lower().endswith('.csv'):
                df = pd.read_csv(file_path, encoding='utf-8')
            else:
                df = pd.read_excel(file_path)
            
            # Chuẩn hóa tên cột (loại bỏ khoảng trắng, chuyển thành lowercase)
            df.columns = df.columns.str.strip().str.lower()
            
            # Lưu dữ liệu
            self.catalog_data = df
            self.catalog_file_path = file_path
            
            return True
            
        except Exception as e:
            raise Exception(f"Lỗi đọc file: {str(e)}")
    
    def search_medicine(self, medicine_name: str) -> list:
        """Tìm kiếm thuốc trong danh mục"""
        if self.catalog_data is None:
            return []
        
        try:
            # Tìm kiếm theo tên thuốc (không phân biệt hoa thường)
            medicine_name_lower = medicine_name.lower().strip()
            
            # Tìm kiếm trong các cột có thể chứa tên thuốc
            search_columns = []
            for col in self.catalog_data.columns:
                if any(keyword in col.lower() for keyword in ['ten', 'name', 'thuoc', 'medicine', 'san_pham', 'product']):
                    search_columns.append(col)
            
            if not search_columns:
                # Nếu không tìm thấy cột tên, sử dụng cột đầu tiên
                search_columns = [self.catalog_data.columns[0]]
            
            # Tìm kiếm
            results = []
            for col in search_columns:
                mask = self.catalog_data[col].astype(str).str.lower().str.contains(medicine_name_lower, na=False)
                matches = self.catalog_data[mask]
                
                for _, row in matches.iterrows():
                    result = {
                        'name': str(row.get(col, '')),
                        'registration_number': '',
                        'manufacturer': '',
                        'active_ingredient': '',
                        'dosage_form': '',
                        'strength': '',
                        'pack_size': '',
                        'other_info': {}
                    }
                    
                    # Tìm số đăng ký
                    for reg_col in self.catalog_data.columns:
                        if any(keyword in reg_col.lower() for keyword in ['so_dang_ky', 'registration', 'dang_ky', 'number']):
                            result['registration_number'] = str(row.get(reg_col, ''))
                            break
                    
                    # Tìm nhà sản xuất
                    for manu_col in self.catalog_data.columns:
                        if any(keyword in manu_col.lower() for keyword in ['nha_san_xuat', 'manufacturer', 'cong_ty', 'company']):
                            result['manufacturer'] = str(row.get(manu_col, ''))
                            break
                    
                    # Tìm hoạt chất
                    for active_col in self.catalog_data.columns:
                        if any(keyword in active_col.lower() for keyword in ['hoat_chat', 'active', 'ingredient', 'thanh_phan']):
                            result['active_ingredient'] = str(row.get(active_col, ''))
                            break
                    
                    # Tìm dạng bào chế
                    for form_col in self.catalog_data.columns:
                        if any(keyword in form_col.lower() for keyword in ['dang_bao_che', 'dosage', 'form', 'bao_che']):
                            result['dosage_form'] = str(row.get(form_col, ''))
                            break
                    
                    # Tìm hàm lượng
                    for strength_col in self.catalog_data.columns:
                        if any(keyword in strength_col.lower() for keyword in ['ham_luong', 'strength', 'nong_do', 'concentration']):
                            result['strength'] = str(row.get(strength_col, ''))
                            break
                    
                    # Tìm quy cách đóng gói
                    for pack_col in self.catalog_data.columns:
                        if any(keyword in pack_col.lower() for keyword in ['quy_cach', 'pack', 'size', 'dong_goi']):
                            result['pack_size'] = str(row.get(pack_col, ''))
                            break
                    
                    # Lưu thông tin khác
                    for col_name in self.catalog_data.columns:
                        if col_name not in [col] + [reg_col for reg_col in self.catalog_data.columns if any(keyword in reg_col.lower() for keyword in ['so_dang_ky', 'registration', 'dang_ky', 'number'])]:
                            result['other_info'][col_name] = str(row.get(col_name, ''))
                    
                    results.append(result)
            
            # Loại bỏ trùng lặp
            unique_results = []
            seen_names = set()
            for result in results:
                if result['name'] not in seen_names:
                    unique_results.append(result)
                    seen_names.add(result['name'])
            
            return unique_results[:10]  # Giới hạn 10 kết quả
            
        except Exception as e:
            raise Exception(f"Lỗi tìm kiếm thuốc: {str(e)}")
    
    def get_catalog_info(self) -> dict:
        """Lấy thông tin về danh mục hiện tại"""
        if self.catalog_data is None:
            return {
                'loaded': False,
                'file_path': None,
                'total_records': 0,
                'columns': []
            }
        
        return {
            'loaded': True,
            'file_path': self.catalog_file_path,
            'total_records': len(self.catalog_data),
            'columns': list(self.catalog_data.columns)
        }
    
    def get_medicine_suggestions(self, query: str, limit: int = 10) -> list:
        """Lấy danh sách gợi ý thuốc nhanh cho autocomplete"""
        if self.catalog_data is None or not query.strip():
            return []
        
        try:
            query_lower = query.lower().strip()
            suggestions = []
            
            # Tìm kiếm trong các cột có thể chứa tên thuốc
            search_columns = []
            for col in self.catalog_data.columns:
                if any(keyword in col.lower() for keyword in ['ten', 'name', 'thuoc', 'medicine', 'san_pham', 'product']):
                    search_columns.append(col)
            
            if not search_columns:
                search_columns = [self.catalog_data.columns[0]]
            
            # Tìm kiếm và tạo gợi ý
            for col in search_columns:
                mask = self.catalog_data[col].astype(str).str.lower().str.contains(query_lower, na=False)
                matches = self.catalog_data[mask].head(limit)
                
                for _, row in matches.iterrows():
                    name = str(row.get(col, ''))
                    if name and name.lower() not in [s['name'].lower() for s in suggestions]:
                        # Tìm số đăng ký
                        reg_number = ''
                        for reg_col in self.catalog_data.columns:
                            if any(keyword in reg_col.lower() for keyword in ['so_dang_ky', 'registration', 'dang_ky', 'number']):
                                reg_number = str(row.get(reg_col, ''))
                                break
                        
                        suggestions.append({
                            'name': name,
                            'registration_number': reg_number,
                            'display_text': f"{name} - {reg_number}" if reg_number else name
                        })
            
            return suggestions[:limit]
            
        except Exception as e:
            return []
    
    def clear_catalog(self):
        """Xóa danh mục hiện tại"""
        self.catalog_data = None
        self.catalog_file_path = None

# ---------------- UI ----------------
class App(tb.Window):
    def __init__(self):
        super().__init__(themename='flatly')  # Tông xanh dương nhạt thanh lịch và phẳng
        self.title(f'{APP_NAME} — v{APP_VERSION}')
        self.geometry('1180x840'); self.minsize(1100, 740)

        self.db = DB(DB_PATH)
        self.backup_manager = BackupManager(DB_PATH, BACKUP_DIR)
        self.report_manager = ReportManager(DB_PATH)
        self.export_manager = ExportManager()
        self.medicine_catalog = MedicineCatalogManager(DB_PATH)
        self.last_sale_items = []; self.cart = []

        self.make_style()
        
        # Thiết lập Server kiểm kho di động
        self.mobile_server = None
        self.after(1000, self.start_mobile_server_bg)
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        
        # Tạo UI sau khi window đã sẵn sàng
        self.after(100, self.initialize_ui)
        
        # Bắt đầu auto backup
        self.backup_manager.start_auto_backup()
    
    def initialize_ui(self):
        """Khởi tạo UI sau khi window đã sẵn sàng"""
        try:
            self.make_ui()
            # Load dữ liệu sau khi UI đã được tạo
            self.after(200, self.on_ready)
        except Exception as e:
            print(f"Lỗi khởi tạo UI: {e}")
            # Fallback: tạo UI cơ bản
            self.create_basic_ui()
    
    def create_basic_ui(self):
        """Tạo UI cơ bản nếu có lỗi"""
        self.title(f'{APP_NAME} — v{APP_VERSION} — Lỗi khởi tạo')
        tb.Label(self, text="Có lỗi xảy ra khi khởi tạo giao diện. Vui lòng khởi động lại ứng dụng.", 
                font=('Segoe UI', 12)).pack(expand=True)

    def start_mobile_server_bg(self):
        """Khởi động Web Server chạy ngầm cho điện thoại di động kết nối"""
        try:
            self.mobile_server = MobileInventoryServer(self, host="0.0.0.0", port=5000)
            self.mobile_server.start()
        except Exception as e:
            print(f"Lỗi khởi động máy chủ di động: {e}")

    def on_close(self):
        """Đóng ứng dụng và giải phóng máy chủ di động"""
        if hasattr(self, 'mobile_server') and self.mobile_server:
            try:
                self.mobile_server.stop()
            except:
                pass
        self.destroy()

    # theme & fonts
    def make_style(self):
        style = tb.Style()
        # Ẩn header Notebook để tránh “trùng thẻ”
        style.layout('TNotebook.Tab', [])
        # Font to hơn
        style.configure('TLabel', font=('Segoe UI', 11))
        style.configure('TButton', font=('Segoe UI', 11))
        style.configure('TEntry',  font=('Segoe UI', 11))
        style.configure('TCombobox', font=('Segoe UI', 11))
        style.configure('Treeview', rowheight=30, font=('Segoe UI', 11))
        style.configure('Treeview.Heading', font=('Segoe UI', 11, 'bold'))

    # helpers
    def _numberize(self, entry: tb.Entry):
        entry.config(justify='right')
        entry.bind('<FocusIn>', lambda e: entry.selection_range(0, 'end'))
    def open_combo(self, combo):
        combo.focus_set()
        combo.event_generate('<Alt-Down>')
    def _open_dropdown(self, combo: tb.Combobox):
        combo.focus_set()
        # kích hoạt menu xổ xuống (Alt+Down)
        combo.event_generate('<Alt-Down>')

    def toast(self, text, ms=1600):
        w = tk.Toplevel(self); w.wm_overrideredirect(True); w.configure(bg='#111')
        tk.Label(w, text='  ' + text + '  ', bg='#111', fg='white', font=('Segoe UI', 10)).pack()
        w.update_idletasks()
        x = self.winfo_x() + self.winfo_width() - w.winfo_width() - 20
        y = self.winfo_y() + self.winfo_height() - w.winfo_height() - 20
        w.geometry(f'+{x}+{y}'); w.after(ms, w.destroy)

    # layout
    def make_ui(self):
        menubar = tk.Menu(self); self.config(menu=menubar)
        helpm = tk.Menu(menubar, tearoff=0); helpm.add_command(label='Phím tắt', command=self.show_shortcuts)
        helpm.add_separator()
        helpm.add_command(label='Mở thư mục dữ liệu', command=self.open_data_folder)
        helpm.add_command(label='Giới thiệu (About)…', command=self.show_about)
        menubar.add_cascade(label='Trợ giúp', menu=helpm)
        
        self.nb = tb.Notebook(self); self.nb.pack(fill=BOTH, expand=True, padx=8, pady=(0,8))
        
        # Tạo các tab frames
        self.tab_products = tb.Frame(self.nb)
        self.tab_purchase = tb.Frame(self.nb)
        self.tab_dispatch = tb.Frame(self.nb)
        self.tab_stock = tb.Frame(self.nb)
        self.tab_alerts = tb.Frame(self.nb)
        self.tab_report = tb.Frame(self.nb)
        self.tab_backup = tb.Frame(self.nb)
        self.tab_advanced_reports = tb.Frame(self.nb)
        self.tab_mobile = tb.Frame(self.nb)
        
        # Thêm tabs với labels đẹp hơn
        tabs_config = [
            (self.tab_products, "🏷️ Sản phẩm"),
            (self.tab_purchase, "📦 Nhập kho"),
            (self.tab_dispatch, "📤 Xuất kho / Cấp phát"),
            (self.tab_stock, "📊 Tồn kho"),
            (self.tab_alerts, "⏰ Hết hạn"),
            (self.tab_report, "📄 Báo cáo XNT"),
            (self.tab_backup, "💾 Backup"),
            (self.tab_advanced_reports, "📈 Báo cáo nâng cao"),
            (self.tab_mobile, "📱 Kiểm kho di động")
        ]
        
        for tab, label in tabs_config:
            self.nb.add(tab, text=label)
        
        # Tạo toolbar sau khi đã có các tabs
        self.create_toolbar()
        
        self.build_products_tab(); self.build_purchase_tab(); self.build_dispatch_tab()
        self.build_stock_tab(); self.build_alerts_tab(); self.build_report_tab()
        self.build_backup_tab(); self.build_advanced_reports_tab(); self.build_mobile_tab()
        # Status bar với thông tin chi tiết hơn
        status_frame = tb.Frame(self)
        status_frame.pack(fill='x', side='bottom', padx=8, pady=4)
        
        # Status chính
        self.status = tb.Label(status_frame, anchor='w', font=('Segoe UI', 9),
            text='Sẵn sàng • F1-F8: Chuyển tab • F9: In phiếu xuất kho • Ctrl+F: Tìm kiếm')
        self.status.pack(side='left')
        
        # Thông tin database
        self.db_status = tb.Label(status_frame, anchor='e', font=('Segoe UI', 9),
            text='Database: Đang kết nối...')
        self.db_status.pack(side='right')
        
        # Cập nhật status database
        self.update_db_status()
        
    def update_db_status(self):
        """Cập nhật trạng thái database"""
        try:
            # Đếm số sản phẩm
            product_count = self.db.conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]
            # Đếm số batch
            batch_count = self.db.conn.execute("SELECT COUNT(*) FROM batches").fetchone()[0]
            # Đếm số đơn hàng
            sale_count = self.db.conn.execute("SELECT COUNT(*) FROM sales").fetchone()[0]
            
            self.db_status.config(text=f"Database: {product_count} sản phẩm • {batch_count} lô • {sale_count} đơn hàng")
        except Exception as e:
            self.db_status.config(text=f"Database: Lỗi - {str(e)}")
        
        # Hotkeys
        self.bind('<F1>', lambda e: self.nb.select(self.tab_products))
        self.bind('<F2>', lambda e: self.nb.select(self.tab_purchase))
        self.bind('<F3>', lambda e: self.nb.select(self.tab_dispatch))
        self.bind('<F4>', lambda e: self.nb.select(self.tab_stock))
        self.bind('<F5>', lambda e: self.nb.select(self.tab_alerts))
        self.bind('<F6>', lambda e: self.nb.select(self.tab_report))
        self.bind('<F7>', lambda e: self.nb.select(self.tab_backup))
        self.bind('<F8>', lambda e: self.nb.select(self.tab_advanced_reports))
        self.bind('<F10>', lambda e: self.nb.select(self.tab_mobile))
        self.bind('<Control-f>', lambda e: self.focus_search())
        self.bind('<F9>', lambda e: self.print_dispatch_note())
        self.bind('<Control-Return>', lambda e: self.confirm_dispatch())

    def create_toolbar(self):
        tbbar = tb.Frame(self); tbbar.pack(fill='x', padx=8, pady=(8,8))
        
        # Tạo frame chứa các nút chính
        main_buttons = tb.Frame(tbbar)
        main_buttons.pack(side='left')
        
        # Các nút chính với style cải thiện
        buttons_config = [
            ('🏷️ Sản phẩm', 'info', self.tab_products, 'F1'),
            ('📦 Nhập hàng', 'info', self.tab_purchase, 'F2'),
            ('📤 Xuất kho', 'info', self.tab_dispatch, 'F3'),
            ('📊 Tồn kho', 'info', self.tab_stock, 'F4'),
            ('⏰ Hết hạn', 'info', self.tab_alerts, 'F5'),
            ('📄 Báo cáo XNT', 'info', self.tab_report, 'F6'),
            ('💾 Backup', 'info', self.tab_backup, 'F7'),
            ('📈 Báo cáo nâng cao', 'info', self.tab_advanced_reports, 'F8'),
            ('📱 Kiểm kho di động', 'info', self.tab_mobile, 'F10')
        ]
        
        for text, style, tab, shortcut in buttons_config:
            btn = tb.Button(main_buttons, text=text, bootstyle=style,
                          command=lambda t=tab: self.nb.select(t))
            btn.pack(side='left', padx=2)
            # Thêm tooltip với phím tắt
            self.create_tooltip(btn, f"{text} ({shortcut})")
        
        # Spacer
        tb.Label(tbbar, text='').pack(side='left', expand=True)
        
        # Nút in phiếu xuất kho
        print_btn = tb.Button(tbbar, text='🖨️ In phiếu xuất kho', bootstyle='outline-primary',
                             command=self.print_dispatch_note)
        print_btn.pack(side='right', padx=4)
        self.create_tooltip(print_btn, "In phiếu xuất kho (F9)")
        
        # Nút backup nhanh
        backup_btn = tb.Button(tbbar, text='💾 Backup nhanh', bootstyle='outline-success',
                              command=self.create_manual_backup)
        backup_btn.pack(side='right', padx=4)
        self.create_tooltip(backup_btn, "Tạo backup ngay lập tức")

    def create_tooltip(self, widget, text):
        """Tạo tooltip cho widget"""
        def show_tooltip(event):
            tooltip = tb.Toplevel()
            tooltip.wm_overrideredirect(True)
            tooltip.wm_geometry(f"+{event.x_root+10}+{event.y_root+10}")
            label = tb.Label(tooltip, text=text, background='#ffffe0', 
                           relief='solid', borderwidth=1, font=('Segoe UI', 9))
            label.pack()
            widget.tooltip = tooltip
        
        def hide_tooltip(event):
            if hasattr(widget, 'tooltip'):
                widget.tooltip.destroy()
                del widget.tooltip
        
        widget.bind('<Enter>', show_tooltip)
        widget.bind('<Leave>', hide_tooltip)

    def show_shortcuts(self):
        messagebox.showinfo('Phím tắt',
            'F1: Tạo sản phẩm\nF2: Nhập hàng (Ctrl+F tìm)\nF3: POS (Ctrl+F tìm, Enter=Thêm vào giỏ, Ctrl+Enter=Thanh toán, F9=In)\nF4: Tồn theo lô\nF5: Sắp hết hạn\nF6: Báo cáo tồn kho')

    def focus_search(self):
        idx = self.nb.index(self.nb.select())
        if idx == 1 and hasattr(self, 'search_purchase'): self.search_purchase.focus_set()
        elif idx == 2 and hasattr(self, 'search_pos'): self.search_pos.focus_set()
    def open_data_folder(self):
        import sys, subprocess, os
        path = APP_DIR  # đã có sẵn từ phần cấu hình đường dẫn
        try:
            if sys.platform.startswith('win'):
                os.startfile(path)  # type: ignore
            elif sys.platform == 'darwin':
                subprocess.call(['open', path])
            else:
                subprocess.call(['xdg-open', path])
        except Exception as e:
            messagebox.showerror('Lỗi', str(e))

    def show_about(self):
        info = (
            f"{APP_NAME} v{APP_VERSION}\n"
            f"Tác giả: {AUTHOR_NAME}\n"
            f"Điện thoại: {AUTHOR_PHONE}\n"
            f"Email: {AUTHOR_EMAIL}\n"
            f"Website: {AUTHOR_SITE}\n\n"
            f"Thư mục dữ liệu: {APP_DIR}"
        )
        if messagebox.askyesno("Giới thiệu", info + "\n\nMở website tác giả?"):
            try:
                import webbrowser
                webbrowser.open(AUTHOR_SITE)
            except Exception:
                pass

    # -------- Products --------
    def build_products_tab(self):
        frm = self.tab_products
        
        # Khung quản lý danh mục thuốc với style cải thiện
        catalog_frame = tb.Labelframe(frm, text='📚 Quản lý danh mục thuốc', bootstyle='info')
        catalog_frame.pack(fill='x', padx=8, pady=8)
        
        catalog_btn_frame = tb.Frame(catalog_frame)
        catalog_btn_frame.pack(fill='x', padx=8, pady=8)
        
        tb.Button(catalog_btn_frame, text='📁 Load danh mục CSV/Excel', bootstyle='info',
                  command=self.load_medicine_catalog).pack(side='left', padx=4)
        tb.Button(catalog_btn_frame, text='📄 Load thuoc.csv', bootstyle='primary',
                  command=self.load_default_csv).pack(side='left', padx=4)
        tb.Button(catalog_btn_frame, text='🔍 Tra cứu thuốc', bootstyle='success',
                  command=self.search_medicine_dialog).pack(side='left', padx=4)
        tb.Button(catalog_btn_frame, text='ℹ️ Thông tin danh mục', bootstyle='secondary',
                  command=self.show_catalog_info).pack(side='left', padx=4)
        
        # Hiển thị thông tin danh mục hiện tại
        self.catalog_info_label = tb.Label(catalog_frame, text='Chưa load danh mục thuốc', 
                                          font=('Segoe UI', 9), bootstyle='secondary')
        self.catalog_info_label.pack(anchor='w', padx=8, pady=(0,4))
        
        # Hướng dẫn sử dụng
        help_label = tb.Label(catalog_frame, 
                             text='💡 Gợi ý: Load danh mục thuốc, sau đó gõ tên thuốc để xem gợi ý tự động', 
                             font=('Segoe UI', 8), bootstyle='info')
        help_label.pack(anchor='w', padx=8, pady=(0,8))
        
        # Khung thông tin sản phẩm với style cải thiện
        f1 = tb.Labelframe(frm, text='📝 Thông tin sản phẩm', bootstyle='light')
        f1.pack(fill='x', padx=8, pady=8)
        
        # Hàng 0: Nguồn sản phẩm
        tb.Label(f1, text='Nguồn sản phẩm:').grid(row=0, column=0, sticky='w', padx=6, pady=6)
        self.product_source_var = tk.StringVar(value='catalog')
        
        source_frame = tb.Frame(f1)
        source_frame.grid(row=0, column=1, columnspan=3, sticky='w', padx=6, pady=6)
        
        self.r_catalog = tb.Radiobutton(source_frame, text='Từ danh mục chuẩn', 
                                         variable=self.product_source_var, value='catalog',
                                         command=self.on_product_source_change)
        self.r_catalog.pack(side='left', padx=(0, 15))
        
        self.r_free = tb.Radiobutton(source_frame, text='Ngoài danh mục (Nhập tự do)', 
                                      variable=self.product_source_var, value='free',
                                      command=self.on_product_source_change)
        self.r_free.pack(side='left')
        
        # Hàng 1: Tên sản phẩm và loại
        tb.Label(f1, text='Tên sản phẩm:').grid(row=1, column=0, sticky='w', padx=6, pady=6)
        self.p_name = tb.Entry(f1, width=35)
        self.p_name.grid(row=1, column=1, padx=6, pady=6)
        self.p_name.bind('<KeyRelease>', self.on_product_name_change)
        self.p_name.bind('<FocusOut>', self.on_product_name_focus_out)
        
        # Tạo frame chứa dropdown gợi ý
        self.suggestions_frame = tb.Frame(f1)
        self.suggestions_frame.grid(row=2, column=1, sticky='ew', padx=6, pady=(0,6))
        self.suggestions_frame.grid_remove()  # Ẩn ban đầu
        
        # Tạo Listbox cho gợi ý
        self.suggestions_listbox = tk.Listbox(self.suggestions_frame, height=6, width=35)
        self.suggestions_listbox.pack(fill='both', expand=True)
        self.suggestions_listbox.bind('<Double-Button-1>', self.on_suggestion_selected)
        self.suggestions_listbox.bind('<Button-1>', self.on_suggestion_click)
        self.suggestions_listbox.bind('<ButtonRelease-1>', self.on_suggestion_click)
        self.suggestions_listbox.bind('<Return>', self.on_suggestion_selected)
        self.suggestions_listbox.bind('<Escape>', self.hide_suggestions)
        
        # Bind keyboard navigation
        self.p_name.bind('<Down>', self.on_arrow_down)
        self.p_name.bind('<Up>', self.on_arrow_up)
        self.suggestions_listbox.bind('<Up>', self.on_suggestion_up)
        self.suggestions_listbox.bind('<Down>', self.on_suggestion_down)
        
        tb.Label(f1, text='Loại sản phẩm:').grid(row=1, column=2, sticky='w', padx=6, pady=6)
        self.p_type = tb.Combobox(f1, values=['thuoc', 'vaccine', 'vtyt', 'khac'], state='readonly', width=12)
        self.p_type.set('thuoc')
        self.p_type.grid(row=1, column=3, padx=6, pady=6)
        self.p_type.bind('<<ComboboxSelected>>', self.on_product_type_change)
        
        # Hàng 2: Đơn vị và Barcode
        tb.Label(f1, text='Đơn vị cơ sở:').grid(row=3, column=0, sticky='w', padx=6, pady=6)
        self.p_base = tb.Entry(f1, width=12)
        self.p_base.insert(0, 'vien')
        self.p_base.grid(row=3, column=1, padx=6, pady=6)
        
        tb.Label(f1, text='Barcode:').grid(row=3, column=2, sticky='w', padx=6, pady=6)
        barcode_frame = tb.Frame(f1)
        barcode_frame.grid(row=3, column=3, padx=6, pady=6, sticky='ew')
        
        self.p_barcode = tb.Entry(barcode_frame, width=16)
        self.p_barcode.pack(side='left')
        
        # Nút quét barcode cho sản phẩm mới
        if BARCODE_AVAILABLE:
            tb.Button(barcode_frame, text='📷', command=self.scan_barcode_for_product, 
                     bootstyle='info', width=3).pack(side='left', padx=(5, 0))
        else:
            tb.Button(barcode_frame, text='📷', command=self.show_barcode_install_info, 
                     bootstyle='secondary', width=3).pack(side='left', padx=(5, 0))
        
        # Hàng 3: Số đăng ký (chỉ hiển thị khi chọn loại thuốc)
        self.p_reg_label = tb.Label(f1, text='Số đăng ký:')
        self.p_reg_label.grid(row=4, column=0, sticky='w', padx=6, pady=6)
        self.p_reg_label.grid_remove()  # Ẩn ban đầu
        
        self.p_reg_number = tb.Entry(f1, width=35)
        self.p_reg_number.grid(row=4, column=1, columnspan=2, padx=6, pady=6)
        self.p_reg_number.grid_remove()  # Ẩn ban đầu
        
        # Nút lưu
        btns = tb.Frame(frm)
        btns.pack(fill='x', padx=8, pady=8)
        tb.Button(btns, text='💾 Lưu sản phẩm', bootstyle='primary', command=self.save_product).pack(side='right')
        tb.Button(btns, text='📋 Tải Excel mẫu', bootstyle='outline-info', command=self.export_import_template).pack(side='left', padx=4)
        tb.Button(btns, text='📥 Nhập hàng loạt (Excel)', bootstyle='success', command=self.bulk_import_from_excel).pack(side='left', padx=4)
        
        # Thiết lập ban đầu
        self.on_product_source_change()

    def _set_entry_val(self, entry, value):
        """Helper gán giá trị cho Entry bất kể trạng thái readonly"""
        try:
            state = entry.cget('state')
            entry.config(state='normal')
            entry.delete(0, tk.END)
            if value:
                entry.insert(0, value)
            entry.config(state=state)
        except Exception as e:
            print(f"Lỗi set value cho entry: {e}")

    def on_product_source_change(self):
        """Xử lý khi thay đổi nguồn sản phẩm (danh mục mẫu / tự do)"""
        source = self.product_source_var.get()
        if source == 'catalog':
            self.p_base.config(state='readonly')
            self.p_reg_number.config(state='readonly')
        else:
            self.p_base.config(state='normal')
            self.p_reg_number.config(state='normal')

    def save_product(self):
        name = self.p_name.get().strip()
        base = self.p_base.get().strip() or 'vien'
        bc = self.p_barcode.get().strip() or None
        product_type = self.p_type.get()
        reg_number = self.p_reg_number.get().strip() or None if product_type in ('thuoc', 'vaccine') else None
        
        if not name: 
            messagebox.showerror('Lỗi','Nhập tên sản phẩm')
            return
        
        pid = self.db.ex("INSERT INTO products(name, defaultUnit, barcode, productType, registrationNumber) VALUES(?,?,?,?,?)", 
                        (name, base, bc, product_type, reg_number))
        
        # bảo đảm product_units base tồn tại
        try:
            self.db.ex("INSERT INTO product_units(productId, unitCode, toBaseQty, price) VALUES(?,?,1,0)", (pid, base))
        except sqlite3.IntegrityError:
            pass
        
        self.toast(f'Đã tạo sản phẩm #{pid}')
        self.refresh_products()
        
        # Clear form
        self.p_name.delete(0, tk.END)
        self.p_barcode.delete(0, tk.END)
        self._set_entry_val(self.p_reg_number, '')
        self._set_entry_val(self.p_base, 'vien')
        self.p_type.set('thuoc')
        self.on_product_type_change()
        self.hide_suggestions()

    def on_product_type_change(self, event=None):
        """Xử lý khi thay đổi loại sản phẩm"""
        product_type = self.p_type.get()
        if product_type in ('thuoc', 'vaccine'):
            self.p_reg_label.grid()
            self.p_reg_number.grid()
        else:
            self.p_reg_label.grid_remove()
            self.p_reg_number.grid_remove()

    def on_product_name_change(self, event=None):
        """Xử lý khi thay đổi tên sản phẩm - hiển thị autocomplete"""
        if self.product_source_var.get() == 'free':
            self.hide_suggestions()
            return
            
        if self.medicine_catalog.catalog_data is None:
            return
        
        query = self.p_name.get().strip()
        
        if len(query) >= 2:  # Chỉ tìm kiếm khi có ít nhất 2 ký tự
            try:
                suggestions = self.medicine_catalog.get_medicine_suggestions(query, 8)
                
                if suggestions:
                    # Hiển thị listbox với gợi ý
                    self.suggestions_listbox.delete(0, tk.END)
                    for suggestion in suggestions:
                        self.suggestions_listbox.insert(tk.END, suggestion['display_text'])
                    
                    # Lưu suggestions để sử dụng sau
                    self.current_suggestions = suggestions
                    
                    # Hiển thị frame gợi ý
                    self.suggestions_frame.grid()
                else:
                    # Không có gợi ý, ẩn frame
                    self.hide_suggestions()
            except Exception as e:
                # Lỗi trong autocomplete, ẩn gợi ý
                self.hide_suggestions()
        else:
            # Ẩn gợi ý khi ít hơn 2 ký tự
            self.hide_suggestions()

    def hide_suggestions(self, event=None):
        """Ẩn danh sách gợi ý"""
        self.suggestions_frame.grid_remove()
        self.current_suggestions = []

    def on_suggestion_click(self, event=None):
        """Xử lý khi click vào gợi ý từ listbox"""
        try:
            print(f"DEBUG: Click vào listbox, event.y = {event.y}")
            # Lấy index của item được click
            index = self.suggestions_listbox.nearest(event.y)
            print(f"DEBUG: Index được chọn = {index}")
            
            self.suggestions_listbox.selection_clear(0, tk.END)
            self.suggestions_listbox.selection_set(index)
            self.suggestions_listbox.activate(index)
            
            # Chọn ngay lập tức
            self.on_suggestion_selected()
            
        except Exception as e:
            print(f"Lỗi khi click gợi ý: {e}")
            import traceback
            traceback.print_exc()

    def on_suggestion_selected(self, event=None):
        """Xử lý khi chọn gợi ý từ listbox"""
        try:
            selection = self.suggestions_listbox.curselection()
            if not selection:
                return
            
            selected_index = selection[0]
            if hasattr(self, 'current_suggestions') and selected_index < len(self.current_suggestions):
                medicine = self.current_suggestions[selected_index]
                
                print(f"DEBUG: Chọn thuốc: {medicine['name']}")
                
                # Điền thông tin vào form
                self.p_name.delete(0, tk.END)
                self.p_name.insert(0, medicine['name'])
                self.p_type.set('thuoc')
                self.on_product_type_change()
                
                # Điền số đăng ký & đơn vị cơ sở chuẩn
                self._set_entry_val(self.p_reg_number, medicine['registration_number'])
                self._set_entry_val(self.p_base, 'vien')
                
                # Ẩn gợi ý và chuyển focus
                self.hide_suggestions()
                self.p_barcode.focus_set()  # Chuyển focus sang barcode
                
        except Exception as e:
            print(f"Lỗi khi chọn gợi ý: {e}")
            import traceback
            traceback.print_exc()

    def on_product_name_focus_out(self, event=None):
        """Xử lý khi mất focus khỏi ô tên sản phẩm"""
        # Delay một chút để cho phép click vào listbox
        self.after(500, self.hide_suggestions)

    def on_arrow_down(self, event=None):
        """Xử lý phím mũi tên xuống"""
        if self.suggestions_frame.winfo_viewable():
            self.suggestions_listbox.focus_set()
            self.suggestions_listbox.selection_set(0)
            return "break"
        return None

    def on_arrow_up(self, event=None):
        """Xử lý phím mũi tên lên"""
        if self.suggestions_frame.winfo_viewable():
            self.suggestions_listbox.focus_set()
            self.suggestions_listbox.selection_set(tk.END)
            return "break"
        return None

    def on_suggestion_up(self, event=None):
        """Xử lý phím mũi tên lên trong listbox"""
        current = self.suggestions_listbox.curselection()
        if current and current[0] > 0:
            self.suggestions_listbox.selection_clear(current[0])
            self.suggestions_listbox.selection_set(current[0] - 1)
        return "break"

    def on_suggestion_down(self, event=None):
        """Xử lý phím mũi tên xuống trong listbox"""
        current = self.suggestions_listbox.curselection()
        if current and current[0] < self.suggestions_listbox.size() - 1:
            self.suggestions_listbox.selection_clear(current[0])
            self.suggestions_listbox.selection_set(current[0] + 1)
        return "break"

    def load_medicine_catalog(self):
        """Load danh mục thuốc từ file Excel hoặc CSV"""
        try:
            file_path = filedialog.askopenfilename(
                title="Chọn file danh mục thuốc",
                filetypes=[
                    ('CSV files', '*.csv'),
                    ('Excel files', '*.xlsx *.xls'),
                    ('All files', '*.*')
                ]
            )
            
            if file_path:
                self.medicine_catalog.load_catalog_from_excel(file_path)
                self.update_catalog_info()
                self.toast('Đã load danh mục thuốc thành công')
                
        except Exception as e:
            messagebox.showerror('Lỗi', str(e))

    def load_default_csv(self):
        """Load file thuoc.csv mặc định"""
        try:
            # Tìm file thuoc.csv trong thư mục hiện tại
            csv_path = os.path.join(os.getcwd(), 'thuoc.csv')
            
            if os.path.exists(csv_path):
                self.medicine_catalog.load_catalog_from_excel(csv_path)
                self.update_catalog_info()
                self.toast('Đã load file thuoc.csv thành công')
            else:
                # Nếu không tìm thấy, mở dialog chọn file
                messagebox.showinfo('Thông báo', 
                    f'Không tìm thấy file thuoc.csv trong thư mục:\n{csv_path}\n\nVui lòng chọn file khác.')
                self.load_medicine_catalog()
                
        except Exception as e:
            messagebox.showerror('Lỗi', str(e))

    def search_medicine_dialog(self):
        """Hiển thị dialog tra cứu thuốc"""
        if self.medicine_catalog.catalog_data is None:
            messagebox.showwarning('Cảnh báo', 'Vui lòng load danh mục thuốc trước')
            return
        
        # Tạo dialog tra cứu
        dialog = tb.Toplevel(self)
        dialog.title("Tra cứu thuốc")
        dialog.geometry("800x600")
        dialog.transient(self)
        dialog.grab_set()
        
        # Center dialog
        dialog.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() // 2) - (800 // 2)
        y = self.winfo_y() + (self.winfo_height() // 2) - (600 // 2)
        dialog.geometry(f"800x600+{x}+{y}")
        
        main_frame = tb.Frame(dialog)
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Title
        tb.Label(main_frame, text="TRA CỨU THUỐC TRONG DANH MỤC", 
                font=('Segoe UI', 14, 'bold'), bootstyle='primary').pack(pady=(0, 15))
        
        # Search frame
        search_frame = tb.Frame(main_frame)
        search_frame.pack(fill='x', pady=(0, 10))
        
        tb.Label(search_frame, text="Tên thuốc:").pack(side='left', padx=(0, 10))
        search_entry = tb.Entry(search_frame, width=40)
        search_entry.pack(side='left', padx=(0, 10))
        search_entry.focus_set()
        
        def search_medicines():
            try:
                query = search_entry.get().strip()
                if not query:
                    messagebox.showwarning('Cảnh báo', 'Nhập tên thuốc cần tìm')
                    return
                
                results = self.medicine_catalog.search_medicine(query)
                display_results(results)
                
            except Exception as e:
                messagebox.showerror('Lỗi', str(e))
        
        tb.Button(search_frame, text="🔍 Tìm kiếm", bootstyle='success',
                  command=search_medicines).pack(side='left')
        
        # Results frame
        results_frame = tb.Labelframe(main_frame, text="Kết quả tìm kiếm", bootstyle='light')
        results_frame.pack(fill='both', expand=True)
        
        # Results tree
        cols = ('name', 'reg_number', 'manufacturer', 'active_ingredient', 'dosage_form')
        results_tree = tb.Treeview(results_frame, columns=cols, show='headings', height=15)
        
        for c, w, t, anchor in [
            ('name', 200, 'Tên thuốc', 'w'),
            ('reg_number', 120, 'Số đăng ký', 'center'),
            ('manufacturer', 150, 'Nhà SX', 'w'),
            ('active_ingredient', 150, 'Hoạt chất', 'w'),
            ('dosage_form', 100, 'Dạng bào chế', 'w')
        ]:
            results_tree.heading(c, text=t)
            results_tree.column(c, width=w, anchor=anchor)
        
        results_tree.tag_configure('odd', background='#f6f8fa')
        results_tree.pack(fill='both', expand=True, padx=8, pady=8)
        
        def display_results(results):
            # Clear tree
            for item in results_tree.get_children():
                results_tree.delete(item)
            
            for idx, result in enumerate(results):
                results_tree.insert('', 'end', values=(
                    result['name'],
                    result['registration_number'],
                    result['manufacturer'],
                    result['active_ingredient'],
                    result['dosage_form']
                ), tags=('odd',) if idx % 2 else ())
        
        def select_medicine():
            selection = results_tree.selection()
            if not selection:
                messagebox.showwarning('Cảnh báo', 'Chọn thuốc từ danh sách')
                return
            
            item = results_tree.item(selection[0])
            medicine_name = item['values'][0]
            reg_number = item['values'][1]
            
            # Điền vào form sản phẩm
            self.p_name.delete(0, tk.END)
            self.p_name.insert(0, medicine_name)
            self.p_type.set('thuoc')
            self.on_product_type_change()
            self.p_reg_number.delete(0, tk.END)
            self.p_reg_number.insert(0, reg_number)
            
            dialog.destroy()
            self.nb.select(self.tab_products)
            self.p_base.focus_set()
        
        # Buttons
        btn_frame = tb.Frame(main_frame)
        btn_frame.pack(fill='x', pady=(10, 0))
        
        tb.Button(btn_frame, text="✅ Chọn thuốc này", bootstyle='success',
                  command=select_medicine).pack(side='left', padx=(0, 10))
        tb.Button(btn_frame, text="❌ Đóng", bootstyle='secondary',
                  command=dialog.destroy).pack(side='left')
        
        # Bind Enter key
        search_entry.bind('<Return>', lambda e: search_medicines())

    def show_catalog_info(self):
        """Hiển thị thông tin danh mục"""
        info = self.medicine_catalog.get_catalog_info()
        
        if not info['loaded']:
            messagebox.showinfo('Thông tin danh mục', 'Chưa load danh mục thuốc')
            return
        
        info_text = f"""Thông tin danh mục thuốc:

📁 File: {info['file_path']}
📊 Tổng số bản ghi: {info['total_records']:,}
📋 Các cột dữ liệu:
"""
        
        for i, col in enumerate(info['columns'], 1):
            info_text += f"  {i}. {col}\n"
        
        messagebox.showinfo('Thông tin danh mục', info_text)

    def update_catalog_info(self):
        """Cập nhật thông tin danh mục hiển thị"""
        info = self.medicine_catalog.get_catalog_info()
        
        if info['loaded']:
            file_name = os.path.basename(info['file_path'])
            self.catalog_info_label.config(
                text=f"📁 {file_name} - {info['total_records']:,} bản ghi",
                bootstyle='success'
            )
        else:
            self.catalog_info_label.config(
                text='Chưa load danh mục thuốc',
                bootstyle='secondary'
            )

    def scan_and_add(self):
        ok = self.fill_product_by_barcode(only_select=True)  # sửa hàm dưới để trả bool
        if ok:
            self.ent_qty_pos.delete(0, tk.END)
            self.ent_qty_pos.insert(0, '1')
            self.add_to_cart()
            self.ent_barcode.delete(0, tk.END)
        # luôn trả focus về ô barcode để quét tiếp
        self.after(50, lambda: self.ent_barcode.focus_set())
    
    def open_barcode_scanner(self):
        """Mở barcode scanner"""
        if not BARCODE_AVAILABLE:
            self.show_barcode_install_info()
            return
            
        try:
            # Tạo callback để xử lý kết quả quét
            def on_barcode_scanned(barcode_data):
                # Điền barcode vào ô input
                self.ent_barcode.delete(0, tk.END)
                self.ent_barcode.insert(0, barcode_data)
                
                # Tự động thêm vào giỏ hàng
                self.after(100, self.scan_and_add)
                
            # Tạo và mở scanner
            self.barcode_scanner = BarcodeScanner(self, callback=on_barcode_scanned)
            self.barcode_scanner.start_scan()
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể mở barcode scanner: {e}")
    
    def show_barcode_install_info(self):
        """Hiển thị thông tin cài đặt thư viện barcode"""
        info = """
📷 Tính năng quét barcode cần cài đặt thêm thư viện:

pip install opencv-python pyzbar Pillow

Sau khi cài đặt, khởi động lại phần mềm để sử dụng tính năng quét barcode bằng camera.

Hiện tại bạn vẫn có thể:
• Nhập barcode thủ công
• Tìm sản phẩm theo tên
        """
        messagebox.showinfo("Cài đặt thư viện barcode", info)
    
    def scan_barcode_for_product(self):
        """Quét barcode cho sản phẩm mới"""
        if not BARCODE_AVAILABLE:
            self.show_barcode_install_info()
            return
            
        try:
            # Tạo callback để xử lý kết quả quét
            def on_barcode_scanned(barcode_data):
                # Điền barcode vào ô input
                self.p_barcode.delete(0, tk.END)
                self.p_barcode.insert(0, barcode_data)
                
                # Chuyển focus sang ô tên sản phẩm
                self.p_name.focus_set()
                
            # Tạo và mở scanner
            self.barcode_scanner = BarcodeScanner(self, callback=on_barcode_scanned)
            self.barcode_scanner.start_scan()
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể mở barcode scanner: {e}")

    # -------- Purchase --------
    def build_purchase_tab(self):
        frm = self.tab_purchase
        self.cart_purchase = []
        self.last_purchase_items = []
        self.last_purchase_info = {}

        # Header với title đẹp
        header_frame = tb.Frame(frm)
        header_frame.pack(fill='x', padx=8, pady=(8,4))
        
        title_label = tb.Label(header_frame, text='📦 Nhập kho / Lập phiếu nhập', 
                              font=('Segoe UI', 14, 'bold'), bootstyle='success')
        title_label.pack(anchor='w')
        
        subtitle_label = tb.Label(header_frame, text='Lập phiếu nhập kho thuốc, vaccine và vật tư y tế', 
                                 font=('Segoe UI', 9), bootstyle='secondary')
        subtitle_label.pack(anchor='w')

        # --- Khung thông tin phiếu nhập
        info_note = tb.Labelframe(frm, text='📝 Thông tin phiếu nhập kho', bootstyle='light')
        info_note.pack(fill='x', padx=8, pady=8)

        # Nhà cung cấp / Nguồn cấp
        tb.Label(info_note, text='Nguồn cấp/Nhà CC:').grid(row=0, column=0, padx=6, pady=6, sticky='w')
        self.cmb_supplier = tb.Combobox(info_note, width=25)
        self.cmb_supplier.grid(row=0, column=1, padx=6, pady=6, sticky='w')
        self.refresh_suppliers_combo()

        # Ngày nhập
        tb.Label(info_note, text='Ngày nhập:').grid(row=0, column=2, padx=6, pady=6, sticky='w')
        self.ent_purchase_date = DateEntry(
            info_note,
            dateformat="%Y-%m-%d",
            firstweekday=0,
            bootstyle='info',
            width=12
        )
        self.ent_purchase_date.grid(row=0, column=3, padx=6, pady=6, sticky='w')

        # Lý do nhập
        tb.Label(info_note, text='Lý do nhập:').grid(row=0, column=4, padx=6, pady=6, sticky='w')
        self.cmb_purchase_reason = tb.Combobox(info_note, values=['Nhận cấp phát tuyến trên', 'Mua sắm đấu thầu', 'Viện trợ - Tài trợ', 'Khác'], width=22, state='readonly')
        self.cmb_purchase_reason.set('Nhận cấp phát tuyến trên')
        self.cmb_purchase_reason.grid(row=0, column=5, padx=6, pady=6, sticky='w')

        # Ghi chú
        tb.Label(info_note, text='Ghi chú:').grid(row=1, column=0, padx=6, pady=6, sticky='w')
        self.ent_purchase_note = tb.Entry(info_note)
        self.ent_purchase_note.grid(row=1, column=1, columnspan=5, padx=6, pady=6, sticky='ew')
        info_note.columnconfigure(1, weight=1)

        # --- Khung nhập sản phẩm
        box = tb.Labelframe(frm, text='➕ Thêm sản phẩm vào phiếu nhập', bootstyle='light')
        box.pack(fill='x', padx=8, pady=4)

        # Cho các cột có thể giãn đều khi thay đổi kích thước cửa sổ
        for i in range(12):
            box.grid_columnconfigure(i, weight=1)

        # ── Hàng 0: Tìm kiếm + Combobox chọn sản phẩm
        tb.Label(box, text='Tìm sản phẩm:').grid(row=0, column=0, sticky='w', padx=6, pady=6)
        self.search_purchase = tb.Entry(box)
        self.search_purchase.grid(row=0, column=1, columnspan=4, sticky='ew', padx=6, pady=6)
        self.search_purchase.bind('<KeyRelease>', lambda e: self.filter_product_list())
        self.search_purchase.bind('<Down>', lambda e: self.open_combo(self.cmb_prod)) 

        tb.Label(box, text='Chọn:').grid(row=0, column=5, sticky='e', padx=6, pady=6)
        self.cmb_prod = tb.Combobox(box, state='readonly')
        self.cmb_prod.grid(row=0, column=6, columnspan=5, sticky='ew', padx=6, pady=6)
        self.cmb_prod.bind('<<ComboboxSelected>>', lambda e: self.update_purchase_unit_and_price())
        self.cmb_prod.bind('<Escape>', lambda e: self.search_purchase.focus_set())  # ESC quay lại ô tìm
        self.cmb_prod.bind('<Return>', lambda e: self.ent_qty.focus_set())

        # ── Hàng 1: Đơn vị, Số lượng, Số lô, HSD
        tb.Label(box, text='Đơn vị tính:').grid(row=1, column=0, sticky='w', padx=6, pady=6)
        self.lbl_unit_purchase = tb.Label(box, text='-')
        self.lbl_unit_purchase.grid(row=1, column=1, sticky='w', padx=6, pady=6)

        tb.Label(box, text='Số lượng:').grid(row=1, column=2, sticky='e', padx=6, pady=6)
        self.ent_qty = tb.Entry(box, width=10)
        self.ent_qty.insert(0, '1')
        self.ent_qty.grid(row=1, column=3, sticky='w', padx=6, pady=6)
        self._numberize(self.ent_qty)

        tb.Label(box, text='Số lô:').grid(row=1, column=4, sticky='e', padx=6, pady=6)
        self.ent_lot = tb.Entry(box, width=14)
        self.ent_lot.insert(0, 'LOT001')
        self.ent_lot.grid(row=1, column=5, sticky='w', padx=6, pady=6)

        tb.Label(box, text='HSD (YYYY-MM-DD):').grid(row=1, column=6, sticky='e', padx=6, pady=6)
        self.ent_exp = DateEntry(
            box,
            dateformat="%Y-%m-%d",
            firstweekday=0,     # Monday
            bootstyle='info'
        )
        self.ent_exp.grid(row=1, column=7, sticky='w', padx=6, pady=6)

        # ── Hàng 2: Đơn giá nhập
        tb.Label(box, text='Đơn giá nhập:').grid(row=2, column=0, sticky='w', padx=6, pady=(6,8))
        self.ent_cost = tb.Entry(box, width=12)
        self.ent_cost.insert(0, '0')
        self.ent_cost.grid(row=2, column=1, sticky='w', padx=6, pady=(6,8))
        self._numberize(self.ent_cost)

        # --- Nút tác vụ
        btns = tb.Frame(frm)
        btns.pack(fill='x', padx=8, pady=8)
        tb.Button(btns, text='+ Thêm vào danh sách nhập', bootstyle='secondary', command=self.add_to_purchase_cart).pack(side='left', padx=4)
        tb.Button(btns, text='Xóa dòng', bootstyle='warning', command=self.remove_selected_purchase_item).pack(side='left', padx=4)
        tb.Button(btns, text='Xóa danh sách', bootstyle='danger', command=self.clear_purchase_cart).pack(side='left', padx=4)
        tb.Button(btns, text='Xác nhận nhập kho', bootstyle='success', command=self.confirm_purchase).pack(side='left', padx=8)
        tb.Button(btns, text='In phiếu nhập kho', bootstyle='info', command=self.print_purchase_note).pack(side='left', padx=8)

        # --- Bảng danh sách hàng nhập tạm thời
        cols = ('product','productName','unit','qty','lot','exp','cost','total')
        self.tree_purchase_cart = tb.Treeview(frm, columns=cols, show='headings')
        for c, w, t, anchor in [
            ('product',70,'PID','center'),('productName',300,'Tên thuốc/vaccine/VTYT','w'),
            ('unit',80,'ĐVT','center'),('qty',100,'SL','e'),('lot',120,'Số lô','w'),
            ('exp',100,'HSD','center'),('cost',120,'Đơn giá nhập','e'),('total',130,'Thành tiền','e')
        ]:
            self.tree_purchase_cart.heading(c, text=t, command=(lambda col=c: self.sort_tree(self.tree_purchase_cart, col)))
            self.tree_purchase_cart.column(c, width=w, anchor=anchor)
        self.tree_purchase_cart.tag_configure('odd', background='#f6f8fa')
        self.tree_purchase_cart.pack(fill='both', expand=True, padx=8, pady=8)

    def update_purchase_unit_and_price(self):
        sel = self.cmb_prod.get()
        if not sel: self.lbl_unit_purchase.config(text='-'); return
        pid = int(sel.split(' — ')[0])
        du = self.db.default_unit_of(pid) or '-'
        self.lbl_unit_purchase.config(text=du)

    def filter_product_list(self):
        kw = (self.search_purchase.get() or '').strip().lower()
        opts = [f"{p['id']} — {p['name']}" for p in self._products if kw in p['name'].lower()]
        self.cmb_prod['values'] = opts
        if opts:
            self.cmb_prod.current(0)
            self.update_purchase_unit_and_price()

    def refresh_suppliers_combo(self):
        try:
            suppliers = self.db.get_suppliers()
            self.cmb_supplier['values'] = suppliers
        except Exception as e:
            print(f"Lỗi tải nhà cung cấp: {e}")

    def add_to_purchase_cart(self):
        sel = self.cmb_prod.get()
        if not sel:
            messagebox.showerror('Lỗi', 'Chọn sản phẩm để nhập'); return
        pid = int(sel.split(' — ')[0])
        unit = self.db.default_unit_of(pid) or 'vien'
        try:
            qty = float(self.ent_qty.get())
        except:
            qty = 0
        if qty <= 0:
            messagebox.showerror('Lỗi', 'Số lượng nhập phải > 0'); return
        
        lot = self.ent_lot.get().strip()
        if not lot:
            messagebox.showerror('Lỗi', 'Vui lòng nhập số lô'); return
            
        exp = (self.ent_exp.entry.get() or '').strip()
        try:
            dt.datetime.strptime(exp, '%Y-%m-%d')
        except:
            messagebox.showerror('Lỗi', 'Hạn sử dụng không hợp lệ (YYYY-MM-DD)'); return
            
        try:
            cost = float((self.ent_cost.get() or '0').replace(',', ''))
        except:
            cost = 0.0
        if cost < 0:
            messagebox.showerror('Lỗi', 'Đơn giá nhập không được âm'); return

        name = self.name_by_id(pid)
        
        merged = False
        for it in self.cart_purchase:
            if it['productId'] == pid and it['lotNo'] == lot:
                it['qty'] = round(it['qty'] + qty, 4)
                it['cost'] = cost
                merged = True
                break
        if not merged:
            self.cart_purchase.append({
                'productId': pid,
                'productName': name,
                'unitCode': unit,
                'qty': qty,
                'lotNo': lot,
                'expiryDate': exp,
                'cost': cost
            })

        self.refresh_purchase_cart_view()
        self.ent_qty.delete(0, tk.END)
        self.ent_qty.insert(0, '1')
        self.search_purchase.focus_set()

    def remove_selected_purchase_item(self):
        sel = self.tree_purchase_cart.selection()
        if not sel:
            return
        idx = self.tree_purchase_cart.index(sel[0])
        if 0 <= idx < len(self.cart_purchase):
            self.cart_purchase.pop(idx)
            self.refresh_purchase_cart_view()

    def clear_purchase_cart(self):
        if self.cart_purchase and messagebox.askyesno('Xác nhận', 'Xóa toàn bộ danh sách hàng nhập kho?'):
            self.cart_purchase = []
            self.refresh_purchase_cart_view()

    def refresh_purchase_cart_view(self):
        for i in self.tree_purchase_cart.get_children():
            self.tree_purchase_cart.delete(i)
        for idx, it in enumerate(self.cart_purchase):
            total_val = it['qty'] * it['cost']
            self.tree_purchase_cart.insert('', 'end',
                values=(
                    it['productId'], 
                    it['productName'], 
                    it['unitCode'], 
                    f"{it['qty']:g}", 
                    it['lotNo'], 
                    it['expiryDate'], 
                    f"{it['cost']:,.0f}", 
                    f"{total_val:,.0f}"
                ),
                tags=('odd',) if idx % 2 else ())

    def confirm_purchase(self):
        if not self.cart_purchase:
            messagebox.showwarning('Chưa có dữ liệu', 'Danh sách nhập kho trống'); return
        
        supplier = self.cmb_supplier.get().strip()
        if not supplier:
            messagebox.showwarning('Thiếu thông tin', 'Vui lòng nhập/chọn Nguồn cấp/Nhà cung cấp'); return
        
        date_str = (self.ent_purchase_date.entry.get() or '').strip()
        if date_str:
            try:
                dt.datetime.strptime(date_str, '%Y-%m-%d')
            except:
                messagebox.showerror('Lỗi', 'Ngày nhập không hợp lệ (định dạng YYYY-MM-DD)'); return
                
        reason = self.cmb_purchase_reason.get().strip()
        note = self.ent_purchase_note.get().strip()

        if not messagebox.askyesno('Xác nhận', f'Bạn có chắc chắn muốn nhập kho từ:\nNguồn cấp: {supplier}\nNgày nhập: {date_str or "Hôm nay"}\nLý do: {reason}?'):
            return

        try:
            purchase_id, note_number, details = self.db.record_purchase(self.cart_purchase, supplier, reason, note, date_str)
            self.last_purchase_items = details
            
            created_at_str = f"{date_str} {dt.datetime.now().strftime('%H:%M:%S')}" if date_str else dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            self.last_purchase_info = {
                'id': purchase_id,
                'noteNumber': note_number,
                'supplier': supplier,
                'reason': reason,
                'note': note,
                'createdAt': created_at_str
            }
            self.cart_purchase = []
            self.refresh_purchase_cart_view()
            self.refresh_stock()
            self.refresh_suppliers_combo()
            messagebox.showinfo('Thành công', f'Đã nhập kho thành công!\nSố phiếu: {note_number}')
            
            # Tự động hỏi in phiếu nhập kho
            if messagebox.askyesno('In phiếu', 'Bạn có muốn in phiếu nhập kho ngay bây giờ?'):
                self.print_purchase_note()
                
        except Exception as e:
            messagebox.showerror('Lỗi', f'Lỗi nhập kho: {str(e)}')

    def print_purchase_note(self):
        if not self.last_purchase_items:
            messagebox.showwarning('Chưa có dữ liệu', 'Hãy thực hiện nhập kho trước khi in phiếu'); return
        
        info = self.last_purchase_info
        
        # Cho phép người dùng chọn vị trí lưu file PDF
        initial_filename = f"Phieu_Nhap_Kho_{info['noteNumber']}.pdf"
        pdf_path = filedialog.asksaveasfilename(
            title="Chọn vị trí lưu phiếu nhập kho PDF",
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            initialfile=initial_filename
        )
        if not pdf_path:
            return

        try:
            import reportlab
        except ImportError:
            response = messagebox.askyesno(
                "Thiếu thư viện", 
                "Hệ thống thiếu thư viện 'reportlab' để xuất PDF.\nBạn có muốn tự động cài đặt không? (Quá trình này mất khoảng vài giây)"
            )
            if response:
                import subprocess
                import sys
                try:
                    self.toast("Đang cài đặt thư viện reportlab, vui lòng đợi...")
                    subprocess.run([sys.executable, "-m", "pip", "install", "reportlab"], check=True)
                    self.toast("Đã cài đặt reportlab thành công!")
                except Exception as ex:
                    messagebox.showerror("Lỗi cài đặt", f"Không thể tự động cài đặt reportlab: {str(ex)}\nHãy chạy lệnh 'pip install reportlab' trong terminal."); return
            else:
                return

        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            
            try:
                pdfmetrics.registerFont(TTFont('TimesNewRoman', "C:\\Windows\\Fonts\\times.ttf"))
                pdfmetrics.registerFont(TTFont('TimesNewRoman-Bold', "C:\\Windows\\Fonts\\timesbd.ttf"))
                pdfmetrics.registerFont(TTFont('TimesNewRoman-Italic', "C:\\Windows\\Fonts\\timesi.ttf"))
                font_normal = 'TimesNewRoman'
                font_bold = 'TimesNewRoman-Bold'
                font_italic = 'TimesNewRoman-Italic'
            except Exception:
                font_normal = 'Helvetica'
                font_bold = 'Helvetica-Bold'
                font_italic = 'Helvetica-Oblique'
                
            doc = SimpleDocTemplate(pdf_path, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
            story = []
            
            styles = getSampleStyleSheet()
            
            style_header_left = ParagraphStyle(
                'HeaderLeft', parent=styles['Normal'], fontName=font_bold, fontSize=10, leading=14, alignment=0
            )
            style_header_right = ParagraphStyle(
                'HeaderRight', parent=styles['Normal'], fontName=font_normal, fontSize=10, leading=14, alignment=2
            )
            style_title = ParagraphStyle(
                'Title', parent=styles['Heading1'], fontName=font_bold, fontSize=16, leading=20, alignment=1, spaceAfter=5
            )
            style_subtitle = ParagraphStyle(
                'Subtitle', parent=styles['Normal'], fontName=font_bold, fontSize=11, leading=14, alignment=1, spaceAfter=15
            )
            style_info = ParagraphStyle(
                'Info', parent=styles['Normal'], fontName=font_normal, fontSize=11, leading=16, alignment=0
            )
            style_table_header = ParagraphStyle(
                'TableHeader', parent=styles['Normal'], fontName=font_bold, fontSize=9, leading=11, alignment=1, textColor=colors.black
            )
            style_cell = ParagraphStyle(
                'Cell', parent=styles['Normal'], fontName=font_normal, fontSize=9, leading=11, alignment=0
            )
            style_cell_center = ParagraphStyle(
                'CellCenter', parent=styles['Normal'], fontName=font_normal, fontSize=9, leading=11, alignment=1
            )
            style_cell_right = ParagraphStyle(
                'CellRight', parent=styles['Normal'], fontName=font_normal, fontSize=9, leading=11, alignment=2
            )
            
            # Header
            header_data = [
                [
                    Paragraph("SỞ Y TẾ THÀNH PHỐ CẦN THƠ<br/>TRUNG TÂM KIỂM SOÁT BỆNH TẬT (CDC)", style_header_left),
                    Paragraph("<b>Mẫu số: C30-HD</b><br/><i>(Ban hành theo Thông tư số 107/2017/TT-BTC)</i>", style_header_right)
                ]
            ]
            header_table = Table(header_data, colWidths=[280, 230])
            header_table.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
                ('BOTTOMPADDING', (0,0), (-1,-1), 10),
            ]))
            story.append(header_table)
            story.append(Spacer(1, 10))
            
            # Title
            story.append(Paragraph("PHIẾU NHẬP KHO", style_title))
            story.append(Paragraph(f"Số: {info['noteNumber']}", style_subtitle))
            
            # Parse Date
            created_str = info['createdAt']
            try:
                created_at_dt = dt.datetime.strptime(created_str, '%Y-%m-%d %H:%M:%S')
            except Exception:
                try:
                    created_at_dt = dt.datetime.strptime(created_str.split(' ')[0], '%Y-%m-%d')
                except Exception:
                    created_at_dt = dt.datetime.now()
            
            info_lines = [
                f"<b>Nguồn cấp / Nhà CC:</b> {info['supplier']}",
                f"<b>Lý do nhập:</b> {info['reason']}",
                f"<b>Kho nhập:</b> Kho Dược CDC Cần Thơ",
                f"<b>Ngày nhập:</b> {created_at_dt.strftime('%d/%m/%Y')}",
                f"<b>Ghi chú:</b> {info['note'] or 'Không'}"
            ]
            for line in info_lines:
                story.append(Paragraph(line, style_info))
                story.append(Spacer(1, 4))
                
            story.append(Spacer(1, 10))
            
            # Table items
            table_data = [
                [
                    Paragraph("STT", style_table_header),
                    Paragraph("Tên thuốc, vaccine, VTYT", style_table_header),
                    Paragraph("ĐVT", style_table_header),
                    Paragraph("Số lượng", style_table_header),
                    Paragraph("Đơn giá", style_table_header),
                    Paragraph("Thành tiền", style_table_header),
                    Paragraph("Số lô", style_table_header),
                    Paragraph("Hạn dùng", style_table_header)
                ]
            ]
            
            total_sum = 0.0
            for idx, it in enumerate(self.last_purchase_items, 1):
                qty = it['qty']
                cost = it['cost']
                sub_total = qty * cost
                total_sum += sub_total
                
                table_data.append([
                    Paragraph(str(idx), style_cell_center),
                    Paragraph(it['productName'], style_cell),
                    Paragraph(it['unitCode'], style_cell_center),
                    Paragraph(f"{qty:g}", style_cell_right),
                    Paragraph(f"{cost:,.0f}", style_cell_right),
                    Paragraph(f"{sub_total:,.0f}", style_cell_right),
                    Paragraph(it['lotNo'] or '', style_cell_center),
                    Paragraph(it['expiryDate'] or '', style_cell_center)
                ])
            
            # Thêm dòng tổng cộng
            table_data.append([
                Paragraph("<b>Tổng cộng</b>", style_cell_center),
                Paragraph("", style_cell),
                Paragraph("", style_cell_center),
                Paragraph("", style_cell_right),
                Paragraph("", style_cell_right),
                Paragraph(f"<b>{total_sum:,.0f}</b>", style_cell_right),
                Paragraph("", style_cell_center),
                Paragraph("", style_cell_center)
            ])
                
            col_widths = [25, 160, 45, 55, 65, 75, 55, 50]
            items_table = Table(table_data, colWidths=col_widths)
            items_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f2f2f2')),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.black),
                ('SPAN', (0, -1), (4, -1)),
                ('TOPPADDING', (0,0), (-1,-1), 5),
                ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ]))
            story.append(items_table)
            story.append(Spacer(1, 15))
            
            # Signatures
            date_right_style = ParagraphStyle(
                'DateRight', parent=styles['Normal'], fontName=font_italic, fontSize=11, alignment=2, spaceAfter=10
            )
            sig_title_style = ParagraphStyle(
                'SigTitle', parent=styles['Normal'], fontName=font_bold, fontSize=11, alignment=1
            )
            sig_sub_style = ParagraphStyle(
                'SigSub', parent=styles['Normal'], fontName=font_italic, fontSize=9, alignment=1
            )
            
            story.append(Paragraph(f"Cần Thơ, ngày {created_at_dt.strftime('%d')} tháng {created_at_dt.strftime('%m')} năm {created_at_dt.strftime('%Y')}", date_right_style))
            
            sig_headers = [
                [
                    Paragraph("<b>Người lập phiếu</b>", sig_title_style),
                    Paragraph("<b>Người giao hàng</b>", sig_title_style),
                    Paragraph("<b>Thủ kho</b>", sig_title_style),
                    Paragraph("<b>Kế toán trưởng</b>", sig_title_style),
                    Paragraph("<b>Lãnh đạo đơn vị</b>", sig_title_style)
                ],
                [
                    Paragraph("(Ký, họ tên)", sig_sub_style),
                    Paragraph("(Ký, họ tên)", sig_sub_style),
                    Paragraph("(Ký, họ tên)", sig_sub_style),
                    Paragraph("(Ký, họ tên)", sig_sub_style),
                    Paragraph("(Ký, đóng dấu)", sig_sub_style)
                ]
            ]
            sig_table = Table(sig_headers, colWidths=[102, 102, 102, 102, 102])
            sig_table.setStyle(TableStyle([
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('BOTTOMPADDING', (0,0), (-1,-1), 2),
            ]))
            story.append(sig_table)
            story.append(Spacer(1, 60))
            
            doc.build(story)
            os.startfile(pdf_path)
            self.toast("Đã in phiếu nhập ra PDF và mở file thành công")
            
        except Exception as e:
            messagebox.showerror("Lỗi in PDF", f"Không thể xuất file PDF: {str(e)}")

    # -------- Xuất kho / Cấp phát (Dispatch) --------
    def build_dispatch_tab(self):
        frm = self.tab_dispatch

        # Header với title đẹp
        header_frame = tb.Frame(frm)
        header_frame.pack(fill='x', padx=8, pady=(8,4))
        
        title_label = tb.Label(header_frame, text='📤 Xuất kho / Cấp phát', 
                              font=('Segoe UI', 14, 'bold'), bootstyle='warning')
        title_label.pack(anchor='w')
        
        subtitle_label = tb.Label(header_frame, text='Cấp phát thuốc, vaccine và vật tư y tế cho các đơn vị tuyến dưới theo nguyên tắc FEFO', 
                                 font=('Segoe UI', 9), bootstyle='secondary')
        subtitle_label.pack(anchor='w')

        # --- Khung thông tin phiếu xuất
        info_note = tb.Labelframe(frm, text='📝 Thông tin phiếu xuất kho', bootstyle='light')
        info_note.pack(fill='x', padx=8, pady=8)

        # Đơn vị nhận
        tb.Label(info_note, text='Đơn vị nhận:').grid(row=0, column=0, padx=6, pady=6, sticky='w')
        self.cmb_receiving_unit = tb.Combobox(info_note, width=25)
        self.cmb_receiving_unit.grid(row=0, column=1, padx=6, pady=6, sticky='w')
        self.refresh_receiving_units_combo()

        # Ngày xuất
        tb.Label(info_note, text='Ngày xuất:').grid(row=0, column=2, padx=6, pady=6, sticky='w')
        self.ent_dispatch_date = DateEntry(
            info_note,
            dateformat="%Y-%m-%d",
            firstweekday=0,
            bootstyle='info',
            width=12
        )
        self.ent_dispatch_date.grid(row=0, column=3, padx=6, pady=6, sticky='w')

        # Lý do xuất
        tb.Label(info_note, text='Lý do xuất:').grid(row=0, column=4, padx=6, pady=6, sticky='w')
        self.cmb_reason = tb.Combobox(info_note, values=['Cấp phát', 'Hủy kho', 'Chuyển kho', 'Khác'], width=15, state='readonly')
        self.cmb_reason.set('Cấp phát')
        self.cmb_reason.grid(row=0, column=5, padx=6, pady=6, sticky='w')

        # Ghi chú (đặt xuống hàng 1 để rộng rãi hơn)
        tb.Label(info_note, text='Ghi chú:').grid(row=1, column=0, padx=6, pady=6, sticky='w')
        self.ent_dispatch_note = tb.Entry(info_note)
        self.ent_dispatch_note.grid(row=1, column=1, columnspan=5, padx=6, pady=6, sticky='ew')
        info_note.columnconfigure(1, weight=1)

        # --- Hàng điều khiển chọn sản phẩm
        top = tb.Frame(frm)
        top.pack(fill='x', padx=8, pady=4)

        tb.Label(top, text='Barcode:').pack(side='left')
        self.ent_barcode = tb.Entry(top, width=18)
        self.ent_barcode.pack(side='left', padx=6)
        self.ent_barcode.bind('<Return>', lambda e: self.scan_and_add_dispatch())
        self.ent_barcode.bind('<KP_Enter>', lambda e: self.scan_and_add_dispatch())
        
        # Nút quét barcode bằng camera
        if BARCODE_AVAILABLE:
            tb.Button(top, text='📷 Quét', command=self.open_barcode_scanner_dispatch, 
                     bootstyle='info', width=8).pack(side='left', padx=6)
        else:
            tb.Button(top, text='📷 Quét', command=self.show_barcode_install_info, 
                     bootstyle='secondary', width=8).pack(side='left', padx=6)

        tb.Label(top, text='Tìm tên:').pack(side='left')
        self.search_pos = tb.Entry(top, width=30)
        self.search_pos.pack(side='left', padx=6)

        tb.Label(top, text='Chọn:').pack(side='left')
        self.cmb_prod_pos = tb.Combobox(top, state='readonly', width=45)
        self.cmb_prod_pos.pack(side='left', padx=6)

        tb.Label(top, text='SL xuất:').pack(side='left')
        self.ent_qty_pos = tb.Entry(top, width=8)
        self.ent_qty_pos.insert(0, '1')
        self.ent_qty_pos.pack(side='left', padx=6)
        self._numberize(self.ent_qty_pos)

        # --- Bind sự kiện
        self.search_pos.bind('<KeyRelease>', lambda e: self.filter_product_list_dispatch())
        self.search_pos.bind('<Down>', lambda e: (self.cmb_prod_pos.focus_set(),
                                                  self.cmb_prod_pos.event_generate('<Alt-Down>')))

        self.cmb_prod_pos.bind('<<ComboboxSelected>>', lambda e: self.update_dispatch_unit_label())
        self.cmb_prod_pos.bind('<Escape>', lambda e: self.search_pos.focus_set())
        self.cmb_prod_pos.bind('<Return>', lambda e: self.ent_qty_pos.focus_set())

        # --- Nút tác vụ
        btns = tb.Frame(frm)
        btns.pack(fill='x', padx=8, pady=8)
        tb.Button(btns, text='+ Thêm vào danh sách xuất', bootstyle='secondary', command=self.add_to_dispatch_cart).pack(side='left', padx=4)
        tb.Button(btns, text='Xóa dòng', bootstyle='warning', command=self.remove_selected_dispatch_item).pack(side='left', padx=4)
        tb.Button(btns, text='Xóa danh sách', bootstyle='danger', command=self.clear_dispatch_cart).pack(side='left', padx=4)
        tb.Button(btns, text='Xác nhận xuất kho', bootstyle='success', command=self.confirm_dispatch).pack(side='left', padx=8)
        tb.Button(btns, text='In phiếu xuất kho', bootstyle='info', command=self.print_dispatch_note).pack(side='left', padx=8)

        # --- Info tổng quan đơn vị
        info = tb.Frame(frm)
        info.pack(fill='x', padx=8, pady=(0, 4))
        self.lbl_unit_pos = tb.Label(info, text='Đơn vị tính: -', font=('Segoe UI', 10))
        self.lbl_unit_pos.pack(side='left', padx=(8, 12))
        
        # Bảng giỏ hàng xuất
        cols = ('productId', 'productName', 'unitCode', 'qty')
        self.tree_cart = tb.Treeview(frm, columns=cols, show='headings', height=10)
        for c, w, t, anchor in [
            ('productId', 70, 'PID', 'center'),
            ('productName', 450, 'Tên hàng hóa', 'w'),
            ('unitCode', 120, 'ĐVT', 'center'),
            ('qty', 150, 'Số lượng xuất', 'e')
        ]:
            self.tree_cart.heading(c, text=t, command=(lambda col=c: self.sort_tree(self.tree_cart, col)))
            self.tree_cart.column(c, width=w, anchor=anchor)
        self.tree_cart.tag_configure('odd', background='#f6f8fa')
        self.tree_cart.pack(fill='both', expand=True, padx=8, pady=8)

        self.ent_qty_pos.bind('<Return>', lambda e: self.add_to_dispatch_cart())
        self.cart_dispatch = []
        self.last_dispatch_items = []
        self.last_dispatch_info = {}

    def refresh_receiving_units_combo(self):
        """Cập nhật danh sách đơn vị nhận vào combobox"""
        try:
            units = self.db.get_receiving_units()
            self.cmb_receiving_unit['values'] = units
        except Exception as e:
            print(f"Lỗi refresh đơn vị nhận: {e}")

    def update_dispatch_unit_label(self):
        sel = self.cmb_prod_pos.get()
        if not sel:
            self.lbl_unit_pos.config(text='Đơn vị tính: -')
            return
        pid = int(sel.split(' — ')[0])
        du = self.db.default_unit_of(pid) or '-'
        self.lbl_unit_pos.config(text=f'Đơn vị tính: {du}')

    def fill_product_by_barcode_dispatch(self, only_select=False):
        bc = self.ent_barcode.get().strip()
        if not bc:
            return False
        row = self.db.q("SELECT id, name FROM products WHERE barcode=?", (bc,))
        if not row:
            if not only_select:
                messagebox.showwarning('Không tìm thấy', 'Barcode không khớp sản phẩm nào')
            return False
        target = f"{row[0]['id']} — {row[0]['name']}"
        if target not in self.cmb_prod_pos['values']:
            self.cmb_prod_pos['values'] = list(self.cmb_prod_pos['values']) + [target]
        self.cmb_prod_pos.set(target)
        self.update_dispatch_unit_label()
        return True

    def scan_and_add_dispatch(self):
        ok = self.fill_product_by_barcode_dispatch(only_select=True)
        if ok:
            self.ent_qty_pos.delete(0, tk.END)
            self.ent_qty_pos.insert(0, '1')
            self.add_to_dispatch_cart()
            self.ent_barcode.delete(0, tk.END)
        self.after(50, lambda: self.ent_barcode.focus_set())

    def open_barcode_scanner_dispatch(self):
        if not BARCODE_AVAILABLE:
            self.show_barcode_install_info()
            return
        try:
            def on_barcode_scanned(barcode_data):
                self.ent_barcode.delete(0, tk.END)
                self.ent_barcode.insert(0, barcode_data)
                self.after(100, self.scan_and_add_dispatch)
            self.barcode_scanner = BarcodeScanner(self, callback=on_barcode_scanned)
            self.barcode_scanner.start_scan()
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể mở barcode scanner: {e}")

    def filter_product_list_dispatch(self):
        kw = (self.search_pos.get() or '').strip().lower()
        opts = [f"{p['id']} — {p['name']}" for p in self._products if kw in p['name'].lower()]
        self.cmb_prod_pos['values'] = opts
        if opts:
            self.cmb_prod_pos.current(0)
            self.update_dispatch_unit_label()

    def add_to_dispatch_cart(self):
        sel = self.cmb_prod_pos.get()
        if not sel:
            messagebox.showerror('Lỗi', 'Chọn sản phẩm để xuất'); return
        pid = int(sel.split(' — ')[0])
        unit = self.db.default_unit_of(pid) or 'vien'
        try:
            qty = float(self.ent_qty_pos.get())
        except:
            qty = 0
        if qty <= 0:
            messagebox.showerror('Lỗi', 'Số lượng xuất phải > 0'); return
        
        name = self.name_by_id(pid)
        
        merged = False
        for it in self.cart_dispatch:
            if it['productId'] == pid and it['unitCode'] == unit:
                it['qty'] = round(it['qty'] + qty, 4)
                merged = True
                break
        if not merged:
            self.cart_dispatch.append({
                'productId': pid,
                'productName': name,
                'unitCode': unit,
                'qty': qty
            })

        self.refresh_dispatch_cart_view()
        self.ent_qty_pos.delete(0, tk.END)
        self.ent_qty_pos.insert(0, '1')
        self.ent_qty_pos.focus_set()

    def remove_selected_dispatch_item(self):
        sel = self.tree_cart.selection()
        if not sel:
            return
        idx = self.tree_cart.index(sel[0])
        if 0 <= idx < len(self.cart_dispatch):
            self.cart_dispatch.pop(idx)
            self.refresh_dispatch_cart_view()

    def clear_dispatch_cart(self):
        if self.cart_dispatch and messagebox.askyesno('Xác nhận', 'Xóa toàn bộ danh sách xuất kho?'):
            self.cart_dispatch = []
            self.refresh_dispatch_cart_view()

    def refresh_dispatch_cart_view(self):
        for i in self.tree_cart.get_children():
            self.tree_cart.delete(i)
        for idx, it in enumerate(self.cart_dispatch):
            self.tree_cart.insert('', 'end',
                values=(it['productId'], it['productName'], it['unitCode'], f"{it['qty']:g}"),
                tags=('odd',) if idx % 2 else ())

    def confirm_dispatch(self):
        if not self.cart_dispatch:
            messagebox.showwarning('Chưa có dữ liệu', 'Danh sách xuất kho trống'); return
        
        receiving_unit = self.cmb_receiving_unit.get().strip()
        if not receiving_unit:
            messagebox.showwarning('Thiếu thông tin', 'Vui lòng nhập Đơn vị nhận'); return
        
        date_str = (self.ent_dispatch_date.entry.get() or '').strip()
        if date_str:
            try:
                dt.datetime.strptime(date_str, '%Y-%m-%d')
            except:
                messagebox.showerror('Lỗi', 'Ngày xuất không hợp lệ (định dạng YYYY-MM-DD)'); return
                
        reason = self.cmb_reason.get().strip()
        note = self.ent_dispatch_note.get().strip()

        if not messagebox.askyesno('Xác nhận', f'Bạn có chắc chắn muốn xuất kho cho:\nĐơn vị: {receiving_unit}\nNgày xuất: {date_str or "Hôm nay"}\nLý do: {reason}?'):
            return

        try:
            dispatch_id, note_number, details = self.db.dispatch(self.cart_dispatch, receiving_unit, reason, note, date_str)
            self.last_dispatch_items = details
            
            created_at_str = f"{date_str} {dt.datetime.now().strftime('%H:%M:%S')}" if date_str else dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            self.last_dispatch_info = {
                'id': dispatch_id,
                'noteNumber': note_number,
                'receivingUnit': receiving_unit,
                'reason': reason,
                'note': note,
                'createdAt': created_at_str
            }
            self.cart_dispatch = []
            self.refresh_dispatch_cart_view()
            self.refresh_stock()
            self.refresh_receiving_units_combo()
            messagebox.showinfo('Thành công', f'Đã xuất kho thành công!\nSố phiếu: {note_number}')
            
            # Tự động hỏi in phiếu xuất kho
            if messagebox.askyesno('In phiếu', 'Bạn có muốn in phiếu xuất kho ngay bây giờ?'):
                self.print_dispatch_note()
                
        except Exception as e:
            messagebox.showerror('Lỗi', f'Lỗi xuất kho: {str(e)}')

    def name_by_id(self, pid):
        for p in self._products:
            if p['id'] == pid:
                return p['name']
        return f'#{pid}'

    def print_dispatch_note(self):
        if not self.last_dispatch_items:
            messagebox.showwarning('Chưa có dữ liệu', 'Hãy thực hiện xuất kho trước khi in phiếu'); return
        
        # Kiểm tra và tự động cài đặt reportlab nếu thiếu
        try:
            import reportlab
        except ImportError:
            response = messagebox.askyesno(
                "Thiếu thư viện", 
                "Hệ thống thiếu thư viện 'reportlab' để xuất PDF.\nBạn có muốn tự động cài đặt không? (Quá trình này mất khoảng vài giây)"
            )
            if response:
                import subprocess
                import sys
                try:
                    self.toast("Đang cài đặt thư viện reportlab, vui lòng đợi...")
                    # Chạy lệnh pip install reportlab bằng python hiện tại
                    subprocess.run([sys.executable, "-m", "pip", "install", "reportlab"], check=True)
                    self.toast("Đã cài đặt reportlab thành công!")
                except Exception as ex:
                    messagebox.showerror("Lỗi cài đặt", f"Không thể tự động cài đặt reportlab: {str(ex)}\nHãy chạy lệnh 'pip install reportlab' trong terminal."); return
            else:
                return
        
        info = self.last_dispatch_info
        
        # Cho phép người dùng chọn vị trí lưu file PDF
        initial_filename = f"Phieu_Xuat_Kho_{info['noteNumber']}.pdf"
        pdf_path = filedialog.asksaveasfilename(
            title="Chọn vị trí lưu phiếu xuất kho PDF",
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            initialfile=initial_filename
        )
        if not pdf_path:
            return  # Người dùng hủy bỏ lưu file

        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            
            # Đăng ký font Times New Roman trên Windows hỗ trợ tiếng Việt
            try:
                pdfmetrics.registerFont(TTFont('TimesNewRoman', "C:\\Windows\\Fonts\\times.ttf"))
                pdfmetrics.registerFont(TTFont('TimesNewRoman-Bold', "C:\\Windows\\Fonts\\timesbd.ttf"))
                pdfmetrics.registerFont(TTFont('TimesNewRoman-Italic', "C:\\Windows\\Fonts\\timesi.ttf"))
                font_normal = 'TimesNewRoman'
                font_bold = 'TimesNewRoman-Bold'
                font_italic = 'TimesNewRoman-Italic'
            except Exception:
                font_normal = 'Helvetica'
                font_bold = 'Helvetica-Bold'
                font_italic = 'Helvetica-Oblique'
                
            doc = SimpleDocTemplate(pdf_path, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
            story = []
            
            # Styles
            styles = getSampleStyleSheet()
            
            style_header_left = ParagraphStyle(
                'HeaderLeft',
                parent=styles['Normal'],
                fontName=font_bold,
                fontSize=10,
                leading=14,
                alignment=0 # Left
            )
            
            style_header_right = ParagraphStyle(
                'HeaderRight',
                parent=styles['Normal'],
                fontName=font_normal,
                fontSize=10,
                leading=14,
                alignment=2 # Right
            )
            
            style_title = ParagraphStyle(
                'Title',
                parent=styles['Heading1'],
                fontName=font_bold,
                fontSize=16,
                leading=20,
                alignment=1, # Center
                spaceAfter=5
            )
            
            style_subtitle = ParagraphStyle(
                'Subtitle',
                parent=styles['Normal'],
                fontName=font_bold,
                fontSize=11,
                leading=14,
                alignment=1, # Center
                spaceAfter=15
            )
            
            style_info = ParagraphStyle(
                'Info',
                parent=styles['Normal'],
                fontName=font_normal,
                fontSize=11,
                leading=16,
                alignment=0
            )
            
            style_table_header = ParagraphStyle(
                'TableHeader',
                parent=styles['Normal'],
                fontName=font_bold,
                fontSize=10,
                leading=12,
                alignment=1, # Center
                textColor=colors.black
            )
            
            style_cell = ParagraphStyle(
                'Cell',
                parent=styles['Normal'],
                fontName=font_normal,
                fontSize=10,
                leading=12,
                alignment=0
            )
            
            style_cell_center = ParagraphStyle(
                'CellCenter',
                parent=styles['Normal'],
                fontName=font_normal,
                fontSize=10,
                leading=12,
                alignment=1
            )
            
            style_cell_right = ParagraphStyle(
                'CellRight',
                parent=styles['Normal'],
                fontName=font_normal,
                fontSize=10,
                leading=12,
                alignment=2
            )
            
            # Header
            header_data = [
                [
                    Paragraph("SỞ Y TẾ THÀNH PHỐ CẦN THƠ<br/>TRUNG TÂM KIỂM SOÁT BỆNH TẬT (CDC)", style_header_left),
                    Paragraph("<b>Mẫu số: C31-HD</b><br/><i>(Ban hành theo Thông tư số 107/2017/TT-BTC)</i>", style_header_right)
                ]
            ]
            header_table = Table(header_data, colWidths=[280, 230])
            header_table.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
                ('BOTTOMPADDING', (0,0), (-1,-1), 10),
            ]))
            story.append(header_table)
            story.append(Spacer(1, 10))
            
            # Title
            story.append(Paragraph("PHIẾU XUẤT KHO", style_title))
            story.append(Paragraph(f"Số: {info['noteNumber']}", style_subtitle))
            
            # Parse createdAt
            created_str = info['createdAt']
            try:
                created_at_dt = dt.datetime.strptime(created_str, '%Y-%m-%d %H:%M:%S')
            except Exception:
                try:
                    created_at_dt = dt.datetime.strptime(created_str.split(' ')[0], '%Y-%m-%d')
                except Exception:
                    created_at_dt = dt.datetime.now()
            
            info_lines = [
                f"<b>Đơn vị nhận:</b> {info['receivingUnit']}",
                f"<b>Lý do xuất:</b> {info['reason']}",
                f"<b>Kho xuất:</b> Kho Dược CDC Cần Thơ",
                f"<b>Ngày xuất:</b> {created_at_dt.strftime('%d/%m/%Y')}",
                f"<b>Ghi chú:</b> {info['note'] or 'Không'}"
            ]
            for line in info_lines:
                story.append(Paragraph(line, style_info))
                story.append(Spacer(1, 4))
                
            story.append(Spacer(1, 10))
            
            # Table of items
            table_data = [
                [
                    Paragraph("STT", style_table_header),
                    Paragraph("Tên thuốc, vaccine, VTYT", style_table_header),
                    Paragraph("ĐVT", style_table_header),
                    Paragraph("Số lượng", style_table_header),
                    Paragraph("Số lô", style_table_header),
                    Paragraph("Hạn dùng", style_table_header),
                    Paragraph("Ghi chú", style_table_header)
                ]
            ]
            
            for idx, it in enumerate(self.last_dispatch_items, 1):
                table_data.append([
                    Paragraph(str(idx), style_cell_center),
                    Paragraph(it['productName'], style_cell),
                    Paragraph(it['unitCode'], style_cell_center),
                    Paragraph(f"{it['qty']:g}", style_cell_right),
                    Paragraph(it['lotNo'] or '', style_cell_center),
                    Paragraph(it['expiryDate'] or '', style_cell_center),
                    Paragraph('', style_cell)
                ])
                
            col_widths = [30, 190, 50, 70, 70, 70, 30]
            items_table = Table(table_data, colWidths=col_widths)
            items_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f2f2f2')),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.black),
                ('TOPPADDING', (0,0), (-1,-1), 6),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ]))
            story.append(items_table)
            story.append(Spacer(1, 15))
            
            # Date & Signatures
            date_right_style = ParagraphStyle(
                'DateRight',
                parent=styles['Normal'],
                fontName=font_italic,
                fontSize=11,
                alignment=2,
                spaceAfter=10
            )
            
            sig_title_style = ParagraphStyle(
                'SigTitle',
                parent=styles['Normal'],
                fontName=font_bold,
                fontSize=11,
                alignment=1
            )
            
            sig_sub_style = ParagraphStyle(
                'SigSub',
                parent=styles['Normal'],
                fontName=font_italic,
                fontSize=9,
                alignment=1
            )
            
            story.append(Paragraph(f"Cần Thơ, ngày {created_at_dt.strftime('%d')} tháng {created_at_dt.strftime('%m')} năm {created_at_dt.strftime('%Y')}", date_right_style))
            
            sig_headers = [
                [
                    Paragraph("<b>Người lập phiếu</b>", sig_title_style),
                    Paragraph("<b>Người nhận hàng</b>", sig_title_style),
                    Paragraph("<b>Thủ kho</b>", sig_title_style),
                    Paragraph("<b>Kế toán trưởng</b>", sig_title_style),
                    Paragraph("<b>Lãnh đạo đơn vị</b>", sig_title_style)
                ],
                [
                    Paragraph("(Ký, họ tên)", sig_sub_style),
                    Paragraph("(Ký, họ tên)", sig_sub_style),
                    Paragraph("(Ký, họ tên)", sig_sub_style),
                    Paragraph("(Ký, họ tên)", sig_sub_style),
                    Paragraph("(Ký, đóng dấu)", sig_sub_style)
                ]
            ]
            sig_table = Table(sig_headers, colWidths=[102, 102, 102, 102, 102])
            sig_table.setStyle(TableStyle([
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('BOTTOMPADDING', (0,0), (-1,-1), 2),
            ]))
            story.append(sig_table)
            story.append(Spacer(1, 60)) # Chừa khoảng trống để ký và ghi tên
            
            doc.build(story)
            
            # Mở file PDF kết quả
            os.startfile(pdf_path)
            self.toast("Đã in phiếu xuất ra PDF và mở file thành công")
            
        except Exception as e:
            messagebox.showerror("Lỗi in PDF", f"Không thể xuất file PDF: {str(e)}")

    # -------- Stock --------
    def build_stock_tab(self):
        frm = self.tab_stock
        cols = ('product','productName','batch','lot','exp','qty')
        self.tree_stock2 = tb.Treeview(frm, columns=cols, show='headings')
        for c, w, t, anchor in [
            ('product',70,'PID','center'),('productName',300,'Tên thuốc','w'),('batch',70,'Batch','center'),
            ('lot',130,'Lot','w'),('exp',100,'HSD','center'),('qty',120,'SL (base)','e')
        ]:
            self.tree_stock2.heading(c, text=t, command=(lambda col=c: self.sort_tree(self.tree_stock2, col)))
            self.tree_stock2.column(c, width=w, anchor=anchor)
        self.tree_stock2.tag_configure('odd', background='#f6f8fa')
        self.tree_stock2.pack(fill='both', expand=True, padx=8, pady=8)

    # -------- Alerts --------
    def build_alerts_tab(self):
        frm = self.tab_alerts
        top = tb.Frame(frm); top.pack(fill='x', padx=8, pady=8)
        tb.Label(top, text='Cảnh báo trong (ngày):').pack(side='left')
        self.ent_warn_days = tb.Entry(top, width=6); self.ent_warn_days.insert(0, '90'); self.ent_warn_days.pack(side='left', padx=6)
        self._numberize(self.ent_warn_days)
        tb.Button(top, text='Làm mới', bootstyle='secondary', command=self.refresh_alerts).pack(side='left', padx=8)

        cols = ('product','productName','batch','lot','exp','qty')
        self.tree_alerts = tb.Treeview(frm, columns=cols, show='headings')
        for c, w, t, anchor in [
            ('product',70,'PID','center'),('productName',300,'Tên thuốc','w'),('batch',70,'Batch','center'),
            ('lot',130,'Lot','w'),('exp',100,'HSD','center'),('qty',120,'SL (base)','e')
        ]:
            self.tree_alerts.heading(c, text=t, command=(lambda col=c: self.sort_tree(self.tree_alerts, col)))
            self.tree_alerts.column(c, width=w, anchor=anchor)
        self.tree_alerts.tag_configure('odd', background='#f6f8fa')
        self.tree_alerts.pack(fill='both', expand=True, padx=8, pady=8)

    # -------- Report --------
    def build_report_tab(self):
        frm = self.tab_report

        # Thanh điều kiện: từ ngày / đến ngày
        top = tb.Frame(frm); top.pack(fill='x', padx=8, pady=8)

        tb.Label(top, text='Từ ngày:').pack(side='left', padx=(0,6))
        # DateEntry đã được bạn thêm trước đó; nếu chưa có, nhớ: from ttkbootstrap.widgets import DateEntry
        self.de_from = DateEntry(top, dateformat="%Y-%m-%d", firstweekday=0, bootstyle='secondary')
        self.de_from.entry.delete(0, 'end')
        self.de_from.entry.insert(0, dt.date.today().replace(day=1).strftime("%Y-%m-%d"))  # đầu tháng
        self.de_from.pack(side='left', padx=(0,12))

        tb.Label(top, text='Đến ngày:').pack(side='left', padx=(0,6))
        self.de_to = DateEntry(top, dateformat="%Y-%m-%d", firstweekday=0, bootstyle='secondary')
        self.de_to.entry.delete(0, 'end')
        self.de_to.entry.insert(0, dt.datetime.now().strftime("%Y-%m-%d"))      # hôm nay
        self.de_to.pack(side='left', padx=(0,12))

        tb.Button(top, text='Làm mới', bootstyle='primary', command=self.refresh_report).pack(side='left', padx=6)
        tb.Button(top, text='Xuất CSV…', bootstyle='info', command=self.export_report_csv).pack(side='left', padx=6)
        tb.Button(top, text='Xuất PDF…', bootstyle='danger', command=self.export_report_pdf).pack(side='left', padx=6)
        tb.Button(top, text='Biên bản kiểm kê (PDF)', bootstyle='warning', command=self.print_inventory_check_pdf).pack(side='left', padx=6)

        # Bảng Xuất–Nhập–Tồn
        cols = ('product','productName','lotNo','expiryDate','opening','inbound','outbound','closing')
        self.tree_report = tb.Treeview(frm, columns=cols, show='headings')
        for c, w, t, anchor in [
            ('product',50,'PID','center'),
            ('productName',300,'Tên thuốc/vaccine/VTYT','w'),
            ('lotNo',100,'Số lô','center'),
            ('expiryDate',100,'Hạn sử dụng','center'),
            ('opening',100,'Tồn đầu kỳ','e'),
            ('inbound',100,'Nhập','e'),
            ('outbound',100,'Xuất','e'),
            ('closing',100,'Tồn cuối kỳ','e'),
        ]:
            self.tree_report.heading(c, text=t, command=(lambda col=c: self.sort_tree(self.tree_report, col)))
            self.tree_report.column(c, width=w, anchor=anchor)

        self.tree_report.tag_configure('odd', background='#f6f8fa')
        self.tree_report.tag_configure('total', background='#e8f5e9')  # dòng tổng
        self.tree_report.pack(fill='both', expand=True, padx=8, pady=8)

    # -------- Backup --------
    def build_backup_tab(self):
        frm = self.tab_backup
        
        # Khung tạo backup
        backup_frame = tb.Labelframe(frm, text='Tạo Backup', bootstyle='light')
        backup_frame.pack(fill='x', padx=8, pady=8)
        
        btn_frame = tb.Frame(backup_frame)
        btn_frame.pack(fill='x', padx=8, pady=8)
        
        tb.Button(btn_frame, text='💾 Tạo Backup Ngay', bootstyle='success',
                  command=self.create_manual_backup).pack(side='left', padx=4)
        tb.Button(btn_frame, text='📤 Export Dữ Liệu', bootstyle='info',
                  command=self.export_data).pack(side='left', padx=4)
        tb.Button(btn_frame, text='📥 Import Dữ Liệu', bootstyle='warning',
                  command=self.import_data).pack(side='left', padx=4)
        
        # Khung khôi phục backup
        restore_frame = tb.Labelframe(frm, text='Khôi Phục Backup', bootstyle='light')
        restore_frame.pack(fill='x', padx=8, pady=8)
        
        tb.Label(restore_frame, text='Chọn backup để khôi phục:').pack(anchor='w', padx=8, pady=(8,4))
        
        # Bảng danh sách backup
        cols = ('file', 'created', 'size', 'version')
        self.tree_backups = tb.Treeview(restore_frame, columns=cols, show='headings', height=8)
        for c, w, t, anchor in [
            ('file', 300, 'Tên File', 'w'),
            ('created', 150, 'Ngày Tạo', 'center'),
            ('size', 100, 'Kích Thước', 'e'),
            ('version', 80, 'Phiên Bản', 'center')
        ]:
            self.tree_backups.heading(c, text=t, command=(lambda col=c: self.sort_tree(self.tree_backups, col)))
            self.tree_backups.column(c, width=w, anchor=anchor)
        self.tree_backups.tag_configure('odd', background='#f6f8fa')
        self.tree_backups.pack(fill='x', padx=8, pady=8)
        
        # Nút khôi phục
        restore_btn_frame = tb.Frame(restore_frame)
        restore_btn_frame.pack(fill='x', padx=8, pady=(0,8))
        
        tb.Button(restore_btn_frame, text='🔄 Khôi Phục Backup', bootstyle='danger',
                  command=self.restore_selected_backup).pack(side='left', padx=4)
        tb.Button(restore_btn_frame, text='🗑️ Xóa Backup', bootstyle='outline-danger',
                  command=self.delete_selected_backup).pack(side='left', padx=4)
        tb.Button(restore_btn_frame, text='🔄 Làm Mới', bootstyle='secondary',
                  command=self.refresh_backup_list).pack(side='left', padx=4)
        
        # Thông tin auto backup
        info_frame = tb.Labelframe(frm, text='Tự Động Backup', bootstyle='info')
        info_frame.pack(fill='x', padx=8, pady=8)
        
        info_text = tb.Text(info_frame, height=4, wrap='word')
        info_text.pack(fill='x', padx=8, pady=8)
        info_text.insert('1.0', 
            "• Tự động backup mỗi ngày lúc 2:00 AM\n"
            "• Giữ tối đa 30 file backup\n"
            "• Backup được lưu trong thư mục: " + BACKUP_DIR + "\n"
            "• Trước khi khôi phục, hệ thống sẽ tự động tạo backup hiện tại")
        info_text.config(state='disabled')
        
        # Load danh sách backup
        self.refresh_backup_list()

    def create_manual_backup(self):
        """Tạo backup thủ công"""
        try:
            backup_path = self.backup_manager.create_backup("manual")
            self.toast(f'Đã tạo backup: {os.path.basename(backup_path)}')
            self.refresh_backup_list()
        except Exception as e:
            messagebox.showerror('Lỗi', str(e))

    def export_data(self):
        """Export dữ liệu ra file JSON"""
        try:
            path = filedialog.asksaveasfilename(
                defaultextension='.json',
                filetypes=[('JSON files', '*.json'), ('All files', '*.*')],
                initialfile=f'export_data_{dt.datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
            )
            if path:
                self.backup_manager.export_data(path)
                self.toast('Đã export dữ liệu thành công')
        except Exception as e:
            messagebox.showerror('Lỗi', str(e))

    def import_data(self):
        """Import dữ liệu từ file JSON"""
        try:
            path = filedialog.askopenfilename(
                filetypes=[('JSON files', '*.json'), ('All files', '*.*')]
            )
            if path:
                if messagebox.askyesno('Xác nhận', 
                    'Import sẽ thay thế toàn bộ dữ liệu hiện tại!\n'
                    'Hệ thống sẽ tự động tạo backup trước khi import.\n'
                    'Bạn có chắc chắn muốn tiếp tục?'):
                    
                    self.backup_manager.import_data(path)
                    self.toast('Đã import dữ liệu thành công')
                    # Refresh tất cả dữ liệu
                    self.refresh_products()
                    self.refresh_stock()
                    self.refresh_alerts()
                    self.refresh_report()
                    self.refresh_backup_list()
        except Exception as e:
            messagebox.showerror('Lỗi', str(e))

    def restore_selected_backup(self):
        """Khôi phục backup được chọn"""
        selection = self.tree_backups.selection()
        if not selection:
            messagebox.showwarning('Cảnh báo', 'Chọn backup để khôi phục')
            return
        
        item = self.tree_backups.item(selection[0])
        backup_file = item['values'][0]
        backup_path = os.path.join(BACKUP_DIR, backup_file)
        
        if messagebox.askyesno('Xác nhận', 
            f'Khôi phục backup: {backup_file}\n'
            'Hệ thống sẽ tự động tạo backup hiện tại trước khi khôi phục.\n'
            'Bạn có chắc chắn muốn tiếp tục?'):
            
            try:
                self.backup_manager.restore_backup(backup_path)
                self.toast('Đã khôi phục backup thành công')
                # Refresh tất cả dữ liệu
                self.refresh_products()
                self.refresh_stock()
                self.refresh_alerts()
                self.refresh_report()
                self.refresh_backup_list()
            except Exception as e:
                messagebox.showerror('Lỗi', str(e))

    def delete_selected_backup(self):
        """Xóa backup được chọn"""
        selection = self.tree_backups.selection()
        if not selection:
            messagebox.showwarning('Cảnh báo', 'Chọn backup để xóa')
            return
        
        item = self.tree_backups.item(selection[0])
        backup_file = item['values'][0]
        
        if messagebox.askyesno('Xác nhận', f'Xóa backup: {backup_file}?'):
            try:
                backup_path = os.path.join(BACKUP_DIR, backup_file)
                os.remove(backup_path)
                metadata_path = backup_path.replace('.db', '.json')
                if os.path.exists(metadata_path):
                    os.remove(metadata_path)
                self.toast('Đã xóa backup')
                self.refresh_backup_list()
            except Exception as e:
                messagebox.showerror('Lỗi', str(e))

    def refresh_backup_list(self):
        """Làm mới danh sách backup"""
        try:
            # Clear tree
            for item in self.tree_backups.get_children():
                self.tree_backups.delete(item)
            
            backups = self.backup_manager.list_backups()
            for idx, backup in enumerate(backups):
                size_mb = backup['size'] / (1024 * 1024)
                created_str = backup['created'].strftime('%Y-%m-%d %H:%M')
                version = backup.get('version', 'N/A')
                
                self.tree_backups.insert('', 'end',
                    values=(backup['file'], created_str, f'{size_mb:.1f} MB', version),
                    tags=('odd',) if idx % 2 else ())
        except Exception as e:
            messagebox.showerror('Lỗi', f'Không thể load danh sách backup: {str(e)}')

    # -------- Advanced Reports --------
    def build_advanced_reports_tab(self):
        frm = self.tab_advanced_reports
        
        # Khung điều khiển
        control_frame = tb.Labelframe(frm, text='Điều kiện báo cáo', bootstyle='light')
        control_frame.pack(fill='x', padx=8, pady=8)
        
        # Hàng 1: Ngày tháng
        date_frame = tb.Frame(control_frame)
        date_frame.pack(fill='x', padx=8, pady=8)
        
        tb.Label(date_frame, text='Từ ngày:').pack(side='left', padx=(0,6))
        self.adv_de_from = DateEntry(date_frame, dateformat="%Y-%m-%d", firstweekday=0, bootstyle='secondary')
        self.adv_de_from.entry.delete(0, 'end')
        self.adv_de_from.entry.insert(0, dt.date.today().replace(day=1).strftime("%Y-%m-%d"))
        self.adv_de_from.pack(side='left', padx=(0,12))
        
        tb.Label(date_frame, text='Đến ngày:').pack(side='left', padx=(0,6))
        self.adv_de_to = DateEntry(date_frame, dateformat="%Y-%m-%d", firstweekday=0, bootstyle='secondary')
        self.adv_de_to.entry.delete(0, 'end')
        self.adv_de_to.entry.insert(0, dt.datetime.now().strftime("%Y-%m-%d"))
        self.adv_de_to.pack(side='left', padx=(0,12))
        
        # Hàng 2: Nút báo cáo
        btn_frame = tb.Frame(control_frame)
        btn_frame.pack(fill='x', padx=8, pady=(0,8))
        
        tb.Button(btn_frame, text='📥 Lịch sử phiếu nhập', bootstyle='secondary',
                  command=self.show_purchase_history).pack(side='left', padx=4)
        tb.Button(btn_frame, text='📋 Lịch sử phiếu xuất', bootstyle='secondary',
                  command=self.show_dispatch_history).pack(side='left', padx=4)
        tb.Button(btn_frame, text='📤 Thống kê phiếu xuất', bootstyle='success',
                  command=self.show_revenue_report).pack(side='left', padx=4)
        tb.Button(btn_frame, text='🏢 Thống kê theo đơn vị nhận', bootstyle='info',
                  command=self.show_profit_report).pack(side='left', padx=4)
        tb.Button(btn_frame, text='🏆 Top sản phẩm cấp phát', bootstyle='warning',
                  command=self.show_top_products_report).pack(side='left', padx=4)
        tb.Button(btn_frame, text='📈 Biểu đồ cấp phát', bootstyle='primary',
                  command=self.show_revenue_chart).pack(side='left', padx=4)
        
        # Hàng 3: Nút xuất báo cáo
        export_frame = tb.Frame(control_frame)
        export_frame.pack(fill='x', padx=8, pady=(0,8))
        
        tb.Label(export_frame, text='Xuất báo cáo:', font=('Segoe UI', 9, 'bold')).pack(side='left', padx=(0,8))
        tb.Button(export_frame, text='📊 Excel', bootstyle='success',
                  command=self.export_current_report_excel).pack(side='left', padx=4)
        tb.Button(export_frame, text='📄 PDF', bootstyle='danger',
                  command=self.export_current_report_pdf).pack(side='left', padx=4)
        tb.Button(export_frame, text='📋 CSV', bootstyle='secondary',
                  command=self.export_current_report_csv).pack(side='left', padx=4)
        
        # Khung hiển thị báo cáo
        self.report_display_frame = tb.Frame(frm)
        self.report_display_frame.pack(fill='both', expand=True, padx=8, pady=8)
        
        # Tạo notebook cho các loại báo cáo
        self.adv_report_nb = tb.Notebook(self.report_display_frame)
        self.adv_report_nb.pack(fill='both', expand=True)
        
        # Tab tóm tắt
        self.adv_summary_tab = tb.Frame(self.adv_report_nb)
        self.adv_report_nb.add(self.adv_summary_tab, text='📋 Tóm tắt')
        
        # Tab báo cáo chi tiết
        self.adv_detail_tab = tb.Frame(self.adv_report_nb)
        self.adv_report_nb.add(self.adv_detail_tab, text='📊 Chi tiết')
        
        # Tab biểu đồ
        self.adv_chart_tab = tb.Frame(self.adv_report_nb)
        self.adv_report_nb.add(self.adv_chart_tab, text='📈 Biểu đồ')
        
        # Thống kê báo cáo hiện tại
        self.current_report_type = 'dispatch'
        self.current_report_group_by = 'day'
        
        # Load tóm tắt ban đầu
        self.load_advanced_summary()

    def load_advanced_summary(self):
        """Load tóm tắt báo cáo cấp phát"""
        try:
            # Clear summary tab
            for widget in self.adv_summary_tab.winfo_children():
                widget.destroy()
            
            start_date = self.adv_de_from.entry.get().strip()
            end_date = self.adv_de_to.entry.get().strip()
            
            if not start_date or not end_date:
                return
            
            # Lấy dữ liệu tóm tắt
            summary_data = self.report_manager.get_daily_sales_summary(start_date, end_date)
            summary = summary_data['summary']
            
            # Tạo layout tóm tắt
            main_frame = tb.Frame(self.adv_summary_tab)
            main_frame.pack(fill='both', expand=True, padx=20, pady=20)
            
            # Title
            title_label = tb.Label(main_frame, text="TÓM TẮT THÔNG TIN CẤP PHÁT KHO", 
                                  font=('Segoe UI', 16, 'bold'), bootstyle='primary')
            title_label.pack(pady=(0, 20))
            
            # KPI Cards
            kpi_frame = tb.Frame(main_frame)
            kpi_frame.pack(fill='x', pady=(0, 20))
            
            # Card 1: Tổng phiếu xuất
            card1 = tb.Labelframe(kpi_frame, text='Tổng số phiếu xuất', bootstyle='info')
            card1.pack(side='left', fill='both', expand=True, padx=(0, 10))
            total_orders = summary.get('total_orders', 0) or 0
            tb.Label(card1, text=f"{total_orders:,}", 
                    font=('Segoe UI', 24, 'bold'), bootstyle='info').pack(pady=10)
            
            # Card 2: Tổng số lượng xuất
            card2 = tb.Labelframe(kpi_frame, text='Tổng số lượng xuất', bootstyle='success')
            card2.pack(side='left', fill='both', expand=True, padx=(0, 10))
            total_revenue = summary.get('total_revenue', 0) or 0
            tb.Label(card2, text=f"{total_revenue:,.0f}", 
                    font=('Segoe UI', 24, 'bold'), bootstyle='success').pack(pady=10)
            
            # Card 3: Số lượng TB/phiếu
            card3 = tb.Labelframe(kpi_frame, text='SL trung bình/phiếu', bootstyle='warning')
            card3.pack(side='left', fill='both', expand=True, padx=(0, 10))
            avg_order = summary.get('avg_order_value', 0) or 0
            tb.Label(card3, text=f"{avg_order:,.1f}", 
                    font=('Segoe UI', 24, 'bold'), bootstyle='warning').pack(pady=10)
            
            # Card 4: Phiếu xuất lớn nhất
            card4 = tb.Labelframe(kpi_frame, text='Phiếu xuất lớn nhất', bootstyle='danger')
            card4.pack(side='left', fill='both', expand=True)
            max_order = summary.get('max_order', 0) or 0
            tb.Label(card4, text=f"{max_order:,.0f}", 
                    font=('Segoe UI', 24, 'bold'), bootstyle='danger').pack(pady=10)
            
            # Bảng cấp phát theo ngày
            daily_frame = tb.Labelframe(main_frame, text='Cấp phát theo ngày', bootstyle='light')
            daily_frame.pack(fill='both', expand=True)
            
            cols = ('date', 'orders', 'revenue')
            daily_tree = tb.Treeview(daily_frame, columns=cols, show='headings', height=8)
            for c, w, t, anchor in [
                ('date', 120, 'Ngày', 'center'),
                ('orders', 100, 'Số phiếu', 'e'),
                ('revenue', 150, 'Tổng số lượng xuất', 'e')
            ]:
                daily_tree.heading(c, text=t)
                daily_tree.column(c, width=w, anchor=anchor)
            
            daily_tree.tag_configure('odd', background='#f6f8fa')
            daily_tree.pack(fill='both', expand=True, padx=8, pady=8)
            
            # Load dữ liệu
            daily_data = summary_data.get('daily_data', [])
            for idx, row in enumerate(daily_data):
                orders = row.get('orders', 0) or 0
                revenue = row.get('revenue', 0) or 0
                daily_tree.insert('', 'end', values=(
                    row.get('sale_date', ''),
                    f"{orders:,}",
                    f"{revenue:,.0f}"
                ), tags=('odd',) if idx % 2 else ())
                
        except Exception as e:
            messagebox.showerror('Lỗi', f'Không thể load tóm tắt: {str(e)}')

    def build_mobile_tab(self):
        """Xây dựng giao diện cho Tab Kiểm kho di động"""
        # Xóa các widget cũ
        for widget in self.tab_mobile.winfo_children():
            widget.destroy()
            
        main_frame = tb.Frame(self.tab_mobile)
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        main_frame.columnconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(0, weight=1)
        
        # --- CỘT TRÁI: ĐIỀU KHIỂN & TRẠNG THÁI ---
        left_frame = tb.LabelFrame(main_frame, text="⚙️ Cấu hình máy chủ di động", padding=15)
        left_frame.grid(row=0, column=0, sticky='nsew', padx=(0, 10))
        
        # Trạng thái máy chủ
        status_lbl_frame = tb.Frame(left_frame)
        status_lbl_frame.pack(fill='x', pady=10)
        
        tb.Label(status_lbl_frame, text="Trạng thái máy chủ:", font=('Segoe UI', 11, 'bold')).pack(side='left')
        
        self.mobile_status_val = tb.Label(status_lbl_frame, font=('Segoe UI', 11, 'bold'))
        self.mobile_status_val.pack(side='left', padx=10)
        
        # Điều khiển nút
        self.btn_toggle_server = tb.Button(left_frame, command=self.toggle_mobile_server, bootstyle='success')
        self.btn_toggle_server.pack(fill='x', pady=10)
        
        # Đường link kết nối
        link_frame = tb.Frame(left_frame)
        link_frame.pack(fill='x', pady=10)
        
        tb.Label(link_frame, text="Địa chỉ kết nối LAN:", font=('Segoe UI', 10, 'bold')).pack(anchor='w')
        
        self.mobile_url_val = tb.Entry(link_frame, font=('Segoe UI', 10), state='readonly')
        self.mobile_url_val.pack(fill='x', pady=5)
        
        def copy_url():
            url = self.mobile_url_val.get()
            if url:
                self.clipboard_clear()
                self.clipboard_append(url)
                self.toast("Đã copy đường dẫn kết nối vào clipboard!")
                
        def open_browser():
            url = self.mobile_url_val.get()
            if url:
                import webbrowser
                webbrowser.open(url)
                
        btn_copy = tb.Button(link_frame, text="📋 Sao chép liên kết", command=copy_url, bootstyle='outline-info')
        btn_copy.pack(side='left', padx=2)
        
        btn_open = tb.Button(link_frame, text="🌐 Mở trên PC (Test)", command=open_browser, bootstyle='outline-secondary')
        btn_open.pack(side='left', padx=2)
        
        # Hướng dẫn chi tiết
        help_text = (
            "💡 Hướng dẫn sử dụng:\n"
            "1. Đảm bảo máy tính và điện thoại di động cùng kết nối chung một mạng Wi-Fi cục bộ.\n"
            "2. Khởi động máy chủ bằng nút phía trên (nếu đang dừng).\n"
            "3. Lấy điện thoại di động quét mã QR ở ô bên phải để mở liên kết.\n"
            "4. Cấp quyền truy cập Camera cho trình duyệt trên điện thoại nếu được hỏi.\n"
            "5. Đưa camera điện thoại quét mã vạch sản phẩm để kiểm tra số tồn tức thì."
        )
        tb.Label(left_frame, text=help_text, font=('Segoe UI', 9), justify='left', 
                 wraplength=450, bootstyle='secondary').pack(anchor='w', pady=15)
                 
        # --- CỘT PHẢI: MÃ QR ---
        right_frame = tb.LabelFrame(main_frame, text="📱 Quét mã QR để kết nối", padding=15)
        right_frame.grid(row=0, column=1, sticky='nsew', padx=(10, 0))
        
        # Canvas vẽ QR Code
        self.qr_canvas = tk.Canvas(right_frame, width=260, height=260, bg='white', relief='ridge', borderwidth=1)
        self.qr_canvas.pack(pady=10)
        
        self.qr_help_lbl = tb.Label(right_frame, text="Đang tạo mã QR kết nối...", font=('Segoe UI', 10), justify='center')
        self.qr_help_lbl.pack(pady=5)
        
        # Cập nhật UI ban đầu
        self.update_mobile_server_ui()

    def toggle_mobile_server(self):
        """Khởi động hoặc dừng máy chủ di động"""
        if self.mobile_server and self.mobile_server.is_running:
            try:
                self.mobile_server.stop()
                self.toast("Đã dừng máy chủ di động")
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể dừng máy chủ: {e}")
        else:
            try:
                self.mobile_server = MobileInventoryServer(self, host="0.0.0.0", port=5000)
                self.mobile_server.start()
                self.toast("Đã khởi động máy chủ di động")
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể khởi động máy chủ: {e}")
        
        self.update_mobile_server_ui()

    def refresh_all_data(self):
        """Làm mới tất cả bảng dữ liệu trên giao diện máy tính"""
        try:
            self.refresh_products()
            self.refresh_stock()
            self.refresh_alerts()
            self.refresh_report()
            # Làm mới lịch sử nếu giao diện lịch sử đang mở
            if hasattr(self, 'show_purchase_history'):
                try: self.show_purchase_history()
                except: pass
            if hasattr(self, 'show_dispatch_history'):
                try: self.show_dispatch_history()
                except: pass
        except Exception as e:
            print(f"Lỗi refresh_all_data: {e}")

    def update_mobile_server_ui(self):
        """Cập nhật trạng thái giao diện và vẽ lại mã QR kết nối"""
        if not hasattr(self, 'mobile_status_val') or not self.mobile_status_val:
            return
        if not hasattr(self, 'mobile_url_val') or not self.mobile_url_val:
            return
        if not hasattr(self, 'btn_toggle_server') or not self.btn_toggle_server:
            return
        if not hasattr(self, 'qr_help_lbl') or not self.qr_help_lbl:
            return
        if not hasattr(self, 'qr_canvas') or not self.qr_canvas:
            return
            
        ip = get_local_ip()
        port = 5000
        if self.mobile_server:
            port = self.mobile_server.port
            
        url = f"http://{ip}:{port}"
        
        self.mobile_url_val.config(state='normal')
        self.mobile_url_val.delete(0, 'end')
        self.mobile_url_val.insert(0, url)
        self.mobile_url_val.config(state='readonly')
        
        if self.mobile_server and self.mobile_server.is_running:
            self.mobile_status_val.config(text="ĐANG CHẠY", bootstyle='success')
            self.btn_toggle_server.config(text="⏹️ Dừng máy chủ di động", bootstyle='danger')
            self.qr_help_lbl.config(text="Mở Zalo hoặc quét QR bằng điện thoại để truy cập", bootstyle='success')
            self.draw_qr_code(url)
        else:
            self.mobile_status_val.config(text="ĐÃ DỪNG", bootstyle='danger')
            self.btn_toggle_server.config(text="▶️ Khởi động máy chủ di động", bootstyle='success')
            self.qr_help_lbl.config(text="Vui lòng khởi động máy chủ để hiển thị mã QR", bootstyle='warning')
            self.qr_canvas.delete("all")
            self.qr_canvas.create_text(130, 130, text="MÁY CHỦ\nĐANG DỪNG", fill='gray', font=('Segoe UI', 14, 'bold'), justify='center')

    def draw_qr_code(self, url):
        """Vẽ mã QR lên Canvas"""
        global QR_CODE_AVAILABLE, qrcode
        self.qr_canvas.delete("all")
        
        if not QR_CODE_AVAILABLE:
            try:
                import qrcode
                QR_CODE_AVAILABLE = True
            except ImportError:
                pass
                
        if not QR_CODE_AVAILABLE:
            self.qr_canvas.create_text(130, 100, text="Thiếu thư viện 'qrcode'\nđể hiển thị mã QR", 
                                       fill='red', font=('Segoe UI', 10, 'bold'), justify='center')
                                       
            def auto_install_qr():
                import subprocess, sys
                try:
                    subprocess.run([sys.executable, "-m", "pip", "install", "qrcode"], check=True)
                    global QR_CODE_AVAILABLE
                    QR_CODE_AVAILABLE = True
                    self.toast("Đã cài đặt thành công thư viện qrcode!")
                    self.update_mobile_server_ui()
                except Exception as ex:
                    messagebox.showerror("Lỗi cài đặt", f"Không thể tự động cài đặt: {str(ex)}")
            
            btn_install = tb.Button(self.qr_canvas, text="🔧 Cài đặt qrcode", command=auto_install_qr, bootstyle='warning-outline')
            self.qr_canvas.create_window(130, 160, window=btn_install)
            return

        try:
            qr = qrcode.QRCode(version=1, box_size=1, border=1)
            qr.add_data(url)
            qr.make(fit=True)
            matrix = qr.get_matrix()
            
            num_rows = len(matrix)
            block_size = min(220 // num_rows, 10)
            offset_x = (260 - num_rows * block_size) // 2
            offset_y = (260 - num_rows * block_size) // 2
            
            for r in range(num_rows):
                for c in range(num_rows):
                    if matrix[r][c]:
                        x1 = offset_x + c * block_size
                        y1 = offset_y + r * block_size
                        x2 = x1 + block_size
                        y2 = y1 + block_size
                        self.qr_canvas.create_rectangle(x1, y1, x2, y2, fill="black", outline="black")
        except Exception as e:
            self.qr_canvas.create_text(130, 130, text=f"Lỗi vẽ QR:\n{str(e)}", fill='red', font=('Segoe UI', 10), justify='center')

    def show_purchase_history(self):
        """Hiển thị lịch sử các phiếu nhập kho đã được tạo"""
        try:
            start_date = self.adv_de_from.entry.get().strip()
            end_date = self.adv_de_to.entry.get().strip()
            
            if not start_date or not end_date:
                messagebox.showwarning('Thiếu thông tin', 'Vui lòng chọn đầy đủ ngày bắt đầu và kết thúc')
                return
            
            # Lấy danh sách phiếu nhập kho từ DB
            notes = self.db.get_purchase_notes(start_date, end_date)
            
            # Clear detail tab
            for widget in self.adv_detail_tab.winfo_children():
                widget.destroy()
            
            main_frame = tb.Frame(self.adv_detail_tab)
            main_frame.pack(fill='both', expand=True, padx=10, pady=10)
            
            # Title
            tb.Label(main_frame, text=f"LỊCH SỬ PHIẾU NHẬP KHO ({start_date} -> {end_date})", 
                     font=('Segoe UI', 14, 'bold'), bootstyle='success').pack(pady=(0, 10))
            
            # Bảng danh sách phiếu nhập
            cols = ('id', 'noteNumber', 'supplier', 'createdAt', 'reason', 'item_count', 'note')
            tree = tb.Treeview(main_frame, columns=cols, show='headings', height=12)
            
            for c, w, t, anchor in [
                ('id', 50, 'ID', 'center'),
                ('noteNumber', 150, 'Số phiếu', 'center'),
                ('supplier', 220, 'Nguồn cấp/Nhà CC', 'w'),
                ('createdAt', 150, 'Ngày nhập', 'center'),
                ('reason', 120, 'Lý do nhập', 'center'),
                ('item_count', 90, 'Số mặt hàng', 'center'),
                ('note', 200, 'Ghi chú', 'w')
            ]:
                tree.heading(c, text=t, command=(lambda col=c: self.sort_tree(tree, col)))
                tree.column(c, width=w, anchor=anchor)
                
            tree.tag_configure('odd', background='#f6f8fa')
            tree.pack(fill='both', expand=True, pady=10)
            
            # Load dữ liệu vào tree
            for idx, n in enumerate(notes):
                created_at = n['createdAt']
                tree.insert('', 'end', values=(
                    n['id'],
                    n['noteNumber'],
                    n['supplier'],
                    created_at,
                    n['reason'] or 'Nhập kho',
                    n['item_count'],
                    n['note'] or ''
                ), tags=('odd',) if idx % 2 else ())
            
            # Frame điều khiển bên dưới
            ctrl_btn_frame = tb.Frame(main_frame)
            ctrl_btn_frame.pack(fill='x', pady=5)
            
            def on_reprint():
                sel = tree.selection()
                if not sel:
                    messagebox.showwarning("Chưa chọn dòng", "Vui lòng chọn một phiếu nhập kho trong danh sách!"); return
                val = tree.item(sel[0])['values']
                purchase_id = int(val[0])
                self.reprint_selected_purchase(purchase_id)
                
            def on_delete_purchase():
                sel = tree.selection()
                if not sel:
                    messagebox.showwarning("Chưa chọn dòng", "Vui lòng chọn một phiếu nhập kho trong danh sách để xóa!"); return
                val = tree.item(sel[0])['values']
                purchase_id = int(val[0])
                note_num = val[1]
                
                confirm = messagebox.askyesno(
                    "Xác nhận xóa", 
                    f"Bạn có chắc chắn muốn xóa phiếu nhập số '{note_num}'?\n\n"
                    "Lưu ý: Hành động này sẽ trừ số lượng tồn kho tương ứng của các sản phẩm trong phiếu này và không thể hoàn tác!"
                )
                if not confirm:
                    return
                
                try:
                    # Bắt đầu transaction
                    self.db.conn.execute("BEGIN TRANSACTION")
                    
                    # Lấy thông tin phiếu nhập
                    note_rows = self.db.q("SELECT createdAt FROM purchase_notes WHERE id=?", (purchase_id,))
                    if not note_rows:
                        raise Exception("Không tìm thấy phiếu nhập kho trong cơ sở dữ liệu")
                    note_created_at = note_rows[0]['createdAt']
                    
                    # Lấy danh sách hàng hóa trong phiếu
                    items = self.db.q("SELECT productId, batchId, unitCode, qty FROM purchase_items WHERE purchaseId=?", (purchase_id,))
                    
                    # Xóa các stock movements tương ứng
                    for it in items:
                        self.db.conn.execute(
                            "DELETE FROM stock_movements WHERE productId=? AND batchId=? AND unitCode=? AND qty=? AND type='PURCHASE' AND createdAt=?",
                            (it['productId'], it['batchId'], it['unitCode'], float(it['qty']), note_created_at)
                        )
                    
                    # Xóa phiếu nhập (foreign keys ON DELETE CASCADE sẽ tự động xóa purchase_items)
                    self.db.conn.execute("DELETE FROM purchase_notes WHERE id=?", (purchase_id,))
                    
                    self.db.conn.commit()
                    
                    # Cập nhật lại UI
                    self.toast(f"Đã xóa phiếu nhập {note_num} thành công")
                    self.refresh_products()
                    self.refresh_stock()
                    self.refresh_alerts()
                    self.refresh_report()
                    
                    # Tải lại lịch sử phiếu nhập
                    self.show_purchase_history()
                    
                except Exception as ex:
                    try:
                        self.db.conn.rollback()
                    except:
                        pass
                    messagebox.showerror("Lỗi", f"Không thể xóa phiếu nhập: {str(ex)}")

            tb.Button(ctrl_btn_frame, text="📄 Xem chi tiết & In lại phiếu PDF", bootstyle='info',
                      command=on_reprint).pack(side='left', padx=5)
            
            tb.Button(ctrl_btn_frame, text="🗑️ Xóa phiếu nhập", bootstyle='danger-outline',
                      command=on_delete_purchase).pack(side='left', padx=5)
            
            # Double click để in luôn
            tree.bind("<Double-1>", lambda e: on_reprint())
            
            # Chuyển tab của notebook sang tab Chi tiết
            self.adv_report_nb.select(self.adv_detail_tab)
            self.current_report_type = 'purchase_history'
            
        except Exception as e:
            messagebox.showerror('Lỗi', f"Không thể tải lịch sử phiếu nhập: {str(e)}")

    def reprint_selected_purchase(self, purchase_id):
        """In lại một phiếu nhập kho đã lưu trong cơ sở dữ liệu"""
        try:
            # Lấy thông tin phiếu
            note_rows = self.db.q("SELECT * FROM purchase_notes WHERE id=?", (purchase_id,))
            if not note_rows:
                messagebox.showerror("Lỗi", "Không tìm thấy phiếu nhập kho này"); return
            note_info = note_rows[0]
            
            # Lấy chi tiết hàng hóa
            items = self.db.get_purchase_detail(purchase_id)
            if not items:
                messagebox.showwarning("Trống", "Phiếu nhập kho này không chứa mặt hàng nào!"); return
            
            # Chuyển đổi tên key để tương thích với print_purchase_note
            purchase_items = []
            for it in items:
                purchase_items.append({
                    'productId': it['productId'],
                    'productName': it['productName'],
                    'unitCode': it['unitCode'],
                    'qty': it['qty'],
                    'lotNo': it['lotNo'],
                    'expiryDate': it['expiryDate'],
                    'cost': it['cost']
                })
                
            self.last_purchase_items = purchase_items
            self.last_purchase_info = {
                'id': note_info['id'],
                'noteNumber': note_info['noteNumber'],
                'supplier': note_info['supplier'],
                'reason': note_info['reason'],
                'note': note_info['note'],
                'createdAt': note_info['createdAt']
            }
            
            # Gọi in phiếu nhập kho PDF
            self.print_purchase_note()
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể in lại phiếu nhập: {str(e)}")

    def show_dispatch_history(self):
        """Hiển thị lịch sử các phiếu xuất kho đã được tạo"""
        try:
            start_date = self.adv_de_from.entry.get().strip()
            end_date = self.adv_de_to.entry.get().strip()
            
            if not start_date or not end_date:
                messagebox.showwarning('Thiếu thông tin', 'Vui lòng chọn đầy đủ ngày bắt đầu và kết thúc')
                return
            
            # Lấy danh sách phiếu xuất kho từ DB
            notes = self.db.get_dispatch_notes(start_date, end_date)
            
            # Clear detail tab
            for widget in self.adv_detail_tab.winfo_children():
                widget.destroy()
            
            main_frame = tb.Frame(self.adv_detail_tab)
            main_frame.pack(fill='both', expand=True, padx=10, pady=10)
            
            # Title
            tb.Label(main_frame, text=f"LỊCH SỬ PHIẾU XUẤT KHO ({start_date} -> {end_date})", 
                     font=('Segoe UI', 14, 'bold'), bootstyle='primary').pack(pady=(0, 10))
            
            # Bảng danh sách phiếu xuất
            cols = ('id', 'noteNumber', 'receivingUnit', 'createdAt', 'reason', 'item_count', 'note')
            tree = tb.Treeview(main_frame, columns=cols, show='headings', height=12)
            
            for c, w, t, anchor in [
                ('id', 50, 'ID', 'center'),
                ('noteNumber', 150, 'Số phiếu', 'center'),
                ('receivingUnit', 220, 'Đơn vị nhận', 'w'),
                ('createdAt', 150, 'Ngày xuất', 'center'),
                ('reason', 120, 'Lý do xuất', 'center'),
                ('item_count', 90, 'Số mặt hàng', 'center'),
                ('note', 200, 'Ghi chú', 'w')
            ]:
                tree.heading(c, text=t, command=(lambda col=c: self.sort_tree(tree, col)))
                tree.column(c, width=w, anchor=anchor)
                
            tree.tag_configure('odd', background='#f6f8fa')
            tree.pack(fill='both', expand=True, pady=10)
            
            # Load dữ liệu vào tree
            for idx, n in enumerate(notes):
                created_at = n['createdAt']
                tree.insert('', 'end', values=(
                    n['id'],
                    n['noteNumber'],
                    n['receivingUnit'],
                    created_at,
                    n['reason'] or 'Cấp phát',
                    n['item_count'],
                    n['note'] or ''
                ), tags=('odd',) if idx % 2 else ())
            
            # Frame điều khiển bên dưới
            ctrl_btn_frame = tb.Frame(main_frame)
            ctrl_btn_frame.pack(fill='x', pady=5)
            
            def on_reprint():
                sel = tree.selection()
                if not sel:
                    messagebox.showwarning("Chưa chọn dòng", "Vui lòng chọn một phiếu xuất kho trong danh sách!"); return
                val = tree.item(sel[0])['values']
                dispatch_id = int(val[0])
                self.reprint_selected_dispatch(dispatch_id)
                
            def on_delete_dispatch():
                sel = tree.selection()
                if not sel:
                    messagebox.showwarning("Chưa chọn dòng", "Vui lòng chọn một phiếu xuất kho trong danh sách để xóa!"); return
                val = tree.item(sel[0])['values']
                dispatch_id = int(val[0])
                note_num = val[1]
                
                confirm = messagebox.askyesno(
                    "Xác nhận xóa", 
                    f"Bạn có chắc chắn muốn xóa phiếu xuất số '{note_num}'?\n\n"
                    "Lưu ý: Hành động này sẽ cộng hoàn lại số lượng tồn kho tương ứng của các sản phẩm trong phiếu này và không thể hoàn tác!"
                )
                if not confirm:
                    return
                
                try:
                    # Bắt đầu transaction
                    self.db.conn.execute("BEGIN TRANSACTION")
                    
                    # Lấy thông tin phiếu xuất
                    note_rows = self.db.q("SELECT createdAt FROM dispatch_notes WHERE id=?", (dispatch_id,))
                    if not note_rows:
                        raise Exception("Không tìm thấy phiếu xuất kho trong cơ sở dữ liệu")
                    note_created_at = note_rows[0]['createdAt']
                    
                    # Lấy danh sách hàng hóa trong phiếu
                    items = self.db.q("SELECT productId, batchId, unitCode, qty FROM dispatch_items WHERE dispatchId=?", (dispatch_id,))
                    
                    # Xóa các stock movements tương ứng
                    for it in items:
                        self.db.conn.execute(
                            "DELETE FROM stock_movements WHERE productId=? AND batchId=? AND unitCode=? AND qty=? AND type='DISPATCH' AND createdAt=?",
                            (it['productId'], it['batchId'], it['unitCode'], -float(it['qty']), note_created_at)
                        )
                    
                    # Xóa phiếu xuất (foreign keys ON DELETE CASCADE sẽ tự động xóa dispatch_items)
                    self.db.conn.execute("DELETE FROM dispatch_notes WHERE id=?", (dispatch_id,))
                    
                    self.db.conn.commit()
                    
                    # Cập nhật lại UI
                    self.toast(f"Đã xóa phiếu xuất {note_num} thành công")
                    self.refresh_products()
                    self.refresh_stock()
                    self.refresh_alerts()
                    self.refresh_report()
                    
                    # Tải lại lịch sử phiếu xuất
                    self.show_dispatch_history()
                    
                except Exception as ex:
                    try:
                        self.db.conn.rollback()
                    except:
                        pass
                    messagebox.showerror("Lỗi", f"Không thể xóa phiếu xuất: {str(ex)}")

            tb.Button(ctrl_btn_frame, text="📄 Xem chi tiết & In lại phiếu PDF", bootstyle='info',
                      command=on_reprint).pack(side='left', padx=5)
            
            tb.Button(ctrl_btn_frame, text="🗑️ Xóa phiếu xuất", bootstyle='danger-outline',
                      command=on_delete_dispatch).pack(side='left', padx=5)
            
            # Double click để in luôn
            tree.bind("<Double-1>", lambda e: on_reprint())
            
            # Chuyển tab của notebook sang tab Chi tiết
            self.adv_report_nb.select(self.adv_detail_tab)
            self.current_report_type = 'dispatch_history'
            
        except Exception as e:
            messagebox.showerror('Lỗi', f"Không thể tải lịch sử phiếu xuất: {str(e)}")

    def reprint_selected_dispatch(self, dispatch_id):
        """In lại một phiếu xuất kho đã lưu trong cơ sở dữ liệu"""
        try:
            # Lấy thông tin phiếu
            note_rows = self.db.q("SELECT * FROM dispatch_notes WHERE id=?", (dispatch_id,))
            if not note_rows:
                messagebox.showerror("Lỗi", "Không tìm thấy phiếu xuất kho này"); return
            note_info = note_rows[0]
            
            # Lấy chi tiết hàng hóa
            items = self.db.get_dispatch_detail(dispatch_id)
            if not items:
                messagebox.showwarning("Trống", "Phiếu xuất kho này không chứa mặt hàng nào!"); return
            
            # Chuyển đổi tên key để tương thích với print_dispatch_note
            dispatch_items = []
            for it in items:
                dispatch_items.append({
                    'productId': it['productId'],
                    'productName': it['productName'],
                    'unitCode': it['unitCode'],
                    'qty': it['qty'],
                    'lotNo': it['lotNo'],
                    'expiryDate': it['expiryDate']
                })
                
            self.last_dispatch_items = dispatch_items
            self.last_dispatch_info = {
                'id': note_info['id'],
                'noteNumber': note_info['noteNumber'],
                'receivingUnit': note_info['receivingUnit'],
                'reason': note_info['reason'],
                'note': note_info['note'],
                'createdAt': note_info['createdAt']
            }
            
            self.print_dispatch_note()
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể in lại phiếu: {str(e)}")

    def show_revenue_report(self):
        """Hiển thị báo cáo cấp phát"""
        try:
            start_date = self.adv_de_from.entry.get().strip()
            end_date = self.adv_de_to.entry.get().strip()
            
            if not start_date or not end_date:
                messagebox.showwarning('Thiếu thông tin', 'Vui lòng chọn đầy đủ ngày bắt đầu và kết thúc')
                return
            
            # Tạo dialog chọn loại báo cáo
            dialog = tb.Toplevel(self)
            dialog.title("Báo cáo cấp phát")
            dialog.geometry("300x200")
            dialog.transient(self)
            dialog.grab_set()
            
            # Center dialog
            dialog.update_idletasks()
            x = self.winfo_x() + (self.winfo_width() // 2) - (300 // 2)
            y = self.winfo_y() + (self.winfo_height() // 2) - (200 // 2)
            dialog.geometry(f"300x200+{x}+{y}")
            
            main_frame = tb.Frame(dialog)
            main_frame.pack(fill='both', expand=True, padx=20, pady=20)
            
            tb.Label(main_frame, text="Chọn loại báo cáo:", 
                    font=('Segoe UI', 12, 'bold')).pack(pady=(0, 15))
            
            group_var = tk.StringVar(value='day')
            tb.Radiobutton(main_frame, text="Theo ngày", variable=group_var, value='day').pack(anchor='w', pady=5)
            tb.Radiobutton(main_frame, text="Theo tháng", variable=group_var, value='month').pack(anchor='w', pady=5)
            tb.Radiobutton(main_frame, text="Theo năm", variable=group_var, value='year').pack(anchor='w', pady=5)
            
            def generate_report():
                try:
                    group_by = group_var.get()
                    data = self.report_manager.get_revenue_report(start_date, end_date, group_by)
                    self.current_report_type = 'dispatch'
                    self.current_report_group_by = group_by
                    self.display_revenue_report(data, group_by)
                    dialog.destroy()
                except Exception as e:
                    messagebox.showerror('Lỗi', str(e))
            
            btn_frame = tb.Frame(main_frame)
            btn_frame.pack(fill='x', pady=(20, 0))
            
            tb.Button(btn_frame, text="Tạo báo cáo", bootstyle='success',
                      command=generate_report).pack(side='left', padx=(0, 10))
            tb.Button(btn_frame, text="Hủy", bootstyle='secondary',
                      command=dialog.destroy).pack(side='left')
            
        except Exception as e:
            messagebox.showerror('Lỗi', str(e))

    def display_revenue_report(self, data, group_by):
        """Hiển thị báo cáo cấp phát"""
        try:
            # Clear detail tab
            for widget in self.adv_detail_tab.winfo_children():
                widget.destroy()
            
            main_frame = tb.Frame(self.adv_detail_tab)
            main_frame.pack(fill='both', expand=True, padx=10, pady=10)
            
            # Title
            period_text = {'day': 'ngày', 'month': 'tháng', 'year': 'năm'}
            title = f"BÁO CÁO CẤP PHÁT KHO THEO {period_text[group_by].upper()}"
            tb.Label(main_frame, text=title, 
                    font=('Segoe UI', 14, 'bold'), bootstyle='primary').pack(pady=(0, 15))
            
            # Bảng dữ liệu
            cols = ('period', 'orders', 'revenue', 'paid', 'avg_order')
            tree = tb.Treeview(main_frame, columns=cols, show='headings', height=15)
            
            for c, w, t, anchor in [
                ('period', 120, 'Thời gian', 'center'),
                ('orders', 120, 'Số phiếu xuất', 'e'),
                ('revenue', 150, 'Tổng số lượng xuất', 'e'),
                ('paid', 150, 'Số loại sản phẩm', 'e'),
                ('avg_order', 150, 'SL trung bình/phiếu', 'e')
            ]:
                tree.heading(c, text=t)
                tree.column(c, width=w, anchor=anchor)
            
            tree.tag_configure('odd', background='#f6f8fa')
            tree.tag_configure('total', background='#e8f5e9', font=('Segoe UI', 10, 'bold'))
            tree.pack(fill='both', expand=True)
            
            # Load dữ liệu
            total_orders = total_revenue = total_paid = 0
            for idx, row in enumerate(data):
                orders = row.get('total_orders', 0) or 0
                revenue = row.get('total_revenue', 0) or 0
                paid = row.get('total_paid', 0) or 0
                avg_order = row.get('avg_order_value', 0) or 0
                
                total_orders += orders
                total_revenue += revenue
                total_paid += paid
                
                tree.insert('', 'end', values=(
                    row.get('period', ''),
                    f"{orders:,}",
                    f"{revenue:,.0f}",
                    f"{paid:,}",
                    f"{avg_order:,.1f}"
                ), tags=('odd',) if idx % 2 else ())
            
            # Dòng tổng
            if data:
                tree.insert('', 'end', values=(
                    'TỔNG',
                    f"{total_orders:,}",
                    f"{total_revenue:,.0f}",
                    f"{total_paid:,}",
                    f"{total_revenue/total_orders:,.1f}" if total_orders > 0 else "0"
                ), tags=('total',))
            
            # Chuyển sang tab chi tiết
            self.adv_report_nb.select(self.adv_detail_tab)
            
        except Exception as e:
            messagebox.showerror('Lỗi', str(e))

    def show_profit_report(self):
        """Hiển thị thống kê cấp phát theo đơn vị nhận"""
        try:
            start_date = self.adv_de_from.entry.get().strip()
            end_date = self.adv_de_to.entry.get().strip()
            
            if not start_date or not end_date:
                messagebox.showwarning('Thiếu thông tin', 'Vui lòng chọn đầy đủ ngày bắt đầu và kết thúc')
                return
            
            data = self.report_manager.get_profit_report(start_date, end_date)
            self.current_report_type = 'receiving_unit'
            self.display_profit_report(data)
            
        except Exception as e:
            messagebox.showerror('Lỗi', str(e))

    def display_profit_report(self, data):
        """Hiển thị thống kê cấp phát theo đơn vị nhận"""
        try:
            # Clear detail tab
            for widget in self.adv_detail_tab.winfo_children():
                widget.destroy()
            
            main_frame = tb.Frame(self.adv_detail_tab)
            main_frame.pack(fill='both', expand=True, padx=10, pady=10)
            
            # Title
            tb.Label(main_frame, text="THỐNG KÊ CẤP PHÁT THEO ĐƠN VỊ NHẬN", 
                    font=('Segoe UI', 14, 'bold'), bootstyle='primary').pack(pady=(0, 15))
            
            # Bảng dữ liệu
            cols = ('date', 'unit', 'notes_count', 'total_qty', 'items_count')
            tree = tb.Treeview(main_frame, columns=cols, show='headings', height=15)
            
            for c, w, t, anchor in [
                ('date', 120, 'Ngày xuất', 'center'),
                ('unit', 300, 'Đơn vị nhận', 'w'),
                ('notes_count', 130, 'Số phiếu nhận', 'e'),
                ('total_qty', 150, 'Tổng số lượng nhận', 'e'),
                ('items_count', 150, 'Số loại sản phẩm nhận', 'e')
            ]:
                tree.heading(c, text=t)
                tree.column(c, width=w, anchor=anchor)
            
            tree.tag_configure('odd', background='#f6f8fa')
            tree.tag_configure('total', background='#e8f5e9', font=('Segoe UI', 10, 'bold'))
            tree.pack(fill='both', expand=True)
            
            # Load dữ liệu
            total_notes = total_qty_sum = 0
            for idx, row in enumerate(data):
                unit_name = row.get('product_name', '') or ''
                notes_count = row.get('qty', 0) or 0
                total_qty = row.get('revenue', 0) or 0
                items_count = row.get('cost', 0) or 0
                
                total_notes += notes_count
                total_qty_sum += total_qty
                
                tags = ['odd'] if idx % 2 else []
                
                tree.insert('', 'end', values=(
                    row.get('sale_date', ''),
                    unit_name,
                    f"{notes_count:,}",
                    f"{total_qty:,.0f}",
                    f"{items_count:,}"
                ), tags=tags)
            
            # Dòng tổng
            if data:
                tree.insert('', 'end', values=(
                    'TỔNG', '',
                    f"{total_notes:,}",
                    f"{total_qty_sum:,.0f}",
                    ''
                ), tags=('total',))
            
            # Chuyển sang tab chi tiết
            self.adv_report_nb.select(self.adv_detail_tab)
            
        except Exception as e:
            messagebox.showerror('Lỗi', str(e))

    def show_top_products_report(self):
        """Hiển thị báo cáo top sản phẩm cấp phát"""
        try:
            start_date = self.adv_de_from.entry.get().strip()
            end_date = self.adv_de_to.entry.get().strip()
            
            if not start_date or not end_date:
                messagebox.showwarning('Thiếu thông tin', 'Vui lòng chọn đầy đủ ngày bắt đầu và kết thúc')
                return
            
            data = self.report_manager.get_top_products(start_date, end_date, 20)
            self.current_report_type = 'top_products'
            self.display_top_products_report(data)
            
        except Exception as e:
            messagebox.showerror('Lỗi', str(e))

    def display_top_products_report(self, data):
        """Hiển thị báo cáo top sản phẩm cấp phát"""
        try:
            # Clear detail tab
            for widget in self.adv_detail_tab.winfo_children():
                widget.destroy()
            
            main_frame = tb.Frame(self.adv_detail_tab)
            main_frame.pack(fill='both', expand=True, padx=10, pady=10)
            
            # Title
            tb.Label(main_frame, text="TOP SẢN PHẨM CẤP PHÁT NHIỀU NHẤT", 
                    font=('Segoe UI', 14, 'bold'), bootstyle='primary').pack(pady=(0, 15))
            
            # Bảng dữ liệu
            cols = ('rank', 'product', 'qty', 'orders')
            tree = tb.Treeview(main_frame, columns=cols, show='headings', height=15)
            
            for c, w, t, anchor in [
                ('rank', 80, 'Hạng', 'center'),
                ('product', 450, 'Sản phẩm', 'w'),
                ('qty', 200, 'Tổng SL cấp', 'e'),
                ('orders', 200, 'Số phiếu xuất', 'e')
            ]:
                tree.heading(c, text=t)
                tree.column(c, width=w, anchor=anchor)
            
            tree.tag_configure('odd', background='#f6f8fa')
            tree.tag_configure('top3', background='#fff3e0')
            tree.pack(fill='both', expand=True)
            
            # Load dữ liệu
            for idx, row in enumerate(data):
                rank = idx + 1
                tags = ['odd'] if idx % 2 else []
                if rank <= 3:
                    tags.append('top3')
                
                product_name = row.get('product_name', '') or ''
                total_qty = row.get('total_qty', 0) or 0
                total_orders = row.get('total_orders', 0) or 0
                
                tree.insert('', 'end', values=(
                    f"#{rank}",
                    product_name,
                    f"{total_qty:,.0f}",
                    f"{total_orders:,}"
                ), tags=tags)
            
            # Chuyển sang tab chi tiết
            self.adv_report_nb.select(self.adv_detail_tab)
            
        except Exception as e:
            messagebox.showerror('Lỗi', str(e))

    def show_revenue_chart(self):
        """Hiển thị biểu đồ cấp phát"""
        if not MATPLOTLIB_AVAILABLE:
            messagebox.showerror('Lỗi', 'Thư viện matplotlib chưa được cài đặt. Vui lòng chạy: pip install matplotlib')
            return
        
        try:
            start_date = self.adv_de_from.entry.get().strip()
            end_date = self.adv_de_to.entry.get().strip()
            
            if not start_date or not end_date:
                messagebox.showwarning('Thiếu thông tin', 'Vui lòng chọn đầy đủ ngày bắt đầu và kết thúc')
                return
            
            # Clear chart tab
            for widget in self.adv_chart_tab.winfo_children():
                widget.destroy()
            
            # Lấy dữ liệu
            data = self.report_manager.get_revenue_report(start_date, end_date, 'day')
            
            if not data:
                messagebox.showinfo('Thông báo', 'Không có dữ liệu để hiển thị biểu đồ')
                return
            
            # Tạo biểu đồ
            fig = Figure(figsize=(12, 6), dpi=100)
            ax = fig.add_subplot(111)
            
            # Chuẩn bị dữ liệu
            dates = []
            revenues = []
            orders = []
            
            for row in data:
                try:
                    period = row.get('period', '')
                    if period:
                        dates.append(dt.datetime.strptime(period, '%Y-%m-%d'))
                        revenues.append(row.get('total_revenue', 0) or 0)
                        orders.append(row.get('total_orders', 0) or 0)
                except ValueError:
                    continue  # Bỏ qua ngày không hợp lệ
            
            if not dates:
                # Không có dữ liệu để vẽ
                ax.text(0.5, 0.5, 'Không có dữ liệu để hiển thị biểu đồ', 
                       ha='center', va='center', transform=ax.transAxes, fontsize=14)
                ax.set_title('Biểu đồ cấp phát theo ngày')
            else:
                # Vẽ biểu đồ cấp phát
                ax.plot(dates, revenues, marker='o', linewidth=2, markersize=6, color='#2e7d32', label='Số lượng cấp')
                ax.set_xlabel('Ngày')
                ax.set_ylabel('Số lượng cấp phát')
                ax.set_title('Biểu đồ cấp phát theo ngày')
                ax.grid(True, alpha=0.3)
                ax.legend()
            
            # Format trục Y
            ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x:,.0f}'))
            
            # Format trục X (chỉ khi có dữ liệu)
            if dates:
                ax.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m'))
                ax.xaxis.set_major_locator(mdates.DayLocator(interval=max(1, len(dates)//10)))
                plt.setp(ax.xaxis.get_majorticklabels(), rotation=45)
            
            # Tạo canvas
            canvas = FigureCanvasTkAgg(fig, self.adv_chart_tab)
            canvas.draw()
            canvas.get_tk_widget().pack(fill='both', expand=True, padx=10, pady=10)
            
            # Chuyển sang tab biểu đồ
            self.adv_report_nb.select(self.adv_chart_tab)
            
        except Exception as e:
            messagebox.showerror('Lỗi', str(e))

    def export_report_csv(self):
        start_s = self.de_from.entry.get().strip() if hasattr(self, 'de_from') else ''
        end_s   = self.de_to.entry.get().strip() if hasattr(self, 'de_to') else ''
        if not start_s or not end_s:
            messagebox.showwarning('Thiếu ngày', 'Chọn đủ Từ ngày và Đến ngày'); return

        rows = self.db.xnt_report(start_s, end_s)
        if not rows:
            messagebox.showinfo('Thông báo','Không có dữ liệu'); return

        path = filedialog.asksaveasfilename(
            defaultextension='.csv',
            filetypes=[('CSV','*.csv')],
            initialfile=f'bao_cao_xuat_nhap_ton_{start_s}_to_{end_s}.csv'
        )
        if not path: return

        import csv
        tot_open = tot_in = tot_out = tot_close = 0.0
        with open(path, 'w', newline='', encoding='utf-8') as f:
            w = csv.writer(f)
            w.writerow(['productId','productName','opening','inbound','outbound','closing'])
            for r in rows:
                w.writerow([r['productId'], r['productName'], r['opening'], r['inbound'], r['outbound'], r['closing']])
                tot_open  += float(r['opening']); tot_in += float(r['inbound'])
                tot_out   += float(r['outbound']); tot_close += float(r['closing'])
            w.writerow([])
            w.writerow(['', 'TOTAL', round(tot_open,4), round(tot_in,4), round(tot_out,4), round(tot_close,4)])

        self.toast('Đã lưu báo cáo X–N–T')

    def export_report_pdf(self):
        start_s = self.de_from.entry.get().strip() if hasattr(self, 'de_from') else ''
        end_s   = self.de_to.entry.get().strip() if hasattr(self, 'de_to') else ''
        if not start_s or not end_s:
            messagebox.showwarning('Thiếu ngày', 'Chọn đủ Từ ngày và Đến ngày'); return

        rows = self.db.xnt_report(start_s, end_s)
        if not rows:
            messagebox.showinfo('Thông báo', 'Không có dữ liệu báo cáo trong khoảng thời gian này'); return

        path = filedialog.asksaveasfilename(
            defaultextension='.pdf',
            filetypes=[('PDF files', '*.pdf'), ('All files', '*.*')],
            initialfile=f'bao_cao_xuat_nhap_ton_{start_s}_to_{end_s}.pdf'
        )
        if not path:
            return

        try:
            import reportlab
        except ImportError:
            response = messagebox.askyesno(
                "Thiếu thư viện", 
                "Hệ thống thiếu thư viện 'reportlab' để xuất PDF.\nBạn có muốn tự động cài đặt không? (Quá trình này mất khoảng vài giây)"
            )
            if response:
                import subprocess
                import sys
                try:
                    self.toast("Đang cài đặt thư viện reportlab, vui lòng đợi...")
                    subprocess.run([sys.executable, "-m", "pip", "install", "reportlab"], check=True)
                    self.toast("Đã cài đặt reportlab thành công!")
                except Exception as ex:
                    messagebox.showerror("Lỗi cài đặt", f"Không thể tự động cài đặt reportlab: {str(ex)}\nHãy chạy lệnh 'pip install reportlab' trong terminal."); return
            else:
                return

        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            
            try:
                pdfmetrics.registerFont(TTFont('TimesNewRoman', "C:\\Windows\\Fonts\\times.ttf"))
                pdfmetrics.registerFont(TTFont('TimesNewRoman-Bold', "C:\\Windows\\Fonts\\timesbd.ttf"))
                pdfmetrics.registerFont(TTFont('TimesNewRoman-Italic', "C:\\Windows\\Fonts\\timesi.ttf"))
                font_normal = 'TimesNewRoman'
                font_bold = 'TimesNewRoman-Bold'
                font_italic = 'TimesNewRoman-Italic'
            except Exception:
                font_normal = 'Helvetica'
                font_bold = 'Helvetica-Bold'
                font_italic = 'Helvetica-Oblique'
                
            # Đặt trang nằm ngang (landscape) để bảng rộng rãi
            doc = SimpleDocTemplate(path, pagesize=landscape(A4), rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
            story = []
            
            styles = getSampleStyleSheet()
            
            style_header_left = ParagraphStyle(
                'HeaderLeft', parent=styles['Normal'], fontName=font_bold, fontSize=10, leading=14, alignment=0
            )
            style_header_right = ParagraphStyle(
                'HeaderRight', parent=styles['Normal'], fontName=font_normal, fontSize=10, leading=14, alignment=2
            )
            style_title = ParagraphStyle(
                'Title', parent=styles['Heading1'], fontName=font_bold, fontSize=16, leading=20, alignment=1, spaceAfter=5
            )
            style_subtitle = ParagraphStyle(
                'Subtitle', parent=styles['Normal'], fontName=font_bold, fontSize=11, leading=14, alignment=1, spaceAfter=15
            )
            style_table_header = ParagraphStyle(
                'TableHeader', parent=styles['Normal'], fontName=font_bold, fontSize=9, leading=11, alignment=1, textColor=colors.black
            )
            style_cell = ParagraphStyle(
                'Cell', parent=styles['Normal'], fontName=font_normal, fontSize=9, leading=11, alignment=0
            )
            style_cell_center = ParagraphStyle(
                'CellCenter', parent=styles['Normal'], fontName=font_normal, fontSize=9, leading=11, alignment=1
            )
            style_cell_right = ParagraphStyle(
                'CellRight', parent=styles['Normal'], fontName=font_normal, fontSize=9, leading=11, alignment=2
            )
            
            # Header
            header_data = [
                [
                    Paragraph("SỞ Y TẾ THÀNH PHỐ CẦN THƠ<br/>TRUNG TÂM KIỂM SOÁT BỆNH TẬT (CDC)", style_header_left),
                    Paragraph("<b>Mẫu số: C30-HD</b><br/><i>(Ban hành theo Thông tư số 107/2017/TT-BTC)</i>", style_header_right)
                ]
            ]
            header_table = Table(header_data, colWidths=[400, 362])
            header_table.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
                ('BOTTOMPADDING', (0,0), (-1,-1), 10),
            ]))
            story.append(header_table)
            story.append(Spacer(1, 10))
            
            # Title
            story.append(Paragraph("BÁO CÁO XUẤT - NHẬP - TỒN KHO THUỐC, VACCINE & VẬT TƯ Y TẾ", style_title))
            
            # Định dạng ngày
            try:
                start_dt = dt.datetime.strptime(start_s, '%Y-%m-%d')
                end_dt = dt.datetime.strptime(end_s, '%Y-%m-%d')
                date_range_str = f"Từ ngày {start_dt.strftime('%d/%m/%Y')} đến ngày {end_dt.strftime('%d/%m/%Y')}"
            except Exception:
                date_range_str = f"Từ ngày {start_s} đến ngày {end_s}"
                
            story.append(Paragraph(date_range_str, style_subtitle))
            story.append(Spacer(1, 10))
            
            # Table items
            table_data = [
                [
                    Paragraph("STT", style_table_header),
                    Paragraph("Mã SP", style_table_header),
                    Paragraph("Tên thuốc, vaccine, VTYT", style_table_header),
                    Paragraph("ĐVT", style_table_header),
                    Paragraph("Số lô", style_table_header),
                    Paragraph("Hạn dùng", style_table_header),
                    Paragraph("Tồn đầu", style_table_header),
                    Paragraph("Nhập", style_table_header),
                    Paragraph("Xuất", style_table_header),
                    Paragraph("Tồn cuối", style_table_header)
                ]
            ]
            
            tot_open = tot_in = tot_out = tot_close = 0.0
            for idx, r in enumerate(rows, 1):
                o_val = float(r['opening'])
                i_val = float(r['inbound'])
                ou_val = float(r['outbound'])
                c_val = float(r['closing'])
                
                tot_open += o_val
                tot_in += i_val
                tot_out += ou_val
                tot_close += c_val
                
                table_data.append([
                    Paragraph(str(idx), style_cell_center),
                    Paragraph(str(r['productId']), style_cell_center),
                    Paragraph(r['productName'], style_cell),
                    Paragraph(r['unit'] or '-', style_cell_center),
                    Paragraph(r['lotNo'] or '', style_cell_center),
                    Paragraph(r['expiryDate'] or '', style_cell_center),
                    Paragraph(f"{o_val:g}", style_cell_right),
                    Paragraph(f"{i_val:g}", style_cell_right),
                    Paragraph(f"{ou_val:g}", style_cell_right),
                    Paragraph(f"{c_val:g}", style_cell_right)
                ])
                
            # Thêm dòng tổng cộng
            table_data.append([
                Paragraph("<b>Tổng cộng</b>", style_cell_center),
                Paragraph("", style_cell_center),
                Paragraph("", style_cell),
                Paragraph("", style_cell_center),
                Paragraph("", style_cell_center),
                Paragraph("", style_cell_center),
                Paragraph(f"<b>{tot_open:g}</b>", style_cell_right),
                Paragraph(f"<b>{tot_in:g}</b>", style_cell_right),
                Paragraph(f"<b>{tot_out:g}</b>", style_cell_right),
                Paragraph(f"<b>{tot_close:g}</b>", style_cell_right)
            ])
            
            col_widths = [25, 45, 257, 45, 65, 65, 65, 65, 65, 65]
            items_table = Table(table_data, colWidths=col_widths)
            items_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f2f2f2')),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.black),
                ('SPAN', (0, -1), (5, -1)),
                ('TOPPADDING', (0,0), (-1,-1), 5),
                ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ]))
            story.append(items_table)
            story.append(Spacer(1, 20))
            
            # Ký tên
            date_right_style = ParagraphStyle(
                'DateRight', parent=styles['Normal'], fontName=font_italic, fontSize=11, alignment=2, spaceAfter=10
            )
            sig_title_style = ParagraphStyle(
                'SigTitle', parent=styles['Normal'], fontName=font_bold, fontSize=11, alignment=1
            )
            sig_sub_style = ParagraphStyle(
                'SigSub', parent=styles['Normal'], fontName=font_italic, fontSize=9, alignment=1
            )
            
            now_dt = dt.datetime.now()
            story.append(Paragraph(f"Cần Thơ, ngày {now_dt.strftime('%d')} tháng {now_dt.strftime('%m')} năm {now_dt.strftime('%Y')}", date_right_style))
            
            sig_headers = [
                [
                    Paragraph("<b>Người lập báo cáo</b>", sig_title_style),
                    Paragraph("<b>Thủ kho</b>", sig_title_style),
                    Paragraph("<b>Kế toán trưởng</b>", sig_title_style),
                    Paragraph("<b>Thủ trưởng đơn vị</b>", sig_title_style)
                ],
                [
                    Paragraph("(Ký, họ tên)", sig_sub_style),
                    Paragraph("(Ký, họ tên)", sig_sub_style),
                    Paragraph("(Ký, họ tên)", sig_sub_style),
                    Paragraph("(Ký, đóng dấu)", sig_sub_style)
                ]
            ]
            sig_table = Table(sig_headers, colWidths=[190, 190, 190, 190])
            sig_table.setStyle(TableStyle([
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('BOTTOMPADDING', (0,0), (-1,-1), 2),
            ]))
            story.append(sig_table)
            story.append(Spacer(1, 60))
            
            doc.build(story)
            os.startfile(path)
            self.toast("Đã xuất báo cáo XNT ra PDF và mở file thành công")
            
        except Exception as e:
            messagebox.showerror("Lỗi in PDF", f"Không thể xuất báo cáo PDF: {str(e)}")

    def print_inventory_check_pdf(self):
        end_s = self.de_to.entry.get().strip() if hasattr(self, 'de_to') else ''
        if not end_s:
            messagebox.showwarning('Thiếu ngày', 'Hãy chọn ngày đến (ngày kết thúc kiểm kê) ở ô Đến ngày'); return

        rows = self.db.xnt_report('2000-01-01', end_s)
        items = [r for r in rows if float(r['closing']) > 0]
        
        if not items:
            messagebox.showinfo('Thông báo', 'Không có sản phẩm nào có số dư tồn kho tại ngày này để kiểm kê.'); return

        path = filedialog.asksaveasfilename(
            defaultextension='.pdf',
            filetypes=[('PDF files', '*.pdf'), ('All files', '*.*')],
            initialfile=f'bien_ban_kiem_ke_kho_{end_s}.pdf'
        )
        if not path:
            return

        try:
            import reportlab
        except ImportError:
            response = messagebox.askyesno(
                "Thiếu thư viện", 
                "Hệ thống thiếu thư viện 'reportlab' để xuất PDF.\nBạn có muốn tự động cài đặt không?"
            )
            if response:
                import subprocess, sys
                try:
                    subprocess.run([sys.executable, "-m", "pip", "install", "reportlab"], check=True)
                    self.toast("Đã cài đặt reportlab thành công!")
                except Exception as ex:
                    messagebox.showerror("Lỗi cài đặt", f"Không thể cài đặt reportlab: {str(ex)}"); return
            else:
                return

        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            
            try:
                pdfmetrics.registerFont(TTFont('TimesNewRoman', "C:\\Windows\\Fonts\\times.ttf"))
                pdfmetrics.registerFont(TTFont('TimesNewRoman-Bold', "C:\\Windows\\Fonts\\timesbd.ttf"))
                pdfmetrics.registerFont(TTFont('TimesNewRoman-Italic', "C:\\Windows\\Fonts\\timesi.ttf"))
                font_normal = 'TimesNewRoman'
                font_bold = 'TimesNewRoman-Bold'
                font_italic = 'TimesNewRoman-Italic'
            except Exception:
                font_normal = 'Helvetica'
                font_bold = 'Helvetica-Bold'
                font_italic = 'Helvetica-Oblique'
                
            doc = SimpleDocTemplate(path, pagesize=landscape(A4), rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
            story = []
            
            styles = getSampleStyleSheet()
            
            style_header_left = ParagraphStyle(
                'HeaderLeft', parent=styles['Normal'], fontName=font_bold, fontSize=10, leading=14, alignment=0
            )
            style_header_right = ParagraphStyle(
                'HeaderRight', parent=styles['Normal'], fontName=font_normal, fontSize=10, leading=14, alignment=2
            )
            style_title = ParagraphStyle(
                'Title', parent=styles['Heading1'], fontName=font_bold, fontSize=16, leading=20, alignment=1, spaceAfter=5
            )
            style_subtitle = ParagraphStyle(
                'Subtitle', parent=styles['Normal'], fontName=font_bold, fontSize=11, leading=14, alignment=1, spaceAfter=10
            )
            style_text_bold = ParagraphStyle(
                'TextBold', parent=styles['Normal'], fontName=font_bold, fontSize=10, leading=14
            )
            style_text_normal = ParagraphStyle(
                'TextNormal', parent=styles['Normal'], fontName=font_normal, fontSize=10, leading=15
            )
            style_table_header = ParagraphStyle(
                'TableHeader', parent=styles['Normal'], fontName=font_bold, fontSize=9, leading=11, alignment=1
            )
            style_cell = ParagraphStyle(
                'Cell', parent=styles['Normal'], fontName=font_normal, fontSize=9, leading=11, alignment=0
            )
            style_cell_center = ParagraphStyle(
                'CellCenter', parent=styles['Normal'], fontName=font_normal, fontSize=9, leading=11, alignment=1
            )
            style_cell_right = ParagraphStyle(
                'CellRight', parent=styles['Normal'], fontName=font_normal, fontSize=9, leading=11, alignment=2
            )
            
            # Header
            header_data = [
                [
                    Paragraph("SỞ Y TẾ THÀNH PHỐ CẦN THƠ<br/>TRUNG TÂM KIỂM SOÁT BỆNH TẬT (CDC)", style_header_left),
                    Paragraph("<b>Mẫu số: C33-HD</b><br/><i>(Ban hành theo Thông tư số 107/2017/TT-BTC)</i>", style_header_right)
                ]
            ]
            header_table = Table(header_data, colWidths=[400, 362])
            header_table.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
                ('BOTTOMPADDING', (0,0), (-1,-1), 10),
            ]))
            story.append(header_table)
            story.append(Spacer(1, 10))
            
            # Title
            story.append(Paragraph("BIÊN BẢN KIỂM KÊ THUỐC, VACCINE & VẬT TƯ Y TẾ", style_title))
            
            try:
                check_dt = dt.datetime.strptime(end_s, '%Y-%m-%d')
                date_str = f"Thời điểm kiểm kê: 24 giờ 00 phút ngày {check_dt.strftime('%d/%m/%Y')}"
            except Exception:
                date_str = f"Thời điểm kiểm kê: ngày {end_s}"
                
            story.append(Paragraph(date_str, style_subtitle))
            
            # Ban kiểm kê
            story.append(Paragraph("<b>BAN KIỂM KÊ GỒM:</b>", style_text_bold))
            story.append(Paragraph(
                "1. Ông/Bà: .................... Chức danh: .................... Đại diện Ban Giám đốc (Trưởng ban)<br/>"
                "2. Ông/Bà: .................... Chức danh: .................... Kế toán trưởng (Thành viên)<br/>"
                "3. Ông/Bà: .................... Chức danh: .................... Thủ kho (Thành viên)<br/>"
                "4. Ông/Bà: .................... Chức danh: .................... Trưởng khoa Dược (Thành viên)",
                style_text_normal
            ))
            story.append(Spacer(1, 10))
            
            # Table items
            table_data = [
                [
                    Paragraph("STT", style_table_header),
                    Paragraph("Mã SP", style_table_header),
                    Paragraph("Tên thuốc, vaccine, VTYT", style_table_header),
                    Paragraph("ĐVT", style_table_header),
                    Paragraph("Số lô", style_table_header),
                    Paragraph("Hạn dùng", style_table_header),
                    Paragraph("Số lượng<br/>sổ sách", style_table_header),
                    Paragraph("Số lượng<br/>thực tế", style_table_header),
                    Paragraph("Chênh lệch<br/>(Thừa/Thiếu)", style_table_header),
                    Paragraph("Ghi chú", style_table_header)
                ]
            ]
            
            tot_books = 0.0
            for idx, r in enumerate(items, 1):
                c_val = float(r['closing'])
                tot_books += c_val
                
                table_data.append([
                    Paragraph(str(idx), style_cell_center),
                    Paragraph(str(r['productId']), style_cell_center),
                    Paragraph(r['productName'], style_cell),
                    Paragraph(r['unit'] or '-', style_cell_center),
                    Paragraph(r['lotNo'] or '', style_cell_center),
                    Paragraph(r['expiryDate'] or '', style_cell_center),
                    Paragraph(f"{c_val:g}", style_cell_right),
                    Paragraph("", style_cell_center),
                    Paragraph("", style_cell_center),
                    Paragraph("", style_cell_center)
                ])
                
            # Dòng tổng cộng
            table_data.append([
                Paragraph("<b>Tổng cộng</b>", style_cell_center),
                Paragraph("", style_cell_center),
                Paragraph("", style_cell),
                Paragraph("", style_cell_center),
                Paragraph("", style_cell_center),
                Paragraph("", style_cell_center),
                Paragraph(f"<b>{tot_books:g}</b>", style_cell_right),
                Paragraph("", style_cell_center),
                Paragraph("", style_cell_center),
                Paragraph("", style_cell_center)
            ])
            
            col_widths = [25, 45, 292, 40, 60, 60, 60, 60, 60, 60]
            items_table = Table(table_data, colWidths=col_widths)
            items_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f2f2f2')),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.black),
                ('SPAN', (0, -1), (5, -1)),
                ('TOPPADDING', (0,0), (-1,-1), 5),
                ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ]))
            story.append(items_table)
            story.append(Spacer(1, 20))
            
            # Ký tên
            date_right_style = ParagraphStyle(
                'DateRight', parent=styles['Normal'], fontName=font_italic, fontSize=11, alignment=2, spaceAfter=10
            )
            sig_title_style = ParagraphStyle(
                'SigTitle', parent=styles['Normal'], fontName=font_bold, fontSize=11, alignment=1
            )
            sig_sub_style = ParagraphStyle(
                'SigSub', parent=styles['Normal'], fontName=font_italic, fontSize=9, alignment=1
            )
            
            now_dt = dt.datetime.now()
            story.append(Paragraph(f"Cần Thơ, ngày {now_dt.strftime('%d')} tháng {now_dt.strftime('%m')} năm {now_dt.strftime('%Y')}", date_right_style))
            
            sig_headers = [
                [
                    Paragraph("<b>Người lập biểu</b>", sig_title_style),
                    Paragraph("<b>Thủ kho</b>", sig_title_style),
                    Paragraph("<b>Kế toán trưởng</b>", sig_title_style),
                    Paragraph("<b>Thủ trưởng đơn vị</b>", sig_title_style)
                ],
                [
                    Paragraph("(Ký, họ tên)", sig_sub_style),
                    Paragraph("(Ký, họ tên)", sig_sub_style),
                    Paragraph("(Ký, họ tên)", sig_sub_style),
                    Paragraph("(Ký, đóng dấu)", sig_sub_style)
                ]
            ]
            sig_table = Table(sig_headers, colWidths=[190, 190, 190, 190])
            sig_table.setStyle(TableStyle([
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('BOTTOMPADDING', (0,0), (-1,-1), 2),
            ]))
            story.append(sig_table)
            story.append(Spacer(1, 60))
            
            doc.build(story)
            os.startfile(path)
            self.toast("Đã in biên bản kiểm kê ra PDF thành công")
            
        except Exception as e:
            messagebox.showerror("Lỗi in PDF", f"Không thể xuất biên bản kiểm kê: {str(e)}")

    def export_current_report_excel(self):
        """Xuất báo cáo hiện tại ra Excel"""
        try:
            # Lấy dữ liệu báo cáo hiện tại
            report_data = self.get_current_report_data()
            if not report_data:
                messagebox.showwarning('Cảnh báo', 'Không có dữ liệu báo cáo để xuất')
                return
            
            # Chọn file để lưu
            start_date = self.adv_de_from.entry.get().strip()
            end_date = self.adv_de_to.entry.get().strip()
            filename = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[('Excel files', '*.xlsx'), ('All files', '*.*')],
                initialfile=f'bao_cao_{start_date}_to_{end_date}.xlsx'
            )
            
            if filename:
                # Xuất ra Excel
                self.export_manager.export_to_excel(
                    data=report_data['data'],
                    filename=filename,
                    sheet_name=report_data['title'],
                    headers=report_data.get('headers')
                )
                messagebox.showinfo('Thành công', f'Đã xuất báo cáo ra Excel:\n{filename}')
                
        except Exception as e:
            messagebox.showerror('Lỗi', f'Không thể xuất Excel: {str(e)}')
    
    def export_current_report_pdf(self):
        """Xuất báo cáo hiện tại ra PDF"""
        try:
            # Lấy dữ liệu báo cáo hiện tại
            report_data = self.get_current_report_data()
            if not report_data:
                messagebox.showwarning('Cảnh báo', 'Không có dữ liệu báo cáo để xuất')
                return
            
            # Chọn file để lưu
            start_date = self.adv_de_from.entry.get().strip()
            end_date = self.adv_de_to.entry.get().strip()
            filename = filedialog.asksaveasfilename(
                defaultextension='.pdf',
                filetypes=[('PDF files', '*.pdf'), ('All files', '*.*')],
                initialfile=f'bao_cao_{start_date}_to_{end_date}.pdf'
            )
            
            if filename:
                # Xuất ra PDF
                self.export_manager.export_to_pdf(
                    data=report_data['data'],
                    filename=filename,
                    title=report_data['title'],
                    headers=report_data.get('headers')
                )
                messagebox.showinfo('Thành công', f'Đã xuất báo cáo ra PDF:\n{filename}')
                
        except Exception as e:
            messagebox.showerror('Lỗi', f'Không thể xuất PDF: {str(e)}')
    
    def export_current_report_csv(self):
        """Xuất báo cáo hiện tại ra CSV"""
        try:
            # Lấy dữ liệu báo cáo hiện tại
            report_data = self.get_current_report_data()
            if not report_data:
                messagebox.showwarning('Cảnh báo', 'Không có dữ liệu báo cáo để xuất')
                return
            
            # Chọn file để lưu
            start_date = self.adv_de_from.entry.get().strip()
            end_date = self.adv_de_to.entry.get().strip()
            filename = filedialog.asksaveasfilename(
                defaultextension='.csv',
                filetypes=[('CSV files', '*.csv'), ('All files', '*.*')],
                initialfile=f'bao_cao_{start_date}_to_{end_date}.csv'
            )
            
            if filename:
                # Xuất ra CSV
                import csv
                with open(filename, 'w', newline='', encoding='utf-8') as f:
                    if report_data.get('headers'):
                        writer = csv.writer(f)
                        writer.writerow(report_data['headers'])
                        
                        for row in report_data['data']:
                            if isinstance(row, dict):
                                writer.writerow([row.get(h, '') for h in report_data['headers']])
                            else:
                                writer.writerow(row)
                    else:
                        # Sử dụng pandas nếu có
                        if PANDAS_AVAILABLE:
                            df = pd.DataFrame(report_data['data'])
                            df.to_csv(filename, index=False, encoding='utf-8')
                        else:
                            # Fallback manual CSV
                            writer = csv.writer(f)
                            for row in report_data['data']:
                                if isinstance(row, dict):
                                    writer.writerow(list(row.values()))
                                else:
                                    writer.writerow(row)
                
                messagebox.showinfo('Thành công', f'Đã xuất báo cáo ra CSV:\n{filename}')
                
        except Exception as e:
            messagebox.showerror('Lỗi', f'Không thể xuất CSV: {str(e)}')
    
    def get_current_report_data(self):
        """Lấy dữ liệu báo cáo hiện tại đang hiển thị"""
        try:
            # Lấy thông tin từ tab hiện tại
            current_tab = self.adv_report_nb.select()
            tab_text = self.adv_report_nb.tab(current_tab, 'text')
            
            start_date = self.adv_de_from.entry.get().strip()
            end_date = self.adv_de_to.entry.get().strip()
            
            if not start_date or not end_date:
                return None
            
            # Xác định loại báo cáo dựa trên tab hiện tại
            if 'Tóm tắt' in tab_text:
                # Báo cáo tóm tắt
                summary_data = self.report_manager.get_daily_sales_summary(start_date, end_date)
                summary = summary_data['summary']
                data_list = [
                    ['Tổng số phiếu xuất', summary.get('total_orders', 0)],
                    ['Tổng số lượng xuất', summary.get('total_revenue', 0)],
                    ['Số lượng TB/phiếu', round(summary.get('avg_order_value', 0), 1)],
                    ['Phiếu xuất lớn nhất', summary.get('max_order', 0)]
                ]
                return {
                    'title': f'Báo cáo tóm tắt từ {start_date} đến {end_date}',
                    'data': data_list,
                    'headers': ['Chỉ số', 'Giá trị']
                }
            elif 'Chi tiết' in tab_text:
                if self.current_report_type == 'dispatch':
                    data = self.report_manager.get_revenue_report(start_date, end_date, self.current_report_group_by)
                    data_list = []
                    for row in data:
                        data_list.append([
                            row.get('period', ''),
                            row.get('total_orders', 0),
                            row.get('total_revenue', 0),
                            row.get('total_paid', 0),
                            round(row.get('avg_order_value', 0), 1)
                        ])
                    return {
                        'title': f'Báo cáo cấp phát chi tiết từ {start_date} đến {end_date}',
                        'data': data_list,
                        'headers': ['Thời gian', 'Số phiếu xuất', 'Tổng số lượng xuất', 'Số loại sản phẩm', 'SL trung bình/phiếu']
                    }
                elif self.current_report_type == 'receiving_unit':
                    data = self.report_manager.get_profit_report(start_date, end_date)
                    data_list = []
                    for row in data:
                        data_list.append([
                            row.get('sale_date', ''),
                            row.get('product_name', ''),
                            row.get('qty', 0),
                            row.get('revenue', 0),
                            row.get('cost', 0)
                        ])
                    return {
                        'title': f'Thống kê cấp phát theo đơn vị nhận từ {start_date} đến {end_date}',
                        'data': data_list,
                        'headers': ['Ngày xuất', 'Đơn vị nhận', 'Số phiếu nhận', 'Tổng số lượng nhận', 'Số loại sản phẩm nhận']
                    }
                elif self.current_report_type == 'top_products':
                    data = self.report_manager.get_top_products(start_date, end_date, 20)
                    data_list = []
                    for idx, row in enumerate(data):
                        data_list.append([
                            f"#{idx+1}",
                            row.get('product_name', ''),
                            row.get('total_qty', 0),
                            row.get('total_orders', 0)
                        ])
                    return {
                        'title': f'Top sản phẩm cấp phát nhiều nhất từ {start_date} đến {end_date}',
                        'data': data_list,
                        'headers': ['Hạng', 'Sản phẩm', 'Tổng SL cấp', 'Số phiếu xuất']
                    }
            else:
                # Mặc định là báo cáo cấp phát
                data = self.report_manager.get_revenue_report(start_date, end_date, 'day')
                data_list = []
                for row in data:
                    data_list.append([
                        row.get('period', ''),
                        row.get('total_orders', 0),
                        row.get('total_revenue', 0),
                        row.get('total_paid', 0),
                        round(row.get('avg_order_value', 0), 1)
                    ])
                return {
                    'title': f'Báo cáo cấp phát từ {start_date} đến {end_date}',
                    'data': data_list,
                    'headers': ['Thời gian', 'Số phiếu xuất', 'Tổng số lượng xuất', 'Số loại sản phẩm', 'SL trung bình/phiếu']
                }
                
        except Exception as e:
            print(f"Lỗi lấy dữ liệu báo cáo: {e}")
            return None

    # helpers chung
    def sort_tree(self, tree: tb.Treeview, col: str):
        data = [(tree.set(k, col), k) for k in tree.get_children('')]
        def to_num(x):
            try:
                if isinstance(x, str): x = x.replace(',', '')
                return float(x)
            except:
                return x.lower() if isinstance(x, str) else x
        data.sort(key=lambda t: to_num(t[0]))
        for i, (_, k) in enumerate(data): tree.move(k, '', i)

    def refresh_products(self):
        self._products = self.db.q('SELECT id, name FROM products ORDER BY name')
        opts = [f"{p['id']} — {p['name']}" for p in self._products]
        self.cmb_prod['values'] = opts; self.cmb_prod_pos['values'] = opts
        if opts:
            self.cmb_prod.current(0); self.cmb_prod_pos.current(0)
            self.update_purchase_unit_and_price()
            self.update_dispatch_unit_label()

    def _fill_tree(self, tree: tb.Treeview, rows):
        for i in tree.get_children(): tree.delete(i)
        keymap = {'product':'productId','productName':'productName','batch':'batchId','lot':'lotNo','exp':'expiryDate',
                  'qty':'qtyBase','cost':'costBase','value':'valueBase'}
        for idx, r in enumerate(rows):
            tree.insert('', 'end', values=[r.get(keymap.get(c, c), '') for c in tree['columns']],
                        tags=('odd',) if idx%2 else ())

    def refresh_stock(self):
        rows = self.db.stock_view()
        if hasattr(self, 'tree_stock') and self.tree_stock:
            self._fill_tree(self.tree_stock, rows)
        if hasattr(self, 'tree_stock2') and self.tree_stock2:
            self._fill_tree(self.tree_stock2, rows)

    def refresh_alerts(self):
        try: days = int(self.ent_warn_days.get())
        except: days = 90
        self._fill_tree(self.tree_alerts, self.db.expiring_view(days))

    def refresh_report(self):
        start_s = self.de_from.entry.get().strip() if hasattr(self, 'de_from') else ''
        end_s   = self.de_to.entry.get().strip() if hasattr(self, 'de_to') else ''

        # Clear
        for i in self.tree_report.get_children():
            self.tree_report.delete(i)

        if not start_s or not end_s:
            messagebox.showwarning('Thiếu ngày', 'Chọn đủ Từ ngày và Đến ngày'); return

        rows = self.db.xnt_report(start_s, end_s)

        tot_open = tot_in = tot_out = tot_close = 0.0
        for idx, r in enumerate(rows):
            tag = 'odd' if idx % 2 else ''
            tot_open  += float(r['opening'])
            tot_in    += float(r['inbound'])
            tot_out   += float(r['outbound'])
            tot_close += float(r['closing'])
            self.tree_report.insert(
                '',
                'end',
                values=(
                    r['productId'], 
                    r['productName'],
                    r['lotNo'] or '',
                    r['expiryDate'] or '',
                    f"{r['opening']:g}", 
                    f"{r['inbound']:g}", 
                    f"{r['outbound']:g}", 
                    f"{r['closing']:g}"
                ),
                tags=(tag,)
            )

        # Dòng tổng
        if rows:
            self.tree_report.insert(
                '', 'end',
                values=('', 'TỔNG CỘNG', '', '', f"{tot_open:g}", f"{tot_in:g}", f"{tot_out:g}", f"{tot_close:g}"),
                tags=('total',)
            )


    def on_ready(self):
        self.refresh_products(); self.refresh_stock(); self.refresh_alerts(); self.refresh_report()
        self.refresh_backup_list()
        # Load advanced reports summary
        if hasattr(self, 'load_advanced_summary'):
            self.load_advanced_summary()
        # Update catalog info
        if hasattr(self, 'update_catalog_info'):
            self.update_catalog_info()
        # Tự động load thuoc.csv nếu có
        self.auto_load_medicine_catalog()
        # Cập nhật status database
        self.update_db_status()

    def auto_load_medicine_catalog(self):
        """Tự động load file thuoc.csv khi khởi động"""
        try:
            # Tìm file thuoc.csv trong thư mục hiện tại
            csv_path = os.path.join(os.getcwd(), 'thuoc.csv')
            
            if os.path.exists(csv_path):
                self.medicine_catalog.load_catalog_from_excel(csv_path)
                self.update_catalog_info()
                print(f"Đã tự động load danh mục thuốc: {os.path.basename(csv_path)}")
            else:
                print("Không tìm thấy file thuoc.csv để tự động load")
                
        except Exception as e:
            print(f"Lỗi khi tự động load danh mục thuốc: {e}")

    def export_import_template(self):
        """Xuất file mẫu Excel để nhập dữ liệu hàng loạt"""
        global pd, PANDAS_AVAILABLE
        if not PANDAS_AVAILABLE:
            response = messagebox.askyesno(
                "Thiếu thư viện", 
                "Hệ thống thiếu thư viện 'pandas' và 'openpyxl' để xử lý Excel.\nBạn có muốn tự động cài đặt không? (Quá trình này mất khoảng vài giây)"
            )
            if response:
                import subprocess, sys
                try:
                    self.toast("Đang cài đặt thư viện pandas và openpyxl...")
                    subprocess.run([sys.executable, "-m", "pip", "install", "pandas", "openpyxl"], check=True)
                    import pandas as pd
                    PANDAS_AVAILABLE = True
                    self.toast("Đã cài đặt thành công!")
                except Exception as ex:
                    messagebox.showerror("Lỗi cài đặt", f"Không thể tự động cài đặt: {str(ex)}\nHãy chạy lệnh 'pip install pandas openpyxl' trong terminal."); return
            else:
                return
        
        try:
            path = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[('Excel files', '*.xlsx'), ('CSV files', '*.csv'), ('All files', '*.*')],
                initialfile='mau_nhap_lieu_hang_loat.xlsx',
                title='Chọn nơi lưu file mẫu'
            )
            if not path:
                return

            headers = [
                'Tên sản phẩm',
                'Đơn vị cơ sở',
                'Mã vạch (Barcode)',
                'Loại sản phẩm (thuoc/vaccine/vtyt/khac)',
                'Số đăng ký',
                'Đơn vị quy đổi 1',
                'Tỷ lệ quy đổi 1',
                'Giá bán đơn vị quy đổi 1',
                'Đơn vị quy đổi 2',
                'Tỷ lệ quy đổi 2',
                'Giá bán đơn vị quy đổi 2',
                'Đơn vị quy đổi 3',
                'Tỷ lệ quy đổi 3',
                'Giá bán đơn vị quy đổi 3',
                'Số lô',
                'Hạn sử dụng (YYYY-MM-DD)',
                'Số lượng tồn (Đơn vị cơ sở)',
                'Giá nhập (Đơn vị cơ sở)'
            ]

            sample_data = [
                [
                    'Paracetamol 500mg',
                    'vien',
                    '8931234567890',
                    'thuoc',
                    'VD-12345-20',
                    'vi',
                    10,
                    15000,
                    'hop',
                    100,
                    140000,
                    '',
                    '',
                    '',
                    'LOT123',
                    '2027-12-31',
                    500,
                    1200
                ],
                [
                    'Vaccine Quinvaxem',
                    'lo',
                    '8930987654321',
                    'vaccine',
                    'QLSP-987-19',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    'B2209',
                    '2026-09-30',
                    50,
                    150000
                ]
            ]

            if path.lower().endswith('.csv'):
                df = pd.DataFrame(sample_data, columns=headers)
                df.to_csv(path, index=False, encoding='utf-8-sig')
            else:
                df = pd.DataFrame(sample_data, columns=headers)
                with pd.ExcelWriter(path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Sheet1', index=False)
                    worksheet = writer.sheets['Sheet1']
                    
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 3, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
                    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
                    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
                    thin_border = Border(
                        left=Side(style='thin', color='DDDDDD'),
                        right=Side(style='thin', color='DDDDDD'),
                        top=Side(style='thin', color='DDDDDD'),
                        bottom=Side(style='thin', color='DDDDDD')
                    )

                    worksheet.row_dimensions[1].height = 28
                    
                    for cell in worksheet[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_align
                        cell.border = thin_border
                    
                    data_font = Font(name='Segoe UI', size=10)
                    for row in range(2, worksheet.max_row + 1):
                        worksheet.row_dimensions[row].height = 20
                        for col in range(1, worksheet.max_column + 1):
                            cell = worksheet.cell(row=row, column=col)
                            cell.font = data_font
                            cell.border = thin_border
                            if col in [7, 8, 10, 11, 13, 14, 17, 18]:
                                cell.alignment = Alignment(horizontal='right')
                            elif col in [2, 4, 15, 16]:
                                cell.alignment = Alignment(horizontal='center')
                            else:
                                cell.alignment = Alignment(horizontal='left')

            self.toast('Đã tải file Excel mẫu thành công')
            
        except Exception as e:
            messagebox.showerror('Lỗi', f'Không thể xuất file mẫu: {str(e)}')

    def bulk_import_from_excel(self):
        """Nhập sản phẩm và tồn kho hàng loạt từ file Excel/CSV"""
        global pd, PANDAS_AVAILABLE
        if not PANDAS_AVAILABLE:
            response = messagebox.askyesno(
                "Thiếu thư viện", 
                "Hệ thống thiếu thư viện 'pandas' và 'openpyxl' để xử lý Excel.\nBạn có muốn tự động cài đặt không? (Quá trình này mất khoảng vài giây)"
            )
            if response:
                import subprocess, sys
                try:
                    self.toast("Đang cài đặt thư viện pandas và openpyxl...")
                    subprocess.run([sys.executable, "-m", "pip", "install", "pandas", "openpyxl"], check=True)
                    import pandas as pd
                    PANDAS_AVAILABLE = True
                    self.toast("Đã cài đặt thành công!")
                except Exception as ex:
                    messagebox.showerror("Lỗi cài đặt", f"Không thể tự động cài đặt: {str(ex)}\nHãy chạy lệnh 'pip install pandas openpyxl' trong terminal."); return
            else:
                return
        
        try:
            path = filedialog.askopenfilename(
                title="Chọn file Excel hoặc CSV để nhập hàng loạt",
                filetypes=[
                    ('Excel/CSV files', '*.xlsx;*.xls;*.csv'),
                    ('Excel files', '*.xlsx;*.xls'),
                    ('CSV files', '*.csv'),
                    ('All files', '*.*')
                ]
            )
            if not path:
                return
            
            if path.lower().endswith('.csv'):
                df = pd.read_csv(path, encoding='utf-8')
            else:
                df = pd.read_excel(path)
            
            df.columns = df.columns.str.strip()
            
            required_cols = ['Tên sản phẩm', 'Đơn vị cơ sở']
            missing = [c for c in required_cols if c not in df.columns]
            if missing:
                messagebox.showerror('Lỗi định dạng', f'File thiếu các cột bắt buộc: {", ".join(missing)}')
                return
            
            if df.empty:
                messagebox.showwarning('Cảnh báo', 'File Excel/CSV không có dữ liệu')
                return
            
            total_rows = len(df)
            imported_products = 0
            updated_products = 0
            imported_units = 0
            imported_stock = 0
            errors = []
            
            def parse_import_date(val):
                if pd.isna(val) or val is None:
                    return None
                if isinstance(val, (dt.datetime, dt.date)):
                    return val.strftime('%Y-%m-%d')
                val_str = str(val).strip()
                if not val_str or val_str.lower() in ('nan', 'none', 'null', ''):
                    return None
                for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d', '%d-%m-%Y'):
                    try:
                        return dt.datetime.strptime(val_str, fmt).strftime('%Y-%m-%d')
                    except ValueError:
                        continue
                try:
                    val_float = float(val_str)
                    d = dt.datetime(1899, 12, 30) + dt.timedelta(days=int(val_float))
                    return d.strftime('%Y-%m-%d')
                except:
                    pass
                return val_str

            self.db.conn.execute("BEGIN TRANSACTION")
            
            purchase_id = None
            note_number = f"PNK-INITIAL-{dt.datetime.now().strftime('%Y%m%d%H%M%S')}"
            created_at = dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            has_stock_data = False
            for idx, row in df.iterrows():
                lot_no = str(row.get('Số lô', '')).strip()
                if lot_no and lot_no.lower() not in ('nan', 'none', ''):
                    has_stock_data = True
                    break
            
            if has_stock_data:
                cur = self.db.conn.execute(
                    "INSERT INTO purchase_notes(noteNumber, supplier, reason, note, createdAt) VALUES(?,?,?,?,?)",
                    (note_number, "Nhập kho ban đầu", "Nhập kho ban đầu", "Nhập hàng loạt từ Excel", created_at)
                )
                purchase_id = cur.lastrowid
            
            for idx, row in df.iterrows():
                row_num = idx + 2
                
                name = str(row.get('Tên sản phẩm', '')).strip()
                if not name or name.lower() == 'nan':
                    errors.append(f"Dòng {row_num}: Tên sản phẩm trống")
                    continue
                
                default_unit = str(row.get('Đơn vị cơ sở', '')).strip()
                if not default_unit or default_unit.lower() == 'nan':
                    errors.append(f"Dòng {row_num}: Đơn vị cơ sở trống")
                    continue
                
                barcode = str(row.get('Mã vạch (Barcode)', '')).strip()
                if not barcode or barcode.lower() in ('nan', 'none', ''):
                    barcode = None
                
                product_type = str(row.get('Loại sản phẩm (thuoc/vaccine/vtyt/khac)', '')).strip().lower()
                if not product_type or product_type not in ('thuoc', 'vaccine', 'vtyt', 'khac'):
                    product_type = 'thuoc'
                
                reg_num = str(row.get('Số đăng ký', '')).strip()
                if not reg_num or reg_num.lower() in ('nan', 'none', ''):
                    reg_num = None
                
                prod_row = self.db.q("SELECT id, defaultUnit FROM products WHERE name = ?", (name,))
                is_new = False
                if prod_row:
                    product_id_db = prod_row[0]['id']
                    self.db.conn.execute(
                        "UPDATE products SET barcode=COALESCE(?, barcode), productType=?, registrationNumber=COALESCE(?, registrationNumber) WHERE id=?",
                        (barcode, product_type, reg_num, product_id_db)
                    )
                    updated_products += 1
                else:
                    if barcode:
                        prod_barcode = self.db.q("SELECT id, defaultUnit FROM products WHERE barcode = ?", (barcode,))
                        if prod_barcode:
                            product_id_db = prod_barcode[0]['id']
                            self.db.conn.execute(
                                "UPDATE products SET name=?, productType=?, registrationNumber=COALESCE(?, registrationNumber) WHERE id=?",
                                (name, product_type, reg_num, product_id_db)
                            )
                            updated_products += 1
                        else:
                            is_new = True
                    else:
                        is_new = True
                
                if is_new:
                    cur = self.db.conn.execute(
                        "INSERT INTO products (name, defaultUnit, barcode, productType, registrationNumber) VALUES (?, ?, ?, ?, ?)",
                        (name, default_unit, barcode, product_type, reg_num)
                    )
                    product_id_db = cur.lastrowid
                    imported_products += 1
                
                self.db.conn.execute(
                    "INSERT OR IGNORE INTO product_units(productId, unitCode, toBaseQty, price) VALUES(?,?,1,0)", 
                    (product_id_db, default_unit)
                )
                
                for i in range(1, 4):
                    unit_name = str(row.get(f'Đơn vị quy đổi {i}', '')).strip()
                    if not unit_name or unit_name.lower() in ('nan', 'none', ''):
                        continue
                    
                    try:
                        ratio_val = row.get(f'Tỷ lệ quy đổi {i}')
                        if pd.isna(ratio_val):
                            continue
                        ratio = float(ratio_val)
                        if ratio <= 0:
                            errors.append(f"Dòng {row_num}: Tỷ lệ quy đổi {i} của '{name}' phải > 0 (bỏ qua đơn vị này)")
                            continue
                    except ValueError:
                        errors.append(f"Dòng {row_num}: Tỷ lệ quy đổi {i} của '{name}' không phải là số (bỏ qua đơn vị này)")
                        continue
                    
                    try:
                        price_val = row.get(f'Giá bán đơn vị quy đổi {i}')
                        price = float(price_val) if not pd.isna(price_val) else 0.0
                        if price < 0:
                            price = 0.0
                    except ValueError:
                        price = 0.0
                    
                    self.db.conn.execute(
                        "INSERT OR REPLACE INTO product_units (productId, unitCode, toBaseQty, price) VALUES (?, ?, ?, ?)",
                        (product_id_db, unit_name, ratio, price)
                    )
                    imported_units += 1
                
                lot_no = str(row.get('Số lô', '')).strip()
                if lot_no and lot_no.lower() not in ('nan', 'none', ''):
                    expiry_val = row.get('Hạn sử dụng (YYYY-MM-DD)')
                    expiry_date = parse_import_date(expiry_val)
                    
                    if not expiry_date:
                        errors.append(f"Dòng {row_num}: Số lô '{lot_no}' cho '{name}' thiếu hoặc sai hạn sử dụng (Bỏ qua nhập lô)")
                        continue
                    
                    try:
                        dt.datetime.strptime(expiry_date, '%Y-%m-%d')
                    except ValueError:
                        errors.append(f"Dòng {row_num}: Hạn sử dụng '{expiry_date}' của lô '{lot_no}' không đúng định dạng YYYY-MM-DD (Bỏ qua nhập lô)")
                        continue
                    
                    try:
                        qty_val = row.get('Số lượng tồn (Đơn vị cơ sở)')
                        if pd.isna(qty_val):
                            errors.append(f"Dòng {row_num}: Thiếu số lượng tồn cho lô '{lot_no}' của '{name}' (Bỏ qua nhập lô)")
                            continue
                        qty = float(qty_val)
                        if qty <= 0:
                            errors.append(f"Dòng {row_num}: Số lượng tồn {qty} cho lô '{lot_no}' của '{name}' phải > 0 (Bỏ qua nhập lô)")
                            continue
                    except ValueError:
                        errors.append(f"Dòng {row_num}: Số lượng tồn cho lô '{lot_no}' của '{name}' không hợp lệ (Bỏ qua nhập lô)")
                        continue
                    
                    try:
                        cost_val = row.get('Giá nhập (Đơn vị cơ sở)')
                        cost = float(cost_val) if not pd.isna(cost_val) else 0.0
                        if cost < 0:
                            cost = 0.0
                    except ValueError:
                        cost = 0.0
                    
                    batch_id = None
                    b_row = self.db.q("SELECT id FROM batches WHERE productId=? AND lotNo=?", (product_id_db, lot_no))
                    if b_row:
                        batch_id = b_row[0]['id']
                    else:
                        cur_b = self.db.conn.execute(
                            "INSERT INTO batches(productId, lotNo, expiryDate) VALUES(?,?,?)",
                            (product_id_db, lot_no, expiry_date)
                        )
                        batch_id = cur_b.lastrowid
                    
                    self.db.conn.execute(
                        "INSERT INTO stock_movements(productId, batchId, unitCode, qty, type, cost, receivingUnit, reason, createdAt) VALUES(?,?,?,?,'PURCHASE',?,'Nhập kho ban đầu','Nhập kho ban đầu',?)",
                        (product_id_db, batch_id, default_unit, qty, cost, created_at)
                    )
                    
                    if purchase_id:
                        self.db.conn.execute(
                            "INSERT INTO purchase_items(purchaseId, productId, batchId, unitCode, qty, lotNo, expiryDate, cost) VALUES(?,?,?,?,?,?,?,?)",
                            (purchase_id, product_id_db, batch_id, default_unit, qty, lot_no, expiry_date, cost)
                        )
                    
                    p_unit_row = self.db.q("SELECT price FROM product_units WHERE productId=? AND unitCode=?", (product_id_db, default_unit))
                    if p_unit_row and (p_unit_row[0]['price'] is None or float(p_unit_row[0]['price']) == 0.0):
                        self.db.conn.execute(
                            "UPDATE product_units SET price=? WHERE productId=? AND unitCode=?",
                            (cost, product_id_db, default_unit)
                        )
                    
                    imported_stock += 1
            
            self.db.conn.commit()
            
            self.refresh_products()
            self.refresh_stock()
            self.refresh_alerts()
            self.refresh_report()
            
            success_msg = f"Đã nhập dữ liệu thành công:\n"
            success_msg += f"- Thêm mới {imported_products} sản phẩm\n"
            success_msg += f"- Cập nhật {updated_products} sản phẩm\n"
            if imported_units > 0:
                success_msg += f"- Thêm {imported_units} đơn vị quy đổi\n"
            if imported_stock > 0:
                success_msg += f"- Nhập {imported_stock} lô tồn kho ban đầu (Số phiếu: {note_number})\n"
            
            if errors:
                self.show_import_log_dialog(success_msg, errors)
            else:
                messagebox.showinfo('Thành công', success_msg)
                self.toast('Đã nhập dữ liệu hàng loạt thành công')
                
        except Exception as e:
            try:
                self.db.conn.rollback()
            except:
                pass
            messagebox.showerror('Lỗi', f'Lỗi trong quá trình nhập dữ liệu: {str(e)}')

    def show_import_log_dialog(self, summary, errors):
        """Hiển thị thông báo kết quả nhập và danh sách lỗi/cảnh báo"""
        dialog = tb.Toplevel(self)
        dialog.title("Kết quả nhập hàng loạt")
        dialog.geometry("650x500")
        dialog.transient(self)
        dialog.grab_set()
        
        dialog.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() // 2) - (650 // 2)
        y = self.winfo_y() + (self.winfo_height() // 2) - (500 // 2)
        dialog.geometry(f"650x500+{x}+{y}")
        
        main_frame = tb.Frame(dialog)
        main_frame.pack(fill='both', expand=True, padx=15, pady=15)
        
        summary_title = tb.Label(main_frame, text="TÓM TẮT KẾT QUẢ", font=('Segoe UI', 11, 'bold'), bootstyle='success')
        summary_title.pack(anchor='w', pady=(0, 5))
        
        summary_box = tb.Label(main_frame, text=summary, font=('Segoe UI', 9), justify='left')
        summary_box.pack(anchor='w', pady=(0, 15))
        
        error_title = tb.Label(main_frame, text=f"DANH SÁCH CHI TIẾT BỎ QUA/CẢNH BÁO ({len(errors)} dòng bị lỗi)", 
                               font=('Segoe UI', 11, 'bold'), bootstyle='danger')
        error_title.pack(anchor='w', pady=(0, 5))
        
        txt_frame = tb.Frame(main_frame)
        txt_frame.pack(fill='both', expand=True)
        
        scrollbar = tb.Scrollbar(txt_frame)
        scrollbar.pack(side='right', fill='y')
        
        text_area = tb.Text(txt_frame, yscrollcommand=scrollbar.set, font=('Consolas', 9), wrap='word')
        text_area.pack(fill='both', expand=True, side='left')
        scrollbar.config(command=text_area.yview)
        
        for err in errors:
            text_area.insert('end', f"• {err}\n")
        
        text_area.config(state='disabled')
        
        tb.Button(main_frame, text="Đóng", bootstyle='secondary', command=dialog.destroy).pack(pady=(15, 0), side='right')


# --- LicenseManager: xác minh license offline (Ed25519, không cần PyNaCl) ---
import os, json, base64
from tkinter import simpledialog, messagebox
from cryptography.hazmat.primitives.asymmetric.ed25519 import Ed25519PublicKey
from cryptography.exceptions import InvalidSignature

try:
    LIC_PATH
except NameError:
    APP_DIR = os.path.join(os.environ.get('LOCALAPPDATA', os.path.expanduser('~')), 'Nhathuoc')
    os.makedirs(APP_DIR, exist_ok=True)
    LIC_PATH = os.path.join(APP_DIR, 'license.lic')

PUBLIC_B64 = "cfCth+TwUKiTYPS0gjUoSmwEBBPM6ElQo82e8RdtWl8="  # Public key từ dist/private_key.pem

class LicenseError(Exception): pass

class LicenseManager:
    def __init__(self):
        self.data = None

    def _verify(self, lic_str: str):
        try:
            pack = json.loads(base64.b64decode(lic_str).decode('utf-8'))
            payload = pack["payload"]; sig_b64 = pack["sig"]
        except Exception as e:
            raise LicenseError("Định dạng license không hợp lệ") from e

        blob = json.dumps(payload, separators=(',', ':'), ensure_ascii=False).encode('utf-8')
        sig  = base64.b64decode(sig_b64)
        pub  = base64.b64decode(PUBLIC_B64)

        try:
            Ed25519PublicKey.from_public_bytes(pub).verify(sig, blob)
        except InvalidSignature:
            raise LicenseError("License không hợp lệ hoặc bị chỉnh sửa")

        if payload.get("product") != "nhathuoc":
            raise LicenseError("License sai sản phẩm")

        # Khớp fingerprint (nếu cấp theo máy)
        try:
            hw_now = machine_fingerprint()
        except Exception:
            hw_now = None
        if payload.get("hw") and hw_now and payload["hw"] != hw_now:
            raise LicenseError("License không thuộc máy này")

        self.data = payload
        return payload

    def load(self):
        if not os.path.exists(LIC_PATH):
            raise LicenseError("Chưa kích hoạt")
        lic_str = open(LIC_PATH, 'r', encoding='utf-8').read().strip()
        return self._verify(lic_str)

    def prompt_and_save(self, tkroot=None):
        lic = simpledialog.askstring("Kích hoạt", "Dán license vĩnh viễn cho máy này:")
        if not lic:
            raise LicenseError("Không có license")
        self._verify(lic)
        with open(LIC_PATH, 'w', encoding='utf-8') as f:
            f.write(lic.strip())
        messagebox.showinfo("Kích hoạt", "Kích hoạt thành công! Khởi chạy đầy đủ tính năng.")
        return self.data
# --- /LicenseManager ---

# --- Barcode Scanner ---
class BarcodeScanner:
    def __init__(self, parent_window, callback=None):
        self.parent = parent_window
        self.callback = callback
        self.cap = None
        self.scanning = False
        self.window = None
        
    def start_scan(self):
        """Bắt đầu quét barcode"""
        if not BARCODE_AVAILABLE:
            messagebox.showerror("Lỗi", 
                "Thư viện quét barcode chưa được cài đặt.\n"
                "Vui lòng chạy: pip install opencv-python pyzbar Pillow")
            return
            
        try:
            # Tạo window mới cho camera
            self.window = tk.Toplevel(self.parent)
            self.window.title("📷 Quét Barcode")
            self.window.geometry("640x480")
            self.window.resizable(False, False)
            
            # Tạo frame chính
            main_frame = tb.Frame(self.window)
            main_frame.pack(fill='both', expand=True, padx=10, pady=10)
            
            # Header
            header_frame = tb.Frame(main_frame)
            header_frame.pack(fill='x', pady=(0, 10))
            
            title_label = tb.Label(header_frame, text="📷 Quét Barcode", 
                                  font=('Segoe UI', 14, 'bold'), bootstyle='primary')
            title_label.pack()
            
            subtitle_label = tb.Label(header_frame, text="Đưa barcode vào khung hình để quét", 
                                     font=('Segoe UI', 10), bootstyle='secondary')
            subtitle_label.pack()
            
            # Video frame
            self.video_frame = tb.Frame(main_frame, relief='sunken', borderwidth=2)
            self.video_frame.pack(fill='both', expand=True, pady=(0, 10))
            
            self.video_label = tb.Label(self.video_frame, text="Đang khởi động camera...", 
                                       font=('Segoe UI', 12), bootstyle='info')
            self.video_label.pack(expand=True)
            
            # Control buttons
            button_frame = tb.Frame(main_frame)
            button_frame.pack(fill='x')
            
            self.start_btn = tb.Button(button_frame, text="▶️ Bắt đầu quét", 
                                      command=self.toggle_scan, bootstyle='success')
            self.start_btn.pack(side='left', padx=(0, 10))
            
            self.stop_btn = tb.Button(button_frame, text="⏹️ Dừng", 
                                     command=self.stop_scan, bootstyle='danger')
            self.stop_btn.pack(side='left', padx=(0, 10))
            
            tb.Button(button_frame, text="❌ Đóng", 
                     command=self.close_scanner, bootstyle='secondary').pack(side='right')
            
            # Status label
            self.status_label = tb.Label(main_frame, text="Sẵn sàng", 
                                        font=('Segoe UI', 10), bootstyle='info')
            self.status_label.pack(pady=(10, 0))
            
            # Khởi tạo camera
            self.cap = cv2.VideoCapture(0)
            if not self.cap.isOpened():
                messagebox.showerror("Lỗi", "Không thể mở camera. Vui lòng kiểm tra kết nối.")
                self.close_scanner()
                return
                
            # Cấu hình camera
            self.cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
            self.cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
            
            # Bắt đầu quét
            self.toggle_scan()
            
            # Bind close event
            self.window.protocol("WM_DELETE_WINDOW", self.close_scanner)
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể khởi động camera: {e}")
            
    def toggle_scan(self):
        """Bật/tắt quét"""
        if not self.scanning:
            self.start_scanning()
        else:
            self.stop_scanning()
            
    def start_scanning(self):
        """Bắt đầu quét"""
        self.scanning = True
        self.start_btn.config(text="⏸️ Tạm dừng", bootstyle='warning')
        self.status_label.config(text="Đang quét...", bootstyle='success')
        self.update_frame()
        
    def stop_scanning(self):
        """Tạm dừng quét"""
        self.scanning = False
        self.start_btn.config(text="▶️ Tiếp tục", bootstyle='success')
        self.status_label.config(text="Đã tạm dừng", bootstyle='warning')
        
    def update_frame(self):
        """Cập nhật frame camera"""
        if not self.scanning or not self.cap:
            return
            
        try:
            ret, frame = self.cap.read()
            if ret:
                # Lật frame để hiển thị đúng
                frame = cv2.flip(frame, 1)
                
                # Quét barcode nếu đang quét
                if self.scanning:
                    barcodes = pyzbar.decode(frame)
                    for barcode in barcodes:
                        # Lấy dữ liệu barcode
                        barcode_data = barcode.data.decode('utf-8')
                        barcode_type = barcode.type
                        
                        # Vẽ khung quanh barcode
                        (x, y, w, h) = barcode.rect
                        cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                        
                        # Hiển thị thông tin
                        text = f"{barcode_type}: {barcode_data}"
                        cv2.putText(frame, text, (x, y - 10), 
                                   cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)
                        
                        # Gọi callback với dữ liệu barcode
                        if self.callback:
                            self.callback(barcode_data)
                            
                        # Dừng quét sau khi tìm thấy
                        self.stop_scanning()
                        self.status_label.config(text=f"Đã quét: {barcode_data}", bootstyle='success')
                        break
                
                # Chuyển đổi frame để hiển thị trong Tkinter
                frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                frame_pil = Image.fromarray(frame_rgb)
                frame_tk = ImageTk.PhotoImage(frame_pil)
                
                # Cập nhật label
                self.video_label.config(image=frame_tk, text="")
                self.video_label.image = frame_tk
                
            # Lên lịch cập nhật tiếp theo
            if self.scanning:
                self.window.after(30, self.update_frame)
                
        except Exception as e:
            print(f"Lỗi cập nhật frame: {e}")
            self.stop_scanning()
            
    def stop_scan(self):
        """Dừng quét"""
        self.stop_scanning()
        
    def close_scanner(self):
        """Đóng scanner"""
        self.scanning = False
        if self.cap:
            self.cap.release()
        if self.window:
            self.window.destroy()
        self.window = None
        self.cap = None

# --- /Barcode Scanner ---

# --- Mobile Inventory Web Server ---
import http.server
import socket
import threading
import json
import urllib.parse

QR_CODE_AVAILABLE = False
try:
    import qrcode
    QR_CODE_AVAILABLE = True
except ImportError:
    pass

def get_local_ip():
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(('10.255.255.255', 1))
        IP = s.getsockname()[0]
    except Exception:
        IP = '127.0.0.1'
    finally:
        s.close()
    return IP

class MobileInventoryRequestHandler(http.server.BaseHTTPRequestHandler):
    def log_message(self, format, *args):
        # Mute logging to keep console clean
        pass
        
    def do_GET(self):
        parsed_url = urllib.parse.urlparse(self.path)
        path = parsed_url.path
        query = urllib.parse.parse_qs(parsed_url.query)
        
        if path == "/" or path == "/index.html":
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.end_headers()
            self.wfile.write(MOBILE_HTML.encode("utf-8"))
            
        elif path == "/api/stock":
            barcode = query.get("barcode", [""])[0].strip()
            if not barcode:
                self.send_json({"success": False, "message": "Mã vạch trống"}, 400)
                return
                
            import sqlite3
            try:
                conn = sqlite3.connect(DB_PATH)
                conn.row_factory = sqlite3.Row
                
                # Tìm kiếm sản phẩm theo barcode hoặc tên gần đúng
                product_rows = conn.execute("""
                    SELECT id, name, defaultUnit, barcode, productType, registrationNumber 
                    FROM products 
                    WHERE barcode=? OR name LIKE ? LIMIT 1
                """, (barcode, f"%{barcode}%")).fetchall()
                
                if not product_rows:
                    self.send_json({"success": False, "message": "Không tìm thấy sản phẩm"}, 404)
                    conn.close()
                    return
                    
                p = product_rows[0]
                pid = p['id']
                
                # Lấy thông tin tồn kho chi tiết từng lô
                batches_rows = conn.execute("""
                    SELECT b.lotNo, b.expiryDate, COALESCE(SUM(sm.qty), 0) as qtyBase
                    FROM batches b
                    LEFT JOIN stock_movements sm ON sm.productId = b.productId AND sm.batchId = b.id
                    WHERE b.productId = ?
                    GROUP BY b.id
                    ORDER BY DATE(b.expiryDate) ASC
                """, (pid,)).fetchall()
                
                batches_list = []
                total_qty = 0
                for b in batches_rows:
                    q_val = float(b["qtyBase"])
                    # Chỉ hiện các lô có số lượng khác 0
                    if abs(q_val) > 0.001:
                        batches_list.append({
                            "lotNo": b["lotNo"],
                            "expiryDate": b["expiryDate"],
                            "qty": q_val
                        })
                        total_qty += q_val
                    
                self.send_json({
                    "success": True,
                    "product": {
                        "id": pid,
                        "name": p["name"],
                        "unit": p["defaultUnit"],
                        "barcode": p["barcode"] or "",
                        "type": p["productType"] or "thuoc",
                        "regNumber": p["registrationNumber"] or ""
                    },
                    "batches": batches_list,
                    "totalQty": total_qty
                })
                conn.close()
            except Exception as e:
                self.send_json({"success": False, "message": f"Database error: {str(e)}"}, 500)
            
        elif path == "/api/products":
            q_term = query.get("q", [""])[0].strip()
            filter_type = query.get("filter", [""])[0].strip()
            import sqlite3
            try:
                conn = sqlite3.connect(DB_PATH)
                conn.row_factory = sqlite3.Row
                
                if filter_type == 'expiring':
                    rows = conn.execute("""
                        SELECT DISTINCT p.id, p.name, p.defaultUnit, p.barcode
                        FROM products p
                        JOIN batches b ON p.id = b.productId
                        JOIN stock_movements sm ON sm.productId = b.productId AND sm.batchId = b.id
                        GROUP BY p.id, b.id
                        HAVING SUM(sm.qty) > 0.001 AND DATE(b.expiryDate) <= DATE('now', '+180 days')
                        ORDER BY p.name ASC
                    """).fetchall()
                elif filter_type == 'outofstock':
                    rows = conn.execute("""
                        SELECT p.id, p.name, p.defaultUnit, p.barcode, COALESCE(SUM(sm.qty), 0) as totalQty
                        FROM products p
                        LEFT JOIN stock_movements sm ON sm.productId = p.id
                        GROUP BY p.id
                        HAVING totalQty <= 0.001
                        ORDER BY p.name ASC
                    """).fetchall()
                elif q_term:
                    rows = conn.execute("""
                        SELECT id, name, defaultUnit, barcode 
                        FROM products 
                        WHERE name LIKE ? OR barcode=? LIMIT 50
                    """, (f"%{q_term}%", q_term)).fetchall()
                else:
                    rows = conn.execute("""
                        SELECT id, name, defaultUnit, barcode 
                        FROM products 
                        ORDER BY name ASC LIMIT 50
                    """).fetchall()
                
                products_list = []
                for r in rows:
                    products_list.append({
                        "id": r["id"],
                        "name": r["name"],
                        "unit": r["defaultUnit"],
                        "barcode": r["barcode"] or ""
                    })
                
                self.send_json({"success": True, "products": products_list})
                conn.close()
            except Exception as e:
                self.send_json({"success": False, "message": f"Database error: {str(e)}"}, 500)

        elif path == "/api/dashboard-stats":
            import sqlite3
            try:
                conn = sqlite3.connect(DB_PATH)
                total_products = conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]
                
                outofstock_products = conn.execute("""
                    SELECT COUNT(*) FROM (
                        SELECT p.id, COALESCE(SUM(sm.qty), 0) as totalQty 
                        FROM products p 
                        LEFT JOIN stock_movements sm ON sm.productId = p.id 
                        GROUP BY p.id 
                        HAVING totalQty <= 0.001
                    )
                """).fetchone()[0]
                
                expiring_products = conn.execute("""
                    SELECT COUNT(*) FROM (
                        SELECT DISTINCT p.id
                        FROM products p
                        JOIN batches b ON p.id = b.productId
                        JOIN stock_movements sm ON sm.productId = b.productId AND sm.batchId = b.id
                        GROUP BY p.id, b.id
                        HAVING SUM(sm.qty) > 0.001 AND DATE(b.expiryDate) <= DATE('now', '+180 days')
                    )
                """).fetchone()[0]
                
                conn.close()
                self.send_json({
                    "success": True,
                    "totalProducts": total_products,
                    "outofstockProducts": outofstock_products,
                    "expiringProducts": expiring_products
                })
            except Exception as e:
                self.send_json({"success": False, "message": f"Database error: {str(e)}"}, 500)
            
        elif path == "/api/print-purchase":
            note_id = query.get("id", [""])[0].strip()
            if not note_id:
                self.send_response(400)
                self.end_headers()
                self.wfile.write(b"Yeu cau ID phieu nhap")
                return
            import sqlite3
            try:
                conn = sqlite3.connect(DB_PATH)
                conn.row_factory = sqlite3.Row
                note = conn.execute("SELECT * FROM purchase_notes WHERE id=?", (note_id,)).fetchone()
                if not note:
                    self.send_response(404)
                    self.end_headers()
                    self.wfile.write(b"Khong tim thay phieu nhap")
                    conn.close()
                    return
                items = conn.execute("""
                    SELECT pi.*, p.name as productName 
                    FROM purchase_items pi
                    JOIN products p ON pi.productId = p.id
                    WHERE pi.purchaseId = ?
                """, (note_id,)).fetchall()
                conn.close()
                html = self.render_print_purchase_html(note, items)
                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.end_headers()
                self.wfile.write(html.encode("utf-8"))
            except Exception as e:
                self.send_response(500)
                self.end_headers()
                self.wfile.write(f"Database error: {str(e)}".encode("utf-8"))

        elif path == "/api/print-dispatch":
            note_id = query.get("id", [""])[0].strip()
            if not note_id:
                self.send_response(400)
                self.end_headers()
                self.wfile.write(b"Yeu cau ID phieu xuat")
                return
            import sqlite3
            try:
                conn = sqlite3.connect(DB_PATH)
                conn.row_factory = sqlite3.Row
                note = conn.execute("SELECT * FROM dispatch_notes WHERE id=?", (note_id,)).fetchone()
                if not note:
                    self.send_response(404)
                    self.end_headers()
                    self.wfile.write(b"Khong tim thay phieu xuat")
                    conn.close()
                    return
                items = conn.execute("""
                    SELECT di.*, p.name as productName 
                    FROM dispatch_items di
                    JOIN products p ON di.productId = p.id
                    WHERE di.dispatchId = ?
                """, (note_id,)).fetchall()
                conn.close()
                html = self.render_print_dispatch_html(note, items)
                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.end_headers()
                self.wfile.write(html.encode("utf-8"))
            except Exception as e:
                self.send_response(500)
                self.end_headers()
                self.wfile.write(f"Database error: {str(e)}".encode("utf-8"))

        elif path == "/api/pc-print":
            note_type = query.get("type", [""])[0].strip() # purchase or dispatch
            note_id = query.get("id", [""])[0].strip()
            if note_type == 'nhap':
                note_type = 'purchase'
            elif note_type == 'xuat':
                note_type = 'dispatch'
                
            if not note_type or not note_id:
                self.send_json({"success": False, "message": "Thiếu thông tin loại phiếu hoặc ID"}, 400)
                return
            
            success, msg = self.print_to_pc_printer(note_type, note_id)
            self.send_json({"success": success, "message": msg})

        elif path == "/api/note-details":
            note_type = query.get("type", [""])[0].strip()
            note_id = query.get("id", [""])[0].strip()
            if note_type == 'nhap':
                note_type = 'purchase'
            elif note_type == 'xuat':
                note_type = 'dispatch'
                
            if not note_type or not note_id:
                self.send_json({"success": False, "message": "Thiếu thông tin loại phiếu hoặc ID"}, 400)
                return
            
            import sqlite3
            try:
                conn = sqlite3.connect(DB_PATH)
                conn.row_factory = sqlite3.Row
                
                if note_type == 'purchase':
                    note = conn.execute("SELECT id, noteNumber, supplier as partner, createdAt, reason, note FROM purchase_notes WHERE id=?", (note_id,)).fetchone()
                    if not note:
                        conn.close()
                        self.send_json({"success": False, "message": "Không tìm thấy phiếu nhập"}, 404)
                        return
                    items = conn.execute("""
                        SELECT pi.qty, pi.lotNo, pi.expiryDate, p.name as productName, p.defaultUnit as unit
                        FROM purchase_items pi
                        JOIN products p ON pi.productId = p.id
                        WHERE pi.purchaseId = ?
                    """, (note_id,)).fetchall()
                else:
                    note = conn.execute("SELECT id, noteNumber, receivingUnit as partner, createdAt, reason, note FROM dispatch_notes WHERE id=?", (note_id,)).fetchone()
                    if not note:
                        conn.close()
                        self.send_json({"success": False, "message": "Không tìm thấy phiếu xuất"}, 404)
                        return
                    items = conn.execute("""
                        SELECT di.qty, di.lotNo, p.name as productName, p.defaultUnit as unit
                        FROM dispatch_items di
                        JOIN products p ON di.productId = p.id
                        WHERE di.dispatchId = ?
                    """, (note_id,)).fetchall()
                
                conn.close()
                
                res_items = []
                for it in items:
                    res_items.append({
                        "productName": it["productName"],
                        "qty": it["qty"],
                        "unit": it["unit"],
                        "lotNo": it["lotNo"],
                        "expiryDate": it["expiryDate"] if "expiryDate" in it.keys() else ""
                    })
                
                self.send_json({
                    "success": True,
                    "type": "nhap" if note_type == 'purchase' else "xuat",
                    "noteNumber": note["noteNumber"],
                    "partner": note["partner"],
                    "createdAt": note["createdAt"],
                    "reason": note["reason"],
                    "note": note["note"],
                    "items": res_items
                })
            except Exception as e:
                self.send_json({"success": False, "message": f"Database error: {str(e)}"}, 500)

        elif path == "/api/recent-activities":
            import sqlite3
            try:
                conn = sqlite3.connect(DB_PATH)
                conn.row_factory = sqlite3.Row
                purchases = conn.execute("""
                    SELECT id, noteNumber, supplier as details, createdAt
                    FROM purchase_notes
                    ORDER BY id DESC LIMIT 5
                """).fetchall()
                dispatches = conn.execute("""
                    SELECT id, noteNumber, receivingUnit as details, createdAt
                    FROM dispatch_notes
                    ORDER BY id DESC LIMIT 5
                """).fetchall()
                conn.close()
                
                # Kết hợp và sắp xếp theo ngày tạo giảm dần
                combined = []
                for r in purchases:
                    combined.append({
                        "type": "nhap",
                        "id": r["id"],
                        "noteNumber": r["noteNumber"],
                        "details": r["details"],
                        "createdAt": r["createdAt"]
                    })
                for r in dispatches:
                    combined.append({
                        "type": "xuat",
                        "id": r["id"],
                        "noteNumber": r["noteNumber"],
                        "details": r["details"],
                        "createdAt": r["createdAt"]
                    })
                combined.sort(key=lambda x: x["createdAt"], reverse=True)
                
                self.send_json({"success": True, "activities": combined[:8]})
            except Exception as e:
                self.send_json({"success": False, "message": str(e)}, 500)

        else:
            self.send_response(404)
            self.end_headers()

    def render_print_purchase_html(self, note, items):
        # Format date
        created_str = note['createdAt']
        try:
            from datetime import datetime
            dt_val = datetime.strptime(created_str, '%Y-%m-%d %H:%M:%S')
            date_formatted = dt_val.strftime('%d/%m/%Y %H:%M:%S')
        except Exception:
            date_formatted = created_str

        # Generate table rows
        rows_html = ""
        total_amount = 0.0
        for idx, it in enumerate(items, 1):
            qty = float(it['qty'])
            cost = float(it['cost'])
            amount = qty * cost
            total_amount += amount
            
            # Format cost and amount
            cost_str = f"{cost:,.1f}".replace(".0", "") if cost > 0 else "0"
            amount_str = f"{amount:,.1f}".replace(".0", "") if amount > 0 else "0"
            qty_str = f"{qty:,.2f}".rstrip('0').rstrip('.')
            
            expiry_str = it['expiryDate']
            try:
                from datetime import datetime as dt_parser
                exp_dt = dt_parser.strptime(expiry_str, '%Y-%m-%d')
                expiry_formatted = exp_dt.strftime('%d/%m/%Y')
            except Exception:
                expiry_formatted = expiry_str
                
            rows_html += f"""
            <tr>
                <td style="text-align: center;">{idx}</td>
                <td>{it['productName']}</td>
                <td style="text-align: center;">{it['unitCode']}</td>
                <td style="text-align: right;">{qty_str}</td>
                <td style="text-align: right;">{cost_str}</td>
                <td style="text-align: right;">{amount_str}</td>
                <td style="text-align: center;">{it['lotNo']}</td>
                <td style="text-align: center;">{expiry_formatted}</td>
            </tr>
            """
            
        total_amount_str = f"{total_amount:,.1f}".replace(".0", "") if total_amount > 0 else "0"
        
        html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>Phieu Nhap Kho {note['noteNumber']}</title>
    <style>
        body {{
            font-family: "Times New Roman", Times, serif;
            font-size: 13pt;
            line-height: 1.3;
            margin: 0;
            padding: 20px;
            color: #000;
            background-color: #fff;
        }}
        .header {{
            display: flex;
            justify-content: space-between;
            margin-bottom: 20px;
        }}
        .header-left {{
            font-weight: bold;
            font-size: 11pt;
            text-align: center;
        }}
        .header-right {{
            font-size: 10pt;
            text-align: center;
        }}
        .title-block {{
            text-align: center;
            margin-bottom: 20px;
        }}
        .title {{
            font-size: 16pt;
            font-weight: bold;
            margin: 0;
        }}
        .subtitle {{
            font-size: 12pt;
            font-weight: bold;
            margin: 5px 0 0 0;
        }}
        .info-table {{
            width: 100%;
            margin-bottom: 15px;
            border-collapse: collapse;
        }}
        .info-table td {{
            padding: 4px 0;
            vertical-align: top;
        }}
        .items-table {{
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 20px;
        }}
        .items-table th, .items-table td {{
            border: 1px solid #000;
            padding: 6px 8px;
            font-size: 11pt;
        }}
        .items-table th {{
            font-weight: bold;
            text-align: center;
            background-color: #f2f2f2;
        }}
        .signatures {{
            display: flex;
            justify-content: space-around;
            margin-top: 40px;
            page-break-inside: avoid;
        }}
        .signature-block {{
            text-align: center;
            width: 30%;
        }}
        .signature-title {{
            font-weight: bold;
            margin-bottom: 60px;
        }}
        .no-print-btn {{
            background: #4f46e5;
            color: white;
            border: none;
            padding: 10px 20px;
            border-radius: 6px;
            font-size: 11pt;
            font-weight: bold;
            cursor: pointer;
            margin-bottom: 20px;
            display: inline-flex;
            align-items: center;
            gap: 8px;
        }}
        @media print {{
            .no-print {{
                display: none !important;
            }}
            body {{
                padding: 0;
            }}
        }}
    </style>
</head>
<body>
    <div class="no-print" style="text-align: right;">
        <button class="no-print-btn" onclick="window.print()">
            🖨️ In Phieu / Luu PDF
        </button>
    </div>
    
    <div class="header">
        <div class="header-left">
            SO Y TE THANH PHO CAN THO<br>
            TRUNG TAM KIEM SOAT BENH TAT (CDC)
        </div>
        <div class="header-right">
            <strong>Mau so: C30-HD</strong><br>
            <em>(Ban hanh theo Thong tu so 107/2017/TT-BTC)</em>
        </div>
    </div>
    
    <div class="title-block">
        <h1 class="title">PHIEU NHAP KHO</h1>
        <div class="subtitle">So: {note['noteNumber']}</div>
    </div>
    
    <table class="info-table">
        <tr>
            <td style="width: 180px;"><strong>Nguon cap / Nha CC:</strong></td>
            <td>{note['supplier']}</td>
        </tr>
        <tr>
            <td><strong>Ly do nhap:</strong></td>
            <td>{note['reason']}</td>
        </tr>
        <tr>
            <td><strong>Kho nhap:</strong></td>
            <td>Kho Duoc CDC Can Tho</td>
        </tr>
        <tr>
            <td><strong>Ngay nhap:</strong></td>
            <td>{date_formatted}</td>
        </tr>
        <tr>
            <td><strong>Ghi chu:</strong></td>
            <td>{note['note'] or 'Khong'}</td>
        </tr>
    </table>
    
    <table class="items-table">
        <thead>
            <tr>
                <th style="width: 5%;">STT</th>
                <th>Ten thuoc, vaccine, VTYT</th>
                <th style="width: 8%;">DVT</th>
                <th style="width: 12%;">So luong</th>
                <th style="width: 12%;">Don gia</th>
                <th style="width: 12%;">Thanh tien</th>
                <th style="width: 12%;">So lo</th>
                <th style="width: 12%;">Han dung</th>
            </tr>
        </thead>
        <tbody>
            {rows_html}
            <tr style="font-weight: bold;">
                <td colspan="5" style="text-align: right;">Tong cong:</td>
                <td style="text-align: right;">{total_amount_str}</td>
                <td colspan="2"></td>
            </tr>
        </tbody>
    </table>
    
    <div class="signatures">
        <div class="signature-block">
            <div class="signature-title">Nguoi giao hang</div>
            <div>(Ky, ho ten)</div>
        </div>
        <div class="signature-block">
            <div class="signature-title">Thu kho</div>
            <div>(Ky, ho ten)</div>
        </div>
        <div class="signature-block">
            <div class="signature-title">Nguoi lap phieu</div>
            <div>(Ky, ho ten)</div>
        </div>
    </div>

    <script>
        window.onload = function() {{
            setTimeout(function() {{
                window.print();
            }}, 500);
        }}
    </script>
</body>
</html>
"""
        return html

    def render_print_dispatch_html(self, note, items):
        # Format date
        created_str = note['createdAt']
        try:
            from datetime import datetime
            dt_val = datetime.strptime(created_str, '%Y-%m-%d %H:%M:%S')
            date_formatted = dt_val.strftime('%d/%m/%Y %H:%M:%S')
        except Exception:
            date_formatted = created_str

        # Generate table rows
        rows_html = ""
        for idx, it in enumerate(items, 1):
            qty = float(it['qty'])
            qty_str = f"{qty:,.2f}".rstrip('0').rstrip('.')
            
            expiry_str = it['expiryDate']
            try:
                from datetime import datetime as dt_parser
                exp_dt = dt_parser.strptime(expiry_str, '%Y-%m-%d')
                expiry_formatted = exp_dt.strftime('%d/%m/%Y')
            except Exception:
                expiry_formatted = expiry_str
                
            rows_html += f"""
            <tr>
                <td style="text-align: center;">{idx}</td>
                <td>{it['productName']}</td>
                <td style="text-align: center;">{it['unitCode']}</td>
                <td style="text-align: right;">{qty_str}</td>
                <td style="text-align: center;">{it['lotNo']}</td>
                <td style="text-align: center;">{expiry_formatted}</td>
                <td></td>
            </tr>
            """
            
        html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>Phieu Xuat Kho {note['noteNumber']}</title>
    <style>
        body {{
            font-family: "Times New Roman", Times, serif;
            font-size: 13pt;
            line-height: 1.3;
            margin: 0;
            padding: 20px;
            color: #000;
            background-color: #fff;
        }}
        .header {{
            display: flex;
            justify-content: space-between;
            margin-bottom: 20px;
        }}
        .header-left {{
            font-weight: bold;
            font-size: 11pt;
            text-align: center;
        }}
        .header-right {{
            font-size: 10pt;
            text-align: center;
        }}
        .title-block {{
            text-align: center;
            margin-bottom: 20px;
        }}
        .title {{
            font-size: 16pt;
            font-weight: bold;
            margin: 0;
        }}
        .subtitle {{
            font-size: 12pt;
            font-weight: bold;
            margin: 5px 0 0 0;
        }}
        .info-table {{
            width: 100%;
            margin-bottom: 15px;
            border-collapse: collapse;
        }}
        .info-table td {{
            padding: 4px 0;
            vertical-align: top;
        }}
        .items-table {{
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 20px;
        }}
        .items-table th, .items-table td {{
            border: 1px solid #000;
            padding: 6px 8px;
            font-size: 11pt;
        }}
        .items-table th {{
            font-weight: bold;
            text-align: center;
            background-color: #f2f2f2;
        }}
        .signatures {{
            display: flex;
            justify-content: space-around;
            margin-top: 40px;
            page-break-inside: avoid;
        }}
        .signature-block {{
            text-align: center;
            width: 30%;
        }}
        .signature-title {{
            font-weight: bold;
            margin-bottom: 60px;
        }}
        .no-print-btn {{
            background: #4f46e5;
            color: white;
            border: none;
            padding: 10px 20px;
            border-radius: 6px;
            font-size: 11pt;
            font-weight: bold;
            cursor: pointer;
            margin-bottom: 20px;
            display: inline-flex;
            align-items: center;
            gap: 8px;
        }}
        @media print {{
            .no-print {{
                display: none !important;
            }}
            body {{
                padding: 0;
            }}
        }}
    </style>
</head>
<body>
    <div class="no-print" style="text-align: right;">
        <button class="no-print-btn" onclick="window.print()">
            🖨️ In Phieu / Luu PDF
        </button>
    </div>
    
    <div class="header">
        <div class="header-left">
            SO Y TE THANH PHO CAN THO<br>
            TRUNG TAM KIEM SOAT BENH TAT (CDC)
        </div>
        <div class="header-right">
            <strong>Mau so: C31-HD</strong><br>
            <em>(Ban hanh theo Thong tu so 107/2017/TT-BTC)</em>
        </div>
    </div>
    
    <div class="title-block">
        <h1 class="title">PHIEU XUAT KHO</h1>
        <div class="subtitle">So: {note['noteNumber']}</div>
    </div>
    
    <table class="info-table">
        <tr>
            <td style="width: 180px;"><strong>Don vi nhan:</strong></td>
            <td>{note['receivingUnit']}</td>
        </tr>
        <tr>
            <td><strong>Ly do xuat:</strong></td>
            <td>{note['reason']}</td>
        </tr>
        <tr>
            <td><strong>Kho xuat:</strong></td>
            <td>Kho Duoc CDC Can Tho</td>
        </tr>
        <tr>
            <td><strong>Ngay xuat:</strong></td>
            <td>{date_formatted}</td>
        </tr>
        <tr>
            <td><strong>Ghi chu:</strong></td>
            <td>{note['note'] or 'Khong'}</td>
        </tr>
    </table>
    
    <table class="items-table">
        <thead>
            <tr>
                <th style="width: 5%;">STT</th>
                <th>Ten thuoc, vaccine, VTYT</th>
                <th style="width: 10%;">DVT</th>
                <th style="width: 15%;">So luong</th>
                <th style="width: 15%;">So lo</th>
                <th style="width: 15%;">Han dung</th>
                <th style="width: 15%;">Ghi chu</th>
            </tr>
        </thead>
        <tbody>
            {rows_html}
        </tbody>
    </table>
    
    <div class="signatures">
        <div class="signature-block">
            <div class="signature-title">Nguoi nhan hang</div>
            <div>(Ky, ho ten)</div>
        </div>
        <div class="signature-block">
            <div class="signature-title">Thu kho</div>
            <div>(Ky, ho ten)</div>
        </div>
        <div class="signature-block">
            <div class="signature-title">Nguoi lap phieu</div>
            <div>(Ky, ho ten)</div>
        </div>
    </div>

    <script>
        window.onload = function() {{
            setTimeout(function() {{
                window.print();
            }}, 500);
        }}
    </script>
</body>
</html>
"""
        return html

    def print_to_pc_printer(self, note_type, note_id):
        import sqlite3
        import tempfile
        import os
        from datetime import datetime
        
        try:
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row
            
            if note_type == 'purchase':
                note = conn.execute("SELECT * FROM purchase_notes WHERE id=?", (note_id,)).fetchone()
                if not note:
                    conn.close()
                    return False, "Không tìm thấy phiếu nhập"
                items = conn.execute("""
                    SELECT pi.*, p.name as productName 
                    FROM purchase_items pi
                    JOIN products p ON pi.productId = p.id
                    WHERE pi.purchaseId = ?
                """, (note_id,)).fetchall()
            else:
                note = conn.execute("SELECT * FROM dispatch_notes WHERE id=?", (note_id,)).fetchone()
                if not note:
                    conn.close()
                    return False, "Không tìm thấy phiếu xuất"
                items = conn.execute("""
                    SELECT di.*, p.name as productName 
                    FROM dispatch_items di
                    JOIN products p ON di.productId = p.id
                    WHERE di.dispatchId = ?
                """, (note_id,)).fetchall()
            conn.close()
            
            # Tạo đường dẫn lưu PDF tạm
            temp_dir = tempfile.gettempdir()
            filename = f"Phieu_{'Nhap' if note_type == 'purchase' else 'Xuat'}_Kho_{note['noteNumber']}.pdf"
            pdf_path = os.path.join(temp_dir, filename)
            
            # Tạo file PDF với ReportLab
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            import datetime as dt_module
            
            try:
                pdfmetrics.registerFont(TTFont('TimesNewRoman', "C:\\Windows\\Fonts\\times.ttf"))
                pdfmetrics.registerFont(TTFont('TimesNewRoman-Bold', "C:\\Windows\\Fonts\\timesbd.ttf"))
                pdfmetrics.registerFont(TTFont('TimesNewRoman-Italic', "C:\\Windows\\Fonts\\timesi.ttf"))
                font_normal = 'TimesNewRoman'
                font_bold = 'TimesNewRoman-Bold'
                font_italic = 'TimesNewRoman-Italic'
            except Exception:
                font_normal = 'Helvetica'
                font_bold = 'Helvetica-Bold'
                font_italic = 'Helvetica-Oblique'
                
            doc = SimpleDocTemplate(pdf_path, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
            story = []
            
            styles = getSampleStyleSheet()
            
            style_header_left = ParagraphStyle(
                'HeaderLeft', parent=styles['Normal'], fontName=font_bold, fontSize=10, leading=14, alignment=0
            )
            style_header_right = ParagraphStyle(
                'HeaderRight', parent=styles['Normal'], fontName=font_normal, fontSize=10, leading=14, alignment=2
            )
            style_title = ParagraphStyle(
                'Title', parent=styles['Heading1'], fontName=font_bold, fontSize=16, leading=20, alignment=1, spaceAfter=5
            )
            style_subtitle = ParagraphStyle(
                'Subtitle', parent=styles['Normal'], fontName=font_bold, fontSize=11, leading=14, alignment=1, spaceAfter=15
            )
            style_info = ParagraphStyle(
                'Info', parent=styles['Normal'], fontName=font_normal, fontSize=11, leading=16, alignment=0
            )
            style_table_header = ParagraphStyle(
                'TableHeader', parent=styles['Normal'], fontName=font_bold, fontSize=9, leading=11, alignment=1, textColor=colors.black
            )
            style_cell = ParagraphStyle(
                'Cell', parent=styles['Normal'], fontName=font_normal, fontSize=9, leading=11, alignment=0
            )
            style_cell_center = ParagraphStyle(
                'CellCenter', parent=styles['Normal'], fontName=font_normal, fontSize=9, leading=11, alignment=1
            )
            style_cell_right = ParagraphStyle(
                'CellRight', parent=styles['Normal'], fontName=font_normal, fontSize=9, leading=11, alignment=2
            )
            
            # Header
            header_data = [
                [
                    Paragraph("SỞ Y TẾ THÀNH PHỐ CẦN THƠ<br/>TRUNG TÂM KIỂM SOÁT BỆNH TẬT (CDC)", style_header_left),
                    Paragraph(f"<b>Mẫu số: {'C30-HD' if note_type == 'purchase' else 'C31-HD'}</b><br/><i>(Ban hành theo Thông tư số 107/2017/TT-BTC)</i>", style_header_right)
                ]
            ]
            header_table = Table(header_data, colWidths=[280, 230])
            header_table.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
                ('BOTTOMPADDING', (0,0), (-1,-1), 10),
            ]))
            story.append(header_table)
            story.append(Spacer(1, 10))
            
            # Tiêu đề phiếu
            story.append(Paragraph("PHIẾU NHẬP KHO" if note_type == 'purchase' else "PHIẾU XUẤT KHO", style_title))
            story.append(Paragraph(f"Số: {note['noteNumber']}", style_subtitle))
            
            created_str = note['createdAt']
            try:
                created_at_dt = dt_module.datetime.strptime(created_str, '%Y-%m-%d %H:%M:%S')
            except Exception:
                try:
                    created_at_dt = dt_module.datetime.strptime(created_str.split(' ')[0], '%Y-%m-%d')
                except Exception:
                    created_at_dt = dt_module.datetime.now()
            
            if note_type == 'purchase':
                info_lines = [
                    f"<b>Nguồn cấp / Nhà CC:</b> {note['supplier']}",
                    f"<b>Lý do nhập:</b> {note['reason']}",
                    f"<b>Kho nhập:</b> Kho Dược CDC Cần Thơ",
                    f"<b>Ngày nhập:</b> {created_at_dt.strftime('%d/%m/%Y')}",
                    f"<b>Ghi chú:</b> {note['note'] or 'Không'}"
                ]
            else:
                info_lines = [
                    f"<b>Đơn vị nhận:</b> {note['receivingUnit']}",
                    f"<b>Lý do xuất:</b> {note['reason']}",
                    f"<b>Kho xuất:</b> Kho Dược CDC Cần Thơ",
                    f"<b>Ngày xuất:</b> {created_at_dt.strftime('%d/%m/%Y')}",
                    f"<b>Ghi chú:</b> {note['note'] or 'Không'}"
                ]
            for line in info_lines:
                story.append(Paragraph(line, style_info))
                story.append(Spacer(1, 4))
                
            story.append(Spacer(1, 10))
            
            # Dữ liệu bảng
            if note_type == 'purchase':
                table_data = [
                    [
                        Paragraph("STT", style_table_header),
                        Paragraph("Tên thuốc, vaccine, VTYT", style_table_header),
                        Paragraph("ĐVT", style_table_header),
                        Paragraph("Số lượng", style_table_header),
                        Paragraph("Đơn giá", style_table_header),
                        Paragraph("Thành tiền", style_table_header),
                        Paragraph("Số lô", style_table_header),
                        Paragraph("Hạn dùng", style_table_header)
                    ]
                ]
                total_sum = 0.0
                for idx, it in enumerate(items, 1):
                    qty = float(it['qty'])
                    cost = float(it['cost'])
                    sub_total = qty * cost
                    total_sum += sub_total
                    table_data.append([
                        Paragraph(str(idx), style_cell_center),
                        Paragraph(it['productName'], style_cell),
                        Paragraph(it['unitCode'], style_cell_center),
                        Paragraph(f"{qty:g}", style_cell_right),
                        Paragraph(f"{cost:,.0f}", style_cell_right),
                        Paragraph(f"{sub_total:,.0f}", style_cell_right),
                        Paragraph(it['lotNo'] or '', style_cell_center),
                        Paragraph(it['expiryDate'] or '', style_cell_center)
                    ])
                table_data.append([
                    Paragraph("<b>Tổng cộng</b>", style_cell_center),
                    Paragraph("", style_cell),
                    Paragraph("", style_cell_center),
                    Paragraph("", style_cell_right),
                    Paragraph("", style_cell_right),
                    Paragraph(f"<b>{total_sum:,.0f}</b>", style_cell_right),
                    Paragraph("", style_cell_center),
                    Paragraph("", style_cell_center)
                ])
                col_widths = [25, 160, 45, 55, 65, 75, 55, 50]
                items_table = Table(table_data, colWidths=col_widths)
                items_table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f2f2f2')),
                    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.black),
                    ('SPAN', (0, -1), (4, -1)),
                    ('TOPPADDING', (0,0), (-1,-1), 5),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 5),
                ]))
            else:
                table_data = [
                    [
                        Paragraph("STT", style_table_header),
                        Paragraph("Tên thuốc, vaccine, VTYT", style_table_header),
                        Paragraph("ĐVT", style_table_header),
                        Paragraph("Số lượng", style_table_header),
                        Paragraph("Số lô", style_table_header),
                        Paragraph("Hạn dùng", style_table_header),
                        Paragraph("Ghi chú", style_table_header)
                    ]
                ]
                for idx, it in enumerate(items, 1):
                    qty = float(it['qty'])
                    table_data.append([
                        Paragraph(str(idx), style_cell_center),
                        Paragraph(it['productName'], style_cell),
                        Paragraph(it['unitCode'], style_cell_center),
                        Paragraph(f"{qty:g}", style_cell_right),
                        Paragraph(it['lotNo'] or '', style_cell_center),
                        Paragraph(it['expiryDate'] or '', style_cell_center),
                        Paragraph('', style_cell)
                    ])
                col_widths = [30, 200, 50, 60, 70, 70, 50]
                items_table = Table(table_data, colWidths=col_widths)
                items_table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f2f2f2')),
                    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.black),
                    ('TOPPADDING', (0,0), (-1,-1), 5),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 5),
                ]))
                
            story.append(items_table)
            story.append(Spacer(1, 15))
            
            # Chữ ký
            date_right_style = ParagraphStyle(
                'DateRight', parent=styles['Normal'], fontName=font_italic, fontSize=11, alignment=2, spaceAfter=10
            )
            sig_title_style = ParagraphStyle(
                'SigTitle', parent=styles['Normal'], fontName=font_bold, fontSize=11, alignment=1
            )
            sig_sub_style = ParagraphStyle(
                'SigSub', parent=styles['Normal'], fontName=font_italic, fontSize=9, alignment=1
            )
            
            story.append(Paragraph(f"Cần Thơ, ngày {created_at_dt.strftime('%d')} tháng {created_at_dt.strftime('%m')} năm {created_at_dt.strftime('%Y')}", date_right_style))
            
            if note_type == 'purchase':
                sig_headers = [
                    [
                        Paragraph("<b>Người lập phiếu</b>", sig_title_style),
                        Paragraph("<b>Người giao hàng</b>", sig_title_style),
                        Paragraph("<b>Thủ kho</b>", sig_title_style),
                        Paragraph("<b>Kế toán trưởng</b>", sig_title_style),
                        Paragraph("<b>Lãnh đạo đơn vị</b>", sig_title_style)
                    ],
                    [
                        Paragraph("(Ký, họ tên)", sig_sub_style),
                        Paragraph("(Ký, họ tên)", sig_sub_style),
                        Paragraph("(Ký, họ tên)", sig_sub_style),
                        Paragraph("(Ký, họ tên)", sig_sub_style),
                        Paragraph("(Ký, đóng dấu)", sig_sub_style)
                    ]
                ]
            else:
                sig_headers = [
                    [
                        Paragraph("<b>Người lập phiếu</b>", sig_title_style),
                        Paragraph("<b>Người nhận hàng</b>", sig_title_style),
                        Paragraph("<b>Thủ kho</b>", sig_title_style),
                        Paragraph("<b>Kế toán trưởng</b>", sig_title_style),
                        Paragraph("<b>Lãnh đạo đơn vị</b>", sig_title_style)
                    ],
                    [
                        Paragraph("(Ký, họ tên)", sig_sub_style),
                        Paragraph("(Ký, họ tên)", sig_sub_style),
                        Paragraph("(Ký, họ tên)", sig_sub_style),
                        Paragraph("(Ký, họ tên)", sig_sub_style),
                        Paragraph("(Ký, đóng dấu)", sig_sub_style)
                    ]
                ]
                
            sig_table = Table(sig_headers, colWidths=[102, 102, 102, 102, 102])
            sig_table.setStyle(TableStyle([
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('BOTTOMPADDING', (0,0), (-1,-1), 2),
            ]))
            story.append(sig_table)
            story.append(Spacer(1, 60))
            
            doc.build(story)
            
            # Gửi lệnh in hệ thống trên Windows
            try:
                os.startfile(pdf_path, 'print')
                return True, "Đã gửi lệnh in đến máy in trên máy tính"
            except Exception as pe:
                os.startfile(pdf_path)
                return True, f"Đã tạo PDF và mở trên máy tính: {str(pe)}"
        except Exception as e:
            return False, f"Lỗi tạo phiếu in: {str(e)}"

    def do_POST(self):
        parsed_url = urllib.parse.urlparse(self.path)
        path = parsed_url.path
        
        # Đọc dữ liệu gửi lên
        content_length = int(self.headers.get('Content-Length', 0))
        post_data = self.rfile.read(content_length)
        
        try:
            data = json.loads(post_data.decode('utf-8')) if post_data else {}
        except Exception:
            self.send_json({"success": False, "message": "Dữ liệu JSON không hợp lệ"}, 400)
            return
            
        import sqlite3
        from datetime import datetime
        
        if path == "/api/create-product":
            name = data.get("name", "").strip()
            default_unit = data.get("defaultUnit", "").strip()
            barcode = data.get("barcode", "").strip() or None
            product_type = data.get("productType", "thuoc").strip()
            reg_number = data.get("registrationNumber", "").strip() or None
            
            if not name or not default_unit:
                self.send_json({"success": False, "message": "Tên sản phẩm và Đơn vị tính không được để trống"}, 400)
                return
                
            try:
                conn = sqlite3.connect(DB_PATH)
                conn.row_factory = sqlite3.Row
                
                # Kiểm tra xem mã vạch này đã tồn tại chưa
                if barcode:
                    existing = conn.execute("SELECT id, name FROM products WHERE barcode=?", (barcode,)).fetchone()
                    if existing:
                        self.send_json({"success": False, "message": f"Mã vạch này đã được sử dụng cho sản phẩm: {existing['name']}"}, 400)
                        conn.close()
                        return
                
                now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                with conn:
                    cur = conn.execute("""
                        INSERT INTO products (name, defaultUnit, barcode, productType, registrationNumber, createdAt)
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (name, default_unit, barcode, product_type, reg_number, now_str))
                    product_id = cur.lastrowid
                
                conn.close()
                
                # Cập nhật UI trên Main Thread
                if hasattr(self.server, 'app_instance') and self.server.app_instance:
                    self.server.app_instance.after(0, self.server.app_instance.refresh_all_data)
                    
                self.send_json({
                    "success": True, 
                    "message": "Đã tạo sản phẩm mới thành công!", 
                    "productId": product_id,
                    "barcode": barcode or str(product_id)
                })
            except Exception as e:
                self.send_json({"success": False, "message": f"Lỗi cơ sở dữ liệu: {str(e)}"}, 500)
                
        elif path == "/api/purchase":
            items = data.get("items")
            supplier = data.get("supplier", "Nhập kho di động").strip() or "Nhập kho di động"
            reason = data.get("reason", "Nhập qua điện thoại").strip() or "Nhập qua điện thoại"
            note = data.get("note", "Tạo tự động từ điện thoại").strip() or "Tạo tự động từ điện thoại"

            if items is not None:
                if not isinstance(items, list) or len(items) == 0:
                    self.send_json({"success": False, "message": "Danh sách mặt hàng nhập trống"}, 400)
                    return
                for item in items:
                    p_id = item.get("productId")
                    qty = item.get("qty")
                    lot_no = str(item.get("lotNo", "")).strip()
                    expiry_date = str(item.get("expiryDate", "")).strip()
                    if not p_id or not qty or not lot_no or not expiry_date:
                        self.send_json({"success": False, "message": "Thông tin mặt hàng nhập không đầy đủ"}, 400)
                        return
                    try:
                        item["qty"] = float(qty)
                        if item["qty"] <= 0: raise ValueError()
                    except ValueError:
                        self.send_json({"success": False, "message": "Số lượng phải là số dương lớn hơn 0"}, 400)
                        return
                
                try:
                    conn = sqlite3.connect(DB_PATH)
                    conn.row_factory = sqlite3.Row
                    conn.execute('PRAGMA foreign_keys = ON')
                    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    note_num = f"PNK-MOB-{datetime.now().strftime('%Y%m%d%H%M%S')}"
                    
                    with conn:
                        cur = conn.execute("""
                            INSERT INTO purchase_notes(noteNumber, supplier, reason, note, createdAt)
                            VALUES(?, ?, ?, ?, ?)
                        """, (note_num, supplier, reason, note, now_str))
                        purchase_id = cur.lastrowid
                        
                        for item in items:
                            p_id = item["productId"]
                            qty = item["qty"]
                            lot_no = str(item["lotNo"]).strip()
                            expiry_date = str(item["expiryDate"]).strip()
                            
                            prod = conn.execute("SELECT defaultUnit FROM products WHERE id=?", (p_id,)).fetchone()
                            if not prod:
                                raise Exception(f"Không tìm thấy sản phẩm ID {p_id}")
                            unit = prod['defaultUnit']
                            
                            batch = conn.execute("SELECT id FROM batches WHERE productId=? AND lotNo=?", (p_id, lot_no)).fetchone()
                            if batch:
                                batch_id = batch['id']
                            else:
                                cur_b = conn.execute("INSERT INTO batches(productId, lotNo, expiryDate) VALUES(?,?,?)", (p_id, lot_no, expiry_date))
                                batch_id = cur_b.lastrowid
                                
                            conn.execute("""
                                INSERT INTO purchase_items(purchaseId, productId, batchId, unitCode, qty, lotNo, expiryDate, cost)
                                VALUES(?, ?, ?, ?, ?, ?, ?, ?)
                            """, (purchase_id, p_id, batch_id, unit, qty, lot_no, expiry_date, 0.0))
                            
                            conn.execute("""
                                INSERT INTO stock_movements(productId, batchId, unitCode, qty, type, cost, reason, createdAt)
                                VALUES(?, ?, ?, ?, ?, ?, ?, ?)
                            """, (p_id, batch_id, unit, qty, 'PURCHASE', 0.0, reason, now_str))
                            
                    conn.close()
                    if hasattr(self.server, 'app_instance') and self.server.app_instance:
                        self.server.app_instance.after(0, self.server.app_instance.refresh_all_data)
                        
                    self.send_json({
                        "success": True, 
                        "message": f"Đã tạo phiếu nhập {note_num} với {len(items)} mặt hàng",
                        "purchaseId": purchase_id,
                        "noteNumber": note_num
                    })
                except Exception as e:
                    self.send_json({"success": False, "message": f"Lỗi cơ sở dữ liệu: {str(e)}"}, 500)
                return

            # Single-item backward-compatible logic
            product_id = data.get("productId")
            qty = data.get("qty")
            lot_no = data.get("lotNo", "").strip()
            expiry_date = data.get("expiryDate", "").strip()
            
            if not product_id or not qty or not lot_no or not expiry_date:
                self.send_json({"success": False, "message": "Vui lòng nhập đầy đủ thông tin"}, 400)
                return
                
            try:
                qty = float(qty)
                if qty <= 0: raise ValueError()
            except ValueError:
                self.send_json({"success": False, "message": "Số lượng phải là số dương lớn hơn 0"}, 400)
                return
                
            try:
                conn = sqlite3.connect(DB_PATH)
                conn.row_factory = sqlite3.Row
                conn.execute('PRAGMA foreign_keys = ON')
                now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                note_num = f"PNK-MOB-{datetime.now().strftime('%Y%m%d%H%M%S')}"
                
                with conn:
                    prod = conn.execute("SELECT name, defaultUnit FROM products WHERE id=?", (product_id,)).fetchone()
                    if not prod:
                        self.send_json({"success": False, "message": "Không tìm thấy sản phẩm"}, 404)
                        conn.close()
                        return
                    unit = prod['defaultUnit']
                    
                    batch = conn.execute("SELECT id FROM batches WHERE productId=? AND lotNo=?", (product_id, lot_no)).fetchone()
                    if batch:
                        batch_id = batch['id']
                    else:
                        cur = conn.execute("INSERT INTO batches(productId, lotNo, expiryDate) VALUES(?,?,?)", (product_id, lot_no, expiry_date))
                        batch_id = cur.lastrowid
                        
                    cur = conn.execute("""
                        INSERT INTO purchase_notes(noteNumber, supplier, reason, note, createdAt)
                        VALUES(?, ?, ?, ?, ?)
                    """, (note_num, supplier, reason, note, now_str))
                    purchase_id = cur.lastrowid
                    
                    conn.execute("""
                        INSERT INTO purchase_items(purchaseId, productId, batchId, unitCode, qty, lotNo, expiryDate, cost)
                        VALUES(?, ?, ?, ?, ?, ?, ?, ?)
                    """, (purchase_id, product_id, batch_id, unit, qty, lot_no, expiry_date, 0.0))
                    
                    conn.execute("""
                        INSERT INTO stock_movements(productId, batchId, unitCode, qty, type, cost, reason, createdAt)
                        VALUES(?, ?, ?, ?, ?, ?, ?, ?)
                    """, (product_id, batch_id, unit, qty, 'PURCHASE', 0.0, reason, now_str))
                    
                conn.close()
                if hasattr(self.server, 'app_instance') and self.server.app_instance:
                    self.server.app_instance.after(0, self.server.app_instance.refresh_all_data)
                    
                self.send_json({
                    "success": True, 
                    "message": f"Đã nhập thành công {qty} {unit} vào lô {lot_no}",
                    "purchaseId": purchase_id,
                    "noteNumber": note_num
                })
            except Exception as e:
                self.send_json({"success": False, "message": f"Lỗi cơ sở dữ liệu: {str(e)}"}, 500)
                
        elif path == "/api/dispatch":
            items = data.get("items")
            receiving_unit = data.get("receivingUnit", "Điện thoại di động").strip() or "Điện thoại di động"
            reason_str = data.get("reason", "Xuất qua điện thoại").strip() or "Xuất qua điện thoại"
            note = data.get("note", "Tạo tự động từ điện thoại").strip() or "Tạo tự động từ điện thoại"

            if items is not None:
                if not isinstance(items, list) or len(items) == 0:
                    self.send_json({"success": False, "message": "Danh sách mặt hàng xuất trống"}, 400)
                    return
                for item in items:
                    p_id = item.get("productId")
                    qty = item.get("qty")
                    lot_no = str(item.get("lotNo", "")).strip()
                    if not p_id or not qty or not lot_no:
                        self.send_json({"success": False, "message": "Thông tin mặt hàng xuất không đầy đủ"}, 400)
                        return
                    try:
                        item["qty"] = float(qty)
                        if item["qty"] <= 0: raise ValueError()
                    except ValueError:
                        self.send_json({"success": False, "message": "Số lượng phải là số dương lớn hơn 0"}, 400)
                        return
                
                try:
                    conn = sqlite3.connect(DB_PATH)
                    conn.row_factory = sqlite3.Row
                    conn.execute('PRAGMA foreign_keys = ON')
                    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    note_num = f"PXK-MOB-{datetime.now().strftime('%Y%m%d%H%M%S')}"
                    
                    with conn:
                        for item in items:
                            p_id = item["productId"]
                            qty = item["qty"]
                            lot_no = item["lotNo"]
                            
                            prod = conn.execute("SELECT name, defaultUnit FROM products WHERE id=?", (p_id,)).fetchone()
                            if not prod:
                                raise Exception(f"Không tìm thấy sản phẩm ID {p_id}")
                            unit = prod['defaultUnit']
                            
                            batch = conn.execute("SELECT id, expiryDate FROM batches WHERE productId=? AND lotNo=?", (p_id, lot_no)).fetchone()
                            if not batch:
                                self.send_json({"success": False, "message": f"Không tìm thấy lô '{lot_no}' của sản phẩm '{prod['name']}'"}, 400)
                                conn.close()
                                return
                            batch_id = batch['id']
                            item["batchId"] = batch_id
                            item["expiryDate"] = batch['expiryDate']
                            item["unit"] = unit
                            
                            existing_qty = conn.execute("""
                                SELECT COALESCE(SUM(qty), 0) as qtyBase
                                FROM stock_movements
                                WHERE productId=? AND batchId=?
                            """, (p_id, batch_id)).fetchone()
                            
                            current_stock = float(existing_qty['qtyBase']) if existing_qty else 0.0
                            if current_stock < qty:
                                self.send_json({"success": False, "message": f"Mặt hàng '{prod['name']}' lô '{lot_no}' không đủ tồn kho (cần {qty}, có {current_stock} {unit})"}, 400)
                                conn.close()
                                return
                        
                        cur = conn.execute("""
                            INSERT INTO dispatch_notes(noteNumber, receivingUnit, reason, note, createdAt)
                            VALUES(?, ?, ?, ?, ?)
                        """, (note_num, receiving_unit, reason_str, note, now_str))
                        dispatch_id = cur.lastrowid
                        
                        for item in items:
                            p_id = item["productId"]
                            qty = item["qty"]
                            lot_no = item["lotNo"]
                            batch_id = item["batchId"]
                            expiry_date = item["expiryDate"]
                            unit = item["unit"]
                            
                            conn.execute("""
                                INSERT INTO dispatch_items(dispatchId, productId, batchId, unitCode, qty, lotNo, expiryDate)
                                VALUES(?, ?, ?, ?, ?, ?, ?)
                            """, (dispatch_id, p_id, batch_id, unit, qty, lot_no, expiry_date))
                            
                            conn.execute("""
                                INSERT INTO stock_movements(productId, batchId, unitCode, qty, type, reason, createdAt)
                                VALUES(?, ?, ?, ?, ?, ?, ?)
                            """, (p_id, batch_id, unit, -qty, 'DISPATCH', reason_str, now_str))
                            
                    conn.close()
                    if hasattr(self.server, 'app_instance') and self.server.app_instance:
                        self.server.app_instance.after(0, self.server.app_instance.refresh_all_data)
                        
                    self.send_json({
                        "success": True, 
                        "message": f"Đã tạo phiếu xuất {note_num} với {len(items)} mặt hàng",
                        "dispatchId": dispatch_id,
                        "noteNumber": note_num
                    })
                except Exception as e:
                    self.send_json({"success": False, "message": f"Lỗi cơ sở dữ liệu: {str(e)}"}, 500)
                return

            # Single-item backward-compatible logic
            product_id = data.get("productId")
            lot_no = data.get("lotNo", "").strip()
            qty = data.get("qty")
            
            if not product_id or not lot_no or not qty:
                self.send_json({"success": False, "message": "Vui lòng nhập đầy đủ thông tin"}, 400)
                return
                
            try:
                qty = float(qty)
                if qty <= 0: raise ValueError()
            except ValueError:
                self.send_json({"success": False, "message": "Số lượng phải là số dương lớn hơn 0"}, 400)
                return
                
            try:
                conn = sqlite3.connect(DB_PATH)
                conn.row_factory = sqlite3.Row
                conn.execute('PRAGMA foreign_keys = ON')
                
                now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                note_num = f"PXK-MOB-{datetime.now().strftime('%Y%m%d%H%M%S')}"
                
                with conn:
                    prod = conn.execute("SELECT name, defaultUnit FROM products WHERE id=?", (product_id,)).fetchone()
                    if not prod:
                        self.send_json({"success": False, "message": "Không tìm thấy sản phẩm"}, 404)
                        conn.close()
                        return
                    unit = prod['defaultUnit']
                    
                    batch = conn.execute("SELECT id, lotNo, expiryDate FROM batches WHERE productId=? AND lotNo=?", (product_id, lot_no)).fetchone()
                    if not batch:
                        self.send_json({"success": False, "message": f"Không tìm thấy lô hàng '{lot_no}' của sản phẩm này"}, 404)
                        conn.close()
                        return
                    batch_id = batch['id']
                    expiry_date = batch['expiryDate']
                    
                    existing_qty = conn.execute("""
                        SELECT COALESCE(SUM(qty), 0) as qtyBase
                        FROM stock_movements
                        WHERE productId=? AND batchId=?
                    """, (product_id, batch_id)).fetchone()
                    
                    current_stock = float(existing_qty['qtyBase']) if existing_qty else 0.0
                    if current_stock < qty:
                        self.send_json({"success": False, "message": f"Số lượng xuất ({qty}) vượt quá số lượng tồn của lô ({current_stock} {unit})"}, 400)
                        conn.close()
                        return
                        
                    cur = conn.execute("""
                        INSERT INTO dispatch_notes(noteNumber, receivingUnit, reason, note, createdAt)
                        VALUES(?, ?, ?, ?, ?)
                    """, (note_num, receiving_unit, reason_str, note, now_str))
                    dispatch_id = cur.lastrowid
                    
                    conn.execute("""
                        INSERT INTO dispatch_items(dispatchId, productId, batchId, unitCode, qty, lotNo, expiryDate)
                        VALUES(?, ?, ?, ?, ?, ?, ?)
                    """, (dispatch_id, product_id, batch_id, unit, qty, lot_no, expiry_date))
                    
                    conn.execute("""
                        INSERT INTO stock_movements(productId, batchId, unitCode, qty, type, reason, createdAt)
                        VALUES(?, ?, ?, ?, ?, ?, ?)
                    """, (product_id, batch_id, unit, -qty, 'DISPATCH', reason_str, now_str))
                    
                conn.close()
                if hasattr(self.server, 'app_instance') and self.server.app_instance:
                    self.server.app_instance.after(0, self.server.app_instance.refresh_all_data)
                    
                self.send_json({
                    "success": True, 
                    "message": f"Đã xuất thành công {qty} {unit} từ lô {lot_no}",
                    "dispatchId": dispatch_id,
                    "noteNumber": note_num
                })
            except Exception as e:
                self.send_json({"success": False, "message": f"Lỗi cơ sở dữ liệu: {str(e)}"}, 500)
                
        elif path == "/api/update-barcode":
            product_id = data.get("productId")
            barcode = data.get("barcode", "").strip()
            
            if not product_id or not barcode:
                self.send_json({"success": False, "message": "Vui lòng nhập đầy đủ thông tin"}, 400)
                return
                
            try:
                conn = sqlite3.connect(DB_PATH)
                conn.row_factory = sqlite3.Row
                
                # Kiểm tra xem mã vạch này đã được sử dụng bởi sản phẩm khác chưa
                existing = conn.execute("SELECT id, name FROM products WHERE barcode=? AND id<>?", (barcode, product_id)).fetchone()
                if existing:
                    self.send_json({"success": False, "message": f"Mã vạch này đã được sử dụng cho sản phẩm: {existing['name']}"}, 400)
                    conn.close()
                    return
                    
                with conn:
                    conn.execute("UPDATE products SET barcode=? WHERE id=?", (barcode, product_id))
                    
                conn.close()
                
                # Cập nhật UI trên Main Thread
                if hasattr(self.server, 'app_instance') and self.server.app_instance:
                    self.server.app_instance.after(0, self.server.app_instance.refresh_all_data)
                    
                self.send_json({"success": True, "message": "Đã liên kết mã vạch thành công!"})
            except Exception as e:
                self.send_json({"success": False, "message": f"Lỗi cơ sở dữ liệu: {str(e)}"}, 500)
                
        else:
            self.send_response(404)
            self.end_headers()
            
    def send_json(self, data, status_code=200):
        self.send_response(status_code)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.end_headers()
        self.wfile.write(json.dumps(data, ensure_ascii=False).encode("utf-8"))

class MobileInventoryServer(threading.Thread):
    def __init__(self, app_instance, host="0.0.0.0", port=5000):
        super().__init__()
        self.app_instance = app_instance
        self.db_instance = app_instance.db
        self.host = host
        self.port = port
        self.server = None
        self.daemon = True
        self.is_running = False
        
    def run(self):
        attempts = 0
        while attempts < 10:
            try:
                self.server = http.server.HTTPServer((self.host, self.port), MobileInventoryRequestHandler)
                self.server.db_instance = self.db_instance
                self.server.app_instance = self.app_instance
                self.is_running = True
                print(f"Mobile inventory server started on http://{self.host}:{self.port}")
                self.server.serve_forever()
                break
            except Exception as e:
                print(f"Failed to start mobile server on port {self.port}: {e}")
                self.port += 1
                attempts += 1
                
    def stop(self):
        if self.server:
            self.server.shutdown()
            self.server.server_close()
            self.is_running = False
            print("Mobile inventory server stopped")

MOBILE_HTML = """<!DOCTYPE html>
<html lang="vi">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
    <title>Kiểm Kho Di Động</title>
    <style>
        :root {
            --primary: #0284c7;
            --primary-hover: #0369a1;
            --bg-grad: linear-gradient(135deg, #f0f9ff 0%, #e0f2fe 100%);
            --glass-bg: rgba(255, 255, 255, 0.65);
            --glass-border: rgba(2, 132, 199, 0.12);
            --text-light: #0f172a;
            --text-muted: #475569;
            --success: #0d9488;
            --warning: #ea580c;
            --danger: #e11d48;
        }
        * {
            box-sizing: border-box;
            margin: 0;
            padding: 0;
            font-family: 'Segoe UI', -apple-system, BlinkMacSystemFont, Roboto, Helvetica, Arial, sans-serif;
            -webkit-tap-highlight-color: transparent;
        }
        body {
            background: var(--bg-grad);
            color: var(--text-light);
            min-height: 100vh;
            padding: 15px;
            display: flex;
            flex-direction: column;
            align-items: center;
        }
        .container {
            width: 100%;
            max-width: 500px;
            display: flex;
            flex-direction: column;
            gap: 12px;
        }
        header {
            text-align: center;
            padding: 5px 0;
            display: flex;
            flex-direction: column;
            align-items: center;
            gap: 4px;
        }
        header h1 {
            font-size: 1.4rem;
            font-weight: 700;
            letter-spacing: 0.5px;
            color: #0369a1;
            display: flex;
            align-items: center;
            gap: 8px;
        }
        header p {
            font-size: 0.8rem;
            color: var(--text-muted);
        }
        .nav-tabs {
            display: flex;
            width: 100%;
            background: rgba(2, 132, 199, 0.04);
            border: 1px solid var(--glass-border);
            border-radius: 12px;
            padding: 4px;
        }
        .tab-btn {
            flex: 1;
            background: transparent;
            border: none;
            color: var(--text-muted);
            padding: 10px;
            font-size: 0.9rem;
            font-weight: 600;
            cursor: pointer;
            border-radius: 8px;
            transition: all 0.3s;
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 6px;
        }
        .tab-btn.active {
            background: var(--primary);
            color: #fff;
            box-shadow: 0 4px 12px rgba(2, 130, 199, 0.3);
        }
        .tab-content {
            display: none;
            flex-direction: column;
            gap: 12px;
        }
        .tab-content.active {
            display: flex;
        }
        .card {
            background: var(--glass-bg);
            border: 1px solid var(--glass-border);
            backdrop-filter: blur(10px);
            -webkit-backdrop-filter: blur(10px);
            border-radius: 16px;
            padding: 15px;
            box-shadow: 0 8px 32px 0 rgba(2, 132, 199, 0.08);
        }
        .scanner-card {
            overflow: hidden;
            display: flex;
            flex-direction: column;
            align-items: center;
        }
        #reader {
            width: 100% !important;
            border: none !important;
            border-radius: 12px;
            overflow: hidden;
            background: #000;
        }
        #reader button {
            background: var(--primary) !important;
            color: #fff !important;
            border: none !important;
            padding: 8px 16px !important;
            border-radius: 8px !important;
            font-size: 0.9rem !important;
            font-weight: 600 !important;
            cursor: pointer !important;
            margin: 10px 0 !important;
            transition: background 0.2s !important;
        }
        #reader button:hover {
            background: var(--primary-hover) !important;
        }
        #reader select {
            background: rgba(255, 255, 255, 0.8) !important;
            color: var(--text-light) !important;
            border: 1px solid var(--glass-border) !important;
            padding: 8px !important;
            border-radius: 8px !important;
            margin: 5px 0 !important;
            width: 90% !important;
        }
        .search-box {
            display: flex;
            gap: 8px;
        }
        .search-box input {
            flex: 1;
            background: rgba(255, 255, 255, 0.8);
            border: 1px solid var(--glass-border);
            border-radius: 8px;
            padding: 10px 12px;
            color: var(--text-light);
            font-size: 0.95rem;
            outline: none;
            transition: border-color 0.2s;
        }
        .search-box input:focus {
            border-color: var(--primary);
        }
        .search-box button {
            background: var(--primary);
            border: none;
            border-radius: 8px;
            color: #fff;
            padding: 0 16px;
            font-weight: 600;
            cursor: pointer;
            transition: background 0.2s;
        }
        .search-box button:hover {
            background: var(--primary-hover);
        }
        .result-title {
            font-size: 1.1rem;
            font-weight: 700;
            margin-bottom: 12px;
            display: flex;
            align-items: center;
            gap: 6px;
            color: var(--text-light);
            border-bottom: 1px solid var(--glass-border);
            padding-bottom: 8px;
        }
        .product-info {
            display: grid;
            grid-template-columns: 1fr;
            gap: 8px;
            margin-bottom: 12px;
        }
        .info-row {
            display: flex;
            justify-content: space-between;
            font-size: 0.9rem;
        }
        .info-label {
            color: var(--text-muted);
        }
        .info-value {
            font-weight: 600;
            color: var(--text-light);
        }
        .batch-list {
            display: flex;
            flex-direction: column;
            gap: 8px;
        }
        .batch-item {
            background: rgba(255, 255, 255, 0.45);
            border: 1px solid rgba(2, 132, 199, 0.08);
            border-radius: 10px;
            padding: 10px 12px;
            display: flex;
            flex-direction: column;
            gap: 4px;
        }
        .batch-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        .batch-lot {
            font-weight: 700;
            color: #0369a1;
            font-size: 0.95rem;
        }
        .batch-qty {
            font-size: 1.1rem;
            font-weight: 800;
            color: var(--success);
        }
        .batch-expiry {
            font-size: 0.8rem;
            color: var(--text-muted);
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        .badge {
            padding: 2px 6px;
            border-radius: 4px;
            font-size: 0.75rem;
            font-weight: 600;
            text-transform: uppercase;
        }
        .badge-expired {
            background: rgba(225, 29, 72, 0.12);
            color: var(--danger);
            border: 1px solid rgba(225, 29, 72, 0.2);
        }
        .badge-warning {
            background: rgba(234, 88, 12, 0.12);
            color: var(--warning);
            border: 1px solid rgba(234, 88, 12, 0.2);
        }
        .badge-ok {
            background: rgba(13, 148, 136, 0.12);
            color: var(--success);
            border: 1px solid rgba(13, 148, 136, 0.2);
        }
        
        /* Modals & Form buttons */
        .action-buttons {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 8px;
            margin-top: 15px;
            border-top: 1px solid var(--glass-border);
            padding-top: 15px;
        }
        .action-btn {
            padding: 10px;
            border: none;
            border-radius: 8px;
            font-weight: bold;
            color: #fff;
            cursor: pointer;
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 6px;
            font-size: 0.85rem;
            transition: opacity 0.2s, transform 0.1s;
        }
        .action-btn:active {
            transform: scale(0.97);
        }
        .btn-purchase { background: var(--success); }
        .btn-dispatch { background: var(--danger); }
        .btn-barcode { background: var(--warning); grid-column: span 2; }
        
        .form-container {
            margin-top: 15px;
            padding: 12px;
            border-radius: 12px;
            background: rgba(255, 255, 255, 0.5);
            border: 1px dashed rgba(2, 132, 199, 0.2);
            display: none;
        }
        .form-title {
            font-size: 0.95rem;
            font-weight: 700;
            color: var(--text-light);
            margin-bottom: 10px;
            display: flex;
            align-items: center;
            gap: 6px;
        }
        .form-group {
            margin-bottom: 10px;
            display: flex;
            flex-direction: column;
            gap: 5px;
        }
        .form-group label {
            font-size: 0.75rem;
            color: var(--text-muted);
            font-weight: 600;
        }
        .form-control {
            background: rgba(255, 255, 255, 0.85);
            border: 1px solid var(--glass-border);
            border-radius: 6px;
            padding: 8px 10px;
            color: var(--text-light);
            font-size: 0.9rem;
            outline: none;
        }
        .form-control:focus {
            border-color: var(--primary);
        }
        .form-actions {
            display: flex;
            gap: 8px;
            margin-top: 12px;
        }
        .form-actions button {
            flex: 1;
            padding: 8px;
            border-radius: 6px;
            border: none;
            font-weight: 600;
            cursor: pointer;
            font-size: 0.85rem;
        }
        .btn-submit { background: var(--primary); color: #fff; }
        .btn-cancel { background: rgba(0, 0, 0, 0.05); color: var(--text-light); }
        
        /* Product catalog */
        .product-list {
            display: flex;
            flex-direction: column;
            gap: 8px;
            max-height: 50vh;
            overflow-y: auto;
            margin-top: 8px;
            padding-right: 2px;
        }
        .product-item {
            background: rgba(255, 255, 255, 0.55);
            border: 1px solid rgba(2, 132, 199, 0.08);
            border-radius: 10px;
            padding: 10px 12px;
            cursor: pointer;
            transition: background 0.2s;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        .product-item:hover {
            background: rgba(255, 255, 255, 0.85);
        }
        .product-item-details {
            display: flex;
            flex-direction: column;
            gap: 2px;
        }
        .product-item-name {
            font-weight: 600;
            color: var(--text-light);
            font-size: 0.9rem;
        }
        .product-item-sub {
            font-size: 0.75rem;
            color: var(--text-muted);
        }
        .product-item-arrow {
            color: var(--text-muted);
            font-size: 1.1rem;
        }
        
        /* Success Print Modal */
        .modal-overlay {
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            background: rgba(15, 23, 42, 0.45);
            backdrop-filter: blur(8px);
            -webkit-backdrop-filter: blur(8px);
            display: flex;
            align-items: center;
            justify-content: center;
            z-index: 10000;
            animation: fadeIn 0.25s ease-out;
        }
        .modal-content {
            background: rgba(255, 255, 255, 0.96);
            border: 1px solid var(--glass-border);
            border-radius: 20px;
            padding: 24px;
            width: 90%;
            max-width: 380px;
            text-align: center;
            box-shadow: 0 20px 40px rgba(2, 132, 199, 0.12);
            animation: scaleIn 0.3s cubic-bezier(0.34, 1.56, 0.64, 1);
        }
        .modal-icon {
            font-size: 3rem;
            margin-bottom: 12px;
        }
        .modal-content h3 {
            font-size: 1.25rem;
            font-weight: 700;
            margin-bottom: 8px;
            color: var(--text-light);
        }
        .modal-content p {
            font-size: 0.9rem;
            color: var(--text-muted);
            margin-bottom: 20px;
            line-height: 1.4;
        }
        .modal-actions {
            display: flex;
            flex-direction: column;
            gap: 10px;
        }
        .btn-modal-print {
            background: var(--primary);
            color: white;
            border: none;
            padding: 12px;
            border-radius: 10px;
            font-weight: bold;
            font-size: 0.95rem;
            cursor: pointer;
            transition: background 0.2s;
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 8px;
        }
        .btn-modal-print:hover {
            background: var(--primary-hover);
        }
        .btn-modal-close {
            background: rgba(0, 0, 0, 0.05);
            color: var(--text-light);
            border: 1px solid var(--glass-border);
            padding: 12px;
            border-radius: 10px;
            font-weight: 600;
            font-size: 0.95rem;
            cursor: pointer;
            transition: background 0.2s;
        }
        .btn-modal-close:hover {
            background: rgba(0, 0, 0, 0.08);
        }
        @keyframes fadeIn {
            from { opacity: 0; }
            to { opacity: 1; }
        }
        @keyframes scaleIn {
            from { transform: scale(0.9); opacity: 0; }
            to { transform: scale(1); opacity: 1; }
        }
        .cart-item-row {
            display: flex;
            justify-content: space-between;
            align-items: center;
            padding: 8px 12px;
            background: rgba(2, 132, 199, 0.04);
            border: 1px solid var(--glass-border);
            border-radius: 8px;
            margin-bottom: 8px;
        }
        .cart-item-details {
            flex: 1;
        }
        .cart-item-name {
            font-weight: 600;
            font-size: 0.85rem;
            color: var(--text-light);
        }
        .cart-item-meta {
            font-size: 0.75rem;
            color: var(--text-muted);
            margin-top: 2px;
        }
        .btn-cart-remove {
            background: none;
            border: none;
            color: #ef4444;
            font-size: 1.1rem;
            cursor: pointer;
            padding: 4px 8px;
            transition: opacity 0.2s;
        }
        .btn-cart-remove:hover {
            opacity: 0.7;
        }
        .preview-table th, .preview-table td {
            padding: 8px 10px;
            border-bottom: 1px solid var(--glass-border);
            color: var(--text-light);
        }
        .preview-table tbody tr:last-child td {
            border-bottom: none;
        }
        
        /* Toast notification system */
        #toast-container {
            position: fixed;
            bottom: 20px;
            left: 50%;
            transform: translateX(-50%);
            display: flex;
            flex-direction: column;
            gap: 8px;
            z-index: 9999;
            width: 90%;
            max-width: 320px;
        }
        .toast {
            background: rgba(15, 23, 42, 0.9);
            border: 1px solid var(--glass-border);
            backdrop-filter: blur(10px);
            border-radius: 8px;
            padding: 10px 14px;
            color: #fff;
            font-size: 0.8rem;
            font-weight: 600;
            display: flex;
            align-items: center;
            gap: 8px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.5);
            animation: slideUp 0.3s ease forwards;
        }
        .toast-success { border-left: 4px solid var(--success); }
        .toast-error { border-left: 4px solid var(--danger); }
        
        .no-result, .loading, .error-msg {
            text-align: center;
            padding: 20px;
            color: var(--text-muted);
            font-size: 0.9rem;
        }
        .error-msg {
            color: var(--danger);
        }
        .loading-spinner {
            border: 3px solid rgba(255,255,255,0.1);
            border-top: 3px solid var(--primary);
            border-radius: 50%;
            width: 26px;
            height: 26px;
            animation: spin 1s linear infinite;
            margin: 10px auto;
        }
        @keyframes spin {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }
        @keyframes slideUp {
            from { transform: translateY(20px); opacity: 0; }
            to { transform: translateY(0); opacity: 1; }
        }
    </style>
</head>
<body>
    <div id="toast-container"></div>

    <!-- Print Success Modal -->
    <div id="print-modal" class="modal-overlay" style="display: none;">
        <div class="modal-content">
            <div class="modal-icon">✅</div>
            <h3 id="print-modal-title">Thành công</h3>
            <p id="print-modal-message">Đã thực hiện thành công.</p>
            <div class="modal-actions">
                <button id="btn-modal-print-pc" class="btn-modal-print" style="background: #10b981;">
                    🖥️ In qua máy tính (PC)
                </button>
                <button id="btn-modal-print-phone" class="btn-modal-print">
                    📱 In/Tải về trên ĐT
                </button>
                <button class="btn-modal-close" onclick="closePrintModal()">Đóng</button>
            </div>
        </div>
    </div>

    <!-- Cart Review Modal -->
    <div id="cart-modal" class="modal-overlay" style="display: none;">
        <div class="modal-content" style="max-width: 440px; text-align: left;">
            <h3 id="cart-modal-title" style="text-align: center; margin-bottom: 12px; color: var(--primary);">🛒 Giỏ Hàng</h3>
            
            <div id="cart-items-container" style="max-height: 200px; overflow-y: auto; margin-bottom: 15px; border-bottom: 1px solid var(--glass-border); padding-bottom: 10px;">
                <!-- Cart items will be generated here -->
            </div>
            
            <div id="cart-form-fields">
                <div class="form-group" style="margin-bottom: 10px;">
                    <label id="cart-partner-label" style="font-weight: 600; font-size: 0.85rem; color: var(--text-muted); display: block; margin-bottom: 4px;">Nhà cung cấp *</label>
                    <input type="text" id="cart-partner-input" class="form-control" placeholder="Nhập tên đối tác..." />
                </div>
                <div class="form-group" style="margin-bottom: 10px;">
                    <label id="cart-reason-label" style="font-weight: 600; font-size: 0.85rem; color: var(--text-muted); display: block; margin-bottom: 4px;">Lý do thực hiện *</label>
                    <input type="text" id="cart-reason-input" class="form-control" placeholder="Lý do..." />
                </div>
                <div class="form-group" style="margin-bottom: 15px;">
                    <label style="font-weight: 600; font-size: 0.85rem; color: var(--text-muted); display: block; margin-bottom: 4px;">Ghi chú</label>
                    <input type="text" id="cart-note-input" class="form-control" placeholder="Ghi chú thêm..." />
                </div>
            </div>
            
            <div class="modal-actions" style="margin-top: 15px; gap: 8px;">
                <button id="btn-cart-submit-pc" class="btn-modal-print" style="background: #10b981;">
                    🖥️ Tạo phiếu & In PC
                </button>
                <button id="btn-cart-submit-phone" class="btn-modal-print">
                    📱 Tạo phiếu & In ĐT
                </button>
                <button class="btn-modal-close" onclick="closeCartModal()">Đóng</button>
            </div>
        </div>
    </div>

    <!-- Note Preview Modal -->
    <div id="preview-modal" class="modal-overlay" style="display: none;">
        <div class="modal-content" style="max-width: 460px; text-align: left;">
            <h3 id="preview-modal-title" style="text-align: center; margin-bottom: 12px; color: var(--primary);">📋 Xem Trước Phiếu</h3>
            
            <div id="preview-info-container" style="font-size: 0.85rem; line-height: 1.5; margin-bottom: 12px; background: rgba(0,0,0,0.02); padding: 10px; border-radius: 10px; border: 1px solid var(--glass-border);">
                <!-- General note metadata will be injected here -->
            </div>
            
            <div style="font-weight: 600; font-size: 0.85rem; color: var(--text-light); margin-bottom: 6px;">Danh sách sản phẩm:</div>
            <div id="preview-items-container" style="max-height: 180px; overflow-y: auto; margin-bottom: 15px; border: 1px solid var(--glass-border); border-radius: 8px; background: #fff;">
                <table class="preview-table" style="width: 100%; border-collapse: collapse; font-size: 0.8rem;">
                    <thead>
                        <tr style="background: rgba(2, 132, 199, 0.08); border-bottom: 1px solid var(--glass-border);">
                            <th style="padding: 6px 8px; text-align: left;">Tên sản phẩm</th>
                            <th style="padding: 6px 8px; text-align: center; width: 65px;">Lô</th>
                            <th style="padding: 6px 8px; text-align: right; width: 65px;">SL</th>
                        </tr>
                    </thead>
                    <tbody id="preview-table-body">
                        <!-- Table rows will be generated here -->
                    </tbody>
                </table>
            </div>
            
            <div class="modal-actions" style="margin-top: 15px; gap: 8px;">
                <button id="btn-preview-submit-pc" class="btn-modal-print" style="background: #10b981;">
                    🖥️ Xác nhận In PC
                </button>
                <button id="btn-preview-submit-phone" class="btn-modal-print">
                    📱 Tải về / Xem PDF ĐT
                </button>
                <button class="btn-modal-close" onclick="closePreviewModal()">Đóng</button>
            </div>
        </div>
    </div>

    <div class="container">
        <header>
            <h1>🏥 Quản Lý Kho Di Động</h1>
            <p>Kiểm kho, Nhập/Xuất & Liên kết mã vạch nhanh chóng</p>
        </header>

        <!-- Quick Dashboard -->
        <div class="card dashboard-card" style="margin-bottom: 15px; padding: 14px; background: linear-gradient(135deg, rgba(2, 132, 199, 0.06) 0%, rgba(13, 148, 136, 0.06) 100%); border: 1px solid rgba(2, 132, 199, 0.15);">
            <div style="font-weight: 700; font-size: 0.95rem; color: var(--primary); display: flex; align-items: center; gap: 6px; margin-bottom: 12px;">
                📊 Tổng Quan Kho Hàng
                <span id="dashboard-refresh" onclick="loadDashboardStats()" style="cursor: pointer; font-size: 0.85rem; font-weight: normal; color: var(--text-muted); margin-left: auto;">🔄 Cập nhật</span>
            </div>
            <div style="display: grid; grid-template-columns: repeat(3, 1fr); gap: 8px; text-align: center;">
                <div style="background: rgba(255,255,255,0.75); border: 1px solid var(--glass-border); padding: 8px; border-radius: 10px;">
                    <div style="font-size: 0.72rem; color: var(--text-muted); font-weight: 600;">Tổng Thuốc</div>
                    <div id="dash-total-products" style="font-size: 1.1rem; font-weight: 800; color: var(--text-light); margin-top: 4px;">0</div>
                </div>
                <div onclick="filterCatalog('outofstock')" style="background: rgba(255,255,255,0.75); border: 1px solid var(--glass-border); padding: 8px; border-radius: 10px; cursor: pointer; transition: transform 0.2s;">
                    <div style="font-size: 0.72rem; color: var(--text-muted); font-weight: 600;">❌ Hết Hàng</div>
                    <div id="dash-outofstock-products" style="font-size: 1.1rem; font-weight: 800; color: #ef4444; margin-top: 4px;">0</div>
                </div>
                <div onclick="filterCatalog('expiring')" style="background: rgba(255,255,255,0.75); border: 1px solid var(--glass-border); padding: 8px; border-radius: 10px; cursor: pointer; transition: transform 0.2s;">
                    <div style="font-size: 0.72rem; color: var(--text-muted); font-weight: 600;">⚠️ Cận Hạn</div>
                    <div id="dash-expiring-products" style="font-size: 1.1rem; font-weight: 800; color: #f59e0b; margin-top: 4px;">0</div>
                </div>
            </div>
        </div>

        <!-- Navigation Tabs -->
        <div class="nav-tabs">
            <button class="tab-btn active" onclick="switchTab('tab-checker')">🔍 Kiểm Kho</button>
            <button class="tab-btn" onclick="switchTab('tab-catalog')">📋 Danh Sách</button>
            <button class="tab-btn" onclick="switchTab('tab-history')">📜 Lịch Sử</button>
        </div>

        <!-- Cart Status Bar -->
        <div id="cart-status-bar" style="display: none; gap: 8px; width: 100%; margin-top: 5px; margin-bottom: 10px;">
            <div id="cart-purchase-btn" onclick="openCartModal('purchase')" style="flex: 1; background: #0d9488; color: #fff; padding: 10px; border-radius: 12px; font-weight: bold; text-align: center; font-size: 0.82rem; cursor: pointer; display: flex; align-items: center; justify-content: center; gap: 6px; box-shadow: 0 4px 12px rgba(13, 148, 136, 0.15);">
                📥 Giỏ Nhập: <span id="cart-purchase-count">0</span> món
            </div>
            <div id="cart-dispatch-btn" onclick="openCartModal('dispatch')" style="flex: 1; background: #e11d48; color: #fff; padding: 10px; border-radius: 12px; font-weight: bold; text-align: center; font-size: 0.82rem; cursor: pointer; display: flex; align-items: center; justify-content: center; gap: 6px; box-shadow: 0 4px 12px rgba(225, 29, 72, 0.15);">
                📤 Giỏ Xuất: <span id="cart-dispatch-count">0</span> món
            </div>
        </div>

        <!-- Tab 1: Kiểm kho và quét mã vạch -->
        <div id="tab-checker" class="tab-content active">
            <div class="card scanner-card">
                <div id="reader"></div>
            </div>

            <div class="card">
                <div class="search-box">
                    <input type="text" id="barcode-input" placeholder="Nhập mã vạch hoặc tên..." />
                    <button id="search-btn">Tìm</button>
                </div>
            </div>

            <div class="card" id="result-card" style="display: none;">
                <div class="result-title">📦 Kết quả truy vấn</div>
                <div id="result-content"></div>
                
                <!-- Quick Action Forms Container -->
                <div id="action-forms">
                    <!-- Nhập kho nhanh form -->
                    <div id="form-purchase" class="form-container">
                        <div class="form-title">📥 Nhập kho nhanh</div>
                        <div class="form-group">
                            <label>Số lượng nhập (Đơn vị tính gốc)</label>
                            <input type="number" id="pur-qty" class="form-control" placeholder="Ví dụ: 10" step="any" required />
                        </div>
                        <div class="form-group">
                            <label>Số lô (Lot No)</label>
                            <input type="text" id="pur-lot" class="form-control" placeholder="Ví dụ: LO1234" required />
                        </div>
                        <div class="form-group">
                            <label>Hạn sử dụng</label>
                            <input type="date" id="pur-expiry" class="form-control" required />
                        </div>
                        <div class="form-actions" style="display: flex; flex-direction: column; gap: 8px;">
                            <div style="display: flex; gap: 8px; width: 100%;">
                                <button class="btn-cancel" onclick="closeForms()" style="flex: 1; margin: 0; padding: 10px;">Hủy</button>
                                <button class="btn-submit" onclick="addToCart('purchase')" style="flex: 2; background: #0d9488; margin: 0; padding: 10px;">📥 Thêm vào giỏ</button>
                            </div>
                            <button class="btn-submit" onclick="submitPurchase()" style="width: 100%; margin: 0; padding: 10px;">Nhập & Tạo phiếu ngay</button>
                        </div>
                    </div>
                    
                    <!-- Xuất kho nhanh form -->
                    <div id="form-dispatch" class="form-container">
                        <div class="form-title">📤 Xuất kho nhanh</div>
                        <div class="form-group">
                            <label>Chọn lô xuất</label>
                            <select id="disp-batch-id" class="form-control"></select>
                        </div>
                        <div class="form-group">
                            <label>Số lượng xuất (Đơn vị tính gốc)</label>
                            <input type="number" id="disp-qty" class="form-control" placeholder="Ví dụ: 5" step="any" required />
                        </div>
                        <div class="form-group">
                            <label>Lý do xuất</label>
                            <input type="text" id="disp-reason" class="form-control" placeholder="Ví dụ: Hao hụt, Cấp phát di động,..." value="Xuất qua điện thoại" />
                        </div>
                        <div class="form-actions" style="display: flex; flex-direction: column; gap: 8px;">
                            <div style="display: flex; gap: 8px; width: 100%;">
                                <button class="btn-cancel" onclick="closeForms()" style="flex: 1; margin: 0; padding: 10px;">Hủy</button>
                                <button class="btn-submit" onclick="addToCart('dispatch')" style="flex: 2; background: #e11d48; margin: 0; padding: 10px;">📤 Thêm vào giỏ</button>
                            </div>
                            <button class="btn-submit" onclick="submitDispatch()" style="width: 100%; margin: 0; padding: 10px;">Xuất & Tạo phiếu ngay</button>
                        </div>
                    </div>

                    <!-- Gán mã vạch mới form -->
                    <div id="form-barcode" class="form-container">
                        <div class="form-title">🏷️ Khai báo mã vạch mới</div>
                        <div class="form-group">
                            <label>Mã vạch liên kết</label>
                            <input type="text" id="link-barcode" class="form-control" placeholder="Quét hoặc điền mã vạch..." required />
                        </div>
                        <div class="form-actions">
                            <button class="btn-cancel" onclick="closeForms()">Hủy</button>
                            <button class="btn-submit" onclick="submitLinkBarcode()">Lưu Liên Kết</button>
                        </div>
                    </div>
                </div>
            </div>
        </div>

        <!-- Tab 2: Danh sách sản phẩm -->
        <div id="tab-catalog" class="tab-content">
            <div class="card">
                <button class="action-btn" style="margin-bottom: 12px; width: 100%; display: flex; align-items: center; justify-content: center; gap: 6px; background: var(--primary); font-size: 0.9rem;" onclick="openCreateProductForm()">➕ Thêm Sản Phẩm Mới</button>
                
                <!-- Thêm sản phẩm mới form -->
                <div id="form-create-product" class="form-container" style="margin-bottom: 15px; border-style: solid; border-color: var(--primary);">
                    <div class="form-title">➕ Thêm sản phẩm mới</div>
                    <div class="form-group">
                        <label>Tên sản phẩm *</label>
                        <input type="text" id="new-name" class="form-control" placeholder="Ví dụ: Paracetamol 500mg" required />
                    </div>
                    <div class="form-group">
                        <label>Đơn vị tính gốc *</label>
                        <input type="text" id="new-unit" class="form-control" placeholder="Ví dụ: Viên, Hộp, Chai" required />
                    </div>
                    <div class="form-group">
                        <label>Phân loại sản phẩm</label>
                        <select id="new-type" class="form-control">
                            <option value="thuoc">Thuốc / Dược phẩm</option>
                            <option value="vaccine">Vaccine</option>
                            <option value="vtyt">Vật tư y tế</option>
                            <option value="khac">Sản phẩm khác</option>
                        </select>
                    </div>
                    <div class="form-group">
                        <label>Mã vạch (Quét hoặc điền)</label>
                        <input type="text" id="new-barcode" class="form-control" placeholder="Để trống nếu chưa có" />
                    </div>
                    <div class="form-group">
                        <label>Số đăng ký (Không bắt buộc)</label>
                        <input type="text" id="new-regnumber" class="form-control" placeholder="Số đăng ký..." />
                    </div>
                    <div class="form-actions">
                        <button class="btn-cancel" onclick="closeCreateProductForm()">Hủy</button>
                        <button class="btn-submit" onclick="submitCreateProduct()">Tạo & Nhập Kho</button>
                    </div>
                </div>

                <div class="search-box">
                    <input type="text" id="catalog-search" placeholder="Nhập tên sản phẩm..." />
                    <button id="catalog-search-btn">Lọc</button>
                </div>
                <div class="filter-toggles" style="display: flex; gap: 6px; margin-bottom: 12px; font-size: 0.8rem; margin-top: -6px;">
                    <button id="filter-btn-all" onclick="filterCatalog('all')" style="flex: 1; padding: 8px 6px; border: 1px solid var(--primary); background: var(--primary); color: #fff; border-radius: 8px; font-weight: bold; cursor: pointer; font-size: 0.75rem; transition: all 0.2s;">Tất cả</button>
                    <button id="filter-btn-outofstock" onclick="filterCatalog('outofstock')" style="flex: 1; padding: 8px 6px; border: 1px solid var(--glass-border); background: #fff; color: #ef4444; border-radius: 8px; font-weight: bold; cursor: pointer; font-size: 0.75rem; transition: all 0.2s;">❌ Hết Hàng</button>
                    <button id="filter-btn-expiring" onclick="filterCatalog('expiring')" style="flex: 1; padding: 8px 6px; border: 1px solid var(--glass-border); background: #fff; color: #f59e0b; border-radius: 8px; font-weight: bold; cursor: pointer; font-size: 0.75rem; transition: all 0.2s;">⚠️ Cận Hạn</button>
                </div>
                <div id="catalog-list" class="product-list"></div>
            </div>
        </div>
        
        <!-- Tab 3: Lịch sử hoạt động -->
        <div id="tab-history" class="tab-content">
            <div class="card">
                <div class="result-title">📜 Lịch sử hoạt động gần đây</div>
                <div id="history-list" class="product-list">
                    <div style="text-align: center; color: var(--text-muted); padding: 20px;">Đang tải...</div>
                </div>
            </div>
        </div>
    </div>

    <script src="https://unpkg.com/html5-qrcode"></script>
    <script>
        const barcodeInput = document.getElementById('barcode-input');
        const searchBtn = document.getElementById('search-btn');
        const resultCard = document.getElementById('result-card');
        const resultContent = document.getElementById('result-content');
        
        let currentProduct = null;
        let currentProductBatches = [];

        function switchTab(tabId) {
            document.querySelectorAll('.tab-content').forEach(el => el.classList.remove('active'));
            document.querySelectorAll('.tab-btn').forEach(el => el.classList.remove('active'));
            
            document.getElementById(tabId).classList.add('active');
            
            let btnIndex = 0;
            if (tabId === 'tab-catalog') btnIndex = 1;
            else if (tabId === 'tab-history') btnIndex = 2;
            document.querySelectorAll('.tab-btn')[btnIndex].classList.add('active');
            
            if (tabId === 'tab-catalog') {
                loadCatalog('');
            } else if (tabId === 'tab-history') {
                loadRecentActivities();
            }
        }

        // Notification Toast System
        function showToast(message, type = 'success') {
            const container = document.getElementById('toast-container');
            const toast = document.createElement('div');
            toast.className = `toast toast-${type}`;
            toast.innerHTML = (type === 'success' ? '✅' : '❌') + ` <span>${message}</span>`;
            container.appendChild(toast);
            
            setTimeout(() => {
                toast.style.animation = 'none';
                toast.offsetHeight; /* trigger reflow */
                toast.style.animation = 'slideUp 0.3s ease reverse forwards';
                setTimeout(() => toast.remove(), 300);
            }, 3000);
        }

        // Print success modal controller
        function showPrintModal(type, noteId, message) {
            document.getElementById('print-modal-message').textContent = message;
            
            const btnPc = document.getElementById('btn-modal-print-pc');
            const btnPhone = document.getElementById('btn-modal-print-phone');
            
            // Cài đặt sự kiện in trên máy tính (PC)
            btnPc.onclick = function() {
                btnPc.disabled = true;
                btnPc.textContent = '⌛ Đang gửi...';
                
                fetch(`/api/pc-print?type=${type}&id=${noteId}`)
                    .then(res => res.json())
                    .then(data => {
                        btnPc.disabled = false;
                        btnPc.innerHTML = '🖥️ In qua máy tính (PC)';
                        if (data.success) {
                            showToast(data.message, "success");
                            closePrintModal();
                        } else {
                            showToast(data.message, "error");
                        }
                    })
                    .catch(err => {
                        btnPc.disabled = false;
                        btnPc.innerHTML = '🖥️ In qua máy tính (PC)';
                        showToast("Lỗi kết nối lệnh in PC", "error");
                    });
            };
            
            // Cài đặt sự kiện in trên điện thoại
            if (type === 'purchase') {
                document.getElementById('print-modal-title').textContent = 'Nhập Kho Thành Công';
                btnPhone.onclick = function() {
                    window.open(`/api/print-purchase?id=${noteId}`, '_blank');
                    closePrintModal();
                };
            } else {
                document.getElementById('print-modal-title').textContent = 'Xuất Kho Thành Công';
                btnPhone.onclick = function() {
                    window.open(`/api/print-dispatch?id=${noteId}`, '_blank');
                    closePrintModal();
                };
            }
            
            document.getElementById('print-modal').style.display = 'flex';
        }
        
        function closePrintModal() {
            document.getElementById('print-modal').style.display = 'none';
        }

        function loadRecentActivities() {
            const historyList = document.getElementById('history-list');
            historyList.innerHTML = `<div style="text-align: center; color: var(--text-muted); padding: 20px;">Đang tải lịch sử...</div>`;
            
            fetch('/api/recent-activities')
                .then(res => res.json())
                .then(data => {
                    if (data.success && data.activities.length > 0) {
                        historyList.innerHTML = '';
                        data.activities.forEach(act => {
                            const dateStr = act.createdAt;
                            let formattedDate = dateStr;
                            try {
                                const parts = dateStr.split(' ');
                                const dateParts = parts[0].split('-');
                                formattedDate = `${dateParts[2]}/${dateParts[1]} ${parts[1].substring(0, 5)}`;
                            } catch(e) {}
                            
                            const item = document.createElement('div');
                            item.className = 'product-item';
                            item.style.cursor = 'default';
                            item.style.flexDirection = 'column';
                            item.style.alignItems = 'stretch';
                            item.style.gap = '8px';
                            
                            const isPurchase = act.type === 'nhap';
                            const badgeColor = isPurchase ? 'var(--success)' : 'var(--danger)';
                            const typeLabel = isPurchase ? 'NHẬP KHO' : 'XUẤT KHO';
                            
                            item.innerHTML = `
                                <div style="display: flex; justify-content: space-between; align-items: center;">
                                    <span style="font-weight: bold; font-size: 0.8rem; padding: 2px 6px; border-radius: 4px; background: ${badgeColor}; color: #fff;">${typeLabel}</span>
                                    <span style="font-size: 0.75rem; color: var(--text-muted); font-weight: 500;">${formattedDate}</span>
                                </div>
                                <div style="font-weight: 600; font-size: 0.9rem; color: var(--text-light); margin: 2px 0;">Số phiếu: ${act.noteNumber}</div>
                                <div style="font-size: 0.8rem; color: var(--text-muted);">${isPurchase ? 'Nhà cung cấp' : 'Đơn vị nhận'}: ${act.details}</div>
                                <div style="display: flex; gap: 8px; margin-top: 6px; border-top: 1px solid var(--glass-border); padding-top: 8px;">
                                    <button onclick="showNotePreview('${act.type}', ${act.id})" style="flex: 1; padding: 8px; background: var(--primary); border: none; border-radius: 8px; color: #fff; font-weight: bold; font-size: 0.8rem; cursor: pointer; display: flex; align-items: center; justify-content: center; gap: 6px;">
                                        🔎 Xem trước & In phiếu
                                    </button>
                                </div>
                            `;
                            historyList.appendChild(item);
                        });
                    } else {
                        historyList.innerHTML = `<div style="text-align: center; color: var(--text-muted); padding: 20px;">Không có hoạt động gần đây</div>`;
                    }
                })
                .catch(err => {
                    historyList.innerHTML = `<div style="text-align: center; color: var(--danger); padding: 20px;">Lỗi tải dữ liệu</div>`;
                });
        }
        
        function printActivityPC(type, noteId, btn) {
            btn.disabled = true;
            const origText = btn.innerHTML;
            btn.textContent = '⌛...';
            
            fetch(`/api/pc-print?type=${type}&id=${noteId}`)
                .then(res => res.json())
                .then(data => {
                    btn.disabled = false;
                    btn.innerHTML = origText;
                    showToast(data.message, data.success ? "success" : "error");
                })
                .catch(err => {
                    btn.disabled = false;
                    btn.innerHTML = origText;
                    showToast("Lỗi gửi lệnh in PC", "error");
                });
        }
        
        function printActivityPhone(type, noteId) {
            const url = type === 'nhap' ? `/api/print-purchase?id=${noteId}` : `/api/print-dispatch?id=${noteId}`;
            window.open(url, '_blank');
        }

        // --- Cart Management Javascript ---
        let purchaseCart = JSON.parse(localStorage.getItem('mob_purchase_cart')) || [];
        let dispatchCart = JSON.parse(localStorage.getItem('mob_dispatch_cart')) || [];

        function updateCartStatus() {
            const pCount = purchaseCart.length;
            const dCount = dispatchCart.length;
            
            document.getElementById('cart-purchase-count').textContent = pCount;
            document.getElementById('cart-dispatch-count').textContent = dCount;
            
            const statusBar = document.getElementById('cart-status-bar');
            const pBtn = document.getElementById('cart-purchase-btn');
            const dBtn = document.getElementById('cart-dispatch-btn');
            
            if (pCount > 0 || dCount > 0) {
                statusBar.style.display = 'flex';
                pBtn.style.display = pCount > 0 ? 'flex' : 'none';
                dBtn.style.display = dCount > 0 ? 'flex' : 'none';
            } else {
                statusBar.style.display = 'none';
            }
        }

        function addToCart(type) {
            if (!currentProduct) return;
            
            if (type === 'purchase') {
                const qty = parseFloat(document.getElementById('pur-qty').value);
                const lotNo = document.getElementById('pur-lot').value.trim();
                const expiry = document.getElementById('pur-expiry').value;
                
                if (!qty || qty <= 0 || !lotNo || !expiry) {
                    showToast("Vui lòng điền đầy đủ thông tin nhập kho!", "error");
                    return;
                }
                
                const existingIdx = purchaseCart.findIndex(item => item.productId === currentProduct.id && item.lotNo === lotNo);
                if (existingIdx > -1) {
                    purchaseCart[existingIdx].qty += qty;
                } else {
                    purchaseCart.push({
                        productId: currentProduct.id,
                        productName: currentProduct.name,
                        unit: currentProduct.unit,
                        qty: qty,
                        lotNo: lotNo,
                        expiryDate: expiry
                    });
                }
                localStorage.setItem('mob_purchase_cart', JSON.stringify(purchaseCart));
                showToast(`Đã thêm ${qty} ${currentProduct.unit} vào giỏ nhập`, "success");
                closeForms();
                updateCartStatus();
            } else if (type === 'dispatch') {
                const lotNo = document.getElementById('disp-batch-id').value;
                const qty = parseFloat(document.getElementById('disp-qty').value);
                
                if (!lotNo || !qty || qty <= 0) {
                    showToast("Vui lòng điền đầy đủ thông tin xuất kho!", "error");
                    return;
                }
                
                const existingIdx = dispatchCart.findIndex(item => item.productId === currentProduct.id && item.lotNo === lotNo);
                if (existingIdx > -1) {
                    dispatchCart[existingIdx].qty += qty;
                } else {
                    dispatchCart.push({
                        productId: currentProduct.id,
                        productName: currentProduct.name,
                        unit: currentProduct.unit,
                        qty: qty,
                        lotNo: lotNo
                    });
                }
                localStorage.setItem('mob_dispatch_cart', JSON.stringify(dispatchCart));
                showToast(`Đã thêm ${qty} ${currentProduct.unit} vào giỏ xuất`, "success");
                closeForms();
                updateCartStatus();
            }
        }

        let activeCartType = null;
        function openCartModal(type) {
            activeCartType = type;
            const container = document.getElementById('cart-items-container');
            container.innerHTML = '';
            
            const title = document.getElementById('cart-modal-title');
            const partnerLabel = document.getElementById('cart-partner-label');
            const partnerInput = document.getElementById('cart-partner-input');
            const reasonInput = document.getElementById('cart-reason-input');
            const noteInput = document.getElementById('cart-note-input');
            
            const cart = type === 'purchase' ? purchaseCart : dispatchCart;
            
            if (type === 'purchase') {
                title.textContent = '📥 Giỏ Hàng Nhập Kho';
                partnerLabel.textContent = 'Nhà cung cấp *';
                partnerInput.placeholder = 'Ví dụ: Công ty Dược CDC, ...';
                reasonInput.value = 'Nhập qua điện thoại';
            } else {
                title.textContent = '📤 Giỏ Hàng Xuất Kho';
                partnerLabel.textContent = 'Đơn vị nhận *';
                partnerInput.placeholder = 'Ví dụ: Khoa dược, CDC chi nhánh, ...';
                reasonInput.value = 'Xuất qua điện thoại';
            }
            noteInput.value = '';
            partnerInput.value = '';
            
            if (cart.length === 0) {
                container.innerHTML = '<div class="no-result">Giỏ hàng trống.</div>';
            } else {
                cart.forEach((item, index) => {
                    const row = document.createElement('div');
                    row.className = 'cart-item-row';
                    row.innerHTML = `
                        <div class="cart-item-details">
                            <div class="cart-item-name">${item.productName}</div>
                            <div class="cart-item-meta">Lô: ${item.lotNo} | SL: ${item.qty} ${item.unit} ${item.expiryDate ? ' | HSD: ' + item.expiryDate : ''}</div>
                        </div>
                        <button class="btn-cart-remove" onclick="removeFromCart('${type}', ${index})">❌</button>
                    `;
                    container.appendChild(row);
                });
            }
            
            document.getElementById('btn-cart-submit-pc').onclick = () => submitCart(type, 'pc');
            document.getElementById('btn-cart-submit-phone').onclick = () => submitCart(type, 'phone');
            
            document.getElementById('cart-modal').style.display = 'flex';
        }
        
        function closeCartModal() {
            document.getElementById('cart-modal').style.display = 'none';
        }
        
        function removeFromCart(type, index) {
            if (type === 'purchase') {
                purchaseCart.splice(index, 1);
                localStorage.setItem('mob_purchase_cart', JSON.stringify(purchaseCart));
            } else {
                dispatchCart.splice(index, 1);
                localStorage.setItem('mob_dispatch_cart', JSON.stringify(dispatchCart));
            }
            updateCartStatus();
            if (document.getElementById('cart-modal').style.display === 'flex') {
                openCartModal(type);
            }
        }

        function submitCart(type, printTarget) {
            const cart = type === 'purchase' ? purchaseCart : dispatchCart;
            if (cart.length === 0) {
                showToast("Giỏ hàng đang trống!", "error");
                return;
            }
            
            const partner = document.getElementById('cart-partner-input').value.trim();
            const reason = document.getElementById('cart-reason-input').value.trim();
            const note = document.getElementById('cart-note-input').value.trim();
            
            if (!partner || !reason) {
                showToast("Vui lòng nhập đầy đủ đối tác và lý do thực hiện!", "error");
                return;
            }
            
            const url = type === 'purchase' ? '/api/purchase' : '/api/dispatch';
            const bodyData = {
                items: cart,
                reason: reason,
                note: note
            };
            if (type === 'purchase') {
                bodyData.supplier = partner;
            } else {
                bodyData.receivingUnit = partner;
            }
            
            const submitBtnPc = document.getElementById('btn-cart-submit-pc');
            const submitBtnPhone = document.getElementById('btn-cart-submit-phone');
            const oldPcText = submitBtnPc.innerHTML;
            const oldPhoneText = submitBtnPhone.innerHTML;
            
            submitBtnPc.disabled = true;
            submitBtnPhone.disabled = true;
            submitBtnPc.innerHTML = 'Đang xử lý...';
            
            fetch(url, {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify(bodyData)
            })
            .then(res => res.json())
            .then(data => {
                submitBtnPc.disabled = false;
                submitBtnPhone.disabled = false;
                submitBtnPc.innerHTML = oldPcText;
                submitBtnPhone.innerHTML = oldPhoneText;
                
                if (data.success) {
                    showToast(data.message, "success");
                    closeCartModal();
                    
                    if (type === 'purchase') {
                        purchaseCart = [];
                        localStorage.removeItem('mob_purchase_cart');
                    } else {
                        dispatchCart = [];
                        localStorage.removeItem('mob_dispatch_cart');
                    }
                    updateCartStatus();
                    loadDashboardStats();
                    
                    const noteId = type === 'purchase' ? data.purchaseId : data.dispatchId;
                    if (printTarget === 'pc') {
                        showToast("Đang gửi lệnh in tới máy tính...", "info");
                        fetch(`/api/pc-print?type=${type}&id=${noteId}`)
                            .then(res => res.json())
                            .then(pdata => {
                                if (pdata.success) {
                                    showToast("Máy tính đã nhận lệnh in!", "success");
                                } else {
                                    showToast("Lỗi in PC: " + pdata.message, "error");
                                }
                            })
                            .catch(err => showToast("Lỗi gửi lệnh in PC", "error"));
                    } else {
                        const printUrl = type === 'purchase' ? `/api/print-purchase?id=${noteId}` : `/api/print-dispatch?id=${noteId}`;
                        window.open(printUrl, '_blank');
                    }
                } else {
                    showToast(data.message, "error");
                }
            })
            .catch(err => {
                submitBtnPc.disabled = false;
                submitBtnPhone.disabled = false;
                submitBtnPc.innerHTML = oldPcText;
                submitBtnPhone.innerHTML = oldPhoneText;
                showToast("Lỗi kết nối máy chủ", "error");
            });
        }
        // --- Note Preview Modal Javascript ---
        function showNotePreview(type, noteId) {
            const modal = document.getElementById('preview-modal');
            const title = document.getElementById('preview-modal-title');
            const infoContainer = document.getElementById('preview-info-container');
            const tableBody = document.getElementById('preview-table-body');
            
            title.textContent = "📋 Đang tải phiếu...";
            infoContainer.innerHTML = '<div style="text-align: center; padding: 10px;">Đang truy vấn thông tin...</div>';
            tableBody.innerHTML = '<tr><td colspan="3" style="text-align: center; padding: 10px;">Đang tải danh sách mặt hàng...</td></tr>';
            
            modal.style.display = 'flex';
            
            fetch(`/api/note-details?type=${type}&id=${noteId}`)
                .then(res => res.json())
                .then(data => {
                    if (!data.success) {
                        showToast(data.message, "error");
                        modal.style.display = 'none';
                        return;
                    }
                    
                    title.textContent = data.type === 'nhap' ? '📥 Xem Trước Phiếu Nhập' : '📤 Xem Trước Phiếu Xuất';
                    
                    infoContainer.innerHTML = `
                        <div style="margin-bottom: 4px;"><b>Số phiếu:</b> <span style="color: var(--primary); font-weight: bold;">${data.noteNumber}</span></div>
                        <div style="margin-bottom: 4px;"><b>Thời gian:</b> ${data.createdAt}</div>
                        <div style="margin-bottom: 4px;"><b>${data.type === 'nhap' ? 'Nhà cung cấp' : 'Đơn vị nhận'}:</b> ${data.partner}</div>
                        <div style="margin-bottom: 4px;"><b>Lý do:</b> ${data.reason}</div>
                        ${data.note ? `<div style="margin-bottom: 4px;"><b>Ghi chú:</b> ${data.note}</div>` : ''}
                    `;
                    
                    let rowsHtml = '';
                    data.items.forEach(item => {
                        rowsHtml += `
                            <tr>
                                <td style="padding: 8px 10px;">${item.productName}</td>
                                <td style="padding: 8px 10px; text-align: center;">${item.lotNo}</td>
                                <td style="padding: 8px 10px; text-align: right; font-weight: 600;">${item.qty} ${item.unit}</td>
                            </tr>
                        `;
                    });
                    tableBody.innerHTML = rowsHtml;
                    
                    // Bind actions
                    document.getElementById('btn-preview-submit-pc').onclick = () => {
                        const btn = document.getElementById('btn-preview-submit-pc');
                        btn.disabled = true;
                        const origText = btn.innerHTML;
                        btn.innerHTML = '⌛...';
                        
                        fetch(`/api/pc-print?type=${type}&id=${noteId}`)
                            .then(res => res.json())
                            .then(pdata => {
                                btn.disabled = false;
                                btn.innerHTML = origText;
                                showToast(pdata.message, pdata.success ? "success" : "error");
                                if (pdata.success) modal.style.display = 'none';
                            })
                            .catch(err => {
                                btn.disabled = false;
                                btn.innerHTML = origText;
                                showToast("Lỗi gửi lệnh in PC", "error");
                            });
                    };
                    
                    document.getElementById('btn-preview-submit-phone').onclick = () => {
                        const url = data.type === 'nhap' ? `/api/print-purchase?id=${noteId}` : `/api/print-dispatch?id=${noteId}`;
                        window.open(url, '_blank');
                        modal.style.display = 'none';
                    };
                })
                .catch(err => {
                    showToast("Lỗi kết nối máy chủ", "error");
                    modal.style.display = 'none';
                });
        }
        
        function closePreviewModal() {
            document.getElementById('preview-modal').style.display = 'none';
        }
        // --- End Note Preview Modal Javascript ---

        // Load Stock Data
        function checkStock(barcode) {
            if (!barcode) return;
            
            resultCard.style.display = 'block';
            closeForms();
            resultContent.innerHTML = `
                <div class="loading">
                    <div class="loading-spinner"></div>
                    Đang truy vấn dữ liệu kho...
                </div>
            `;

            fetch(`/api/stock?barcode=${encodeURIComponent(barcode)}`)
                .then(res => {
                    if (!res.ok) {
                        return res.json().then(err => { throw new Error(err.message || 'Không tìm thấy sản phẩm') });
                    }
                    return res.json();
                })
                .then(data => {
                    if (!data.success) {
                        showNoResult();
                        return;
                    }
                    currentProduct = data.product;
                    currentProductBatches = data.batches;
                    displayResult(data);
                })
                .catch(err => {
                    showError(err.message);
                });
        }

        function showNoResult() {
            currentProduct = null;
            currentProductBatches = [];
            resultContent.innerHTML = `
                <div class="no-result">
                    ❌ Không tìm thấy sản phẩm trùng khớp.
                </div>
            `;
        }

        function showError(msg) {
            currentProduct = null;
            currentProductBatches = [];
            resultContent.innerHTML = `
                <div class="error-msg">
                    ⚠ Lỗi: ${msg}
                </div>
            `;
        }

        function displayResult(data) {
            const p = data.product;
            const batches = data.batches;
            
            let typeText = "Thuốc / Dược phẩm";
            if (p.type === 'vaccine') typeText = "Vaccine";
            else if (p.type === 'vtyt') typeText = "Vật tư y tế";
            else if (p.type === 'khac') typeText = "Sản phẩm khác";

            let batchesHtml = '';
            if (batches.length === 0) {
                batchesHtml = '<div class="no-result" style="padding: 10px;">Sản phẩm hiện hết hàng hoặc chưa nhập lô.</div>';
            } else {
                batches.forEach(b => {
                    const expDate = new Date(b.expiryDate);
                    const today = new Date();
                    const diffTime = expDate - today;
                    const diffDays = Math.ceil(diffTime / (1000 * 60 * 60 * 24));
                    
                    let badgeHtml = '';
                    if (diffDays <= 0) {
                        badgeHtml = '<span class="badge badge-expired">Hết hạn</span>';
                    } else if (diffDays <= 90) {
                        badgeHtml = `<span class="badge badge-warning">Cận hạn (${diffDays} ngày)</span>`;
                    } else {
                        badgeHtml = '<span class="badge badge-ok">Hạn tốt</span>';
                    }

                    batchesHtml += `
                        <div class="batch-item">
                            <div class="batch-header">
                                <span class="batch-lot">Lô: ${b.lotNo}</span>
                                <span class="batch-qty">${b.qty} ${p.unit}</span>
                            </div>
                            <div class="batch-expiry">
                                <span>Hạn dùng: ${b.expiryDate}</span>
                                ${badgeHtml}
                            </div>
                        </div>
                    `;
                });
            }

            resultContent.innerHTML = `
                <div class="product-info">
                    <div class="info-row">
                        <span class="info-label">Tên sản phẩm</span>
                        <span class="info-value" style="color: #a5b4fc; text-align: right; max-width: 65%;">${p.name}</span>
                    </div>
                    <div class="info-row">
                        <span class="info-label">Mã vạch</span>
                        <span class="info-value">${p.barcode || 'Chưa gán'}</span>
                    </div>
                    <div class="info-row">
                        <span class="info-label">Phân loại</span>
                        <span class="info-value">${typeText}</span>
                    </div>
                    <div class="info-row" style="margin-top: 5px; border-top: 1px solid rgba(255,255,255,0.05); padding-top: 8px;">
                        <span class="info-label" style="font-weight: bold; color: #fff;">Tổng tồn kho</span>
                        <span class="info-value" style="color: var(--success); font-size: 1.1rem;">${data.totalQty} ${p.unit}</span>
                    </div>
                </div>
                
                <div class="result-title" style="font-size: 0.95rem; border: none; margin-top: 12px; margin-bottom: 5px; padding: 0;">📦 Chi tiết tồn kho theo lô</div>
                <div class="batch-list" style="max-height: 200px; overflow-y: auto;">
                    ${batchesHtml}
                </div>
                
                <!-- Action Buttons Grid -->
                <div class="action-buttons">
                    <button class="action-btn btn-purchase" onclick="openForm('purchase')">📥 Nhập Kho</button>
                    <button class="action-btn btn-dispatch" onclick="openForm('dispatch')">📤 Xuất Kho</button>
                    <button class="action-btn btn-barcode" onclick="openForm('barcode')">🏷️ Gán / Liên Kết Mã Vạch</button>
                </div>
            `;
        }

        // Quick forms logic
        function closeForms() {
            document.querySelectorAll('.form-container').forEach(el => el.style.display = 'none');
        }

        function openForm(type) {
            closeForms();
            const form = document.getElementById(`form-${type}`);
            form.style.display = 'block';
            form.scrollIntoView({ behavior: 'smooth', block: 'end' });
            
            if (type === 'purchase') {
                document.getElementById('pur-qty').value = '';
                document.getElementById('pur-lot').value = '';
                document.getElementById('pur-expiry').value = '';
            } else if (type === 'dispatch') {
                document.getElementById('disp-qty').value = '';
                const select = document.getElementById('disp-batch-id');
                select.innerHTML = '';
                
                if (currentProductBatches.length === 0) {
                    select.innerHTML = '<option value="">(Không có lô hàng nào còn tồn)</option>';
                } else {
                    currentProductBatches.forEach(b => {
                        const opt = document.createElement('option');
                        opt.value = b.lotNo;
                        opt.textContent = `Lô: ${b.lotNo} (Còn tồn: ${b.qty})`;
                        select.appendChild(opt);
                    });
                        // Ah! Let's check what the API /api/stock returns for batches list:
                        // "SELECT b.id, b.lotNo, b.expiryDate, COALESCE(SUM(sm.qty), 0) as qtyBase" - yes! Let's ensure the backend returned `id`.
                        // Yes! Let's check python query for batches:
                        // `SELECT b.lotNo, b.expiryDate, COALESCE(SUM(sm.qty), 0) as qtyBase` - Ah! We did NOT query `b.id` in `batches_rows`!
                        // Let's fix that below if needed, or query by lotNo. Since lotNo + productId is UNIQUE, we can query by lotNo as well, but querying by batchId is cleaner.
                        // Let's check the python batches query: it uses `SELECT b.lotNo, b.expiryDate` but we can also use lotNo as value or query lotNo.
                        // Let's use lotNo as select value. In python `POST /api/dispatch`, the client can pass lotNo!
                        // Wait, our `POST /api/dispatch` API endpoint currently reads: `batch_id = data.get("batchId")`.
                        // Let's modify the python handler to read:
                        // `batch_id = data.get("batchId")` or look up by `lotNo` if `lotNo` is provided.
                        // Actually, let's look at `POST /api/dispatch` logic in python:
                        // `batch = conn.execute("SELECT id, lotNo, expiryDate FROM batches WHERE productId=? AND lotNo=?", ...)` or similar.
                        // Let's modify python's `POST /api/dispatch` to accept `lotNo` instead of `batchId` so it's extremely robust and simple!
                        // Yes, we will read `lotNo = data.get("lotNo")` or `batchId`. If they choose lotNo, we find `id` from `batches` where `productId=? AND lotNo=?`.
                        // Let's support BOTH in python so the HTML client is extremely easy to write.
                    // });
                }
            } else if (type === 'barcode') {
                document.getElementById('link-barcode').value = barcodeInput.value || '';
            }
        }

        // Submit Quick Purchase
        function submitPurchase() {
            if (!currentProduct) return;
            const qty = parseFloat(document.getElementById('pur-qty').value);
            const lotNo = document.getElementById('pur-lot').value.trim();
            const expiry = document.getElementById('pur-expiry').value;
            
            if (!qty || qty <= 0 || !lotNo || !expiry) {
                showToast("Vui lòng điền đầy đủ và chính xác thông tin!", "error");
                return;
            }
            
            fetch('/api/purchase', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({
                    productId: currentProduct.id,
                    qty: qty,
                    lotNo: lotNo,
                    expiryDate: expiry
                })
            })
            .then(res => res.json())
            .then(data => {
                if (data.success) {
                    showToast(data.message, "success");
                    closeForms();
                    checkStock(currentProduct.barcode || currentProduct.id); // Reload
                    showNotePreview('purchase', data.purchaseId);
                    loadDashboardStats();
                } else {
                    showToast(data.message, "error");
                }
            })
            .catch(err => showToast("Lỗi kết nối máy chủ", "error"));
        }

        // Submit Quick Dispatch
        function submitDispatch() {
            if (!currentProduct) return;
            const lotNo = document.getElementById('disp-batch-id').value;
            const qty = parseFloat(document.getElementById('disp-qty').value);
            const reason = document.getElementById('disp-reason').value.trim();
            
            if (!lotNo || !qty || qty <= 0) {
                showToast("Vui lòng nhập đầy đủ thông tin!", "error");
                return;
            }
            
            fetch('/api/dispatch', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({
                    productId: currentProduct.id,
                    lotNo: lotNo, // We will send lotNo
                    qty: qty,
                    reason: reason
                })
            })
            .then(res => res.json())
            .then(data => {
                if (data.success) {
                    showToast(data.message, "success");
                    closeForms();
                    checkStock(currentProduct.barcode || currentProduct.id); // Reload
                    showNotePreview('dispatch', data.dispatchId);
                    loadDashboardStats();
                } else {
                    showToast(data.message, "error");
                }
            })
            .catch(err => showToast("Lỗi kết nối máy chủ", "error"));
        }

        // Submit Link Barcode
        function submitLinkBarcode() {
            if (!currentProduct) return;
            const barcode = document.getElementById('link-barcode').value.trim();
            
            if (!barcode) {
                showToast("Vui lòng nhập hoặc quét mã vạch!", "error");
                return;
            }
            
            fetch('/api/update-barcode', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({
                    productId: currentProduct.id,
                    barcode: barcode
                })
            })
            .then(res => res.json())
            .then(data => {
                if (data.success) {
                    showToast(data.message, "success");
                    closeForms();
                    barcodeInput.value = barcode; // Update input
                    checkStock(barcode); // Reload product details
                } else {
                    showToast(data.message, "error");
                }
            })
            .catch(err => showToast("Lỗi kết nối máy chủ", "error"));
        }

        // Load Catalog List
        let activeCatalogFilter = 'all';
        function loadCatalog(query = '', filterType = 'all') {
            activeCatalogFilter = filterType;
            
            // Update filter buttons UI styling
            const btnAll = document.getElementById('filter-btn-all');
            const btnOutOfStock = document.getElementById('filter-btn-outofstock');
            const btnExpiring = document.getElementById('filter-btn-expiring');
            
            if (btnAll && btnOutOfStock && btnExpiring) {
                // Reset styles
                btnAll.style.background = activeCatalogFilter === 'all' ? 'var(--primary)' : '#fff';
                btnAll.style.color = activeCatalogFilter === 'all' ? '#fff' : 'var(--text-muted)';
                btnAll.style.border = activeCatalogFilter === 'all' ? '1px solid var(--primary)' : '1px solid var(--glass-border)';
                
                btnOutOfStock.style.background = activeCatalogFilter === 'outofstock' ? '#ef4444' : '#fff';
                btnOutOfStock.style.color = activeCatalogFilter === 'outofstock' ? '#fff' : '#ef4444';
                btnOutOfStock.style.border = activeCatalogFilter === 'outofstock' ? '1px solid #ef4444' : '1px solid var(--glass-border)';
                
                btnExpiring.style.background = activeCatalogFilter === 'expiring' ? '#f59e0b' : '#fff';
                btnExpiring.style.color = activeCatalogFilter === 'expiring' ? '#fff' : '#f59e0b';
                btnExpiring.style.border = activeCatalogFilter === 'expiring' ? '1px solid #f59e0b' : '1px solid var(--glass-border)';
            }
            
            const list = document.getElementById('catalog-list');
            list.innerHTML = `
                <div class="loading">
                    <div class="loading-spinner"></div>
                    Đang tải danh sách...
                </div>
            `;
            
            let url = `/api/products?q=${encodeURIComponent(query)}`;
            if (filterType !== 'all') {
                url += `&filter=${filterType}`;
            }
            
            fetch(url)
                .then(res => res.json())
                .then(data => {
                    if (!data.success || data.products.length === 0) {
                        list.innerHTML = '<div class="no-result">Không tìm thấy sản phẩm nào.</div>';
                        return;
                    }
                    
                    let html = '';
                    data.products.forEach(p => {
                        html += `
                            <div class="product-item" onclick="selectProductFromCatalog('${p.barcode || p.id}')">
                                <div class="product-item-details">
                                    <span class="product-item-name">${p.name}</span>
                                    <span class="product-item-sub">ĐVT: ${p.unit} ${p.barcode ? ' | Mã vạch: ' + p.barcode : ''}</span>
                                </div>
                                <span class="product-item-arrow">➔</span>
                            </div>
                        `;
                    });
                    list.innerHTML = html;
                })
                .catch(err => {
                    list.innerHTML = '<div class="error-msg">Không thể tải danh sách sản phẩm.</div>';
                });
        }

        function filterCatalog(type) {
            switchTab('tab-catalog');
            document.getElementById('catalog-search').value = '';
            loadCatalog('', type);
        }

        function loadDashboardStats() {
            const refreshBtn = document.getElementById('dashboard-refresh');
            if (refreshBtn) refreshBtn.textContent = '⌛...';
            
            fetch('/api/dashboard-stats')
                .then(res => res.json())
                .then(data => {
                    if (refreshBtn) refreshBtn.innerHTML = '🔄 Cập nhật';
                    if (data.success) {
                        document.getElementById('dash-total-products').textContent = data.totalProducts;
                        document.getElementById('dash-outofstock-products').textContent = data.outofstockProducts;
                        document.getElementById('dash-expiring-products').textContent = data.expiringProducts;
                    }
                })
                .catch(err => {
                    if (refreshBtn) refreshBtn.innerHTML = '🔄 Cập nhật';
                    console.error("Lỗi tải dashboard stats:", err);
                });
        }

        function selectProductFromCatalog(identifier) {
            barcodeInput.value = identifier;
            switchTab('tab-checker');
            checkStock(identifier);
        }

        // Create New Product functions
        function openCreateProductForm() {
            const form = document.getElementById('form-create-product');
            form.style.display = 'block';
            document.getElementById('new-name').value = '';
            document.getElementById('new-unit').value = '';
            document.getElementById('new-type').value = 'thuoc';
            document.getElementById('new-barcode').value = '';
            document.getElementById('new-regnumber').value = '';
            form.scrollIntoView({ behavior: 'smooth', block: 'end' });
        }

        function closeCreateProductForm() {
            document.getElementById('form-create-product').style.display = 'none';
        }

        function submitCreateProduct() {
            const name = document.getElementById('new-name').value.trim();
            const unit = document.getElementById('new-unit').value.trim();
            const type = document.getElementById('new-type').value;
            const barcode = document.getElementById('new-barcode').value.trim();
            const regNumber = document.getElementById('new-regnumber').value.trim();
            
            if (!name || !unit) {
                showToast("Vui lòng nhập tên và đơn vị tính gốc!", "error");
                return;
            }
            
            fetch('/api/create-product', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({
                    name: name,
                    defaultUnit: unit,
                    productType: type,
                    barcode: barcode,
                    registrationNumber: regNumber
                })
            })
            .then(res => res.json())
            .then(data => {
                if (data.success) {
                    showToast(data.message, "success");
                    closeCreateProductForm();
                    loadDashboardStats();
                    
                    // Tự động chuyển qua tab Kiểm kho và tra cứu sản phẩm vừa tạo
                    const identifier = data.barcode || barcode || data.productId;
                    barcodeInput.value = identifier;
                    switchTab('tab-checker');
                    checkStock(identifier);
                    
                    // Tự động kích hoạt màn hình nhập kho
                    setTimeout(() => {
                        openForm('purchase');
                    }, 600);
                } else {
                    showToast(data.message, "error");
                }
            })
            .catch(err => showToast("Lỗi kết nối máy chủ", "error"));
        }

        // Setup Event Listeners
        searchBtn.addEventListener('click', () => {
            checkStock(barcodeInput.value.trim());
        });

        barcodeInput.addEventListener('keypress', (e) => {
            if (e.key === 'Enter') {
                checkStock(barcodeInput.value.trim());
            }
        });

        document.getElementById('catalog-search-btn').addEventListener('click', () => {
            loadCatalog(document.getElementById('catalog-search').value.trim());
        });

        document.getElementById('catalog-search').addEventListener('keypress', (e) => {
            if (e.key === 'Enter') {
                loadCatalog(document.getElementById('catalog-search').value.trim());
            }
        });

        // Setup Html5QrcodeScanner
        let lastScannedCode = "";
        let scanTime = 0;

        function onScanSuccess(decodedText, decodedResult) {
            const now = Date.now();
            if (decodedText === lastScannedCode && (now - scanTime < 2500)) {
                return;
            }
            lastScannedCode = decodedText;
            scanTime = now;
            
            // Prefill barcode update if form open, otherwise check stock
            const barcodeForm = document.getElementById('form-barcode');
            const createForm = document.getElementById('form-create-product');
            if (barcodeForm.style.display === 'block') {
                document.getElementById('link-barcode').value = decodedText;
                showToast(`Đã quét mã mới: ${decodedText}`);
                if (navigator.vibrate) navigator.vibrate(100);
            } else if (createForm && createForm.style.display === 'block') {
                document.getElementById('new-barcode').value = decodedText;
                showToast(`Đã quét mã sản phẩm mới: ${decodedText}`);
                if (navigator.vibrate) navigator.vibrate(100);
            } else {
                barcodeInput.value = decodedText;
                if (navigator.vibrate) navigator.vibrate(100);
                checkStock(decodedText);
            }
        }

        function onScanFailure(error) {}

        const html5QrcodeScanner = new Html5QrcodeScanner(
            "reader", 
            { 
                fps: 10, 
                qrbox: function(width, height) {
                    const size = Math.min(width, height) * 0.65;
                    return { width: size, height: size * 0.6 };
                },
                aspectRatio: 1.0,
                supportedScanTypes: [Html5QrcodeScanType.SCAN_TYPE_CAMERA]
            },
            false
        );
        html5QrcodeScanner.render(onScanSuccess, onScanFailure);
        updateCartStatus();
        loadDashboardStats();
    </script>
</body>
</html>"""

# --- /Mobile Inventory Web Server ---

if __name__ == '__main__':
    app = App()
    app.mainloop()

