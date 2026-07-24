# -*- coding: utf-8 -*-
# ui.py — Giao diện Desktop chính (Tkinter / ttkbootstrap)
import sqlite3
import tkinter as tk
from tkinter import messagebox, filedialog, simpledialog
import datetime as dt
import datetime
import os, webbrowser, tempfile
import shutil
import json
import threading
import time
from collections import defaultdict

import ttkbootstrap as tb
from ttkbootstrap.widgets import DateEntry
from ttkbootstrap.constants import *

# Import matplotlib for charts
try:
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    from matplotlib.figure import Figure
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

# Import pandas for Excel processing
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

# Import barcode scanner libraries
try:
    import cv2
    from pyzbar import pyzbar
    from PIL import Image, ImageTk
    BARCODE_AVAILABLE = True
except ImportError:
    BARCODE_AVAILABLE = False

# Import PDF export libraries
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

# Import config, database, managers, scanner, server
from config import (
    APP_NAME, APP_VERSION, AUTHOR_NAME, AUTHOR_EMAIL, AUTHOR_PHONE, AUTHOR_SITE,
    APP_DIR, DB_PATH, LOG_PATH, BACKUP_DIR,
    BARCODE_AVAILABLE, MATPLOTLIB_AVAILABLE, PANDAS_AVAILABLE, PDF_AVAILABLE
)
from database import DB
from managers import BackupManager, ExportManager, ReportManager, MedicineCatalogManager
from scanner import BarcodeScanner
from server import MobileInventoryServer, get_local_ip

# Thêm biến toàn cục cho qrcode
QR_CODE_AVAILABLE = False
try:
    import qrcode
    QR_CODE_AVAILABLE = True
except ImportError:
    pass

class App(tb.Window):
    def __init__(self):
        super().__init__(themename='flatly')  # Tông xanh dương nhạt thanh lịch và phẳng
        self.title(f'{APP_NAME} — v{APP_VERSION}')
        self.geometry('1180x840'); self.minsize(1100, 740)

        self.db = DB(DB_PATH)
        self.backup_manager = BackupManager(DB_PATH, BACKUP_DIR)
        self.report_manager = ReportManager(DB_PATH)
        self.export_manager = ExportManager()
        self.medicine_catalog = MedicineCatalogManager(DB_PATH)
        self.last_sale_items = []; self.cart = []

        self.make_style()
        
        # Thiết lập Server kiểm kho di động (Lỗi 3: Không tự động bật server)
        self.mobile_server = None
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        
        # Tạo UI sau khi window đã sẵn sàng
        self.after(100, self.initialize_ui)
        
        # Bắt đầu auto backup
        self.backup_manager.start_auto_backup()
    
    def initialize_ui(self):
        """Khởi tạo UI sau khi window đã sẵn sàng"""
        try:
            self.make_ui()
            # Load dữ liệu sau khi UI đã được tạo
            self.after(200, self.on_ready)
        except Exception as e:
            print(f"Lỗi khởi tạo UI: {e}")
            # Fallback: tạo UI cơ bản
            self.create_basic_ui()
    
    def create_basic_ui(self):
        """Tạo UI cơ bản nếu có lỗi"""
        self.title(f'{APP_NAME} — v{APP_VERSION} — Lỗi khởi tạo')
        tb.Label(self, text="Có lỗi xảy ra khi khởi tạo giao diện. Vui lòng khởi động lại ứng dụng.", 
                font=('Segoe UI', 12)).pack(expand=True)

    def start_mobile_server_bg(self):
        """Khởi động Web Server chạy ngầm cho điện thoại di động kết nối"""
        try:
            self.mobile_server = MobileInventoryServer(self, host="0.0.0.0", port=5000)
            self.mobile_server.start()
        except Exception as e:
            print(f"Lỗi khởi động máy chủ di động: {e}")

    def on_close(self):
        """Đóng ứng dụng và giải phóng máy chủ di động"""
        if hasattr(self, 'mobile_server') and self.mobile_server:
            try:
                self.mobile_server.stop()
            except:
                pass
        self.destroy()

    # theme & fonts
    def make_style(self):
        style = tb.Style()
        # Ẩn header Notebook để tránh “trùng thẻ”
        style.layout('TNotebook.Tab', [])
        # Font to hơn
        style.configure('TLabel', font=('Segoe UI', 11))
        style.configure('TButton', font=('Segoe UI', 11))
        style.configure('TEntry',  font=('Segoe UI', 11))
        style.configure('TCombobox', font=('Segoe UI', 11))
        style.configure('Treeview', rowheight=30, font=('Segoe UI', 11))
        style.configure('Treeview.Heading', font=('Segoe UI', 11, 'bold'))

    # helpers
    def _numberize(self, entry: tb.Entry):
        entry.config(justify='right')
        entry.bind('<FocusIn>', lambda e: entry.selection_range(0, 'end'))
    def open_combo(self, combo):
        combo.focus_set()
        combo.event_generate('<Alt-Down>')
    def _open_dropdown(self, combo: tb.Combobox):
        combo.focus_set()
        # kích hoạt menu xổ xuống (Alt+Down)
        combo.event_generate('<Alt-Down>')

    def toast(self, text, ms=1600):
        w = tk.Toplevel(self); w.wm_overrideredirect(True); w.configure(bg='#111')
        tk.Label(w, text='  ' + text + '  ', bg='#111', fg='white', font=('Segoe UI', 10)).pack()
        w.update_idletasks()
        x = self.winfo_x() + self.winfo_width() - w.winfo_width() - 20
        y = self.winfo_y() + self.winfo_height() - w.winfo_height() - 20
        w.geometry(f'+{x}+{y}'); w.after(ms, w.destroy)

    # layout
    def make_ui(self):
        menubar = tk.Menu(self); self.config(menu=menubar)
        helpm = tk.Menu(menubar, tearoff=0); helpm.add_command(label='Phím tắt', command=self.show_shortcuts)
        helpm.add_separator()
        helpm.add_command(label='Mở thư mục dữ liệu', command=self.open_data_folder)
        helpm.add_command(label='Giới thiệu (About)…', command=self.show_about)
        menubar.add_cascade(label='Trợ giúp', menu=helpm)
        
        self.nb = tb.Notebook(self); self.nb.pack(fill=BOTH, expand=True, padx=8, pady=(0,8))
        
        # Tạo các tab frames
        self.tab_products = tb.Frame(self.nb)
        self.tab_purchase = tb.Frame(self.nb)
        self.tab_dispatch = tb.Frame(self.nb)
        self.tab_stock = tb.Frame(self.nb)
        self.tab_alerts = tb.Frame(self.nb)
        self.tab_report = tb.Frame(self.nb)
        self.tab_backup = tb.Frame(self.nb)
        self.tab_advanced_reports = tb.Frame(self.nb)
        self.tab_mobile = tb.Frame(self.nb)
        self.tab_temp_log = tb.Frame(self.nb)
        
        # Thêm tabs với labels đẹp hơn
        tabs_config = [
            (self.tab_products, "🏷️ Sản phẩm"),
            (self.tab_purchase, "📦 Nhập kho"),
            (self.tab_dispatch, "📤 Xuất kho / Cấp phát"),
            (self.tab_stock, "📊 Tồn kho"),
            (self.tab_alerts, "⏰ Hết hạn"),
            (self.tab_report, "📄 Báo cáo XNT"),
            (self.tab_backup, "💾 Backup"),
            (self.tab_advanced_reports, "📈 Báo cáo nâng cao"),
            (self.tab_mobile, "📱 Kiểm kho di động"),
            (self.tab_temp_log, "🌡️ Nhật ký nhiệt độ")
        ]
        
        for tab, label in tabs_config:
            self.nb.add(tab, text=label)
        
        # Tạo toolbar sau khi đã có các tabs
        self.create_toolbar()
        self.nb.bind("<<NotebookTabChanged>>", self.on_tab_changed)
        
        self.build_products_tab(); self.build_purchase_tab(); self.build_dispatch_tab()
        self.build_stock_tab(); self.build_alerts_tab(); self.build_report_tab()
        self.build_backup_tab(); self.build_advanced_reports_tab(); self.build_mobile_tab()
        self.build_temp_log_tab()
        # Status bar với thông tin chi tiết hơn
        status_frame = tb.Frame(self)
        status_frame.pack(fill='x', side='bottom', padx=8, pady=4)
        
        # Status chính
        self.status = tb.Label(status_frame, anchor='w', font=('Segoe UI', 9),
            text='Sẵn sàng • F1-F8, F10-F11: Chuyển tab • F9: In phiếu xuất kho • Ctrl+F: Tìm kiếm')
        self.status.pack(side='left')
        
        # Thông tin database
        self.db_status = tb.Label(status_frame, anchor='e', font=('Segoe UI', 9),
            text='Database: Đang kết nối...')
        self.db_status.pack(side='right')
        
        # Hotkeys
        self.bind('<F1>', lambda e: self.nb.select(self.tab_products))
        self.bind('<F2>', lambda e: self.nb.select(self.tab_purchase))
        self.bind('<F3>', lambda e: self.nb.select(self.tab_dispatch))
        self.bind('<F4>', lambda e: self.nb.select(self.tab_stock))
        self.bind('<F5>', lambda e: self.nb.select(self.tab_alerts))
        self.bind('<F6>', lambda e: self.nb.select(self.tab_report))
        self.bind('<F7>', lambda e: self.nb.select(self.tab_backup))
        self.bind('<F8>', lambda e: self.nb.select(self.tab_advanced_reports))
        self.bind('<F10>', lambda e: self.nb.select(self.tab_mobile))
        self.bind('<F11>', lambda e: self.nb.select(self.tab_temp_log))
        self.bind('<Control-f>', lambda e: self.focus_search())
        self.bind('<F9>', lambda e: self.print_dispatch_note())
        self.bind('<Control-Return>', lambda e: self.confirm_dispatch())

        # Cập nhật status database
        self.update_db_status()
        
    def update_db_status(self):
        """Cập nhật trạng thái database"""
        try:
            # Đếm số sản phẩm
            product_count = self.db.conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]
            # Đếm số batch
            batch_count = self.db.conn.execute("SELECT COUNT(*) FROM batches").fetchone()[0]
            # Đếm số đơn hàng
            sale_count = self.db.conn.execute("SELECT COUNT(*) FROM sales").fetchone()[0]
            
            self.db_status.config(text=f"Database: {product_count} sản phẩm • {batch_count} lô • {sale_count} đơn hàng")
        except Exception as e:
            self.db_status.config(text=f"Database: Lỗi - {str(e)}")

    def create_toolbar(self):
        tbbar = tb.Frame(self); tbbar.pack(fill='x', padx=8, pady=(8,8))
        
        # Tạo frame chứa các nút chính
        main_buttons = tb.Frame(tbbar)
        main_buttons.pack(side='left')
        
        # Các nút chính với style cải thiện
        buttons_config = [
            ('🏷️ Sản phẩm', self.tab_products, 'F1'),
            ('📦 Nhập hàng', self.tab_purchase, 'F2'),
            ('📤 Xuất kho', self.tab_dispatch, 'F3'),
            ('📊 Tồn kho', self.tab_stock, 'F4'),
            ('⏰ Hết hạn', self.tab_alerts, 'F5'),
            ('📄 Báo cáo XNT', self.tab_report, 'F6'),
            ('💾 Backup', self.tab_backup, 'F7'),
            ('📈 Báo cáo nâng cao', self.tab_advanced_reports, 'F8'),
            ('📱 Kiểm kho di động', self.tab_mobile, 'F10'),
            ('🌡️ Nhật ký nhiệt độ', self.tab_temp_log, 'F11')
        ]

        self.toolbar_buttons = {}
        for text, tab, shortcut in buttons_config:
            btn = tb.Button(main_buttons, text=text, bootstyle='outline-info',
                          command=lambda t=tab: self.nb.select(t))
            btn.pack(side='left', padx=2)
            self.toolbar_buttons[tab] = btn
            # Thêm tooltip với phím tắt
            self.create_tooltip(btn, f"{text} ({shortcut})")
        
        # Spacer
        tb.Label(tbbar, text='').pack(side='left', expand=True)
        
        # Nút in phiếu xuất kho
        print_btn = tb.Button(tbbar, text='🖨️ In phiếu xuất kho', bootstyle='outline-primary',
                             command=self.print_dispatch_note)
        print_btn.pack(side='right', padx=4)
        self.create_tooltip(print_btn, "In phiếu xuất kho (F9)")
        
        # Nút backup nhanh
        backup_btn = tb.Button(tbbar, text='💾 Backup nhanh', bootstyle='outline-success',
                               command=self.create_manual_backup)
        backup_btn.pack(side='right', padx=4)
        self.create_tooltip(backup_btn, "Tạo backup ngay lập tức")

    def on_tab_changed(self, event=None):
        try:
            selected_tab = self.nb.select()
            if not selected_tab:
                return
            selected_widget = self.nametowidget(selected_tab)
            for tab, btn in self.toolbar_buttons.items():
                if tab == selected_widget:
                    btn.configure(bootstyle='info')
                else:
                    btn.configure(bootstyle='outline-info')
        except Exception as e:
            print("Error updating active tab highlight:", e)

    def create_tooltip(self, widget, text):
        """Tạo tooltip cho widget"""
        def show_tooltip(event):
            tooltip = tb.Toplevel()
            tooltip.wm_overrideredirect(True)
            tooltip.wm_geometry(f"+{event.x_root+10}+{event.y_root+10}")
            label = tb.Label(tooltip, text=text, background='#ffffe0', 
                           relief='solid', borderwidth=1, font=('Segoe UI', 9))
            label.pack()
            widget.tooltip = tooltip
        
        def hide_tooltip(event):
            if hasattr(widget, 'tooltip'):
                widget.tooltip.destroy()
                del widget.tooltip
        
        widget.bind('<Enter>', show_tooltip)
        widget.bind('<Leave>', hide_tooltip)

    def show_shortcuts(self):
        messagebox.showinfo('Phím tắt',
            'F1: Tạo sản phẩm\nF2: Nhập hàng (Ctrl+F tìm)\nF3: POS (Ctrl+F tìm, Enter=Thêm vào giỏ, Ctrl+Enter=Thanh toán, F9=In)\nF4: Tồn theo lô\nF5: Sắp hết hạn\nF6: Báo cáo tồn kho')

    def focus_search(self):
        idx = self.nb.index(self.nb.select())
        if idx == 1 and hasattr(self, 'search_purchase'): self.search_purchase.focus_set()
        elif idx == 2 and hasattr(self, 'search_pos'): self.search_pos.focus_set()
    def open_data_folder(self):
        import sys, subprocess, os
        path = APP_DIR  # đã có sẵn từ phần cấu hình đường dẫn
        try:
            if sys.platform.startswith('win'):
                os.startfile(path)  # type: ignore
            elif sys.platform == 'darwin':
                subprocess.call(['open', path])
            else:
                subprocess.call(['xdg-open', path])
        except Exception as e:
            messagebox.showerror('Lỗi', str(e))

    def show_about(self):
        info = (
            f"{APP_NAME} v{APP_VERSION}\n"
            f"Tác giả: {AUTHOR_NAME}\n"
            f"Điện thoại: {AUTHOR_PHONE}\n"
            f"Email: {AUTHOR_EMAIL}\n"
            f"Website: {AUTHOR_SITE}\n\n"
            f"Thư mục dữ liệu: {APP_DIR}"
        )
        if messagebox.askyesno("Giới thiệu", info + "\n\nMở website tác giả?"):
            try:
                import webbrowser
                webbrowser.open(AUTHOR_SITE)
            except Exception:
                pass

    # -------- Products --------
    def build_products_tab(self):
        frm = self.tab_products
        
        # Khung quản lý danh mục thuốc với style cải thiện
        catalog_frame = tb.Labelframe(frm, text='📚 Quản lý danh mục thuốc', bootstyle='info')
        catalog_frame.pack(fill='x', padx=8, pady=8)
        
        catalog_btn_frame = tb.Frame(catalog_frame)
        catalog_btn_frame.pack(fill='x', padx=8, pady=8)
        
        tb.Button(catalog_btn_frame, text='📁 Load danh mục CSV/Excel', bootstyle='info',
                  command=self.load_medicine_catalog).pack(side='left', padx=4)
        tb.Button(catalog_btn_frame, text='📄 Load thuoc.csv', bootstyle='primary',
                  command=self.load_default_csv).pack(side='left', padx=4)
        tb.Button(catalog_btn_frame, text='🔍 Tra cứu thuốc', bootstyle='success',
                  command=self.search_medicine_dialog).pack(side='left', padx=4)
        tb.Button(catalog_btn_frame, text='ℹ️ Thông tin danh mục', bootstyle='secondary',
                  command=self.show_catalog_info).pack(side='left', padx=4)
        
        # Hiển thị thông tin danh mục hiện tại
        self.catalog_info_label = tb.Label(catalog_frame, text='Chưa load danh mục thuốc', 
                                          font=('Segoe UI', 9), bootstyle='secondary')
        self.catalog_info_label.pack(anchor='w', padx=8, pady=(0,4))
        
        # Hướng dẫn sử dụng
        help_label = tb.Label(catalog_frame, 
                             text='💡 Gợi ý: Load danh mục thuốc, sau đó gõ tên thuốc để xem gợi ý tự động', 
                             font=('Segoe UI', 8), bootstyle='info')
        help_label.pack(anchor='w', padx=8, pady=(0,8))
        
        # Khung thông tin sản phẩm với style cải thiện
        f1 = tb.Labelframe(frm, text='📝 Thông tin sản phẩm', bootstyle='light')
        f1.pack(fill='x', padx=8, pady=8)
        
        # Hàng 0: Nguồn sản phẩm
        tb.Label(f1, text='Nguồn sản phẩm:').grid(row=0, column=0, sticky='w', padx=6, pady=6)
        self.product_source_var = tk.StringVar(value='catalog')
        
        source_frame = tb.Frame(f1)
        source_frame.grid(row=0, column=1, columnspan=3, sticky='w', padx=6, pady=6)
        
        self.r_catalog = tb.Radiobutton(source_frame, text='Từ danh mục chuẩn', 
                                         variable=self.product_source_var, value='catalog',
                                         command=self.on_product_source_change)
        self.r_catalog.pack(side='left', padx=(0, 15))
        
        self.r_free = tb.Radiobutton(source_frame, text='Ngoài danh mục (Nhập tự do)', 
                                      variable=self.product_source_var, value='free',
                                      command=self.on_product_source_change)
        self.r_free.pack(side='left')
        
        # Hàng 1: Tên sản phẩm và loại
        tb.Label(f1, text='Tên sản phẩm:').grid(row=1, column=0, sticky='w', padx=6, pady=6)
        self.p_name = tb.Entry(f1, width=35)
        self.p_name.grid(row=1, column=1, padx=6, pady=6)
        self.p_name.bind('<KeyRelease>', self.on_product_name_change)
        self.p_name.bind('<FocusOut>', self.on_product_name_focus_out)
        
        # Tạo frame chứa dropdown gợi ý
        self.suggestions_frame = tb.Frame(f1)
        self.suggestions_frame.grid(row=2, column=1, sticky='ew', padx=6, pady=(0,6))
        self.suggestions_frame.grid_remove()  # Ẩn ban đầu
        
        # Tạo Listbox cho gợi ý
        self.suggestions_listbox = tk.Listbox(self.suggestions_frame, height=6, width=35)
        self.suggestions_listbox.pack(fill='both', expand=True)
        self.suggestions_listbox.bind('<Double-Button-1>', self.on_suggestion_selected)
        self.suggestions_listbox.bind('<Button-1>', self.on_suggestion_click)
        self.suggestions_listbox.bind('<ButtonRelease-1>', self.on_suggestion_click)
        self.suggestions_listbox.bind('<Return>', self.on_suggestion_selected)
        self.suggestions_listbox.bind('<Escape>', self.hide_suggestions)
        
        # Bind keyboard navigation
        self.p_name.bind('<Down>', self.on_arrow_down)
        self.p_name.bind('<Up>', self.on_arrow_up)
        self.suggestions_listbox.bind('<Up>', self.on_suggestion_up)
        self.suggestions_listbox.bind('<Down>', self.on_suggestion_down)
        
        tb.Label(f1, text='Loại sản phẩm:').grid(row=1, column=2, sticky='w', padx=6, pady=6)
        self.p_type = tb.Combobox(f1, values=['thuoc', 'vaccine', 'vtyt', 'khac'], state='readonly', width=12)
        self.p_type.set('thuoc')
        self.p_type.grid(row=1, column=3, padx=6, pady=6)
        self.p_type.bind('<<ComboboxSelected>>', self.on_product_type_change)
        
        # Hàng 2: Đơn vị và Barcode
        tb.Label(f1, text='Đơn vị cơ sở:').grid(row=3, column=0, sticky='w', padx=6, pady=6)
        self.p_base = tb.Entry(f1, width=12)
        self.p_base.insert(0, 'vien')
        self.p_base.grid(row=3, column=1, padx=6, pady=6)
        
        tb.Label(f1, text='Barcode:').grid(row=3, column=2, sticky='w', padx=6, pady=6)
        barcode_frame = tb.Frame(f1)
        barcode_frame.grid(row=3, column=3, padx=6, pady=6, sticky='ew')
        
        self.p_barcode = tb.Entry(barcode_frame, width=16)
        self.p_barcode.pack(side='left')
        
        # Nút quét barcode cho sản phẩm mới
        if BARCODE_AVAILABLE:
            tb.Button(barcode_frame, text='📷', command=self.scan_barcode_for_product, 
                     bootstyle='info', width=3).pack(side='left', padx=(5, 0))
        else:
            tb.Button(barcode_frame, text='📷', command=self.show_barcode_install_info, 
                     bootstyle='secondary', width=3).pack(side='left', padx=(5, 0))
        
        # Hàng 3: Số đăng ký (chỉ hiển thị khi chọn loại thuốc)
        self.p_reg_label = tb.Label(f1, text='Số đăng ký:')
        self.p_reg_label.grid(row=4, column=0, sticky='w', padx=6, pady=6)
        self.p_reg_label.grid_remove()  # Ẩn ban đầu
        
        self.p_reg_number = tb.Entry(f1, width=35)
        self.p_reg_number.grid(row=4, column=1, columnspan=2, padx=6, pady=6)
        self.p_reg_number.grid_remove()  # Ẩn ban đầu
        
        # Nút lưu
        btns = tb.Frame(frm)
        btns.pack(fill='x', padx=8, pady=8)
        tb.Button(btns, text='💾 Lưu sản phẩm', bootstyle='primary', command=self.save_product).pack(side='right')
        tb.Button(btns, text='📋 Tải Excel mẫu', bootstyle='outline-info', command=self.export_import_template).pack(side='left', padx=4)
        tb.Button(btns, text='📥 Nhập hàng loạt (Excel)', bootstyle='success', command=self.bulk_import_from_excel).pack(side='left', padx=4)
        
        # Thiết lập ban đầu
        self.on_product_source_change()

    def _set_entry_val(self, entry, value):
        """Helper gán giá trị cho Entry bất kể trạng thái readonly"""
        try:
            state = entry.cget('state')
            entry.config(state='normal')
            entry.delete(0, tk.END)
            if value:
                entry.insert(0, value)
            entry.config(state=state)
        except Exception as e:
            print(f"Lỗi set value cho entry: {e}")

    def on_product_source_change(self):
        """Xử lý khi thay đổi nguồn sản phẩm (danh mục mẫu / tự do)"""
        source = self.product_source_var.get()
        if source == 'catalog':
            self.p_base.config(state='readonly')
            self.p_reg_number.config(state='readonly')
        else:
            self.p_base.config(state='normal')
            self.p_reg_number.config(state='normal')

    def save_product(self):
        name = self.p_name.get().strip()
        base = self.p_base.get().strip() or 'vien'
        bc = self.p_barcode.get().strip() or None
        product_type = self.p_type.get()
        reg_number = self.p_reg_number.get().strip() or None if product_type in ('thuoc', 'vaccine') else None
        
        if not name: 
            messagebox.showerror('Lỗi','Nhập tên sản phẩm')
            return
        
        pid = self.db.ex("INSERT INTO products(name, defaultUnit, barcode, productType, registrationNumber) VALUES(?,?,?,?,?)", 
                        (name, base, bc, product_type, reg_number))
        
        # bảo đảm product_units base tồn tại
        try:
            self.db.ex("INSERT INTO product_units(productId, unitCode, toBaseQty, price) VALUES(?,?,1,0)", (pid, base))
        except sqlite3.IntegrityError:
            pass
        
        self.toast(f'Đã tạo sản phẩm #{pid}')
        try:
            self.db.add_audit_log(
                action="TAO_SAN_PHAM",
                details=f"Tạo sản phẩm mới thành công: {name} (#{pid}), ĐVCS: {base}, Loại: {product_type}, SDK: {reg_number}"
            )
        except Exception as log_err:
            print(f"Lỗi ghi log tao san pham: {log_err}")
        self.refresh_products()
        
        # Clear form
        self.p_name.delete(0, tk.END)
        self.p_barcode.delete(0, tk.END)
        self._set_entry_val(self.p_reg_number, '')
        self._set_entry_val(self.p_base, 'vien')
        self.p_type.set('thuoc')
        self.on_product_type_change()
        self.hide_suggestions()

    def on_product_type_change(self, event=None):
        """Xử lý khi thay đổi loại sản phẩm"""
        product_type = self.p_type.get()
        if product_type in ('thuoc', 'vaccine'):
            self.p_reg_label.grid()
            self.p_reg_number.grid()
        else:
            self.p_reg_label.grid_remove()
            self.p_reg_number.grid_remove()

    def on_product_name_change(self, event=None):
        """Xử lý khi thay đổi tên sản phẩm - hiển thị autocomplete"""
        if self.product_source_var.get() == 'free':
            self.hide_suggestions()
            return
            
        if self.medicine_catalog.catalog_data is None:
            return
        
        query = self.p_name.get().strip()
        
        if len(query) >= 2:  # Chỉ tìm kiếm khi có ít nhất 2 ký tự
            try:
                suggestions = self.medicine_catalog.get_medicine_suggestions(query, 8)
                
                if suggestions:
                    # Hiển thị listbox với gợi ý
                    self.suggestions_listbox.delete(0, tk.END)
                    for suggestion in suggestions:
                        self.suggestions_listbox.insert(tk.END, suggestion['display_text'])
                    
                    # Lưu suggestions để sử dụng sau
                    self.current_suggestions = suggestions
                    
                    # Hiển thị frame gợi ý
                    self.suggestions_frame.grid()
                else:
                    # Không có gợi ý, ẩn frame
                    self.hide_suggestions()
            except Exception as e:
                # Lỗi trong autocomplete, ẩn gợi ý
                self.hide_suggestions()
        else:
            # Ẩn gợi ý khi ít hơn 2 ký tự
            self.hide_suggestions()

    def hide_suggestions(self, event=None):
        """Ẩn danh sách gợi ý"""
        self.suggestions_frame.grid_remove()
        self.current_suggestions = []

    def on_suggestion_click(self, event=None):
        """Xử lý khi click vào gợi ý từ listbox"""
        try:
            # Lấy index của item được click
            index = self.suggestions_listbox.nearest(event.y)
            
            self.suggestions_listbox.selection_clear(0, tk.END)
            self.suggestions_listbox.selection_set(index)
            self.suggestions_listbox.activate(index)
            
            # Chọn ngay lập tức
            self.on_suggestion_selected()
            
        except Exception:
            pass

    def on_suggestion_selected(self, event=None):
        """Xử lý khi chọn gợi ý từ listbox"""
        try:
            selection = self.suggestions_listbox.curselection()
            if not selection:
                return
            
            selected_index = selection[0]
            if hasattr(self, 'current_suggestions') and selected_index < len(self.current_suggestions):
                medicine = self.current_suggestions[selected_index]
                
                # Điền thông tin vào form
                self.p_name.delete(0, tk.END)
                self.p_name.insert(0, medicine['name'])
                self.p_type.set('thuoc')
                self.on_product_type_change()
                
                # Điền số đăng ký & đơn vị cơ sở chuẩn
                self._set_entry_val(self.p_reg_number, medicine['registration_number'])
                self._set_entry_val(self.p_base, 'vien')
                
                # Ẩn gợi ý và chuyển focus
                self.hide_suggestions()
                self.p_barcode.focus_set()  # Chuyển focus sang barcode
                
        except Exception as e:
            print(f"Lỗi khi chọn gợi ý: {e}")
            import traceback
            traceback.print_exc()

    def on_product_name_focus_out(self, event=None):
        """Xử lý khi mất focus khỏi ô tên sản phẩm"""
        # Delay một chút để cho phép click vào listbox
        self.after(500, self.hide_suggestions)

    def on_arrow_down(self, event=None):
        """Xử lý phím mũi tên xuống"""
        if self.suggestions_frame.winfo_viewable():
            self.suggestions_listbox.focus_set()
            self.suggestions_listbox.selection_set(0)
            return "break"
        return None

    def on_arrow_up(self, event=None):
        """Xử lý phím mũi tên lên"""
        if self.suggestions_frame.winfo_viewable():
            self.suggestions_listbox.focus_set()
            self.suggestions_listbox.selection_set(tk.END)
            return "break"
        return None

    def on_suggestion_up(self, event=None):
        """Xử lý phím mũi tên lên trong listbox"""
        current = self.suggestions_listbox.curselection()
        if current and current[0] > 0:
            self.suggestions_listbox.selection_clear(current[0])
            self.suggestions_listbox.selection_set(current[0] - 1)
        return "break"

    def on_suggestion_down(self, event=None):
        """Xử lý phím mũi tên xuống trong listbox"""
        current = self.suggestions_listbox.curselection()
        if current and current[0] < self.suggestions_listbox.size() - 1:
            self.suggestions_listbox.selection_clear(current[0])
            self.suggestions_listbox.selection_set(current[0] + 1)
        return "break"

    def load_medicine_catalog(self):
        """Load danh mục thuốc từ file Excel hoặc CSV"""
        try:
            file_path = filedialog.askopenfilename(
                title="Chọn file danh mục thuốc",
                filetypes=[
                    ('CSV files', '*.csv'),
                    ('Excel files', '*.xlsx *.xls'),
                    ('All files', '*.*')
                ]
            )
            
            if file_path:
                self.medicine_catalog.load_catalog_from_excel(file_path)
                self.update_catalog_info()
                self.toast('Đã load danh mục thuốc thành công')
                
        except Exception as e:
            messagebox.showerror('Lỗi', str(e))

    def load_default_csv(self):
        """Load file thuoc.csv mặc định"""
        try:
            # Tìm file thuoc.csv trong thư mục hiện tại
            csv_path = os.path.join(os.getcwd(), 'thuoc.csv')
            
            if os.path.exists(csv_path):
                self.medicine_catalog.load_catalog_from_excel(csv_path)
                self.update_catalog_info()
                self.toast('Đã load file thuoc.csv thành công')
            else:
                # Nếu không tìm thấy, mở dialog chọn file
                messagebox.showinfo('Thông báo', 
                    f'Không tìm thấy file thuoc.csv trong thư mục:\n{csv_path}\n\nVui lòng chọn file khác.')
                self.load_medicine_catalog()
                
        except Exception as e:
            messagebox.showerror('Lỗi', str(e))

    def search_medicine_dialog(self):
        """Hiển thị dialog tra cứu thuốc"""
        if self.medicine_catalog.catalog_data is None:
            messagebox.showwarning('Cảnh báo', 'Vui lòng load danh mục thuốc trước')
            return
        
        # Tạo dialog tra cứu
        dialog = tb.Toplevel(self)
        dialog.title("Tra cứu thuốc")
        dialog.geometry("800x600")
        dialog.transient(self)
        dialog.grab_set()
        
        # Center dialog
        dialog.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() // 2) - (800 // 2)
        y = self.winfo_y() + (self.winfo_height() // 2) - (600 // 2)
        dialog.geometry(f"800x600+{x}+{y}")
        
        main_frame = tb.Frame(dialog)
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Title
        tb.Label(main_frame, text="TRA CỨU THUỐC TRONG DANH MỤC", 
                font=('Segoe UI', 14, 'bold'), bootstyle='primary').pack(pady=(0, 15))
        
        # Search frame
        search_frame = tb.Frame(main_frame)
        search_frame.pack(fill='x', pady=(0, 10))
        
        tb.Label(search_frame, text="Tên thuốc:").pack(side='left', padx=(0, 10))
        search_entry = tb.Entry(search_frame, width=40)
        search_entry.pack(side='left', padx=(0, 10))
        search_entry.focus_set()
        
        def search_medicines():
            try:
                query = search_entry.get().strip()
                if not query:
                    messagebox.showwarning('Cảnh báo', 'Nhập tên thuốc cần tìm')
                    return
                
                results = self.medicine_catalog.search_medicine(query)
                display_results(results)
                
            except Exception as e:
                messagebox.showerror('Lỗi', str(e))
        
        tb.Button(search_frame, text="🔍 Tìm kiếm", bootstyle='success',
                  command=search_medicines).pack(side='left')
        
        # Results frame
        results_frame = tb.Labelframe(main_frame, text="Kết quả tìm kiếm", bootstyle='light')
        results_frame.pack(fill='both', expand=True)
        
        # Results tree
        cols = ('name', 'reg_number', 'manufacturer', 'active_ingredient', 'dosage_form')
        results_tree = tb.Treeview(results_frame, columns=cols, show='headings', height=15)
        
        for c, w, t, anchor in [
            ('name', 200, 'Tên thuốc', 'w'),
            ('reg_number', 120, 'Số đăng ký', 'center'),
            ('manufacturer', 150, 'Nhà SX', 'w'),
            ('active_ingredient', 150, 'Hoạt chất', 'w'),
            ('dosage_form', 100, 'Dạng bào chế', 'w')
        ]:
            results_tree.heading(c, text=t)
            results_tree.column(c, width=w, anchor=anchor)
        
        results_tree.tag_configure('odd', background='#f6f8fa')
        results_tree.pack(fill='both', expand=True, padx=8, pady=8)
        
        def display_results(results):
            # Clear tree
            for item in results_tree.get_children():
                results_tree.delete(item)
            
            for idx, result in enumerate(results):
                results_tree.insert('', 'end', values=(
                    result['name'],
                    result['registration_number'],
                    result['manufacturer'],
                    result['active_ingredient'],
                    result['dosage_form']
                ), tags=('odd',) if idx % 2 else ())
        
        def select_medicine():
            selection = results_tree.selection()
            if not selection:
                messagebox.showwarning('Cảnh báo', 'Chọn thuốc từ danh sách')
                return
            
            item = results_tree.item(selection[0])
            medicine_name = item['values'][0]
            reg_number = item['values'][1]
            
            # Điền vào form sản phẩm
            self.p_name.delete(0, tk.END)
            self.p_name.insert(0, medicine_name)
            self.p_type.set('thuoc')
            self.on_product_type_change()
            self.p_reg_number.delete(0, tk.END)
            self.p_reg_number.insert(0, reg_number)
            
            dialog.destroy()
            self.nb.select(self.tab_products)
            self.p_base.focus_set()
        
        # Buttons
        btn_frame = tb.Frame(main_frame)
        btn_frame.pack(fill='x', pady=(10, 0))
        
        tb.Button(btn_frame, text="✅ Chọn thuốc này", bootstyle='success',
                  command=select_medicine).pack(side='left', padx=(0, 10))
        tb.Button(btn_frame, text="❌ Đóng", bootstyle='secondary',
                  command=dialog.destroy).pack(side='left')
        
        # Bind Enter key
        search_entry.bind('<Return>', lambda e: search_medicines())

    def show_catalog_info(self):
        """Hiển thị thông tin danh mục"""
        info = self.medicine_catalog.get_catalog_info()
        
        if not info['loaded']:
            messagebox.showinfo('Thông tin danh mục', 'Chưa load danh mục thuốc')
            return
        
        info_text = f"""Thông tin danh mục thuốc:

📁 File: {info['file_path']}
📊 Tổng số bản ghi: {info['total_records']:,}
📋 Các cột dữ liệu:
"""
        
        for i, col in enumerate(info['columns'], 1):
            info_text += f"  {i}. {col}\n"
        
        messagebox.showinfo('Thông tin danh mục', info_text)

    def update_catalog_info(self):
        """Cập nhật thông tin danh mục hiển thị"""
        info = self.medicine_catalog.get_catalog_info()
        
        if info['loaded']:
            file_name = os.path.basename(info['file_path'])
            self.catalog_info_label.config(
                text=f"📁 {file_name} - {info['total_records']:,} bản ghi",
                bootstyle='success'
            )
        else:
            self.catalog_info_label.config(
                text='Chưa load danh mục thuốc',
                bootstyle='secondary'
            )

    def scan_and_add(self):
        ok = self.fill_product_by_barcode(only_select=True)  # sửa hàm dưới để trả bool
        if ok:
            self.ent_qty_pos.delete(0, tk.END)
            self.ent_qty_pos.insert(0, '1')
            self.add_to_cart()
            self.ent_barcode.delete(0, tk.END)
        # luôn trả focus về ô barcode để quét tiếp
        self.after(50, lambda: self.ent_barcode.focus_set())
    
    def open_barcode_scanner(self):
        """Mở barcode scanner"""
        if not BARCODE_AVAILABLE:
            self.show_barcode_install_info()
            return
            
        try:
            # Tạo callback để xử lý kết quả quét
            def on_barcode_scanned(barcode_data):
                # Điền barcode vào ô input
                self.ent_barcode.delete(0, tk.END)
                self.ent_barcode.insert(0, barcode_data)
                
                # Tự động thêm vào giỏ hàng
                self.after(100, self.scan_and_add)
                
            # Tạo và mở scanner
            self.barcode_scanner = BarcodeScanner(self, callback=on_barcode_scanned)
            self.barcode_scanner.start_scan()
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể mở barcode scanner: {e}")
    
    def show_barcode_install_info(self):
        """Hiển thị thông tin cài đặt thư viện barcode"""
        info = """
📷 Tính năng quét barcode cần cài đặt thêm thư viện:

pip install opencv-python pyzbar Pillow

Sau khi cài đặt, khởi động lại phần mềm để sử dụng tính năng quét barcode bằng camera.

Hiện tại bạn vẫn có thể:
• Nhập barcode thủ công
• Tìm sản phẩm theo tên
        """
        messagebox.showinfo("Cài đặt thư viện barcode", info)
    
    def scan_barcode_for_product(self):
        """Quét barcode cho sản phẩm mới"""
        if not BARCODE_AVAILABLE:
            self.show_barcode_install_info()
            return
            
        try:
            # Tạo callback để xử lý kết quả quét
            def on_barcode_scanned(barcode_data):
                # Điền barcode vào ô input
                self.p_barcode.delete(0, tk.END)
                self.p_barcode.insert(0, barcode_data)
                
                # Chuyển focus sang ô tên sản phẩm
                self.p_name.focus_set()
                
            # Tạo và mở scanner
            self.barcode_scanner = BarcodeScanner(self, callback=on_barcode_scanned)
            self.barcode_scanner.start_scan()
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể mở barcode scanner: {e}")

    # -------- Purchase --------
    def build_purchase_tab(self):
        frm = self.tab_purchase
        self.cart_purchase = []
        self.last_purchase_items = []
        self.last_purchase_info = {}

        # Header với title đẹp
        header_frame = tb.Frame(frm)
        header_frame.pack(fill='x', padx=8, pady=(8,4))
        
        title_label = tb.Label(header_frame, text='📦 Nhập kho / Lập phiếu nhập', 
                              font=('Segoe UI', 14, 'bold'), bootstyle='success')
        title_label.pack(anchor='w')
        
        subtitle_label = tb.Label(header_frame, text='Lập phiếu nhập kho thuốc, vaccine và vật tư y tế', 
                                 font=('Segoe UI', 9), bootstyle='secondary')
        subtitle_label.pack(anchor='w')

        # --- Khung thông tin phiếu nhập
        info_note = tb.Labelframe(frm, text='📝 Thông tin phiếu nhập kho', bootstyle='light')
        info_note.pack(fill='x', padx=8, pady=8)

        # Nhà cung cấp / Nguồn cấp
        tb.Label(info_note, text='Nguồn cấp/Nhà CC:').grid(row=0, column=0, padx=6, pady=6, sticky='w')
        self.cmb_supplier = tb.Combobox(info_note, width=25)
        self.cmb_supplier.grid(row=0, column=1, padx=6, pady=6, sticky='w')
        self.refresh_suppliers_combo()

        # Ngày nhập
        tb.Label(info_note, text='Ngày nhập:').grid(row=0, column=2, padx=6, pady=6, sticky='w')
        self.ent_purchase_date = DateEntry(
            info_note,
            dateformat="%Y-%m-%d",
            firstweekday=0,
            bootstyle='info',
            width=12
        )
        self.ent_purchase_date.grid(row=0, column=3, padx=6, pady=6, sticky='w')

        # Lý do nhập
        tb.Label(info_note, text='Lý do nhập:').grid(row=0, column=4, padx=6, pady=6, sticky='w')
        self.cmb_purchase_reason = tb.Combobox(info_note, values=['Nhận cấp phát tuyến trên', 'Mua sắm đấu thầu', 'Viện trợ - Tài trợ', 'Khác'], width=22, state='readonly')
        self.cmb_purchase_reason.set('Nhận cấp phát tuyến trên')
        self.cmb_purchase_reason.grid(row=0, column=5, padx=6, pady=6, sticky='w')

        # Ghi chú
        tb.Label(info_note, text='Ghi chú:').grid(row=1, column=0, padx=6, pady=6, sticky='w')
        self.ent_purchase_note = tb.Entry(info_note)
        self.ent_purchase_note.grid(row=1, column=1, columnspan=5, padx=6, pady=6, sticky='ew')
        info_note.columnconfigure(1, weight=1)

        # --- Khung nhập sản phẩm
        box = tb.Labelframe(frm, text='➕ Thêm sản phẩm vào phiếu nhập', bootstyle='light')
        box.pack(fill='x', padx=8, pady=4)

        # Cho các cột có thể giãn đều khi thay đổi kích thước cửa sổ
        for i in range(12):
            box.grid_columnconfigure(i, weight=1)

        # ── Hàng 0: Tìm kiếm + Combobox chọn sản phẩm
        tb.Label(box, text='Tìm sản phẩm:').grid(row=0, column=0, sticky='w', padx=6, pady=6)
        self.search_purchase = tb.Entry(box)
        self.search_purchase.grid(row=0, column=1, columnspan=4, sticky='ew', padx=6, pady=6)
        self.search_purchase.bind('<KeyRelease>', lambda e: self.filter_product_list())
        self.search_purchase.bind('<Down>', lambda e: self.open_combo(self.cmb_prod)) 

        tb.Label(box, text='Chọn:').grid(row=0, column=5, sticky='e', padx=6, pady=6)
        self.cmb_prod = tb.Combobox(box, state='readonly')
        self.cmb_prod.grid(row=0, column=6, columnspan=5, sticky='ew', padx=6, pady=6)
        self.cmb_prod.bind('<<ComboboxSelected>>', lambda e: self.update_purchase_unit_and_price())
        self.cmb_prod.bind('<Escape>', lambda e: self.search_purchase.focus_set())  # ESC quay lại ô tìm
        self.cmb_prod.bind('<Return>', lambda e: self.ent_qty.focus_set())

        # ── Hàng 1: Đơn vị, Số lượng, Số lô, HSD
        tb.Label(box, text='Đơn vị tính:').grid(row=1, column=0, sticky='w', padx=6, pady=6)
        self.lbl_unit_purchase = tb.Label(box, text='-')
        self.lbl_unit_purchase.grid(row=1, column=1, sticky='w', padx=6, pady=6)

        tb.Label(box, text='Số lượng:').grid(row=1, column=2, sticky='e', padx=6, pady=6)
        self.ent_qty = tb.Entry(box, width=10)
        self.ent_qty.insert(0, '1')
        self.ent_qty.grid(row=1, column=3, sticky='w', padx=6, pady=6)
        self._numberize(self.ent_qty)

        tb.Label(box, text='Số lô:').grid(row=1, column=4, sticky='e', padx=6, pady=6)
        self.ent_lot = tb.Entry(box, width=14)
        self.ent_lot.insert(0, 'LOT001')
        self.ent_lot.grid(row=1, column=5, sticky='w', padx=6, pady=6)

        tb.Label(box, text='HSD (YYYY-MM-DD):').grid(row=1, column=6, sticky='e', padx=6, pady=6)
        self.ent_exp = DateEntry(
            box,
            dateformat="%Y-%m-%d",
            firstweekday=0,     # Monday
            bootstyle='info'
        )
        self.ent_exp.grid(row=1, column=7, sticky='w', padx=6, pady=6)

        # ── Hàng 2: Đơn giá nhập
        tb.Label(box, text='Đơn giá nhập:').grid(row=2, column=0, sticky='w', padx=6, pady=(6,8))
        self.ent_cost = tb.Entry(box, width=12)
        self.ent_cost.insert(0, '0')
        self.ent_cost.grid(row=2, column=1, sticky='w', padx=6, pady=(6,8))
        self._numberize(self.ent_cost)

        tb.Label(box, text='Nguồn kinh phí:').grid(row=2, column=2, sticky='e', padx=6, pady=(6,8))
        self.cmb_item_fund = tb.Combobox(
            box,
            values=[
                'TCMR (Tiêm chủng mở rộng)',
                'Ngân sách địa phương',
                'Dự án viện trợ',
                'Mua sắm đấu thầu',
                'Nguồn khác'
            ],
            width=25
        )
        self.cmb_item_fund.set('TCMR (Tiêm chủng mở rộng)')
        self.cmb_item_fund.grid(row=2, column=3, columnspan=3, sticky='w', padx=6, pady=(6,8))

        # --- Nút tác vụ
        btns = tb.Frame(frm)
        btns.pack(fill='x', padx=8, pady=8)
        
        # Nhóm chính bên trái
        tb.Button(btns, text='+ Thêm vào giỏ hàng', bootstyle='primary', command=self.add_to_purchase_cart).pack(side='left', padx=4)
        tb.Button(btns, text='Xác nhận nhập kho', bootstyle='success', command=self.confirm_purchase).pack(side='left', padx=4)
        tb.Button(btns, text='In phiếu nhập kho', bootstyle='info', command=self.print_purchase_note).pack(side='left', padx=4)
        
        # Nhóm xóa hủy bên phải (phòng tránh bấm nhầm)
        tb.Button(btns, text='Xóa danh sách', bootstyle='danger-outline', command=self.clear_purchase_cart).pack(side='right', padx=4)
        tb.Button(btns, text='Xóa dòng', bootstyle='warning', command=self.remove_selected_purchase_item).pack(side='right', padx=4)

        # --- Bảng danh sách hàng nhập tạm thời
        cols = ('product','productName','unit','qty','lot','exp','cost','fundSource','total')
        self.tree_purchase_cart = tb.Treeview(frm, columns=cols, show='headings')
        for c, w, t, anchor in [
            ('product',50,'PID','center'),('productName',220,'Tên thuốc/vaccine/VTYT','w'),
            ('unit',60,'ĐVT','center'),('qty',60,'SL','e'),('lot',80,'Số lô','w'),
            ('exp',80,'HSD','center'),('cost',95,'Đơn giá nhập','e'),
            ('fundSource',135,'Nguồn kinh phí','w'),('total',105,'Thành tiền','e')
        ]:
            self.tree_purchase_cart.heading(c, text=t, command=(lambda col=c: self.sort_tree(self.tree_purchase_cart, col)))
            self.tree_purchase_cart.column(c, width=w, anchor=anchor)
        self.tree_purchase_cart.tag_configure('odd', background='#f6f8fa')
        self.tree_purchase_cart.pack(fill='both', expand=True, padx=8, pady=8)

    def update_purchase_unit_and_price(self):
        sel = self.cmb_prod.get()
        if not sel: self.lbl_unit_purchase.config(text='-'); return
        pid = int(sel.split(' — ')[0])
        du = self.db.default_unit_of(pid) or '-'
        self.lbl_unit_purchase.config(text=du)

    def filter_product_list(self):
        kw = (self.search_purchase.get() or '').strip().lower()
        opts = [f"{p['id']} — {p['name']}" for p in self._products if kw in p['name'].lower()]
        self.cmb_prod['values'] = opts
        if opts:
            self.cmb_prod.current(0)
            self.update_purchase_unit_and_price()

    def refresh_suppliers_combo(self):
        try:
            suppliers = self.db.get_suppliers()
            self.cmb_supplier['values'] = suppliers
        except Exception as e:
            print(f"Lỗi tải nhà cung cấp: {e}")

    def add_to_purchase_cart(self):
        sel = self.cmb_prod.get()
        if not sel:
            messagebox.showerror('Lỗi', 'Chọn sản phẩm để nhập'); return
        pid = int(sel.split(' — ')[0])
        unit = self.db.default_unit_of(pid) or 'vien'
        try:
            qty = float(self.ent_qty.get())
        except:
            qty = 0
        if qty <= 0:
            messagebox.showerror('Lỗi', 'Số lượng nhập phải > 0'); return
        
        lot = self.ent_lot.get().strip()
        if not lot:
            messagebox.showerror('Lỗi', 'Vui lòng nhập số lô'); return
            
        exp = (self.ent_exp.entry.get() or '').strip()
        try:
            dt.datetime.strptime(exp, '%Y-%m-%d')
        except:
            messagebox.showerror('Lỗi', 'Hạn sử dụng không hợp lệ (YYYY-MM-DD)'); return
            
        try:
            cost = float((self.ent_cost.get() or '0').replace(',', ''))
        except:
            cost = 0.0
        if cost < 0:
            messagebox.showerror('Lỗi', 'Đơn giá nhập không được âm'); return

        name = self.name_by_id(pid)
        fund_source = self.cmb_item_fund.get().strip()
        
        merged = False
        for it in self.cart_purchase:
            if it['productId'] == pid and it['lotNo'] == lot and it.get('fundSource') == fund_source:
                it['qty'] = round(it['qty'] + qty, 4)
                it['cost'] = cost
                merged = True
                break
        if not merged:
            self.cart_purchase.append({
                'productId': pid,
                'productName': name,
                'unitCode': unit,
                'qty': qty,
                'lotNo': lot,
                'expiryDate': exp,
                'cost': cost,
                'fundSource': fund_source
            })

        self.refresh_purchase_cart_view()
        self.ent_qty.delete(0, tk.END)
        self.ent_qty.insert(0, '1')
        self.search_purchase.focus_set()

    def remove_selected_purchase_item(self):
        sel = self.tree_purchase_cart.selection()
        if not sel:
            return
        idx = self.tree_purchase_cart.index(sel[0])
        if 0 <= idx < len(self.cart_purchase):
            self.cart_purchase.pop(idx)
            self.refresh_purchase_cart_view()

    def clear_purchase_cart(self):
        if self.cart_purchase and messagebox.askyesno('Xác nhận', 'Xóa toàn bộ danh sách hàng nhập kho?'):
            self.cart_purchase = []
            self.refresh_purchase_cart_view()

    def refresh_purchase_cart_view(self):
        for i in self.tree_purchase_cart.get_children():
            self.tree_purchase_cart.delete(i)
        for idx, it in enumerate(self.cart_purchase):
            total_val = it['qty'] * it['cost']
            self.tree_purchase_cart.insert('', 'end',
                values=(
                    it['productId'], 
                    it['productName'], 
                    it['unitCode'], 
                    f"{it['qty']:g}", 
                    it['lotNo'], 
                    it['expiryDate'], 
                    f"{it['cost']:,.0f}", 
                    it.get('fundSource', ''),
                    f"{total_val:,.0f}"
                ),
                tags=('odd',) if idx % 2 else ())

    def confirm_purchase(self):
        if not self.cart_purchase:
            messagebox.showwarning('Chưa có dữ liệu', 'Danh sách nhập kho trống'); return
        
        supplier = self.cmb_supplier.get().strip()
        if not supplier:
            messagebox.showwarning('Thiếu thông tin', 'Vui lòng nhập/chọn Nguồn cấp/Nhà cung cấp'); return
        
        date_str = (self.ent_purchase_date.entry.get() or '').strip()
        if date_str:
            try:
                dt.datetime.strptime(date_str, '%Y-%m-%d')
            except:
                messagebox.showerror('Lỗi', 'Ngày nhập không hợp lệ (định dạng YYYY-MM-DD)'); return
                
        reason = self.cmb_purchase_reason.get().strip()
        note = self.ent_purchase_note.get().strip()

        if not messagebox.askyesno('Xác nhận', f'Bạn có chắc chắn muốn nhập kho từ:\nNguồn cấp: {supplier}\nNgày nhập: {date_str or "Hôm nay"}\nLý do: {reason}?'):
            return

        try:
            purchase_id, note_number, details = self.db.record_purchase(self.cart_purchase, supplier, reason, note, date_str)
            self.last_purchase_items = details
            
            created_at_str = f"{date_str} {dt.datetime.now().strftime('%H:%M:%S')}" if date_str else dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            self.last_purchase_info = {
                'id': purchase_id,
                'noteNumber': note_number,
                'supplier': supplier,
                'reason': reason,
                'note': note,
                'createdAt': created_at_str
            }
            self.cart_purchase = []
            self.refresh_purchase_cart_view()
            self.refresh_stock()
            self.refresh_suppliers_combo()
            if hasattr(self, 'refresh_report_funds_combo'):
                self.refresh_report_funds_combo()
            messagebox.showinfo('Thành công', f'Đã nhập kho thành công!\nSố phiếu: {note_number}')
            
            # Tự động hỏi in phiếu nhập kho
            if messagebox.askyesno('In phiếu', 'Bạn có muốn in phiếu nhập kho ngay bây giờ?'):
                self.print_purchase_note()
                
        except Exception as e:
            messagebox.showerror('Lỗi', f'Lỗi nhập kho: {str(e)}')

    def print_purchase_note(self):
        if not self.last_purchase_items:
            messagebox.showwarning('Chưa có dữ liệu', 'Hãy thực hiện nhập kho trước khi in phiếu'); return
        
        info = self.last_purchase_info
        
        # Cho phép người dùng chọn vị trí lưu file PDF
        initial_filename = f"Phieu_Nhap_Kho_{info['noteNumber']}.pdf"
        pdf_path = filedialog.asksaveasfilename(
            title="Chọn vị trí lưu phiếu nhập kho PDF",
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            initialfile=initial_filename
        )
        if not pdf_path:
            return

        try:
            import reportlab
        except ImportError:
            response = messagebox.askyesno(
                "Thiếu thư viện", 
                "Hệ thống thiếu thư viện 'reportlab' để xuất PDF.\nBạn có muốn tự động cài đặt không? (Quá trình này mất khoảng vài giây)"
            )
            if response:
                import subprocess
                import sys
                try:
                    self.toast("Đang cài đặt thư viện reportlab, vui lòng đợi...")
                    subprocess.run([sys.executable, "-m", "pip", "install", "reportlab"], check=True)
                    self.toast("Đã cài đặt reportlab thành công!")
                except Exception as ex:
                    messagebox.showerror("Lỗi cài đặt", f"Không thể tự động cài đặt reportlab: {str(ex)}\nHãy chạy lệnh 'pip install reportlab' trong terminal."); return
            else:
                return

        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            
            try:
                pdfmetrics.registerFont(TTFont('TimesNewRoman', "C:\\Windows\\Fonts\\times.ttf"))
                pdfmetrics.registerFont(TTFont('TimesNewRoman-Bold', "C:\\Windows\\Fonts\\timesbd.ttf"))
                pdfmetrics.registerFont(TTFont('TimesNewRoman-Italic', "C:\\Windows\\Fonts\\timesi.ttf"))
                font_normal = 'TimesNewRoman'
                font_bold = 'TimesNewRoman-Bold'
                font_italic = 'TimesNewRoman-Italic'
            except Exception:
                font_normal = 'Helvetica'
                font_bold = 'Helvetica-Bold'
                font_italic = 'Helvetica-Oblique'
                
            doc = SimpleDocTemplate(pdf_path, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
            story = []
            
            styles = getSampleStyleSheet()
            
            style_header_left = ParagraphStyle(
                'HeaderLeft', parent=styles['Normal'], fontName=font_bold, fontSize=10, leading=14, alignment=0
            )
            style_header_right = ParagraphStyle(
                'HeaderRight', parent=styles['Normal'], fontName=font_normal, fontSize=10, leading=14, alignment=2
            )
            style_title = ParagraphStyle(
                'Title', parent=styles['Heading1'], fontName=font_bold, fontSize=16, leading=20, alignment=1, spaceAfter=5
            )
            style_subtitle = ParagraphStyle(
                'Subtitle', parent=styles['Normal'], fontName=font_bold, fontSize=11, leading=14, alignment=1, spaceAfter=15
            )
            style_info = ParagraphStyle(
                'Info', parent=styles['Normal'], fontName=font_normal, fontSize=11, leading=16, alignment=0
            )
            style_table_header = ParagraphStyle(
                'TableHeader', parent=styles['Normal'], fontName=font_bold, fontSize=9, leading=11, alignment=1, textColor=colors.black
            )
            style_cell = ParagraphStyle(
                'Cell', parent=styles['Normal'], fontName=font_normal, fontSize=9, leading=11, alignment=0
            )
            style_cell_center = ParagraphStyle(
                'CellCenter', parent=styles['Normal'], fontName=font_normal, fontSize=9, leading=11, alignment=1
            )
            style_cell_right = ParagraphStyle(
                'CellRight', parent=styles['Normal'], fontName=font_normal, fontSize=9, leading=11, alignment=2
            )
            
            # Header
            header_data = [
                [
                    Paragraph("SỞ Y TẾ THÀNH PHỐ CẦN THƠ<br/>TRUNG TÂM KIỂM SOÁT BỆNH TẬT (CDC)", style_header_left),
                    Paragraph("<b>Mẫu số: C30-HD</b><br/><i>(Ban hành theo Thông tư số 107/2017/TT-BTC)</i>", style_header_right)
                ]
            ]
            header_table = Table(header_data, colWidths=[280, 230])
            header_table.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
                ('BOTTOMPADDING', (0,0), (-1,-1), 10),
            ]))
            story.append(header_table)
            story.append(Spacer(1, 10))
            
            # Title
            story.append(Paragraph("PHIẾU NHẬP KHO", style_title))
            story.append(Paragraph(f"Số: {info['noteNumber']}", style_subtitle))
            
            # Parse Date
            created_str = info['createdAt']
            try:
                created_at_dt = dt.datetime.strptime(created_str, '%Y-%m-%d %H:%M:%S')
            except Exception:
                try:
                    created_at_dt = dt.datetime.strptime(created_str.split(' ')[0], '%Y-%m-%d')
                except Exception:
                    created_at_dt = dt.datetime.now()
            
            info_lines = [
                f"<b>Nguồn cấp / Nhà CC:</b> {info['supplier']}",
                f"<b>Lý do nhập:</b> {info['reason']}",
                f"<b>Kho nhập:</b> Kho Dược CDC Cần Thơ",
                f"<b>Ngày nhập:</b> {created_at_dt.strftime('%d/%m/%Y')}",
                f"<b>Ghi chú:</b> {info['note'] or 'Không'}"
            ]
            for line in info_lines:
                story.append(Paragraph(line, style_info))
                story.append(Spacer(1, 4))
                
            story.append(Spacer(1, 10))
            
            # Table items
            table_data = [
                [
                    Paragraph("STT", style_table_header),
                    Paragraph("Tên thuốc, vaccine, VTYT", style_table_header),
                    Paragraph("ĐVT", style_table_header),
                    Paragraph("Số lượng", style_table_header),
                    Paragraph("Đơn giá", style_table_header),
                    Paragraph("Thành tiền", style_table_header),
                    Paragraph("Số lô", style_table_header),
                    Paragraph("Hạn dùng", style_table_header)
                ]
            ]
            
            total_sum = 0.0
            for idx, it in enumerate(self.last_purchase_items, 1):
                qty = it['qty']
                cost = it['cost']
                sub_total = qty * cost
                total_sum += sub_total
                
                table_data.append([
                    Paragraph(str(idx), style_cell_center),
                    Paragraph(it['productName'], style_cell),
                    Paragraph(it['unitCode'], style_cell_center),
                    Paragraph(f"{qty:g}", style_cell_right),
                    Paragraph(f"{cost:,.0f}", style_cell_right),
                    Paragraph(f"{sub_total:,.0f}", style_cell_right),
                    Paragraph(it['lotNo'] or '', style_cell_center),
                    Paragraph(it['expiryDate'] or '', style_cell_center)
                ])
            
            # Thêm dòng tổng cộng
            table_data.append([
                Paragraph("<b>Tổng cộng</b>", style_cell_center),
                Paragraph("", style_cell),
                Paragraph("", style_cell_center),
                Paragraph("", style_cell_right),
                Paragraph("", style_cell_right),
                Paragraph(f"<b>{total_sum:,.0f}</b>", style_cell_right),
                Paragraph("", style_cell_center),
                Paragraph("", style_cell_center)
            ])
                
            col_widths = [25, 160, 45, 55, 65, 75, 55, 50]
            items_table = Table(table_data, colWidths=col_widths)
            items_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f2f2f2')),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.black),
                ('SPAN', (0, -1), (4, -1)),
                ('TOPPADDING', (0,0), (-1,-1), 5),
                ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ]))
            story.append(items_table)
            story.append(Spacer(1, 15))
            
            # Signatures
            date_right_style = ParagraphStyle(
                'DateRight', parent=styles['Normal'], fontName=font_italic, fontSize=11, alignment=2, spaceAfter=10
            )
            sig_title_style = ParagraphStyle(
                'SigTitle', parent=styles['Normal'], fontName=font_bold, fontSize=11, alignment=1
            )
            sig_sub_style = ParagraphStyle(
                'SigSub', parent=styles['Normal'], fontName=font_italic, fontSize=9, alignment=1
            )
            
            story.append(Paragraph(f"Cần Thơ, ngày {created_at_dt.strftime('%d')} tháng {created_at_dt.strftime('%m')} năm {created_at_dt.strftime('%Y')}", date_right_style))
            
            sig_headers = [
                [
                    Paragraph("<b>Người lập phiếu</b>", sig_title_style),
                    Paragraph("<b>Người giao hàng</b>", sig_title_style),
                    Paragraph("<b>Thủ kho</b>", sig_title_style),
                    Paragraph("<b>Kế toán trưởng</b>", sig_title_style),
                    Paragraph("<b>Lãnh đạo đơn vị</b>", sig_title_style)
                ],
                [
                    Paragraph("(Ký, họ tên)", sig_sub_style),
                    Paragraph("(Ký, họ tên)", sig_sub_style),
                    Paragraph("(Ký, họ tên)", sig_sub_style),
                    Paragraph("(Ký, họ tên)", sig_sub_style),
                    Paragraph("(Ký, đóng dấu)", sig_sub_style)
                ]
            ]
            sig_table = Table(sig_headers, colWidths=[102, 102, 102, 102, 102])
            sig_table.setStyle(TableStyle([
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('BOTTOMPADDING', (0,0), (-1,-1), 2),
            ]))
            story.append(sig_table)
            story.append(Spacer(1, 60))
            
            doc.build(story)
            os.startfile(pdf_path)
            self.toast("Đã in phiếu nhập ra PDF và mở file thành công")
        except Exception as e:
            messagebox.showerror("Lỗi in PDF", f"Không thể xuất file PDF: {str(e)}")

    # -------- Xuất kho / Cấp phát (Dispatch) --------
    def build_dispatch_tab(self):
        frm = self.tab_dispatch

        # Header với title đẹp
        header_frame = tb.Frame(frm)
        header_frame.pack(fill='x', padx=8, pady=(8,4))
        
        title_label = tb.Label(header_frame, text='📤 Xuất kho / Cấp phát', 
                              font=('Segoe UI', 14, 'bold'), bootstyle='warning')
        title_label.pack(anchor='w')
        
        subtitle_label = tb.Label(header_frame, text='Cấp phát thuốc, vaccine và vật tư y tế cho các đơn vị tuyến dưới theo nguyên tắc FEFO', 
                                 font=('Segoe UI', 9), bootstyle='secondary')
        subtitle_label.pack(anchor='w')

        # --- Khung thông tin phiếu xuất
        info_note = tb.Labelframe(frm, text='📝 Thông tin phiếu xuất kho', bootstyle='light')
        info_note.pack(fill='x', padx=8, pady=8)

        # Đơn vị nhận
        tb.Label(info_note, text='Đơn vị nhận:').grid(row=0, column=0, padx=6, pady=6, sticky='w')
        self.cmb_receiving_unit = tb.Combobox(info_note, width=25)
        self.cmb_receiving_unit.grid(row=0, column=1, padx=6, pady=6, sticky='w')
        self.refresh_receiving_units_combo()

        # Ngày xuất
        tb.Label(info_note, text='Ngày xuất:').grid(row=0, column=2, padx=6, pady=6, sticky='w')
        self.ent_dispatch_date = DateEntry(
            info_note,
            dateformat="%Y-%m-%d",
            firstweekday=0,
            bootstyle='info',
            width=12
        )
        self.ent_dispatch_date.grid(row=0, column=3, padx=6, pady=6, sticky='w')

        # Lý do xuất
        tb.Label(info_note, text='Lý do xuất:').grid(row=0, column=4, padx=6, pady=6, sticky='w')
        self.cmb_reason = tb.Combobox(info_note, values=['Cấp phát', 'Hủy kho', 'Chuyển kho', 'Khác'], width=15, state='readonly')
        self.cmb_reason.set('Cấp phát')
        self.cmb_reason.grid(row=0, column=5, padx=6, pady=6, sticky='w')

        # Ghi chú (đặt xuống hàng 1 để rộng rãi hơn)
        tb.Label(info_note, text='Ghi chú:').grid(row=1, column=0, padx=6, pady=6, sticky='w')
        self.ent_dispatch_note = tb.Entry(info_note)
        self.ent_dispatch_note.grid(row=1, column=1, columnspan=5, padx=6, pady=6, sticky='ew')
        info_note.columnconfigure(1, weight=1)

        # --- Khung thao tác chọn hàng xuất
        op_frame = tb.Labelframe(frm, text='🛒 Thao tác chọn hàng xuất', bootstyle='light')
        op_frame.pack(fill='x', padx=8, pady=4)

        # Hàng 0: Barcode & Quét, Tìm nhanh, Chọn sản phẩm
        tb.Label(op_frame, text='Barcode:').grid(row=0, column=0, padx=6, pady=6, sticky='w')
        barcode_inner = tb.Frame(op_frame)
        barcode_inner.grid(row=0, column=1, padx=6, pady=6, sticky='ew')
        
        self.ent_barcode = tb.Entry(barcode_inner, width=12)
        self.ent_barcode.pack(side='left', fill='x', expand=True, padx=(0, 4))
        self.ent_barcode.bind('<Return>', lambda e: self.scan_and_add_dispatch())
        self.ent_barcode.bind('<KP_Enter>', lambda e: self.scan_and_add_dispatch())
        
        if BARCODE_AVAILABLE:
            btn_scan = tb.Button(barcode_inner, text='📷 Quét', command=self.open_barcode_scanner_dispatch, 
                                 bootstyle='info', width=6)
        else:
            btn_scan = tb.Button(barcode_inner, text='📷 Quét', command=self.show_barcode_install_info, 
                                 bootstyle='secondary', width=6)
        btn_scan.pack(side='right')

        tb.Label(op_frame, text='Tìm nhanh:').grid(row=0, column=2, padx=6, pady=6, sticky='w')
        self.search_pos = tb.Entry(op_frame, width=15)
        self.search_pos.grid(row=0, column=3, padx=6, pady=6, sticky='ew')

        tb.Label(op_frame, text='Chọn sản phẩm:').grid(row=0, column=4, padx=6, pady=6, sticky='w')
        self.cmb_prod_pos = tb.Combobox(op_frame, state='readonly', width=35)
        self.cmb_prod_pos.grid(row=0, column=5, padx=6, pady=6, sticky='ew')

        # Hàng 1: Chọn lô, Nguồn xuất, SL xuất
        tb.Label(op_frame, text='Chọn lô:').grid(row=1, column=0, padx=6, pady=6, sticky='w')
        self.cmb_lot_pos = tb.Combobox(op_frame, state='readonly', width=20)
        self.cmb_lot_pos.grid(row=1, column=1, padx=6, pady=6, sticky='ew')

        tb.Label(op_frame, text='Nguồn xuất:').grid(row=1, column=2, padx=6, pady=6, sticky='w')
        self.cmb_fund_pos = tb.Combobox(op_frame, state='readonly', width=20)
        self.cmb_fund_pos.grid(row=1, column=3, padx=6, pady=6, sticky='ew')

        tb.Label(op_frame, text='SL xuất:').grid(row=1, column=4, padx=6, pady=6, sticky='w')
        qty_inner = tb.Frame(op_frame)
        qty_inner.grid(row=1, column=5, padx=6, pady=6, sticky='w')
        
        self.ent_qty_pos = tb.Entry(qty_inner, width=10)
        self.ent_qty_pos.insert(0, '1')
        self.ent_qty_pos.pack(side='left', padx=(0, 10))
        self._numberize(self.ent_qty_pos)

        # Cấu hình co giãn các cột trong op_frame
        op_frame.columnconfigure(1, weight=1)
        op_frame.columnconfigure(3, weight=1)
        op_frame.columnconfigure(5, weight=2)

        # --- Bind sự kiện
        self.search_pos.bind('<KeyRelease>', lambda e: self.filter_product_list_dispatch())
        self.search_pos.bind('<Down>', lambda e: (self.cmb_prod_pos.focus_set(),
                                                   self.cmb_prod_pos.event_generate('<Alt-Down>')))

        self.cmb_prod_pos.bind('<<ComboboxSelected>>', lambda e: self.update_dispatch_unit_label())
        self.cmb_prod_pos.bind('<Escape>', lambda e: self.search_pos.focus_set())
        self.cmb_prod_pos.bind('<Return>', lambda e: (self.cmb_lot_pos.focus_set(),
                                                      self.cmb_lot_pos.event_generate('<Alt-Down>')))
        self.cmb_lot_pos.bind('<<ComboboxSelected>>', lambda e: self.update_dispatch_funds())
        self.cmb_lot_pos.bind('<Return>', lambda e: (self.cmb_fund_pos.focus_set(),
                                                      self.cmb_fund_pos.event_generate('<Alt-Down>')))
        self.cmb_fund_pos.bind('<Return>', lambda e: self.ent_qty_pos.focus_set())

        # --- Nút tác vụ
        btns = tb.Frame(frm)
        btns.pack(fill='x', padx=8, pady=8)
        
        # Nhóm chính bên trái
        tb.Button(btns, text='+ Thêm vào giỏ hàng', bootstyle='primary', command=self.add_to_dispatch_cart).pack(side='left', padx=4)
        tb.Button(btns, text='Xác nhận xuất kho', bootstyle='success', command=self.confirm_dispatch).pack(side='left', padx=4)
        tb.Button(btns, text='In phiếu xuất kho', bootstyle='info', command=self.print_dispatch_note).pack(side='left', padx=4)
        
        # Nhóm xóa hủy bên phải (phòng tránh bấm nhầm)
        tb.Button(btns, text='Xóa danh sách', bootstyle='danger-outline', command=self.clear_dispatch_cart).pack(side='right', padx=4)
        tb.Button(btns, text='Xóa dòng', bootstyle='warning', command=self.remove_selected_dispatch_item).pack(side='right', padx=4)

        # --- Info tổng quan đơn vị
        info = tb.Frame(frm)
        info.pack(fill='x', padx=8, pady=(0, 4))
        self.lbl_unit_pos = tb.Label(info, text='Đơn vị tính: -', font=('Segoe UI', 10))
        self.lbl_unit_pos.pack(side='left', padx=(8, 12))

        # Báº£ng giÃ³ hÃ ng xuáº¥t
        cols = ('productId', 'productName', 'lotNo', 'expiryDate', 'fundSource', 'unitCode', 'price', 'qty', 'amount')
        self.tree_cart = tb.Treeview(frm, columns=cols, show='headings', height=10)
        for c, w, t, anchor in [
            ('productId', 60, 'PID', 'center'),
            ('productName', 220, 'Tên hàng hóa', 'w'),
            ('lotNo', 100, 'Số lô', 'center'),
            ('expiryDate', 100, 'Hạn dùng', 'center'),
            ('fundSource', 120, 'Nguồn kinh phí', 'w'),
            ('unitCode', 60, 'ĐVT', 'center'),
            ('price', 80, 'Đơn giá', 'e'),
            ('qty', 80, 'SL xuất', 'e'),
            ('amount', 100, 'Thành tiền', 'e')
        ]:
            self.tree_cart.heading(c, text=t, command=(lambda col=c: self.sort_tree(self.tree_cart, col)))
            self.tree_cart.column(c, width=w, anchor=anchor)
        self.tree_cart.tag_configure('odd', background='#f6f8fa')
        self.tree_cart.pack(fill='both', expand=True, padx=8, pady=8)

        self.ent_qty_pos.bind('<Return>', lambda e: self.add_to_dispatch_cart())
        self.cart_dispatch = []
        self.last_dispatch_items = []
        self.last_dispatch_info = {}

    def refresh_receiving_units_combo(self):
        """Cập nhật danh sách đơn vị nhận vào combobox"""
        try:
            units = self.db.get_receiving_units()
            self.cmb_receiving_unit['values'] = units
        except Exception as e:
            print(f"Lỗi refresh đơn vị nhận: {e}")

    def update_dispatch_unit_label(self):
        sel = self.cmb_prod_pos.get()
        if not sel:
            self.lbl_unit_pos.config(text='Đơn vị tính: -')
            self.cmb_lot_pos['values'] = []
            self.cmb_lot_pos.set('')
            return
        pid = int(sel.split(' — ')[0])
        du = self.db.default_unit_of(pid) or '-'
        
        # Lấy các lô hàng khả dụng theo FEFO
        lots = self.db.q('''
          SELECT b.lotNo, b.expiryDate, SUM(sm.qty*1) AS qtyBase
          FROM stock_movements sm
          JOIN batches b ON b.id = sm.batchId
          WHERE sm.productId=?
          GROUP BY sm.batchId
          HAVING qtyBase > 0 AND DATE(b.expiryDate) >= DATE('now')
          ORDER BY DATE(b.expiryDate)
        ''', (pid,))
        
        lot_info_strs = []
        for lot in lots[:3]:
            lot_info_strs.append(f"{lot['lotNo']} (HSD: {lot['expiryDate']}) - Còn: {lot['qtyBase']:g}")
        if len(lots) > 3:
            lot_info_strs.append("...")
            
        if lot_info_strs:
            lots_str = "  |  Lô khả dụng trong kho (FEFO): " + ", ".join(lot_info_strs)
        else:
            lots_str = "  |  HẾT HÀNG TRONG KHO"
            
        self.lbl_unit_pos.config(text=f'Đơn vị tính: {du}{lots_str}')
        
        # Cập nhật danh sách chọn lô thủ công
        lot_options = ["[Tự động - FEFO]"]
        for lot in lots:
            lot_options.append(f"{lot['lotNo']} (HSD: {lot['expiryDate']}) - Tồn: {lot['qtyBase']:g}")
        self.cmb_lot_pos['values'] = lot_options
        self.cmb_lot_pos.current(0)
        self.update_dispatch_funds()

    def update_dispatch_funds(self):
        sel = self.cmb_prod_pos.get()
        if not sel:
            self.cmb_fund_pos['values'] = []
            self.cmb_fund_pos.set('')
            return
        pid = int(sel.split(' — ')[0])
        
        # Lấy lô được chọn
        chosen_val = self.cmb_lot_pos.get()
        chosen_lot = None
        if chosen_val and chosen_val != "[Tự động - FEFO]":
            chosen_lot = chosen_val.split(" (HSD:")[0].strip()
            
        if chosen_lot:
            # Lọc nguồn theo lô cụ thể
            funds = self.db.q('''
                SELECT sm.fundSource, SUM(sm.qty*1) AS qtyBase
                FROM stock_movements sm
                JOIN batches b ON b.id = sm.batchId
                WHERE sm.productId=? AND b.lotNo=?
                GROUP BY sm.fundSource
                HAVING qtyBase > 0
            ''', (pid, chosen_lot))
        else:
            # Lọc nguồn theo sản phẩm nói chung (tổng tất cả các lô)
            funds = self.db.q('''
                SELECT fundSource, SUM(qty*1) AS qtyBase
                FROM stock_movements
                WHERE productId=?
                GROUP BY fundSource
                HAVING qtyBase > 0
            ''', (pid,))
            
        fund_options = ["[Tự động trừ kho]"]
        for f in funds:
            f_name = f['fundSource'] or 'Không rõ nguồn'
            fund_options.append(f"{f_name} - Tồn: {f['qtyBase']:g}")
            
        self.cmb_fund_pos['values'] = fund_options
        self.cmb_fund_pos.current(0)

    def fill_product_by_barcode_dispatch(self, only_select=False):
        bc = self.ent_barcode.get().strip()
        if not bc:
            return False
        row = self.db.q("SELECT id, name FROM products WHERE barcode=?", (bc,))
        if not row:
            if not only_select:
                messagebox.showwarning('Không tìm thấy', 'Barcode không khớp sản phẩm nào')
            return False
        target = f"{row[0]['id']} — {row[0]['name']}"
        if target not in self.cmb_prod_pos['values']:
            self.cmb_prod_pos['values'] = list(self.cmb_prod_pos['values']) + [target]
        self.cmb_prod_pos.set(target)
        self.update_dispatch_unit_label()
        return True

    def scan_and_add_dispatch(self):
        ok = self.fill_product_by_barcode_dispatch(only_select=True)
        if ok:
            self.ent_qty_pos.delete(0, tk.END)
            self.ent_qty_pos.insert(0, '1')
            self.add_to_dispatch_cart()
            self.ent_barcode.delete(0, tk.END)
        self.after(50, lambda: self.ent_barcode.focus_set())

    def open_barcode_scanner_dispatch(self):
        if not BARCODE_AVAILABLE:
            self.show_barcode_install_info()
            return
        try:
            def on_barcode_scanned(barcode_data):
                self.ent_barcode.delete(0, tk.END)
                self.ent_barcode.insert(0, barcode_data)
                self.after(100, self.scan_and_add_dispatch)
            self.barcode_scanner = BarcodeScanner(self, callback=on_barcode_scanned)
            self.barcode_scanner.start_scan()
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể mở barcode scanner: {e}")

    def filter_product_list_dispatch(self):
        kw = (self.search_pos.get() or '').strip().lower()
        opts = [f"{p['id']} — {p['name']}" for p in self._products if kw in p['name'].lower()]
        self.cmb_prod_pos['values'] = opts
        if opts:
            self.cmb_prod_pos.current(0)
            self.update_dispatch_unit_label()

    def add_to_dispatch_cart(self):
        sel = self.cmb_prod_pos.get()
        if not sel:
            messagebox.showerror('Lỗi', 'Chọn sản phẩm để xuất'); return
        pid = int(sel.split(' — ')[0])
        unit = self.db.default_unit_of(pid) or 'vien'
        try:
            qty = float(self.ent_qty_pos.get())
        except:
            qty = 0
        if qty <= 0:
            messagebox.showerror('Lỗi', 'Số lượng xuất phải > 0'); return
        
        name = self.name_by_id(pid)
        
        # Lấy thông tin lô hàng chọn thủ công
        chosen_val = self.cmb_lot_pos.get()
        chosen_lot = None
        if chosen_val and chosen_val != "[Tự động - FEFO]":
            chosen_lot = chosen_val.split(" (HSD:")[0].strip()
            
        # Kiểm tra xem có lô hàng nào cận hạn hơn lô được chọn hay không
        if chosen_lot:
            lots = self.db.q('''
              SELECT b.lotNo, b.expiryDate, SUM(sm.qty*1) AS qtyBase
              FROM stock_movements sm
              JOIN batches b ON b.id = sm.batchId
              WHERE sm.productId=?
              GROUP BY sm.batchId
              HAVING qtyBase > 0 AND DATE(b.expiryDate) >= DATE('now')
              ORDER BY DATE(b.expiryDate)
            ''', (pid,))
            
            chosen_expiry = None
            for lot in lots:
                if lot['lotNo'] == chosen_lot:
                    chosen_expiry = lot['expiryDate']
                    break
                    
            if chosen_expiry:
                earlier_lots = []
                for lot in lots:
                    if lot['lotNo'] != chosen_lot and lot['expiryDate'] < chosen_expiry:
                        earlier_lots.append(f"Lô {lot['lotNo']} (HSD: {lot['expiryDate']}) - Còn: {lot['qtyBase']:g}")
                
                if earlier_lots:
                    warning_msg = f"Cảnh báo: Có lô hàng cận hạn dùng hơn so với lô bạn chọn:\n\n"
                    warning_msg += "\n".join(earlier_lots[:3])
                    if len(earlier_lots) > 3:
                        warning_msg += "\n..."
                    warning_msg += "\n\nHãy cân nhắc kỹ càng! Bạn có chắc chắn muốn tiếp tục chọn lô đã chọn?"
                    if not messagebox.askyesno("Cảnh báo cận hạn", warning_msg):
                        return
                        
        # Lấy nguồn kinh phí chọn thủ công
        chosen_fund_val = self.cmb_fund_pos.get()
        chosen_fund = None
        if chosen_fund_val and chosen_fund_val != "[Tự động trừ kho]":
            chosen_fund = chosen_fund_val.split(" - Tồn:")[0].strip()
            if chosen_fund == "Không rõ nguồn":
                chosen_fund = ""
                         
        merged = False
        for it in self.cart_dispatch:
            if (it['productId'] == pid 
                and it['unitCode'] == unit 
                and it.get('lotNo') == chosen_lot 
                and it.get('fundSource') == chosen_fund):
                it['qty'] = round(it['qty'] + qty, 4)
                merged = True
                break
        if not merged:
            self.cart_dispatch.append({
                'productId': pid,
                'productName': name,
                'unitCode': unit,
                'qty': qty,
                'lotNo': chosen_lot,
                'fundSource': chosen_fund
            })

        self.refresh_dispatch_cart_view()
        self.ent_qty_pos.delete(0, tk.END)
        self.ent_qty_pos.insert(0, '1')
        self.ent_qty_pos.focus_set()

    def remove_selected_dispatch_item(self):
        sel = self.tree_cart.selection()
        if not sel:
            return
        item_vals = self.tree_cart.item(sel[0])['values']
        if not item_vals:
            return
        pid = int(item_vals[0])
        lot = item_vals[2]   # Cột Số lô có chỉ số là 2
        fund = item_vals[4]  # Cột Nguồn kinh phí có chỉ số là 4
        unit = item_vals[5]  # Cột ĐVT có chỉ số là 5
        
        if fund == '[Tự động]':
            fund = None
            
        # Tìm xem sản phẩm có trong giỏ hàng xuất không
        exact_exists = any(it['productId'] == pid and it['unitCode'] == unit and it.get('lotNo') == lot and it.get('fundSource') == fund for it in self.cart_dispatch)
        
        if exact_exists:
            self.cart_dispatch = [it for it in self.cart_dispatch if not (it['productId'] == pid and it['unitCode'] == unit and it.get('lotNo') == lot and it.get('fundSource') == fund)]
        else:
            # Nếu không tìm thấy trùng khớp Số lô chính xác (có thể do xuất tự động FEFO nên trong cart lotNo=None), xóa cả sản phẩm với ĐVT đó
            self.cart_dispatch = [it for it in self.cart_dispatch if not (it['productId'] == pid and it['unitCode'] == unit)]
            
        self.refresh_dispatch_cart_view()

    def clear_dispatch_cart(self):
        if self.cart_dispatch and messagebox.askyesno('Xác nhận', 'Xóa toàn bộ danh sách xuất kho?'):
            self.cart_dispatch = []
            self.refresh_dispatch_cart_view()

    def refresh_dispatch_cart_view(self):
        for i in self.tree_cart.get_children():
            self.tree_cart.delete(i)
        
        idx = 0
        for it in self.cart_dispatch:
            pid = it['productId']
            name = it['productName']
            unitCode = it['unitCode']
            qty = it['qty']
            
            # Lấy thông tin tồn kho của các lô sắp xếp theo HSD (FEFO) hoặc theo lô được chọn
            to_base, _ = self.db.unit_info(pid, unitCode)
            if to_base is None:
                to_base = 1.0
            need_base = qty * to_base
            
            if it.get('lotNo'):
                lots = self.db.q('''
                  SELECT v.batchId, v.qtyBase, b.expiryDate, b.lotNo,
                         COALESCE((
                             SELECT sm2.cost FROM stock_movements sm2
                             WHERE sm2.productId=v.productId AND sm2.batchId=v.batchId
                               AND sm2.type='PURCHASE' AND sm2.cost IS NOT NULL
                             ORDER BY sm2.id DESC LIMIT 1
                         ), 0) AS costBase
                  FROM (
                    SELECT sm.productId, sm.batchId, SUM(sm.qty*1) AS qtyBase
                    FROM stock_movements sm 
                    WHERE sm.productId=? AND sm.batchId=(
                        SELECT id FROM batches WHERE productId=? AND lotNo=? LIMIT 1
                    )
                    GROUP BY sm.batchId
                  ) v JOIN batches b ON b.id=v.batchId
                ''', (pid, pid, it['lotNo']))
            else:
                lots = self.db.q('''
                  SELECT v.batchId, v.qtyBase, b.expiryDate, b.lotNo,
                         COALESCE((
                             SELECT sm2.cost FROM stock_movements sm2
                             WHERE sm2.productId=v.productId AND sm2.batchId=v.batchId
                               AND sm2.type='PURCHASE' AND sm2.cost IS NOT NULL
                             ORDER BY sm2.id DESC LIMIT 1
                         ), 0) AS costBase
                  FROM (
                    SELECT sm.productId, sm.batchId, SUM(sm.qty*1) AS qtyBase
                    FROM stock_movements sm WHERE sm.productId=? GROUP BY sm.batchId
                  ) v JOIN batches b ON b.id=v.batchId
                  WHERE v.qtyBase>0 AND DATE(b.expiryDate) >= DATE('now')
                  ORDER BY DATE(b.expiryDate)
                ''', (pid,))
            
            allocated = []
            for lot in lots:
                if need_base <= 0:
                    break
                take_base = min(need_base, float(lot['qtyBase']))
                take_in_unit = take_base / to_base
                cost_in_unit = float(lot['costBase']) * to_base
                sub_total = take_in_unit * cost_in_unit
                allocated.append({
                    'lotNo': lot['lotNo'],
                    'expiryDate': lot['expiryDate'],
                    'cost': cost_in_unit,
                    'qty': take_in_unit,
                    'total': sub_total
                })
                need_base -= take_base
                
            if need_base > 0:
                # Trường hợp không đủ tồn kho
                price = self.db.unit_price(pid, unitCode)
                allocated.append({
                    'lotNo': 'KHÔNG ĐỦ KHO',
                    'expiryDate': '-',
                    'cost': price,
                    'qty': need_base / to_base,
                    'total': (need_base / to_base) * price
                })
                
            for alloc in allocated:
                self.tree_cart.insert('', 'end',
                    values=(
                        pid,
                        name,
                        alloc['lotNo'],
                        alloc['expiryDate'],
                        it.get('fundSource') or '[Tự động]',
                        unitCode,
                        f"{alloc['cost']:,.0f}",
                        f"{alloc['qty']:g}",
                        f"{alloc['total']:,.0f}"
                    ),
                    tags=('odd',) if idx % 2 else ())
                idx += 1

    def confirm_dispatch(self):
        if not self.cart_dispatch:
            messagebox.showwarning('Chưa có dữ liệu', 'Danh sách xuất kho trống'); return
        
        receiving_unit = self.cmb_receiving_unit.get().strip()
        if not receiving_unit:
            messagebox.showwarning('Thiếu thông tin', 'Vui lòng nhập Đơn vị nhận'); return
        
        date_str = (self.ent_dispatch_date.entry.get() or '').strip()
        if date_str:
            try:
                dt.datetime.strptime(date_str, '%Y-%m-%d')
            except:
                messagebox.showerror('Lỗi', 'Ngày xuất không hợp lệ (định dạng YYYY-MM-DD)'); return
                
        reason = self.cmb_reason.get().strip()
        note = self.ent_dispatch_note.get().strip()

        if not messagebox.askyesno('Xác nhận', f'Bạn có chắc chắn muốn xuất kho cho:\nĐơn vị: {receiving_unit}\nNgày xuất: {date_str or "Hôm nay"}\nLý do: {reason}?'):
            return

        try:
            dispatch_id, note_number, details = self.db.dispatch(self.cart_dispatch, receiving_unit, reason, note, date_str)
            self.last_dispatch_items = details
            
            created_at_str = f"{date_str} {dt.datetime.now().strftime('%H:%M:%S')}" if date_str else dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            self.last_dispatch_info = {
                'id': dispatch_id,
                'noteNumber': note_number,
                'receivingUnit': receiving_unit,
                'reason': reason,
                'note': note,
                'createdAt': created_at_str
            }
            self.cart_dispatch = []
            self.refresh_dispatch_cart_view()
            self.refresh_stock()
            self.refresh_receiving_units_combo()
            messagebox.showinfo('Thành công', f'Đã xuất kho thành công!\nSố phiếu: {note_number}')
            
            # Tự động hỏi in phiếu xuất kho
            if messagebox.askyesno('In phiếu', 'Bạn có muốn in phiếu xuất kho ngay bây giờ?'):
                self.print_dispatch_note()
                
        except Exception as e:
            messagebox.showerror('Lỗi', f'Lỗi xuất kho: {str(e)}')

    def name_by_id(self, pid):
        for p in self._products:
            if p['id'] == pid:
                return p['name']
        return f'#{pid}'

    def print_dispatch_note(self):
        if not self.last_dispatch_items:
            messagebox.showwarning('Chưa có dữ liệu', 'Hãy thực hiện xuất kho trước khi in phiếu'); return
        
        # Kiểm tra và tự động cài đặt reportlab nếu thiếu
        try:
            import reportlab
        except ImportError:
            response = messagebox.askyesno(
                "Thiếu thư viện", 
                "Hệ thống thiếu thư viện 'reportlab' để xuất PDF.\nBạn có muốn tự động cài đặt không? (Quá trình này mất khoảng vài giây)"
            )
            if response:
                import subprocess
                import sys
                try:
                    self.toast("Đang cài đặt thư viện reportlab, vui lòng đợi...")
                    # Chạy lệnh pip install reportlab bằng python hiện tại
                    subprocess.run([sys.executable, "-m", "pip", "install", "reportlab"], check=True)
                    self.toast("Đã cài đặt reportlab thành công!")
                except Exception as ex:
                    messagebox.showerror("Lỗi cài đặt", f"Không thể tự động cài đặt reportlab: {str(ex)}\nHãy chạy lệnh 'pip install reportlab' trong terminal."); return
            else:
                return
        
        info = self.last_dispatch_info
        
        # Cho phép người dùng chọn vị trí lưu file PDF
        initial_filename = f"Phieu_Xuat_Kho_{info['noteNumber']}.pdf"
        pdf_path = filedialog.asksaveasfilename(
            title="Chọn vị trí lưu phiếu xuất kho PDF",
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            initialfile=initial_filename
        )
        if not pdf_path:
            return  # Người dùng hủy bỏ lưu file

        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            
            # Đăng ký font Times New Roman trên Windows hỗ trợ tiếng Việt
            try:
                pdfmetrics.registerFont(TTFont('TimesNewRoman', "C:\\Windows\\Fonts\\times.ttf"))
                pdfmetrics.registerFont(TTFont('TimesNewRoman-Bold', "C:\\Windows\\Fonts\\timesbd.ttf"))
                pdfmetrics.registerFont(TTFont('TimesNewRoman-Italic', "C:\\Windows\\Fonts\\timesi.ttf"))
                font_normal = 'TimesNewRoman'
                font_bold = 'TimesNewRoman-Bold'
                font_italic = 'TimesNewRoman-Italic'
            except Exception:
                font_normal = 'Helvetica'
                font_bold = 'Helvetica-Bold'
                font_italic = 'Helvetica-Oblique'
                
            doc = SimpleDocTemplate(pdf_path, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
            story = []
            
            # Styles
            styles = getSampleStyleSheet()
            
            style_header_left = ParagraphStyle(
                'HeaderLeft',
                parent=styles['Normal'],
                fontName=font_bold,
                fontSize=10,
                leading=14,
                alignment=0 # Left
            )
            
            style_header_right = ParagraphStyle(
                'HeaderRight',
                parent=styles['Normal'],
                fontName=font_normal,
                fontSize=10,
                leading=14,
                alignment=2 # Right
            )
            
            style_title = ParagraphStyle(
                'Title',
                parent=styles['Heading1'],
                fontName=font_bold,
                fontSize=16,
                leading=20,
                alignment=1, # Center
                spaceAfter=5
            )
            
            style_subtitle = ParagraphStyle(
                'Subtitle',
                parent=styles['Normal'],
                fontName=font_bold,
                fontSize=11,
                leading=14,
                alignment=1, # Center
                spaceAfter=15
            )
            
            style_info = ParagraphStyle(
                'Info',
                parent=styles['Normal'],
                fontName=font_normal,
                fontSize=11,
                leading=16,
                alignment=0
            )
            
            style_table_header = ParagraphStyle(
                'TableHeader',
                parent=styles['Normal'],
                fontName=font_bold,
                fontSize=10,
                leading=12,
                alignment=1, # Center
                textColor=colors.black
            )
            
            style_cell = ParagraphStyle(
                'Cell',
                parent=styles['Normal'],
                fontName=font_normal,
                fontSize=10,
                leading=12,
                alignment=0
            )
            
            style_cell_center = ParagraphStyle(
                'CellCenter',
                parent=styles['Normal'],
                fontName=font_normal,
                fontSize=10,
                leading=12,
                alignment=1
            )
            
            style_cell_right = ParagraphStyle(
                'CellRight',
                parent=styles['Normal'],
                fontName=font_normal,
                fontSize=10,
                leading=12,
                alignment=2
            )
            
            # Header
            header_data = [
                [
                    Paragraph("SỞ Y TẾ THÀNH PHỐ CẦN THƠ<br/>TRUNG TÂM KIỂM SOÁT BỆNH TẬT (CDC)", style_header_left),
                    Paragraph("<b>Mẫu số: C31-HD</b><br/><i>(Ban hành theo Thông tư số 107/2017/TT-BTC)</i>", style_header_right)
                ]
            ]
            header_table = Table(header_data, colWidths=[280, 230])
            header_table.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
                ('BOTTOMPADDING', (0,0), (-1,-1), 10),
            ]))
            story.append(header_table)
            story.append(Spacer(1, 10))
            
            # Title
            story.append(Paragraph("PHIẾU XUẤT KHO", style_title))
            story.append(Paragraph(f"Số: {info['noteNumber']}", style_subtitle))
            
            # Parse createdAt
            created_str = info['createdAt']
            try:
                created_at_dt = dt.datetime.strptime(created_str, '%Y-%m-%d %H:%M:%S')
            except Exception:
                try:
                    created_at_dt = dt.datetime.strptime(created_str.split(' ')[0], '%Y-%m-%d')
                except Exception:
                    created_at_dt = dt.datetime.now()
            
            info_lines = [
                f"<b>Đơn vị nhận:</b> {info['receivingUnit']}",
                f"<b>Lý do xuất:</b> {info['reason']}",
                f"<b>Kho xuất:</b> Kho Dược CDC Cần Thơ",
                f"<b>Ngày xuất:</b> {created_at_dt.strftime('%d/%m/%Y')}",
                f"<b>Ghi chú:</b> {info['note'] or 'Không'}"
            ]
            for line in info_lines:
                story.append(Paragraph(line, style_info))
                story.append(Spacer(1, 4))
                
            story.append(Spacer(1, 10))
            
            # Table of items
            table_data = [
                [
                    Paragraph("STT", style_table_header),
                    Paragraph("Tên thuốc, vaccine, VTYT", style_table_header),
                    Paragraph("ĐVT", style_table_header),
                    Paragraph("Số lượng", style_table_header),
                    Paragraph("Đơn giá", style_table_header),
                    Paragraph("Thành tiền", style_table_header),
                    Paragraph("Số lô", style_table_header),
                    Paragraph("Hạn dùng", style_table_header),
                    Paragraph("Nguồn", style_table_header)
                ]
            ]
            
            total_sum = 0.0
            for idx, it in enumerate(self.last_dispatch_items, 1):
                qty = it['qty']
                cost = it.get('cost') or 0.0
                sub_total = qty * cost
                total_sum += sub_total
                
                table_data.append([
                    Paragraph(str(idx), style_cell_center),
                    Paragraph(it['productName'], style_cell),
                    Paragraph(it['unitCode'], style_cell_center),
                    Paragraph(f"{qty:g}", style_cell_right),
                    Paragraph(f"{cost:,.0f}" if cost > 0 else "0", style_cell_right),
                    Paragraph(f"{sub_total:,.0f}" if sub_total > 0 else "0", style_cell_right),
                    Paragraph(it['lotNo'] or '', style_cell_center),
                    Paragraph(it['expiryDate'] or '', style_cell_center),
                    Paragraph(it.get('fundSource') or '', style_cell)
                ])
                
            # Thêm dòng tổng cộng
            table_data.append([
                Paragraph("<b>Tổng cộng</b>", style_cell_center),
                Paragraph("", style_cell),
                Paragraph("", style_cell_center),
                Paragraph("", style_cell_right),
                Paragraph("", style_cell_right),
                Paragraph(f"<b>{total_sum:,.0f}</b>", style_cell_right),
                Paragraph("", style_cell_center),
                Paragraph("", style_cell_center),
                Paragraph("", style_cell)
            ])
                
            col_widths = [20, 130, 30, 40, 50, 60, 50, 50, 80]
            items_table = Table(table_data, colWidths=col_widths)
            items_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f2f2f2')),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.black),
                ('SPAN', (0, -1), (4, -1)),
                ('TOPPADDING', (0,0), (-1,-1), 6),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ]))
            story.append(items_table)
            story.append(Spacer(1, 15))
            
            # Date & Signatures
            date_right_style = ParagraphStyle(
                'DateRight',
                parent=styles['Normal'],
                fontName=font_italic,
                fontSize=11,
                alignment=2,
                spaceAfter=10
            )
            
            sig_title_style = ParagraphStyle(
                'SigTitle',
                parent=styles['Normal'],
                fontName=font_bold,
                fontSize=11,
                alignment=1
            )
            
            sig_sub_style = ParagraphStyle(
                'SigSub',
                parent=styles['Normal'],
                fontName=font_italic,
                fontSize=9,
                alignment=1
            )
            
            story.append(Paragraph(f"Cần Thơ, ngày {created_at_dt.strftime('%d')} tháng {created_at_dt.strftime('%m')} năm {created_at_dt.strftime('%Y')}", date_right_style))
            
            sig_headers = [
                [
                    Paragraph("<b>Người lập phiếu</b>", sig_title_style),
                    Paragraph("<b>Người nhận hàng</b>", sig_title_style),
                    Paragraph("<b>Thủ kho</b>", sig_title_style),
                    Paragraph("<b>Kế toán trưởng</b>", sig_title_style),
                    Paragraph("<b>Lãnh đạo đơn vị</b>", sig_title_style)
                ],
                [
                    Paragraph("(Ký, họ tên)", sig_sub_style),
                    Paragraph("(Ký, họ tên)", sig_sub_style),
                    Paragraph("(Ký, họ tên)", sig_sub_style),
                    Paragraph("(Ký, họ tên)", sig_sub_style),
                    Paragraph("(Ký, đóng dấu)", sig_sub_style)
                ]
            ]
            sig_table = Table(sig_headers, colWidths=[102, 102, 102, 102, 102])
            sig_table.setStyle(TableStyle([
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('BOTTOMPADDING', (0,0), (-1,-1), 2),
            ]))
            story.append(sig_table)
            story.append(Spacer(1, 60)) # Chừa khoảng trống để ký và ghi tên
            
            doc.build(story)
            
            # Mở file PDF kết quả
            os.startfile(pdf_path)
            self.toast("Đã in phiếu xuất ra PDF và mở file thành công")
            
        except Exception as e:
            messagebox.showerror("Lỗi in PDF", f"Không thể xuất file PDF: {str(e)}")

    # -------- Stock --------
    def build_stock_tab(self):
        frm = self.tab_stock
        cols = ('product','productName','batch','lot','exp','qty')
        self.tree_stock2 = tb.Treeview(frm, columns=cols, show='headings')
        for c, w, t, anchor in [
            ('product',70,'PID','center'),('productName',300,'Tên thuốc','w'),('batch',70,'Batch','center'),
            ('lot',130,'Lot','w'),('exp',100,'HSD','center'),('qty',120,'SL (base)','e')
        ]:
            self.tree_stock2.heading(c, text=t, command=(lambda col=c: self.sort_tree(self.tree_stock2, col)))
            self.tree_stock2.column(c, width=w, anchor=anchor)
        self.tree_stock2.tag_configure('odd', background='#f6f8fa')
        self.tree_stock2.pack(fill='both', expand=True, padx=8, pady=8)

    # -------- Alerts --------
    def build_alerts_tab(self):
        frm = self.tab_alerts
        top = tb.Frame(frm); top.pack(fill='x', padx=8, pady=8)
        tb.Label(top, text='Cảnh báo trong (ngày):').pack(side='left')
        self.ent_warn_days = tb.Entry(top, width=6); self.ent_warn_days.insert(0, '180'); self.ent_warn_days.pack(side='left', padx=6)
        self._numberize(self.ent_warn_days)
        tb.Button(top, text='Làm mới', bootstyle='secondary', command=self.refresh_alerts).pack(side='left', padx=8)

        cols = ('product','productName','batch','lot','exp','qty')
        self.tree_alerts = tb.Treeview(frm, columns=cols, show='headings')
        for c, w, t, anchor in [
            ('product',70,'PID','center'),('productName',300,'Tên thuốc','w'),('batch',70,'Batch','center'),
            ('lot',130,'Lot','w'),('exp',100,'HSD','center'),('qty',120,'SL (base)','e')
        ]:
            self.tree_alerts.heading(c, text=t, command=(lambda col=c: self.sort_tree(self.tree_alerts, col)))
            self.tree_alerts.column(c, width=w, anchor=anchor)
        self.tree_alerts.tag_configure('odd', background='#f6f8fa')
        self.tree_alerts.pack(fill='both', expand=True, padx=8, pady=8)

    # -------- Report --------
    def build_report_tab(self):
        frm = self.tab_report

        # Thanh điều kiện: từ ngày / đến ngày
        top = tb.Frame(frm); top.pack(fill='x', padx=8, pady=8)

        tb.Label(top, text='Từ ngày:').pack(side='left', padx=(0,6))
        # DateEntry đã được bạn thêm trước đó; nếu chưa có, nhớ: from ttkbootstrap.widgets import DateEntry
        self.de_from = DateEntry(top, dateformat="%Y-%m-%d", firstweekday=0, bootstyle='secondary')
        self.de_from.entry.delete(0, 'end')
        self.de_from.entry.insert(0, dt.date.today().replace(day=1).strftime("%Y-%m-%d"))  # đầu tháng
        self.de_from.pack(side='left', padx=(0,12))

        tb.Label(top, text='Đến ngày:').pack(side='left', padx=(0,6))
        self.de_to = DateEntry(top, dateformat="%Y-%m-%d", firstweekday=0, bootstyle='secondary')
        self.de_to.entry.delete(0, 'end')
        self.de_to.entry.insert(0, dt.datetime.now().strftime("%Y-%m-%d"))      # hôm nay
        self.de_to.pack(side='left', padx=(0,12))

        tb.Label(top, text='Nguồn:').pack(side='left', padx=(0,6))
        self.cmb_report_fund = tb.Combobox(top, width=22, state='readonly')
        self.cmb_report_fund.pack(side='left', padx=(0,12))
        self.cmb_report_fund.bind('<<ComboboxSelected>>', lambda e: self.refresh_report())

        tb.Button(top, text='Làm mới', bootstyle='primary', command=self.refresh_report).pack(side='left', padx=6)
        tb.Button(top, text='Xuất CSV…', bootstyle='info', command=self.export_report_csv).pack(side='left', padx=6)
        tb.Button(top, text='Xuất PDF…', bootstyle='danger', command=self.export_report_pdf).pack(side='left', padx=6)
        tb.Button(top, text='Biên bản kiểm kê (PDF)', bootstyle='warning', command=self.print_inventory_check_pdf).pack(side='left', padx=6)

        # Bảng Xuất–Nhập–Tồn
        cols = ('product','productName','lotNo','expiryDate','fundSource','opening','inbound','outbound','closing')
        self.tree_report = tb.Treeview(frm, columns=cols, show='headings')
        for c, w, t, anchor in [
            ('product',50,'PID','center'),
            ('productName',220,'Tên thuốc/vaccine/VTYT','w'),
            ('lotNo',80,'Số lô','center'),
            ('expiryDate',90,'Hạn sử dụng','center'),
            ('fundSource',110,'Nguồn kinh phí','w'),
            ('opening',80,'Tồn đầu','e'),
            ('inbound',80,'Nhập','e'),
            ('outbound',80,'Xuất','e'),
            ('closing',80,'Tồn cuối','e'),
        ]:
            self.tree_report.heading(c, text=t, command=(lambda col=c: self.sort_tree(self.tree_report, col)))
            self.tree_report.column(c, width=w, anchor=anchor)

        self.tree_report.tag_configure('odd', background='#f6f8fa')
        self.tree_report.tag_configure('total', background='#e8f5e9')  # dòng tổng
        self.tree_report.pack(fill='both', expand=True, padx=8, pady=8)

    # -------- Backup --------
    def build_backup_tab(self):
        frm = self.tab_backup
        
        # Khung tạo backup
        backup_frame = tb.Labelframe(frm, text='Tạo Backup', bootstyle='light')
        backup_frame.pack(fill='x', padx=8, pady=8)
        
        btn_frame = tb.Frame(backup_frame)
        btn_frame.pack(fill='x', padx=8, pady=8)
        
        tb.Button(btn_frame, text='💾 Tạo Backup Ngay', bootstyle='success',
                  command=self.create_manual_backup).pack(side='left', padx=4)
        tb.Button(btn_frame, text='📤 Export Dữ Liệu', bootstyle='info',
                  command=self.export_data).pack(side='left', padx=4)
        tb.Button(btn_frame, text='📥 Import Dữ Liệu', bootstyle='warning',
                  command=self.import_data).pack(side='left', padx=4)
        
        # Khung khôi phục backup
        restore_frame = tb.Labelframe(frm, text='Khôi Phục Backup', bootstyle='light')
        restore_frame.pack(fill='x', padx=8, pady=8)
        
        tb.Label(restore_frame, text='Chọn backup để khôi phục:').pack(anchor='w', padx=8, pady=(8,4))
        
        # Bảng danh sách backup
        cols = ('file', 'created', 'size', 'version')
        self.tree_backups = tb.Treeview(restore_frame, columns=cols, show='headings', height=8)
        for c, w, t, anchor in [
            ('file', 300, 'Tên File', 'w'),
            ('created', 150, 'Ngày Tạo', 'center'),
            ('size', 100, 'Kích Thước', 'e'),
            ('version', 80, 'Phiên Bản', 'center')
        ]:
            self.tree_backups.heading(c, text=t, command=(lambda col=c: self.sort_tree(self.tree_backups, col)))
            self.tree_backups.column(c, width=w, anchor=anchor)
        self.tree_backups.tag_configure('odd', background='#f6f8fa')
        self.tree_backups.pack(fill='x', padx=8, pady=8)
        
        # Nút khôi phục
        restore_btn_frame = tb.Frame(restore_frame)
        restore_btn_frame.pack(fill='x', padx=8, pady=(0,8))
        
        tb.Button(restore_btn_frame, text='🔄 Khôi Phục Backup', bootstyle='danger',
                  command=self.restore_selected_backup).pack(side='left', padx=4)
        tb.Button(restore_btn_frame, text='🗑️ Xóa Backup', bootstyle='outline-danger',
                  command=self.delete_selected_backup).pack(side='left', padx=4)
        tb.Button(restore_btn_frame, text='🔄 Làm Mới', bootstyle='secondary',
                  command=self.refresh_backup_list).pack(side='left', padx=4)
        
        # Thông tin auto backup
        info_frame = tb.Labelframe(frm, text='Tự Động Backup', bootstyle='info')
        info_frame.pack(fill='x', padx=8, pady=8)
        
        info_text = tb.Text(info_frame, height=4, wrap='word')
        info_text.pack(fill='x', padx=8, pady=8)
        info_text.insert('1.0', 
            "• Tự động backup mỗi ngày lúc 2:00 AM\n"
            "• Giữ tối đa 30 file backup\n"
            "• Backup được lưu trong thư mục: " + BACKUP_DIR + "\n"
            "• Trước khi khôi phục, hệ thống sẽ tự động tạo backup hiện tại")
        info_text.config(state='disabled')
        
        # Load danh sách backup
        self.refresh_backup_list()

    def create_manual_backup(self):
        """Tạo backup thủ công"""
        try:
            backup_path = self.backup_manager.create_backup("manual")
            self.toast(f'Đã tạo backup: {os.path.basename(backup_path)}')
            self.refresh_backup_list()
        except Exception as e:
            messagebox.showerror('Lỗi', str(e))

    def export_data(self):
        """Export dữ liệu ra file JSON"""
        try:
            path = filedialog.asksaveasfilename(
                defaultextension='.json',
                filetypes=[('JSON files', '*.json'), ('All files', '*.*')],
                initialfile=f'export_data_{dt.datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
            )
            if path:
                self.backup_manager.export_data(path)
                self.toast('Đã export dữ liệu thành công')
        except Exception as e:
            messagebox.showerror('Lỗi', str(e))

    def import_data(self):
        """Import dữ liệu từ file JSON"""
        try:
            path = filedialog.askopenfilename(
                filetypes=[('JSON files', '*.json'), ('All files', '*.*')]
            )
            if path:
                if messagebox.askyesno('Xác nhận', 
                    'Import sẽ thay thế toàn bộ dữ liệu hiện tại!\n'
                    'Hệ thống sẽ tự động tạo backup trước khi import.\n'
                    'Bạn có chắc chắn muốn tiếp tục?'):
                    
                    self.backup_manager.import_data(path)
                    self.toast('Đã import dữ liệu thành công')
                    # Refresh tất cả dữ liệu
                    self.refresh_products()
                    self.refresh_stock()
                    self.refresh_alerts()
                    self.refresh_report()
                    self.refresh_backup_list()
        except Exception as e:
            messagebox.showerror('Lỗi', str(e))

    def restore_selected_backup(self):
        """Khôi phục backup được chọn"""
        selection = self.tree_backups.selection()
        if not selection:
            messagebox.showwarning('Cảnh báo', 'Chọn backup để khôi phục')
            return
        
        item = self.tree_backups.item(selection[0])
        backup_file = item['values'][0]
        backup_path = os.path.join(BACKUP_DIR, backup_file)
        
        if messagebox.askyesno('Xác nhận', 
            f'Khôi phục backup: {backup_file}\n'
            'Hệ thống sẽ tự động tạo backup hiện tại trước khi khôi phục.\n'
            'Bạn có chắc chắn muốn tiếp tục?'):
            
            try:
                # 1. Kiểm tra trạng thái máy chủ di động và dừng nếu đang chạy
                server_was_running = False
                if self.mobile_server and self.mobile_server.is_running:
                    server_was_running = True
                    try:
                        self.mobile_server.stop()
                    except Exception as server_err:
                        print(f"Lỗi khi dừng server trước khi restore: {server_err}")
                
                # 2. Đóng kết nối cơ sở dữ liệu chính đang mở
                try:
                    self.db.conn.close()
                except Exception as conn_err:
                    print(f"Lỗi khi đóng kết nối DB: {conn_err}")
                
                # 3. Tiến hành khôi phục tệp tin cơ sở dữ liệu
                self.backup_manager.restore_backup(backup_path)
                
                # 4. Khởi tạo lại kết nối DB và gán lại cho các quản lý
                self.db = DB(DB_PATH)
                self.backup_manager = BackupManager(DB_PATH, BACKUP_DIR)
                self.report_manager = ReportManager(DB_PATH)
                self.medicine_catalog = MedicineCatalogManager(DB_PATH)
                
                # 5. Khởi động lại máy chủ di động nếu trước đó nó đang chạy
                if server_was_running:
                    from server import MobileInventoryServer
                    self.mobile_server = MobileInventoryServer(self, host="0.0.0.0", port=5000)
                    self.mobile_server.start()
                
                self.toast('Đã khôi phục backup thành công')
                
                # 6. Làm mới lại toàn bộ giao diện
                self.refresh_all_data()
                self.refresh_backup_list()
                self.update_mobile_server_ui()
                
            except Exception as e:
                # Trong trường hợp có lỗi xảy ra, đảm bảo DB được khởi tạo lại để ứng dụng không bị treo
                try:
                    self.db = DB(DB_PATH)
                    self.backup_manager = BackupManager(DB_PATH, BACKUP_DIR)
                    self.report_manager = ReportManager(DB_PATH)
                    self.medicine_catalog = MedicineCatalogManager(DB_PATH)
                except:
                    pass
                messagebox.showerror('Lỗi', str(e))

    def delete_selected_backup(self):
        """Xóa backup được chọn"""
        selection = self.tree_backups.selection()
        if not selection:
            messagebox.showwarning('Cảnh báo', 'Chọn backup để xóa')
            return
        
        item = self.tree_backups.item(selection[0])
        backup_file = item['values'][0]
        
        if messagebox.askyesno('Xác nhận', f'Xóa backup: {backup_file}?'):
            try:
                backup_path = os.path.join(BACKUP_DIR, backup_file)
                os.remove(backup_path)
                metadata_path = backup_path.replace('.db', '.json')
                if os.path.exists(metadata_path):
                    os.remove(metadata_path)
                self.toast('Đã xóa backup')
                self.refresh_backup_list()
            except Exception as e:
                messagebox.showerror('Lỗi', str(e))

    def refresh_backup_list(self):
        """Làm mới danh sách backup"""
        try:
            # Clear tree
            for item in self.tree_backups.get_children():
                self.tree_backups.delete(item)
            
            backups = self.backup_manager.list_backups()
            for idx, backup in enumerate(backups):
                size_mb = backup['size'] / (1024 * 1024)
                created_str = backup['created'].strftime('%Y-%m-%d %H:%M')
                version = backup.get('version', 'N/A')
                
                self.tree_backups.insert('', 'end',
                    values=(backup['file'], created_str, f'{size_mb:.1f} MB', version),
                    tags=('odd',) if idx % 2 else ())
        except Exception as e:
            messagebox.showerror('Lỗi', f'Không thể load danh sách backup: {str(e)}')

    # -------- Advanced Reports --------
    def build_advanced_reports_tab(self):
        frm = self.tab_advanced_reports
        
        # Khung điều khiển
        control_frame = tb.Labelframe(frm, text='Điều kiện báo cáo', bootstyle='light')
        control_frame.pack(fill='x', padx=8, pady=8)
        
        # Hàng 1: Ngày tháng
        date_frame = tb.Frame(control_frame)
        date_frame.pack(fill='x', padx=8, pady=8)
        
        tb.Label(date_frame, text='Từ ngày:').pack(side='left', padx=(0,6))
        self.adv_de_from = DateEntry(date_frame, dateformat="%Y-%m-%d", firstweekday=0, bootstyle='secondary')
        self.adv_de_from.entry.delete(0, 'end')
        self.adv_de_from.entry.insert(0, dt.date.today().replace(day=1).strftime("%Y-%m-%d"))
        self.adv_de_from.pack(side='left', padx=(0,12))
        
        tb.Label(date_frame, text='Đến ngày:').pack(side='left', padx=(0,6))
        self.adv_de_to = DateEntry(date_frame, dateformat="%Y-%m-%d", firstweekday=0, bootstyle='secondary')
        self.adv_de_to.entry.delete(0, 'end')
        self.adv_de_to.entry.insert(0, dt.datetime.now().strftime("%Y-%m-%d"))
        self.adv_de_to.pack(side='left', padx=(0,12))
        
        # Hàng 2: Nút báo cáo
        btn_frame = tb.Frame(control_frame)
        btn_frame.pack(fill='x', padx=8, pady=(0,8))
        
        tb.Button(btn_frame, text='📥 Lịch sử phiếu nhập', bootstyle='secondary',
                  command=self.show_purchase_history).pack(side='left', padx=4)
        tb.Button(btn_frame, text='📋 Lịch sử phiếu xuất', bootstyle='secondary',
                  command=self.show_dispatch_history).pack(side='left', padx=4)
        tb.Button(btn_frame, text='📤 Thống kê phiếu xuất', bootstyle='success',
                  command=self.show_revenue_report).pack(side='left', padx=4)
        tb.Button(btn_frame, text='🏢 Thống kê theo đơn vị nhận', bootstyle='info',
                  command=self.show_profit_report).pack(side='left', padx=4)
        tb.Button(btn_frame, text='🏆 Top sản phẩm cấp phát', bootstyle='warning',
                  command=self.show_top_products_report).pack(side='left', padx=4)
        tb.Button(btn_frame, text='📈 Biểu đồ cấp phát', bootstyle='primary',
                  command=self.show_revenue_chart).pack(side='left', padx=4)
        
        # Hàng 3: Nút xuất báo cáo
        export_frame = tb.Frame(control_frame)
        export_frame.pack(fill='x', padx=8, pady=(0,8))
        
        tb.Label(export_frame, text='Xuất báo cáo:', font=('Segoe UI', 9, 'bold')).pack(side='left', padx=(0,8))
        tb.Button(export_frame, text='📊 Excel', bootstyle='success',
                  command=self.export_current_report_excel).pack(side='left', padx=4)
        tb.Button(export_frame, text='📄 PDF', bootstyle='danger',
                  command=self.export_current_report_pdf).pack(side='left', padx=4)
        tb.Button(export_frame, text='📋 CSV', bootstyle='secondary',
                  command=self.export_current_report_csv).pack(side='left', padx=4)
        
        # Khung hiển thị báo cáo
        self.report_display_frame = tb.Frame(frm)
        self.report_display_frame.pack(fill='both', expand=True, padx=8, pady=8)
        
        # Tạo notebook cho các loại báo cáo
        self.adv_report_nb = tb.Notebook(self.report_display_frame)
        self.adv_report_nb.pack(fill='both', expand=True)
        
        # Tab tóm tắt
        self.adv_summary_tab = tb.Frame(self.adv_report_nb)
        self.adv_report_nb.add(self.adv_summary_tab, text='📋 Tóm tắt')
        
        # Tab báo cáo chi tiết
        self.adv_detail_tab = tb.Frame(self.adv_report_nb)
        self.adv_report_nb.add(self.adv_detail_tab, text='📊 Chi tiết')
        
        # Tab biểu đồ
        self.adv_chart_tab = tb.Frame(self.adv_report_nb)
        self.adv_report_nb.add(self.adv_chart_tab, text='📈 Biểu đồ')
        
        # Thống kê báo cáo hiện tại
        self.current_report_type = 'dispatch'
        self.current_report_group_by = 'day'
        
        # Load tóm tắt ban đầu
        self.load_advanced_summary()

    def load_advanced_summary(self):
        """Load tóm tắt báo cáo cấp phát"""
        try:
            # Clear summary tab
            for widget in self.adv_summary_tab.winfo_children():
                widget.destroy()
            
            start_date = self.adv_de_from.entry.get().strip()
            end_date = self.adv_de_to.entry.get().strip()
            
            if not start_date or not end_date:
                return
            
            # Lấy dữ liệu tóm tắt
            summary_data = self.report_manager.get_daily_sales_summary(start_date, end_date)
            summary = summary_data['summary']
            
            # Tạo layout tóm tắt
            main_frame = tb.Frame(self.adv_summary_tab)
            main_frame.pack(fill='both', expand=True, padx=20, pady=20)
            
            # Title
            title_label = tb.Label(main_frame, text="TÓM TẮT THÔNG TIN CẤP PHÁT KHO", 
                                  font=('Segoe UI', 16, 'bold'), bootstyle='primary')
            title_label.pack(pady=(0, 20))
            
            # KPI Cards
            kpi_frame = tb.Frame(main_frame)
            kpi_frame.pack(fill='x', pady=(0, 20))
            
            # Card 1: Tổng phiếu xuất
            card1 = tb.Labelframe(kpi_frame, text='Tổng số phiếu xuất', bootstyle='info')
            card1.pack(side='left', fill='both', expand=True, padx=(0, 10))
            total_orders = summary.get('total_orders', 0) or 0
            tb.Label(card1, text=f"{total_orders:,}", 
                    font=('Segoe UI', 24, 'bold'), bootstyle='info').pack(pady=10)
            
            # Card 2: Tổng số lượng xuất
            card2 = tb.Labelframe(kpi_frame, text='Tổng số lượng xuất', bootstyle='success')
            card2.pack(side='left', fill='both', expand=True, padx=(0, 10))
            total_revenue = summary.get('total_revenue', 0) or 0
            tb.Label(card2, text=f"{total_revenue:,.0f}", 
                    font=('Segoe UI', 24, 'bold'), bootstyle='success').pack(pady=10)
            
            # Card 3: Số lượng TB/phiếu
            card3 = tb.Labelframe(kpi_frame, text='SL trung bình/phiếu', bootstyle='warning')
            card3.pack(side='left', fill='both', expand=True, padx=(0, 10))
            avg_order = summary.get('avg_order_value', 0) or 0
            tb.Label(card3, text=f"{avg_order:,.1f}", 
                    font=('Segoe UI', 24, 'bold'), bootstyle='warning').pack(pady=10)
            
            # Card 4: Phiếu xuất lớn nhất
            card4 = tb.Labelframe(kpi_frame, text='Phiếu xuất lớn nhất', bootstyle='danger')
            card4.pack(side='left', fill='both', expand=True)
            max_order = summary.get('max_order', 0) or 0
            tb.Label(card4, text=f"{max_order:,.0f}", 
                    font=('Segoe UI', 24, 'bold'), bootstyle='danger').pack(pady=10)
            
            # Bảng cấp phát theo ngày
            daily_frame = tb.Labelframe(main_frame, text='Cấp phát theo ngày', bootstyle='light')
            daily_frame.pack(fill='both', expand=True)
            
            cols = ('date', 'orders', 'revenue')
            daily_tree = tb.Treeview(daily_frame, columns=cols, show='headings', height=8)
            for c, w, t, anchor in [
                ('date', 120, 'Ngày', 'center'),
                ('orders', 100, 'Số phiếu', 'e'),
                ('revenue', 150, 'Tổng số lượng xuất', 'e')
            ]:
                daily_tree.heading(c, text=t)
                daily_tree.column(c, width=w, anchor=anchor)
            
            daily_tree.tag_configure('odd', background='#f6f8fa')
            daily_tree.pack(fill='both', expand=True, padx=8, pady=8)
            
            # Load dữ liệu
            daily_data = summary_data.get('daily_data', [])
            for idx, row in enumerate(daily_data):
                orders = row.get('orders', 0) or 0
                revenue = row.get('revenue', 0) or 0
                daily_tree.insert('', 'end', values=(
                    row.get('sale_date', ''),
                    f"{orders:,}",
                    f"{revenue:,.0f}"
                ), tags=('odd',) if idx % 2 else ())
                
        except Exception as e:
            messagebox.showerror('Lỗi', f'Không thể load tóm tắt: {str(e)}')

    def build_mobile_tab(self):
        """Xây dựng giao diện cho Tab Kiểm kho di động"""
        # Xóa các widget cũ
        for widget in self.tab_mobile.winfo_children():
            widget.destroy()
            
        main_frame = tb.Frame(self.tab_mobile)
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        main_frame.columnconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(0, weight=1)
        
        # --- CỘT TRÁI: ĐIỀU KHIỂN & TRẠNG THÁI ---
        left_frame = tb.LabelFrame(main_frame, text="⚙️ Cấu hình máy chủ di động", padding=15)
        left_frame.grid(row=0, column=0, sticky='nsew', padx=(0, 10))
        
        # Trạng thái máy chủ
        status_lbl_frame = tb.Frame(left_frame)
        status_lbl_frame.pack(fill='x', pady=10)
        
        tb.Label(status_lbl_frame, text="Trạng thái máy chủ:", font=('Segoe UI', 11, 'bold')).pack(side='left')
        
        self.mobile_status_val = tb.Label(status_lbl_frame, font=('Segoe UI', 11, 'bold'))
        self.mobile_status_val.pack(side='left', padx=10)
        
        # Điều khiển nút
        self.btn_toggle_server = tb.Button(left_frame, command=self.toggle_mobile_server, bootstyle='success')
        self.btn_toggle_server.pack(fill='x', pady=10)
        
        # Đường link kết nối
        link_frame = tb.Frame(left_frame)
        link_frame.pack(fill='x', pady=10)
        
        tb.Label(link_frame, text="Địa chỉ kết nối LAN:", font=('Segoe UI', 10, 'bold')).pack(anchor='w')
        
        self.mobile_url_val = tb.Entry(link_frame, font=('Segoe UI', 10), state='readonly')
        self.mobile_url_val.pack(fill='x', pady=5)
        
        def copy_url():
            url = self.mobile_url_val.get()
            if url:
                self.clipboard_clear()
                self.clipboard_append(url)
                self.toast("Đã copy đường dẫn kết nối vào clipboard!")
                
        def open_browser():
            url = self.mobile_url_val.get()
            if url:
                import webbrowser
                webbrowser.open(url)
                
        btn_copy = tb.Button(link_frame, text="📋 Sao chép liên kết", command=copy_url, bootstyle='outline-info')
        btn_copy.pack(side='left', padx=2)
        
        btn_open = tb.Button(link_frame, text="🌐 Mở trên PC (Test)", command=open_browser, bootstyle='outline-secondary')
        btn_open.pack(side='left', padx=2)
        
        # Hướng dẫn chi tiết
        help_text = (
            "💡 Hướng dẫn sử dụng:\n"
            "1. Đảm bảo máy tính và điện thoại di động cùng kết nối chung một mạng Wi-Fi cục bộ.\n"
            "2. Khởi động máy chủ bằng nút phía trên (nếu đang dừng).\n"
            "3. Lấy điện thoại di động quét mã QR ở ô bên phải để mở liên kết.\n"
            "4. Cấp quyền truy cập Camera cho trình duyệt trên điện thoại nếu được hỏi.\n"
            "5. Đưa camera điện thoại quét mã vạch sản phẩm để kiểm tra số tồn tức thì."
        )
        tb.Label(left_frame, text=help_text, font=('Segoe UI', 9), justify='left', 
                 wraplength=450, bootstyle='secondary').pack(anchor='w', pady=15)
                 
        # --- CỘT PHẢI: MÃ QR ---
        right_frame = tb.LabelFrame(main_frame, text="📱 Quét mã QR để kết nối", padding=15)
        right_frame.grid(row=0, column=1, sticky='nsew', padx=(10, 0))
        
        # Canvas vẽ QR Code
        self.qr_canvas = tk.Canvas(right_frame, width=260, height=260, bg='white', relief='ridge', borderwidth=1)
        self.qr_canvas.pack(pady=10)
        
        self.qr_help_lbl = tb.Label(right_frame, text="Đang tạo mã QR kết nối...", font=('Segoe UI', 10), justify='center')
        self.qr_help_lbl.pack(pady=5)
        
        # Cập nhật UI ban đầu
        self.update_mobile_server_ui()

    def toggle_mobile_server(self):
        """Khởi động hoặc dừng máy chủ di động"""
        if self.mobile_server and self.mobile_server.is_running:
            try:
                self.mobile_server.stop()
                self.toast("Đã dừng máy chủ di động")
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể dừng máy chủ: {e}")
        else:
            try:
                self.mobile_server = MobileInventoryServer(self, host="0.0.0.0", port=5000)
                self.mobile_server.start()
                self.toast("Đã khởi động máy chủ di động")
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể khởi động máy chủ: {e}")
        
        self.update_mobile_server_ui()

    def refresh_all_data(self):
        """Làm mới tất cả bảng dữ liệu trên giao diện máy tính"""
        try:
            self.refresh_products()
            self.refresh_stock()
            self.refresh_alerts()
            self.refresh_report()
            # Làm mới lịch sử nếu giao diện lịch sử đang mở
            if hasattr(self, 'show_purchase_history'):
                try: self.show_purchase_history()
                except: pass
            if hasattr(self, 'show_dispatch_history'):
                try: self.show_dispatch_history()
                except: pass
        except Exception as e:
            print(f"Lỗi refresh_all_data: {e}")

    def update_mobile_server_ui(self):
        """Cập nhật trạng thái giao diện và vẽ lại mã QR kết nối"""
        if not hasattr(self, 'mobile_status_val') or not self.mobile_status_val:
            return
        if not hasattr(self, 'mobile_url_val') or not self.mobile_url_val:
            return
        if not hasattr(self, 'btn_toggle_server') or not self.btn_toggle_server:
            return
        if not hasattr(self, 'qr_help_lbl') or not self.qr_help_lbl:
            return
        if not hasattr(self, 'qr_canvas') or not self.qr_canvas:
            return
            
        ip = get_local_ip()
        port = 5000
        if self.mobile_server:
            port = self.mobile_server.port
            
        url = f"http://{ip}:{port}"
        
        self.mobile_url_val.config(state='normal')
        self.mobile_url_val.delete(0, 'end')
        self.mobile_url_val.insert(0, url)
        self.mobile_url_val.config(state='readonly')
        
        if self.mobile_server and self.mobile_server.is_running:
            import server
            self.mobile_status_val.config(text=f"ĐANG CHẠY | PIN: {server.SERVER_PIN}", bootstyle='success')
            self.btn_toggle_server.config(text="⏹️ Dừng máy chủ di động", bootstyle='danger')
            self.qr_help_lbl.config(text=f"PIN: {server.SERVER_PIN}\nMở Zalo hoặc quét QR để truy cập", bootstyle='success')
            self.draw_qr_code(url)
        else:
            self.mobile_status_val.config(text="ĐÃ DỪNG", bootstyle='danger')
            self.btn_toggle_server.config(text="▶️ Khởi động máy chủ di động", bootstyle='success')
            self.qr_help_lbl.config(text="Vui lòng khởi động máy chủ để hiển thị mã QR", bootstyle='warning')
            self.qr_canvas.delete("all")
            self.qr_canvas.create_text(130, 130, text="MÁY CHỦ\nĐANG DỪNG", fill='gray', font=('Segoe UI', 14, 'bold'), justify='center')

    def draw_qr_code(self, url):
        """Vẽ mã QR lên Canvas"""
        global QR_CODE_AVAILABLE, qrcode
        self.qr_canvas.delete("all")
        
        if not QR_CODE_AVAILABLE:
            try:
                import qrcode
                QR_CODE_AVAILABLE = True
            except ImportError:
                pass
                
        if not QR_CODE_AVAILABLE:
            self.qr_canvas.create_text(130, 100, text="Thiếu thư viện 'qrcode'\nđể hiển thị mã QR", 
                                       fill='red', font=('Segoe UI', 10, 'bold'), justify='center')
                                       
            def auto_install_qr():
                import subprocess, sys
                try:
                    subprocess.run([sys.executable, "-m", "pip", "install", "qrcode"], check=True)
                    global QR_CODE_AVAILABLE
                    QR_CODE_AVAILABLE = True
                    self.toast("Đã cài đặt thành công thư viện qrcode!")
                    self.update_mobile_server_ui()
                except Exception as ex:
                    messagebox.showerror("Lỗi cài đặt", f"Không thể tự động cài đặt: {str(ex)}")
            
            btn_install = tb.Button(self.qr_canvas, text="🔧 Cài đặt qrcode", command=auto_install_qr, bootstyle='warning-outline')
            self.qr_canvas.create_window(130, 160, window=btn_install)
            return

        try:
            qr = qrcode.QRCode(version=1, box_size=1, border=1)
            qr.add_data(url)
            qr.make(fit=True)
            matrix = qr.get_matrix()
            
            num_rows = len(matrix)
            block_size = min(220 // num_rows, 10)
            offset_x = (260 - num_rows * block_size) // 2
            offset_y = (260 - num_rows * block_size) // 2
            
            for r in range(num_rows):
                for c in range(num_rows):
                    if matrix[r][c]:
                        x1 = offset_x + c * block_size
                        y1 = offset_y + r * block_size
                        x2 = x1 + block_size
                        y2 = y1 + block_size
                        self.qr_canvas.create_rectangle(x1, y1, x2, y2, fill="black", outline="black")
        except Exception as e:
            self.qr_canvas.create_text(130, 130, text=f"Lỗi vẽ QR:\n{str(e)}", fill='red', font=('Segoe UI', 10), justify='center')

    def show_purchase_history(self):
        """Hiển thị lịch sử các phiếu nhập kho đã được tạo"""
        try:
            start_date = self.adv_de_from.entry.get().strip()
            end_date = self.adv_de_to.entry.get().strip()
            
            if not start_date or not end_date:
                messagebox.showwarning('Thiếu thông tin', 'Vui lòng chọn đầy đủ ngày bắt đầu và kết thúc')
                return
            
            # Lấy danh sách phiếu nhập kho từ DB
            notes = self.db.get_purchase_notes(start_date, end_date)
            
            # Clear detail tab
            for widget in self.adv_detail_tab.winfo_children():
                widget.destroy()
            
            main_frame = tb.Frame(self.adv_detail_tab)
            main_frame.pack(fill='both', expand=True, padx=10, pady=10)
            
            # Title
            tb.Label(main_frame, text=f"LỊCH SỬ PHIẾU NHẬP KHO ({start_date} -> {end_date})", 
                     font=('Segoe UI', 14, 'bold'), bootstyle='success').pack(pady=(0, 10))
            
            # Bảng danh sách phiếu nhập
            cols = ('id', 'noteNumber', 'supplier', 'createdAt', 'reason', 'item_count', 'note')
            tree = tb.Treeview(main_frame, columns=cols, show='headings', height=12)
            
            for c, w, t, anchor in [
                ('id', 50, 'ID', 'center'),
                ('noteNumber', 150, 'Số phiếu', 'center'),
                ('supplier', 220, 'Nguồn cấp/Nhà CC', 'w'),
                ('createdAt', 150, 'Ngày nhập', 'center'),
                ('reason', 120, 'Lý do nhập', 'center'),
                ('item_count', 90, 'Số mặt hàng', 'center'),
                ('note', 200, 'Ghi chú', 'w')
            ]:
                tree.heading(c, text=t, command=(lambda col=c: self.sort_tree(tree, col)))
                tree.column(c, width=w, anchor=anchor)
                
            tree.tag_configure('odd', background='#f6f8fa')
            tree.pack(fill='both', expand=True, pady=10)
            
            # Load dữ liệu vào tree
            for idx, n in enumerate(notes):
                created_at = n['createdAt']
                tree.insert('', 'end', values=(
                    n['id'],
                    n['noteNumber'],
                    n['supplier'],
                    created_at,
                    n['reason'] or 'Nhập kho',
                    n['item_count'],
                    n['note'] or ''
                ), tags=('odd',) if idx % 2 else ())
            
            # Frame điều khiển bên dưới
            ctrl_btn_frame = tb.Frame(main_frame)
            ctrl_btn_frame.pack(fill='x', pady=5)
            
            def on_reprint():
                sel = tree.selection()
                if not sel:
                    messagebox.showwarning("Chưa chọn dòng", "Vui lòng chọn một phiếu nhập kho trong danh sách!"); return
                val = tree.item(sel[0])['values']
                purchase_id = int(val[0])
                self.reprint_selected_purchase(purchase_id)
                
            def on_delete_purchase():
                sel = tree.selection()
                if not sel:
                    messagebox.showwarning("Chưa chọn dòng", "Vui lòng chọn một phiếu nhập kho trong danh sách để xóa!"); return
                val = tree.item(sel[0])['values']
                purchase_id = int(val[0])
                note_num = val[1]
                
                confirm = messagebox.askyesno(
                    "Xác nhận xóa", 
                    f"Bạn có chắc chắn muốn xóa phiếu nhập số '{note_num}'?\n\n"
                    "Lưu ý: Hành động này sẽ trừ số lượng tồn kho tương ứng của các sản phẩm trong phiếu này và không thể hoàn tác!"
                )
                if not confirm:
                    return
                
                try:
                    # Bắt đầu transaction
                    self.db.conn.execute("BEGIN TRANSACTION")
                    
                    # Kiểm tra xem có stock_movements liên kết hay không (Tránh xóa phiếu cũ)
                    has_ref = self.db.conn.execute(
                        "SELECT COUNT(*) FROM stock_movements WHERE referenceType='PURCHASE' AND referenceId=?",
                        (purchase_id,)
                    ).fetchone()[0]
                    if has_ref == 0:
                        self.db.conn.rollback()
                        messagebox.showerror("Từ chối xóa", "Đây là phiếu nhập cũ (không có liên kết chi tiết). Để đảm bảo an toàn dữ liệu lịch sử, hệ thống từ chối xóa phiếu này.")
                        return
                    
                    # Lỗi 6: Xóa stock_movements theo referenceId (không dùng createdAt nữa)
                    self.db.conn.execute(
                        "DELETE FROM stock_movements WHERE referenceType='PURCHASE' AND referenceId=?",
                        (purchase_id,)
                    )
                    
                    # Xóa phiếu nhập (foreign keys ON DELETE CASCADE sẽ tự động xóa purchase_items)
                    self.db.conn.execute("DELETE FROM purchase_notes WHERE id=?", (purchase_id,))
                    
                    self.db.conn.commit()
                    
                    # Cập nhật lại UI
                    self.toast(f"Đã xóa phiếu nhập {note_num} thành công")
                    self.refresh_products()
                    self.refresh_stock()
                    self.refresh_alerts()
                    self.refresh_report()
                    
                    # Tải lại lịch sử phiếu nhập
                    self.show_purchase_history()
                    
                except Exception as ex:
                    try:
                        self.db.conn.rollback()
                    except:
                        pass
                    messagebox.showerror("Lỗi", f"Không thể xóa phiếu nhập: {str(ex)}")

            tb.Button(ctrl_btn_frame, text="📄 Xem chi tiết & In lại phiếu PDF", bootstyle='info',
                      command=on_reprint).pack(side='left', padx=5)
            
            tb.Button(ctrl_btn_frame, text="🗑️ Xóa phiếu nhập", bootstyle='danger-outline',
                      command=on_delete_purchase).pack(side='left', padx=5)
            
            # Double click để in luôn
            tree.bind("<Double-1>", lambda e: on_reprint())
            
            # Chuyển tab của notebook sang tab Chi tiết
            self.adv_report_nb.select(self.adv_detail_tab)
            self.current_report_type = 'purchase_history'
            
        except Exception as e:
            messagebox.showerror('Lỗi', f"Không thể tải lịch sử phiếu nhập: {str(e)}")

    def reprint_selected_purchase(self, purchase_id):
        """In lại một phiếu nhập kho đã lưu trong cơ sở dữ liệu"""
        try:
            # Lấy thông tin phiếu
            note_rows = self.db.q("SELECT * FROM purchase_notes WHERE id=?", (purchase_id,))
            if not note_rows:
                messagebox.showerror("Lỗi", "Không tìm thấy phiếu nhập kho này"); return
            note_info = note_rows[0]
            
            # Lấy chi tiết hàng hóa
            items = self.db.get_purchase_detail(purchase_id)
            if not items:
                messagebox.showwarning("Trống", "Phiếu nhập kho này không chứa mặt hàng nào!"); return
            
            # Chuyển đổi tên key để tương thích với print_purchase_note
            purchase_items = []
            for it in items:
                purchase_items.append({
                    'productId': it['productId'],
                    'productName': it['productName'],
                    'unitCode': it['unitCode'],
                    'qty': it['qty'],
                    'lotNo': it['lotNo'],
                    'expiryDate': it['expiryDate'],
                    'cost': it['cost']
                })
                
            self.last_purchase_items = purchase_items
            self.last_purchase_info = {
                'id': note_info['id'],
                'noteNumber': note_info['noteNumber'],
                'supplier': note_info['supplier'],
                'reason': note_info['reason'],
                'note': note_info['note'],
                'createdAt': note_info['createdAt']
            }
            
            # Gọi in phiếu nhập kho PDF
            self.print_purchase_note()
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể in lại phiếu nhập: {str(e)}")

    def show_dispatch_history(self):
        """Hiển thị lịch sử các phiếu xuất kho đã được tạo"""
        try:
            start_date = self.adv_de_from.entry.get().strip()
            end_date = self.adv_de_to.entry.get().strip()
            
            if not start_date or not end_date:
                messagebox.showwarning('Thiếu thông tin', 'Vui lòng chọn đầy đủ ngày bắt đầu và kết thúc')
                return
            
            # Lấy danh sách phiếu xuất kho từ DB
            notes = self.db.get_dispatch_notes(start_date, end_date)
            
            # Clear detail tab
            for widget in self.adv_detail_tab.winfo_children():
                widget.destroy()
            
            main_frame = tb.Frame(self.adv_detail_tab)
            main_frame.pack(fill='both', expand=True, padx=10, pady=10)
            
            # Title
            tb.Label(main_frame, text=f"LỊCH SỬ PHIẾU XUẤT KHO ({start_date} -> {end_date})", 
                     font=('Segoe UI', 14, 'bold'), bootstyle='primary').pack(pady=(0, 10))
            
            # Bảng danh sách phiếu xuất
            cols = ('id', 'noteNumber', 'receivingUnit', 'createdAt', 'reason', 'item_count', 'note')
            tree = tb.Treeview(main_frame, columns=cols, show='headings', height=12)
            
            for c, w, t, anchor in [
                ('id', 50, 'ID', 'center'),
                ('noteNumber', 150, 'Số phiếu', 'center'),
                ('receivingUnit', 220, 'Đơn vị nhận', 'w'),
                ('createdAt', 150, 'Ngày xuất', 'center'),
                ('reason', 120, 'Lý do xuất', 'center'),
                ('item_count', 90, 'Số mặt hàng', 'center'),
                ('note', 200, 'Ghi chú', 'w')
            ]:
                tree.heading(c, text=t, command=(lambda col=c: self.sort_tree(tree, col)))
                tree.column(c, width=w, anchor=anchor)
                
            tree.tag_configure('odd', background='#f6f8fa')
            tree.pack(fill='both', expand=True, pady=10)
            
            # Load dữ liệu vào tree
            for idx, n in enumerate(notes):
                created_at = n['createdAt']
                tree.insert('', 'end', values=(
                    n['id'],
                    n['noteNumber'],
                    n['receivingUnit'],
                    created_at,
                    n['reason'] or 'Cấp phát',
                    n['item_count'],
                    n['note'] or ''
                ), tags=('odd',) if idx % 2 else ())
            
            # Frame điều khiển bên dưới
            ctrl_btn_frame = tb.Frame(main_frame)
            ctrl_btn_frame.pack(fill='x', pady=5)
            
            def on_reprint():
                sel = tree.selection()
                if not sel:
                    messagebox.showwarning("Chưa chọn dòng", "Vui lòng chọn một phiếu xuất kho trong danh sách!"); return
                val = tree.item(sel[0])['values']
                dispatch_id = int(val[0])
                self.reprint_selected_dispatch(dispatch_id)
                
            def on_delete_dispatch():
                sel = tree.selection()
                if not sel:
                    messagebox.showwarning("Chưa chọn dòng", "Vui lòng chọn một phiếu xuất kho trong danh sách để xóa!"); return
                val = tree.item(sel[0])['values']
                dispatch_id = int(val[0])
                note_num = val[1]
                
                confirm = messagebox.askyesno(
                    "Xác nhận xóa", 
                    f"Bạn có chắc chắn muốn xóa phiếu xuất số '{note_num}'?\n\n"
                    "Lưu ý: Hành động này sẽ cộng hoàn lại số lượng tồn kho tương ứng của các sản phẩm trong phiếu này và không thể hoàn tác!"
                )
                if not confirm:
                    return
                
                try:
                    # Bắt đầu transaction
                    self.db.conn.execute("BEGIN TRANSACTION")
                    
                    # Kiểm tra xem có stock_movements liên kết hay không (Tránh xóa phiếu cũ)
                    has_ref = self.db.conn.execute(
                        "SELECT COUNT(*) FROM stock_movements WHERE referenceType='DISPATCH' AND referenceId=?",
                        (dispatch_id,)
                    ).fetchone()[0]
                    if has_ref == 0:
                        self.db.conn.rollback()
                        messagebox.showerror("Từ chối xóa", "Đây là phiếu xuất cũ (không có liên kết chi tiết). Để đảm bảo an toàn dữ liệu lịch sử, hệ thống từ chối xóa phiếu này.")
                        return
                    
                    # Lỗi 6: Xóa stock_movements theo referenceId (không dùng createdAt nữa)
                    self.db.conn.execute(
                        "DELETE FROM stock_movements WHERE referenceType='DISPATCH' AND referenceId=?",
                        (dispatch_id,)
                    )
                    
                    # Xóa phiếu xuất (foreign keys ON DELETE CASCADE sẽ tự động xóa dispatch_items)
                    self.db.conn.execute("DELETE FROM dispatch_notes WHERE id=?", (dispatch_id,))
                    
                    self.db.conn.commit()
                    
                    # Cập nhật lại UI
                    self.toast(f"Đã xóa phiếu xuất {note_num} thành công")
                    self.refresh_products()
                    self.refresh_stock()
                    self.refresh_alerts()
                    self.refresh_report()
                    
                    # Tải lại lịch sử phiếu xuất
                    self.show_dispatch_history()
                    
                except Exception as ex:
                    try:
                        self.db.conn.rollback()
                    except:
                        pass
                    messagebox.showerror("Lỗi", f"Không thể xóa phiếu xuất: {str(ex)}")

            tb.Button(ctrl_btn_frame, text="📄 Xem chi tiết & In lại phiếu PDF", bootstyle='info',
                      command=on_reprint).pack(side='left', padx=5)
            
            tb.Button(ctrl_btn_frame, text="🗑️ Xóa phiếu xuất", bootstyle='danger-outline',
                      command=on_delete_dispatch).pack(side='left', padx=5)
            
            # Double click để in luôn
            tree.bind("<Double-1>", lambda e: on_reprint())
            
            # Chuyển tab của notebook sang tab Chi tiết
            self.adv_report_nb.select(self.adv_detail_tab)
            self.current_report_type = 'dispatch_history'
            
        except Exception as e:
            messagebox.showerror('Lỗi', f"Không thể tải lịch sử phiếu xuất: {str(e)}")

    def reprint_selected_dispatch(self, dispatch_id):
        """In lại một phiếu xuất kho đã lưu trong cơ sở dữ liệu"""
        try:
            # Lấy thông tin phiếu
            note_rows = self.db.q("SELECT * FROM dispatch_notes WHERE id=?", (dispatch_id,))
            if not note_rows:
                messagebox.showerror("Lỗi", "Không tìm thấy phiếu xuất kho này"); return
            note_info = note_rows[0]
            
            # Lấy chi tiết hàng hóa
            items = self.db.get_dispatch_detail(dispatch_id)
            if not items:
                messagebox.showwarning("Trống", "Phiếu xuất kho này không chứa mặt hàng nào!"); return
            
            # Chuyển đổi tên key để tương thích với print_dispatch_note
            dispatch_items = []
            for it in items:
                dispatch_items.append({
                    'productId': it['productId'],
                    'productName': it['productName'],
                    'unitCode': it['unitCode'],
                    'qty': it['qty'],
                    'lotNo': it['lotNo'],
                    'expiryDate': it['expiryDate'],
                    'cost': it.get('cost') or 0.0
                })
                
            self.last_dispatch_items = dispatch_items
            self.last_dispatch_info = {
                'id': note_info['id'],
                'noteNumber': note_info['noteNumber'],
                'receivingUnit': note_info['receivingUnit'],
                'reason': note_info['reason'],
                'note': note_info['note'],
                'createdAt': note_info['createdAt']
            }
            
            self.print_dispatch_note()
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể in lại phiếu: {str(e)}")

    def show_revenue_report(self):
        """Hiển thị báo cáo cấp phát"""
        try:
            start_date = self.adv_de_from.entry.get().strip()
            end_date = self.adv_de_to.entry.get().strip()
            
            if not start_date or not end_date:
                messagebox.showwarning('Thiếu thông tin', 'Vui lòng chọn đầy đủ ngày bắt đầu và kết thúc')
                return
            
            # Tạo dialog chọn loại báo cáo
            dialog = tb.Toplevel(self)
            dialog.title("Báo cáo cấp phát")
            dialog.geometry("300x200")
            dialog.transient(self)
            dialog.grab_set()
            
            # Center dialog
            dialog.update_idletasks()
            x = self.winfo_x() + (self.winfo_width() // 2) - (300 // 2)
            y = self.winfo_y() + (self.winfo_height() // 2) - (200 // 2)
            dialog.geometry(f"300x200+{x}+{y}")
            
            main_frame = tb.Frame(dialog)
            main_frame.pack(fill='both', expand=True, padx=20, pady=20)
            
            tb.Label(main_frame, text="Chọn loại báo cáo:", 
                    font=('Segoe UI', 12, 'bold')).pack(pady=(0, 15))
            
            group_var = tk.StringVar(value='day')
            tb.Radiobutton(main_frame, text="Theo ngày", variable=group_var, value='day').pack(anchor='w', pady=5)
            tb.Radiobutton(main_frame, text="Theo tháng", variable=group_var, value='month').pack(anchor='w', pady=5)
            tb.Radiobutton(main_frame, text="Theo năm", variable=group_var, value='year').pack(anchor='w', pady=5)
            
            def generate_report():
                try:
                    group_by = group_var.get()
                    data = self.report_manager.get_revenue_report(start_date, end_date, group_by)
                    self.current_report_type = 'dispatch'
                    self.current_report_group_by = group_by
                    self.display_revenue_report(data, group_by)
                    dialog.destroy()
                except Exception as e:
                    messagebox.showerror('Lỗi', str(e))
            
            btn_frame = tb.Frame(main_frame)
            btn_frame.pack(fill='x', pady=(20, 0))
            
            tb.Button(btn_frame, text="Tạo báo cáo", bootstyle='success',
                      command=generate_report).pack(side='left', padx=(0, 10))
            tb.Button(btn_frame, text="Hủy", bootstyle='secondary',
                      command=dialog.destroy).pack(side='left')
            
        except Exception as e:
            messagebox.showerror('Lỗi', str(e))

    def display_revenue_report(self, data, group_by):
        """Hiển thị báo cáo cấp phát"""
        try:
            # Clear detail tab
            for widget in self.adv_detail_tab.winfo_children():
                widget.destroy()
            
            main_frame = tb.Frame(self.adv_detail_tab)
            main_frame.pack(fill='both', expand=True, padx=10, pady=10)
            
            # Title
            period_text = {'day': 'ngày', 'month': 'tháng', 'year': 'năm'}
            title = f"BÁO CÁO CẤP PHÁT KHO THEO {period_text[group_by].upper()}"
            tb.Label(main_frame, text=title, 
                    font=('Segoe UI', 14, 'bold'), bootstyle='primary').pack(pady=(0, 15))
            
            # Bảng dữ liệu
            cols = ('period', 'orders', 'revenue', 'paid', 'avg_order')
            tree = tb.Treeview(main_frame, columns=cols, show='headings', height=15)
            
            for c, w, t, anchor in [
                ('period', 120, 'Thời gian', 'center'),
                ('orders', 120, 'Số phiếu xuất', 'e'),
                ('revenue', 150, 'Tổng số lượng xuất', 'e'),
                ('paid', 150, 'Số loại sản phẩm', 'e'),
                ('avg_order', 150, 'SL trung bình/phiếu', 'e')
            ]:
                tree.heading(c, text=t)
                tree.column(c, width=w, anchor=anchor)
            
            tree.tag_configure('odd', background='#f6f8fa')
            tree.tag_configure('total', background='#e8f5e9', font=('Segoe UI', 10, 'bold'))
            tree.pack(fill='both', expand=True)
            
            # Load dữ liệu
            total_orders = total_revenue = total_paid = 0
            for idx, row in enumerate(data):
                orders = row.get('total_orders', 0) or 0
                revenue = row.get('total_revenue', 0) or 0
                paid = row.get('total_paid', 0) or 0
                avg_order = row.get('avg_order_value', 0) or 0
                
                total_orders += orders
                total_revenue += revenue
                total_paid += paid
                
                tree.insert('', 'end', values=(
                    row.get('period', ''),
                    f"{orders:,}",
                    f"{revenue:,.0f}",
                    f"{paid:,}",
                    f"{avg_order:,.1f}"
                ), tags=('odd',) if idx % 2 else ())
            
            # Dòng tổng
            if data:
                tree.insert('', 'end', values=(
                    'TỔNG',
                    f"{total_orders:,}",
                    f"{total_revenue:,.0f}",
                    f"{total_paid:,}",
                    f"{total_revenue/total_orders:,.1f}" if total_orders > 0 else "0"
                ), tags=('total',))
            
            # Chuyển sang tab chi tiết
            self.adv_report_nb.select(self.adv_detail_tab)
            
        except Exception as e:
            messagebox.showerror('Lỗi', str(e))

    def show_profit_report(self):
        """Hiển thị thống kê cấp phát theo đơn vị nhận"""
        try:
            start_date = self.adv_de_from.entry.get().strip()
            end_date = self.adv_de_to.entry.get().strip()
            
            if not start_date or not end_date:
                messagebox.showwarning('Thiếu thông tin', 'Vui lòng chọn đầy đủ ngày bắt đầu và kết thúc')
                return
            
            data = self.db.get_dispatch_stats_by_unit(start_date, end_date)
            self.current_report_type = 'receiving_unit'
            self.display_profit_report(data)
            
        except Exception as e:
            messagebox.showerror('Lỗi', str(e))

    def display_profit_report(self, data):
        """Hiển thị thống kê cấp phát theo đơn vị nhận"""
        try:
            # Clear detail tab
            for widget in self.adv_detail_tab.winfo_children():
                widget.destroy()
            
            main_frame = tb.Frame(self.adv_detail_tab)
            main_frame.pack(fill='both', expand=True, padx=10, pady=10)
            
            # Title
            tb.Label(main_frame, text="THỐNG KÊ CẤP PHÁT THEO ĐƠN VỊ NHẬN", 
                    font=('Segoe UI', 14, 'bold'), bootstyle='primary').pack(pady=(0, 5))
            tb.Label(main_frame, text="💡 Nhấp đúp vào một đơn vị để xem chi tiết các mặt hàng đã cấp phát", 
                    font=('Segoe UI', 9, 'italic'), bootstyle='secondary').pack(pady=(0, 10))
            
            # Bảng dữ liệu
            cols = ('unit', 'notes_count', 'total_qty', 'total_value')
            tree = tb.Treeview(main_frame, columns=cols, show='headings', height=15)
            
            for c, w, t, anchor in [
                ('unit', 350, 'Đơn vị nhận', 'w'),
                ('notes_count', 130, 'Số phiếu nhận', 'e'),
                ('total_qty', 150, 'Tổng số lượng nhận', 'e'),
                ('total_value', 180, 'Tổng giá trị nhận (VNĐ)', 'e')
            ]:
                tree.heading(c, text=t)
                tree.column(c, width=w, anchor=anchor)
            
            tree.tag_configure('odd', background='#f6f8fa')
            tree.tag_configure('total', background='#e8f5e9', font=('Segoe UI', 10, 'bold'))
            tree.pack(fill='both', expand=True)
            
            # Load dữ liệu
            total_notes = 0
            total_qty_sum = 0.0
            total_val_sum = 0.0
            for idx, row in enumerate(data):
                unit_name = row.get('receivingUnit', '') or ''
                notes_count = row.get('noteCount', 0) or 0
                total_qty = row.get('totalQty', 0) or 0
                total_value = row.get('totalValue', 0.0) or 0.0
                
                total_notes += notes_count
                total_qty_sum += total_qty
                total_val_sum += total_value
                
                tags = ['odd'] if idx % 2 else []
                
                tree.insert('', 'end', values=(
                    unit_name,
                    f"{notes_count:,}",
                    f"{total_qty:,.0f}",
                    f"{total_value:,.0f}"
                ), tags=tags)
            
            # Dòng tổng
            if data:
                tree.insert('', 'end', values=(
                    'TỔNG CỘNG',
                    f"{total_notes:,}",
                    f"{total_qty_sum:,.0f}",
                    f"{total_val_sum:,.0f}"
                ), tags=('total',))
            
            # Bind Double-Click Event
            tree.bind('<Double-1>', lambda event: self.show_receiving_unit_drilldown(tree))
            
            # Chuyển sang tab chi tiết
            self.adv_report_nb.select(self.adv_detail_tab)
            
        except Exception as e:
            messagebox.showerror('Lỗi', str(e))

    def show_receiving_unit_drilldown(self, tree):
        sel = tree.selection()
        if not sel:
            return
        
        values = tree.item(sel[0], 'values')
        receiving_unit = values[0]
        if receiving_unit == 'TỔNG CỘNG' or not receiving_unit:
            return
            
        start_date = self.adv_de_from.entry.get().strip()
        end_date = self.adv_de_to.entry.get().strip()
        
        try:
            details = self.db.get_dispatch_detail_by_unit(receiving_unit, start_date, end_date)
            if not details:
                messagebox.showinfo("Thông báo", "Không có chi tiết cấp phát cho đơn vị này.")
                return
                
            # Tạo popup window
            pop = tk.Toplevel(self)
            pop.title(f"Chi tiết cấp phát: {receiving_unit}")
            pop.geometry("950x500")
            pop.grab_set()  # Modal window
            
            # Main frame
            main_frm = tb.Frame(pop)
            main_frm.pack(fill='both', expand=True, padx=10, pady=10)
            
            # Header
            tb.Label(main_frm, text=f"CHI TIẾT CẤP PHÁT CHO ĐƠN VỊ", font=('Segoe UI', 14, 'bold'), bootstyle='primary').pack(anchor='w', pady=(0, 5))
            tb.Label(main_frm, text=f"Đơn vị: {receiving_unit} ({start_date} -> {end_date})", font=('Segoe UI', 11, 'bold')).pack(anchor='w', pady=(0, 10))
            
            # Treeview
            cols = ('note_no', 'date', 'reason', 'product', 'qty', 'lot', 'exp', 'fund', 'cost', 'total')
            tree_detail = tb.Treeview(main_frm, columns=cols, show='headings')
            
            for c, w, t, anchor in [
                ('note_no', 100, 'Số phiếu', 'center'),
                ('date', 80, 'Ngày xuất', 'center'),
                ('reason', 90, 'Lý do', 'w'),
                ('product', 180, 'Tên sản phẩm', 'w'),
                ('qty', 50, 'SL', 'e'),
                ('lot', 70, 'Số lô', 'center'),
                ('exp', 80, 'HSD', 'center'),
                ('fund', 90, 'Nguồn kinh phí', 'w'),
                ('cost', 80, 'Đơn giá', 'e'),
                ('total', 90, 'Thành tiền', 'e')
            ]:
                tree_detail.heading(c, text=t)
                tree_detail.column(c, width=w, anchor=anchor)
                
            tree_detail.tag_configure('odd', background='#f6f8fa')
            tree_detail.pack(fill='both', expand=True, pady=(0, 10))
            
            # Insert data
            tot_qty = 0
            tot_val = 0.0
            for idx, r in enumerate(details):
                qty = float(r['qty'])
                cost = float(r['cost'] or 0)
                sub_total = qty * cost
                tot_qty += qty
                tot_val += sub_total
                
                tags = ('odd',) if idx % 2 else ()
                # format date
                created_date = r['createdAt']
                if len(created_date) > 10:
                    created_date = created_date[:10]
                    
                tree_detail.insert('', 'end', values=(
                    r['noteNumber'],
                    created_date,
                    r['reason'] or '',
                    r['productName'],
                    f"{qty:g}",
                    r['lotNo'] or '',
                    r['expiryDate'] or '',
                    r.get('fundSource') or '',
                    f"{cost:,.0f}",
                    f"{sub_total:,.0f}"
                ), tags=tags)
                
            # Footer summary label
            tb.Label(main_frm, text=f"Tổng số lượng: {tot_qty:g}  |  Tổng giá trị: {tot_val:,.0f} VNĐ", font=('Segoe UI', 10, 'bold'), bootstyle='success').pack(anchor='e')
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể lấy chi tiết cấp phát: {str(e)}")
            
            # Chuyển sang tab chi tiết
            self.adv_report_nb.select(self.adv_detail_tab)
            
        except Exception as e:
            messagebox.showerror('Lỗi', str(e))

    def show_top_products_report(self):
        """Hiển thị báo cáo top sản phẩm cấp phát"""
        try:
            start_date = self.adv_de_from.entry.get().strip()
            end_date = self.adv_de_to.entry.get().strip()
            
            if not start_date or not end_date:
                messagebox.showwarning('Thiếu thông tin', 'Vui lòng chọn đầy đủ ngày bắt đầu và kết thúc')
                return
            
            data = self.report_manager.get_top_products(start_date, end_date, 20)
            self.current_report_type = 'top_products'
            self.display_top_products_report(data)
            
        except Exception as e:
            messagebox.showerror('Lỗi', str(e))

    def display_top_products_report(self, data):
        """Hiển thị báo cáo top sản phẩm cấp phát"""
        try:
            # Clear detail tab
            for widget in self.adv_detail_tab.winfo_children():
                widget.destroy()
            
            main_frame = tb.Frame(self.adv_detail_tab)
            main_frame.pack(fill='both', expand=True, padx=10, pady=10)
            
            # Title
            tb.Label(main_frame, text="TOP SẢN PHẨM CẤP PHÁT NHIỀU NHẤT", 
                    font=('Segoe UI', 14, 'bold'), bootstyle='primary').pack(pady=(0, 15))
            
            # Bảng dữ liệu
            cols = ('rank', 'product', 'qty', 'orders')
            tree = tb.Treeview(main_frame, columns=cols, show='headings', height=15)
            
            for c, w, t, anchor in [
                ('rank', 80, 'Hạng', 'center'),
                ('product', 450, 'Sản phẩm', 'w'),
                ('qty', 200, 'Tổng SL cấp', 'e'),
                ('orders', 200, 'Số phiếu xuất', 'e')
            ]:
                tree.heading(c, text=t)
                tree.column(c, width=w, anchor=anchor)
            
            tree.tag_configure('odd', background='#f6f8fa')
            tree.tag_configure('top3', background='#fff3e0')
            tree.pack(fill='both', expand=True)
            
            # Load dữ liệu
            for idx, row in enumerate(data):
                rank = idx + 1
                tags = ['odd'] if idx % 2 else []
                if rank <= 3:
                    tags.append('top3')
                
                product_name = row.get('product_name', '') or ''
                total_qty = row.get('total_qty', 0) or 0
                total_orders = row.get('total_orders', 0) or 0
                
                tree.insert('', 'end', values=(
                    f"#{rank}",
                    product_name,
                    f"{total_qty:,.0f}",
                    f"{total_orders:,}"
                ), tags=tags)
            
            # Chuyển sang tab chi tiết
            self.adv_report_nb.select(self.adv_detail_tab)
            
        except Exception as e:
            messagebox.showerror('Lỗi', str(e))

    def show_revenue_chart(self):
        """Hiển thị biểu đồ cấp phát"""
        if not MATPLOTLIB_AVAILABLE:
            messagebox.showerror('Lỗi', 'Thư viện matplotlib chưa được cài đặt. Vui lòng chạy: pip install matplotlib')
            return
        
        try:
            start_date = self.adv_de_from.entry.get().strip()
            end_date = self.adv_de_to.entry.get().strip()
            
            if not start_date or not end_date:
                messagebox.showwarning('Thiếu thông tin', 'Vui lòng chọn đầy đủ ngày bắt đầu và kết thúc')
                return
            
            # Clear chart tab
            for widget in self.adv_chart_tab.winfo_children():
                widget.destroy()
            
            # Lấy dữ liệu
            data = self.report_manager.get_revenue_report(start_date, end_date, 'day')
            
            if not data:
                messagebox.showinfo('Thông báo', 'Không có dữ liệu để hiển thị biểu đồ')
                return
            
            # Tạo biểu đồ
            fig = Figure(figsize=(12, 6), dpi=100)
            ax = fig.add_subplot(111)
            
            # Chuẩn bị dữ liệu
            dates = []
            revenues = []
            orders = []
            
            for row in data:
                try:
                    period = row.get('period', '')
                    if period:
                        dates.append(dt.datetime.strptime(period, '%Y-%m-%d'))
                        revenues.append(row.get('total_revenue', 0) or 0)
                        orders.append(row.get('total_orders', 0) or 0)
                except ValueError:
                    continue  # Bỏ qua ngày không hợp lệ
            
            if not dates:
                # Không có dữ liệu để vẽ
                ax.text(0.5, 0.5, 'Không có dữ liệu để hiển thị biểu đồ', 
                       ha='center', va='center', transform=ax.transAxes, fontsize=14)
                ax.set_title('Biểu đồ cấp phát theo ngày')
            else:
                # Vẽ biểu đồ cấp phát
                ax.plot(dates, revenues, marker='o', linewidth=2, markersize=6, color='#2e7d32', label='Số lượng cấp')
                ax.set_xlabel('Ngày')
                ax.set_ylabel('Số lượng cấp phát')
                ax.set_title('Biểu đồ cấp phát theo ngày')
                ax.grid(True, alpha=0.3)
                ax.legend()
            
            # Format trục Y
            ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x:,.0f}'))
            
            # Format trục X (chỉ khi có dữ liệu)
            if dates:
                ax.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m'))
                ax.xaxis.set_major_locator(mdates.DayLocator(interval=max(1, len(dates)//10)))
                plt.setp(ax.xaxis.get_majorticklabels(), rotation=45)
            
            # Tạo canvas
            canvas = FigureCanvasTkAgg(fig, self.adv_chart_tab)
            canvas.draw()
            canvas.get_tk_widget().pack(fill='both', expand=True, padx=10, pady=10)
            
            # Chuyển sang tab biểu đồ
            self.adv_report_nb.select(self.adv_chart_tab)
            
        except Exception as e:
            messagebox.showerror('Lỗi', str(e))

    def export_report_csv(self):
        start_s = self.de_from.entry.get().strip() if hasattr(self, 'de_from') else ''
        end_s   = self.de_to.entry.get().strip() if hasattr(self, 'de_to') else ''
        if not start_s or not end_s:
            messagebox.showwarning('Thiếu ngày', 'Chọn đủ Từ ngày và Đến ngày'); return

        fund_source = self.cmb_report_fund.get().strip() if hasattr(self, 'cmb_report_fund') else 'Tất cả'
        rows = self.db.xnt_report(start_s, end_s, fund_source)
        if not rows:
            messagebox.showinfo('Thông báo','Không có dữ liệu'); return

        path = filedialog.asksaveasfilename(
            defaultextension='.csv',
            filetypes=[('CSV','*.csv')],
            initialfile=f'bao_cao_xuat_nhap_ton_{start_s}_to_{end_s}.csv'
        )
        if not path: return

        import csv
        tot_open = tot_in = tot_out = tot_close = 0.0
        with open(path, 'w', newline='', encoding='utf-8') as f:
            w = csv.writer(f)
            w.writerow(['Mã SP', 'Tên sản phẩm', 'Số lô', 'Hạn sử dụng', 'Nguồn kinh phí', 'Tồn đầu', 'Nhập', 'Xuất', 'Tồn cuối'])
            for r in rows:
                w.writerow([
                    r['productId'], 
                    r['productName'], 
                    r.get('lotNo') or '', 
                    r.get('expiryDate') or '', 
                    r.get('fundSource') or '',
                    r['opening'], 
                    r['inbound'], 
                    r['outbound'], 
                    r['closing']
                ])
                tot_open  += float(r['opening']); tot_in += float(r['inbound'])
                tot_out   += float(r['outbound']); tot_close += float(r['closing'])
            w.writerow([])
            w.writerow(['', 'TỔNG CỘNG', '', '', '', round(tot_open,4), round(tot_in,4), round(tot_out,4), round(tot_close,4)])

        self.toast('Đã lưu báo cáo X–N–T')

    def export_report_pdf(self):
        start_s = self.de_from.entry.get().strip() if hasattr(self, 'de_from') else ''
        end_s   = self.de_to.entry.get().strip() if hasattr(self, 'de_to') else ''
        if not start_s or not end_s:
            messagebox.showwarning('Thiếu ngày', 'Chọn đủ Từ ngày và Đến ngày'); return

        fund_source = self.cmb_report_fund.get().strip() if hasattr(self, 'cmb_report_fund') else 'Tất cả'
        rows = self.db.xnt_report(start_s, end_s, fund_source)
        if not rows:
            messagebox.showinfo('Thông báo', 'Không có dữ liệu báo cáo trong khoảng thời gian này'); return

        path = filedialog.asksaveasfilename(
            defaultextension='.pdf',
            filetypes=[('PDF files', '*.pdf'), ('All files', '*.*')],
            initialfile=f'bao_cao_xuat_nhap_ton_{start_s}_to_{end_s}.pdf'
        )
        if not path:
            return

        try:
            import reportlab
        except ImportError:
            response = messagebox.askyesno(
                "Thiếu thư viện", 
                "Hệ thống thiếu thư viện 'reportlab' để xuất PDF.\nBạn có muốn tự động cài đặt không? (Quá trình này mất khoảng vài giây)"
            )
            if response:
                import subprocess
                import sys
                try:
                    self.toast("Đang cài đặt thư viện reportlab, vui lòng đợi...")
                    subprocess.run([sys.executable, "-m", "pip", "install", "reportlab"], check=True)
                    self.toast("Đã cài đặt reportlab thành công!")
                except Exception as ex:
                    messagebox.showerror("Lỗi cài đặt", f"Không thể tự động cài đặt reportlab: {str(ex)}\nHãy chạy lệnh 'pip install reportlab' trong terminal."); return
            else:
                return

        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            
            try:
                pdfmetrics.registerFont(TTFont('TimesNewRoman', "C:\\Windows\\Fonts\\times.ttf"))
                pdfmetrics.registerFont(TTFont('TimesNewRoman-Bold', "C:\\Windows\\Fonts\\timesbd.ttf"))
                pdfmetrics.registerFont(TTFont('TimesNewRoman-Italic', "C:\\Windows\\Fonts\\timesi.ttf"))
                font_normal = 'TimesNewRoman'
                font_bold = 'TimesNewRoman-Bold'
                font_italic = 'TimesNewRoman-Italic'
            except Exception:
                font_normal = 'Helvetica'
                font_bold = 'Helvetica-Bold'
                font_italic = 'Helvetica-Oblique'
                
            # Đặt trang nằm ngang (landscape) để bảng rộng rãi
            doc = SimpleDocTemplate(path, pagesize=landscape(A4), rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
            story = []
            
            styles = getSampleStyleSheet()
            
            style_header_left = ParagraphStyle(
                'HeaderLeft', parent=styles['Normal'], fontName=font_bold, fontSize=10, leading=14, alignment=0
            )
            style_header_right = ParagraphStyle(
                'HeaderRight', parent=styles['Normal'], fontName=font_normal, fontSize=10, leading=14, alignment=2
            )
            style_title = ParagraphStyle(
                'Title', parent=styles['Heading1'], fontName=font_bold, fontSize=16, leading=20, alignment=1, spaceAfter=5
            )
            style_subtitle = ParagraphStyle(
                'Subtitle', parent=styles['Normal'], fontName=font_bold, fontSize=11, leading=14, alignment=1, spaceAfter=15
            )
            style_table_header = ParagraphStyle(
                'TableHeader', parent=styles['Normal'], fontName=font_bold, fontSize=9, leading=11, alignment=1, textColor=colors.black
            )
            style_cell = ParagraphStyle(
                'Cell', parent=styles['Normal'], fontName=font_normal, fontSize=9, leading=11, alignment=0
            )
            style_cell_center = ParagraphStyle(
                'CellCenter', parent=styles['Normal'], fontName=font_normal, fontSize=9, leading=11, alignment=1
            )
            style_cell_right = ParagraphStyle(
                'CellRight', parent=styles['Normal'], fontName=font_normal, fontSize=9, leading=11, alignment=2
            )
            
            # Header
            header_data = [
                [
                    Paragraph("SỞ Y TẾ THÀNH PHỐ CẦN THƠ<br/>TRUNG TÂM KIỂM SOÁT BỆNH TẬT (CDC)", style_header_left),
                    Paragraph("<b>Mẫu số: S12-H</b><br/><i>(Sổ thẻ kho ban hành theo Thông tư số 107/2017/TT-BTC)</i>", style_header_right)
                ]
            ]
            header_table = Table(header_data, colWidths=[400, 362])
            header_table.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
                ('BOTTOMPADDING', (0,0), (-1,-1), 10),
            ]))
            story.append(header_table)
            story.append(Spacer(1, 10))
            
            # Title
            story.append(Paragraph("BÁO CÁO XUẤT - NHẬP - TỒN KHO THUỐC, VACCINE & VẬT TƯ Y TẾ", style_title))
            
            # Định dạng ngày
            try:
                start_dt = dt.datetime.strptime(start_s, '%Y-%m-%d')
                end_dt = dt.datetime.strptime(end_s, '%Y-%m-%d')
                date_range_str = f"Từ ngày {start_dt.strftime('%d/%m/%Y')} đến ngày {end_dt.strftime('%d/%m/%Y')}"
            except Exception:
                date_range_str = f"Từ ngày {start_s} đến ngày {end_s}"
                
            if fund_source and fund_source != 'Tất cả':
                date_range_str += f" — Nguồn: {fund_source}"

            story.append(Paragraph(date_range_str, style_subtitle))
            story.append(Spacer(1, 10))
            
            # Table items
            table_data = [
                [
                    Paragraph("STT", style_table_header),
                    Paragraph("Mã SP", style_table_header),
                    Paragraph("Tên thuốc, vaccine, VTYT", style_table_header),
                    Paragraph("ĐVT", style_table_header),
                    Paragraph("Số lô", style_table_header),
                    Paragraph("Hạn dùng", style_table_header),
                    Paragraph("Nguồn kinh phí", style_table_header),
                    Paragraph("Tồn đầu", style_table_header),
                    Paragraph("Nhập", style_table_header),
                    Paragraph("Xuất", style_table_header),
                    Paragraph("Tồn cuối", style_table_header)
                ]
            ]
            
            tot_open = tot_in = tot_out = tot_close = 0.0
            for idx, r in enumerate(rows, 1):
                o_val = float(r['opening'])
                i_val = float(r['inbound'])
                ou_val = float(r['outbound'])
                c_val = float(r['closing'])
                
                tot_open += o_val
                tot_in += i_val
                tot_out += ou_val
                tot_close += c_val
                
                table_data.append([
                    Paragraph(str(idx), style_cell_center),
                    Paragraph(str(r['productId']), style_cell_center),
                    Paragraph(r['productName'], style_cell),
                    Paragraph(r['unit'] or '-', style_cell_center),
                    Paragraph(r['lotNo'] or '', style_cell_center),
                    Paragraph(r['expiryDate'] or '', style_cell_center),
                    Paragraph(r.get('fundSource') or '', style_cell),
                    Paragraph(f"{o_val:g}", style_cell_right),
                    Paragraph(f"{i_val:g}", style_cell_right),
                    Paragraph(f"{ou_val:g}", style_cell_right),
                    Paragraph(f"{c_val:g}", style_cell_right)
                ])
                
            # Thêm dòng tổng cộng
            table_data.append([
                Paragraph("<b>Tổng cộng</b>", style_cell_center),
                Paragraph("", style_cell_center),
                Paragraph("", style_cell),
                Paragraph("", style_cell_center),
                Paragraph("", style_cell_center),
                Paragraph("", style_cell_center),
                Paragraph("", style_cell),
                Paragraph(f"<b>{tot_open:g}</b>", style_cell_right),
                Paragraph(f"<b>{tot_in:g}</b>", style_cell_right),
                Paragraph(f"<b>{tot_out:g}</b>", style_cell_right),
                Paragraph(f"<b>{tot_close:g}</b>", style_cell_right)
            ])
            
            col_widths = [25, 45, 197, 45, 65, 65, 60, 65, 65, 65, 65]
            items_table = Table(table_data, colWidths=col_widths)
            items_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f2f2f2')),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.black),
                ('SPAN', (0, -1), (6, -1)),
                ('TOPPADDING', (0,0), (-1,-1), 5),
                ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ]))
            story.append(items_table)
            story.append(Spacer(1, 20))
            
            # Ký tên
            date_right_style = ParagraphStyle(
                'DateRight', parent=styles['Normal'], fontName=font_italic, fontSize=11, alignment=2, spaceAfter=10
            )
            sig_title_style = ParagraphStyle(
                'SigTitle', parent=styles['Normal'], fontName=font_bold, fontSize=11, alignment=1
            )
            sig_sub_style = ParagraphStyle(
                'SigSub', parent=styles['Normal'], fontName=font_italic, fontSize=9, alignment=1
            )
            
            now_dt = dt.datetime.now()
            story.append(Paragraph(f"Cần Thơ, ngày {now_dt.strftime('%d')} tháng {now_dt.strftime('%m')} năm {now_dt.strftime('%Y')}", date_right_style))
            
            sig_headers = [
                [
                    Paragraph("<b>Người lập báo cáo</b>", sig_title_style),
                    Paragraph("<b>Thủ kho</b>", sig_title_style),
                    Paragraph("<b>Kế toán trưởng</b>", sig_title_style),
                    Paragraph("<b>Thủ trưởng đơn vị</b>", sig_title_style)
                ],
                [
                    Paragraph("(Ký, họ tên)", sig_sub_style),
                    Paragraph("(Ký, họ tên)", sig_sub_style),
                    Paragraph("(Ký, họ tên)", sig_sub_style),
                    Paragraph("(Ký, đóng dấu)", sig_sub_style)
                ]
            ]
            sig_table = Table(sig_headers, colWidths=[190, 190, 190, 190])
            sig_table.setStyle(TableStyle([
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('BOTTOMPADDING', (0,0), (-1,-1), 2),
            ]))
            story.append(sig_table)
            story.append(Spacer(1, 60))
            
            doc.build(story)
            os.startfile(path)
            self.toast("Đã xuất báo cáo XNT ra PDF và mở file thành công")
            
        except Exception as e:
            messagebox.showerror("Lỗi in PDF", f"Không thể xuất báo cáo PDF: {str(e)}")

    def print_inventory_check_pdf(self):
        end_s = self.de_to.entry.get().strip() if hasattr(self, 'de_to') else ''
        if not end_s:
            messagebox.showwarning('Thiếu ngày', 'Hãy chọn ngày đến (ngày kết thúc kiểm kê) ở ô Đến ngày'); return

        fund_source = self.cmb_report_fund.get().strip() if hasattr(self, 'cmb_report_fund') else 'Tất cả'
        rows = self.db.xnt_report('2000-01-01', end_s, fund_source)
        items = [r for r in rows if float(r['closing']) > 0]
        
        if not items:
            messagebox.showinfo('Thông báo', 'Không có sản phẩm nào có số dư tồn kho tại ngày này để kiểm kê.'); return

        path = filedialog.asksaveasfilename(
            defaultextension='.pdf',
            filetypes=[('PDF files', '*.pdf'), ('All files', '*.*')],
            initialfile=f'bien_ban_kiem_ke_kho_{end_s}.pdf'
        )
        if not path:
            return

        try:
            import reportlab
        except ImportError:
            response = messagebox.askyesno(
                "Thiếu thư viện", 
                "Hệ thống thiếu thư viện 'reportlab' để xuất PDF.\nBạn có muốn tự động cài đặt không?"
            )
            if response:
                import subprocess, sys
                try:
                    subprocess.run([sys.executable, "-m", "pip", "install", "reportlab"], check=True)
                    self.toast("Đã cài đặt reportlab thành công!")
                except Exception as ex:
                    messagebox.showerror("Lỗi cài đặt", f"Không thể cài đặt reportlab: {str(ex)}"); return
            else:
                return

        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            
            try:
                pdfmetrics.registerFont(TTFont('TimesNewRoman', "C:\\Windows\\Fonts\\times.ttf"))
                pdfmetrics.registerFont(TTFont('TimesNewRoman-Bold', "C:\\Windows\\Fonts\\timesbd.ttf"))
                pdfmetrics.registerFont(TTFont('TimesNewRoman-Italic', "C:\\Windows\\Fonts\\timesi.ttf"))
                font_normal = 'TimesNewRoman'
                font_bold = 'TimesNewRoman-Bold'
                font_italic = 'TimesNewRoman-Italic'
            except Exception:
                font_normal = 'Helvetica'
                font_bold = 'Helvetica-Bold'
                font_italic = 'Helvetica-Oblique'
                
            doc = SimpleDocTemplate(path, pagesize=landscape(A4), rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
            story = []
            
            styles = getSampleStyleSheet()
            
            style_header_left = ParagraphStyle(
                'HeaderLeft', parent=styles['Normal'], fontName=font_bold, fontSize=10, leading=14, alignment=0
            )
            style_header_right = ParagraphStyle(
                'HeaderRight', parent=styles['Normal'], fontName=font_normal, fontSize=10, leading=14, alignment=2
            )
            style_title = ParagraphStyle(
                'Title', parent=styles['Heading1'], fontName=font_bold, fontSize=16, leading=20, alignment=1, spaceAfter=5
            )
            style_subtitle = ParagraphStyle(
                'Subtitle', parent=styles['Normal'], fontName=font_bold, fontSize=11, leading=14, alignment=1, spaceAfter=10
            )
            style_text_bold = ParagraphStyle(
                'TextBold', parent=styles['Normal'], fontName=font_bold, fontSize=10, leading=14
            )
            style_text_normal = ParagraphStyle(
                'TextNormal', parent=styles['Normal'], fontName=font_normal, fontSize=10, leading=15
            )
            style_table_header = ParagraphStyle(
                'TableHeader', parent=styles['Normal'], fontName=font_bold, fontSize=9, leading=11, alignment=1
            )
            style_cell = ParagraphStyle(
                'Cell', parent=styles['Normal'], fontName=font_normal, fontSize=9, leading=11, alignment=0
            )
            style_cell_center = ParagraphStyle(
                'CellCenter', parent=styles['Normal'], fontName=font_normal, fontSize=9, leading=11, alignment=1
            )
            style_cell_right = ParagraphStyle(
                'CellRight', parent=styles['Normal'], fontName=font_normal, fontSize=9, leading=11, alignment=2
            )
            
            # Header
            header_data = [
                [
                    Paragraph("SỞ Y TẾ THÀNH PHỐ CẦN THƠ<br/>TRUNG TÂM KIỂM SOÁT BỆNH TẬT (CDC)", style_header_left),
                    Paragraph("<b>Mẫu số: C33-HD</b><br/><i>(Ban hành theo Thông tư số 107/2017/TT-BTC)</i>", style_header_right)
                ]
            ]
            header_table = Table(header_data, colWidths=[400, 362])
            header_table.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
                ('BOTTOMPADDING', (0,0), (-1,-1), 10),
            ]))
            story.append(header_table)
            story.append(Spacer(1, 10))
            
            # Title
            story.append(Paragraph("BIÊN BẢN KIỂM KÊ THUỐC, VACCINE & VẬT TƯ Y TẾ", style_title))
            
            try:
                check_dt = dt.datetime.strptime(end_s, '%Y-%m-%d')
                date_str = f"Thời điểm kiểm kê: 24 giờ 00 phút ngày {check_dt.strftime('%d/%m/%Y')}"
            except Exception:
                date_str = f"Thời điểm kiểm kê: ngày {end_s}"
                
            if fund_source and fund_source != 'Tất cả':
                date_str += f" — Nguồn: {fund_source}"

            story.append(Paragraph(date_str, style_subtitle))
            
            # Ban kiểm kê
            story.append(Paragraph("<b>BAN KIỂM KÊ GỒM:</b>", style_text_bold))
            story.append(Paragraph(
                "1. Ông/Bà: .................... Chức danh: .................... Đại diện Ban Giám đốc (Trưởng ban)<br/>"
                "2. Ông/Bà: .................... Chức danh: .................... Kế toán trưởng (Thành viên)<br/>"
                "3. Ông/Bà: .................... Chức danh: .................... Thủ kho (Thành viên)<br/>"
                "4. Ông/Bà: .................... Chức danh: .................... Trưởng khoa Dược (Thành viên)",
                style_text_normal
            ))
            story.append(Spacer(1, 10))
            
            # Table items
            table_data = [
                [
                    Paragraph("STT", style_table_header),
                    Paragraph("Mã SP", style_table_header),
                    Paragraph("Tên thuốc, vaccine, VTYT", style_table_header),
                    Paragraph("ĐVT", style_table_header),
                    Paragraph("Số lô", style_table_header),
                    Paragraph("Hạn dùng", style_table_header),
                    Paragraph("Nguồn kinh phí", style_table_header),
                    Paragraph("Số lượng<br/>sổ sách", style_table_header),
                    Paragraph("Số lượng<br/>thực tế", style_table_header),
                    Paragraph("Chênh lệch<br/>(Thừa/Thiếu)", style_table_header),
                    Paragraph("Ghi chú", style_table_header)
                ]
            ]
            
            tot_books = 0.0
            for idx, r in enumerate(items, 1):
                c_val = float(r['closing'])
                tot_books += c_val
                
                table_data.append([
                    Paragraph(str(idx), style_cell_center),
                    Paragraph(str(r['productId']), style_cell_center),
                    Paragraph(r['productName'], style_cell),
                    Paragraph(r['unit'] or '-', style_cell_center),
                    Paragraph(r['lotNo'] or '', style_cell_center),
                    Paragraph(r['expiryDate'] or '', style_cell_center),
                    Paragraph(r.get('fundSource') or '', style_cell),
                    Paragraph(f"{c_val:g}", style_cell_right),
                    Paragraph("", style_cell_center),
                    Paragraph("", style_cell_center),
                    Paragraph("", style_cell_center)
                ])
                
            # Dòng tổng cộng
            table_data.append([
                Paragraph("<b>Tổng cộng</b>", style_cell_center),
                Paragraph("", style_cell_center),
                Paragraph("", style_cell),
                Paragraph("", style_cell_center),
                Paragraph("", style_cell_center),
                Paragraph("", style_cell_center),
                Paragraph("", style_cell),
                Paragraph(f"<b>{tot_books:g}</b>", style_cell_right),
                Paragraph("", style_cell_center),
                Paragraph("", style_cell_center),
                Paragraph("", style_cell_center)
            ])
            
            col_widths = [25, 45, 232, 40, 60, 60, 60, 60, 60, 60, 60]
            items_table = Table(table_data, colWidths=col_widths)
            items_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f2f2f2')),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.black),
                ('SPAN', (0, -1), (6, -1)),
                ('TOPPADDING', (0,0), (-1,-1), 5),
                ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ]))
            story.append(items_table)
            story.append(Spacer(1, 20))
            
            # Ký tên
            date_right_style = ParagraphStyle(
                'DateRight', parent=styles['Normal'], fontName=font_italic, fontSize=11, alignment=2, spaceAfter=10
            )
            sig_title_style = ParagraphStyle(
                'SigTitle', parent=styles['Normal'], fontName=font_bold, fontSize=11, alignment=1
            )
            sig_sub_style = ParagraphStyle(
                'SigSub', parent=styles['Normal'], fontName=font_italic, fontSize=9, alignment=1
            )
            
            now_dt = dt.datetime.now()
            story.append(Paragraph(f"Cần Thơ, ngày {now_dt.strftime('%d')} tháng {now_dt.strftime('%m')} năm {now_dt.strftime('%Y')}", date_right_style))
            
            sig_headers = [
                [
                    Paragraph("<b>Người lập biểu</b>", sig_title_style),
                    Paragraph("<b>Thủ kho</b>", sig_title_style),
                    Paragraph("<b>Kế toán trưởng</b>", sig_title_style),
                    Paragraph("<b>Thủ trưởng đơn vị</b>", sig_title_style)
                ],
                [
                    Paragraph("(Ký, họ tên)", sig_sub_style),
                    Paragraph("(Ký, họ tên)", sig_sub_style),
                    Paragraph("(Ký, họ tên)", sig_sub_style),
                    Paragraph("(Ký, đóng dấu)", sig_sub_style)
                ]
            ]
            sig_table = Table(sig_headers, colWidths=[190, 190, 190, 190])
            sig_table.setStyle(TableStyle([
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('BOTTOMPADDING', (0,0), (-1,-1), 2),
            ]))
            story.append(sig_table)
            story.append(Spacer(1, 60))
            
            doc.build(story)
            os.startfile(path)
            self.toast("Đã in biên bản kiểm kê ra PDF thành công")
            
        except Exception as e:
            messagebox.showerror("Lỗi in PDF", f"Không thể xuất biên bản kiểm kê: {str(e)}")

    def export_current_report_excel(self):
        """Xuất báo cáo hiện tại ra Excel"""
        try:
            # Lấy dữ liệu báo cáo hiện tại
            report_data = self.get_current_report_data()
            if not report_data:
                messagebox.showwarning('Cảnh báo', 'Không có dữ liệu báo cáo để xuất')
                return
            
            # Chọn file để lưu
            start_date = self.adv_de_from.entry.get().strip()
            end_date = self.adv_de_to.entry.get().strip()
            filename = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[('Excel files', '*.xlsx'), ('All files', '*.*')],
                initialfile=f'bao_cao_{start_date}_to_{end_date}.xlsx'
            )
            
            if filename:
                # Xuất ra Excel
                self.export_manager.export_to_excel(
                    data=report_data['data'],
                    filename=filename,
                    sheet_name=report_data['title'],
                    headers=report_data.get('headers')
                )
                messagebox.showinfo('Thành công', f'Đã xuất báo cáo ra Excel:\n{filename}')
                
        except Exception as e:
            messagebox.showerror('Lỗi', f'Không thể xuất Excel: {str(e)}')
    
    def export_current_report_pdf(self):
        """Xuất báo cáo hiện tại ra PDF"""
        try:
            # Lấy dữ liệu báo cáo hiện tại
            report_data = self.get_current_report_data()
            if not report_data:
                messagebox.showwarning('Cảnh báo', 'Không có dữ liệu báo cáo để xuất')
                return
            
            # Chọn file để lưu
            start_date = self.adv_de_from.entry.get().strip()
            end_date = self.adv_de_to.entry.get().strip()
            filename = filedialog.asksaveasfilename(
                defaultextension='.pdf',
                filetypes=[('PDF files', '*.pdf'), ('All files', '*.*')],
                initialfile=f'bao_cao_{start_date}_to_{end_date}.pdf'
            )
            
            if filename:
                # Xuất ra PDF
                self.export_manager.export_to_pdf(
                    data=report_data['data'],
                    filename=filename,
                    title=report_data['title'],
                    headers=report_data.get('headers')
                )
                messagebox.showinfo('Thành công', f'Đã xuất báo cáo ra PDF:\n{filename}')
                
        except Exception as e:
            messagebox.showerror('Lỗi', f'Không thể xuất PDF: {str(e)}')
    
    def export_current_report_csv(self):
        """Xuất báo cáo hiện tại ra CSV"""
        try:
            # Lấy dữ liệu báo cáo hiện tại
            report_data = self.get_current_report_data()
            if not report_data:
                messagebox.showwarning('Cảnh báo', 'Không có dữ liệu báo cáo để xuất')
                return
            
            # Chọn file để lưu
            start_date = self.adv_de_from.entry.get().strip()
            end_date = self.adv_de_to.entry.get().strip()
            filename = filedialog.asksaveasfilename(
                defaultextension='.csv',
                filetypes=[('CSV files', '*.csv'), ('All files', '*.*')],
                initialfile=f'bao_cao_{start_date}_to_{end_date}.csv'
            )
            
            if filename:
                # Xuất ra CSV
                import csv
                with open(filename, 'w', newline='', encoding='utf-8') as f:
                    if report_data.get('headers'):
                        writer = csv.writer(f)
                        writer.writerow(report_data['headers'])
                        
                        for row in report_data['data']:
                            if isinstance(row, dict):
                                writer.writerow([row.get(h, '') for h in report_data['headers']])
                            else:
                                writer.writerow(row)
                    else:
                        # Sử dụng pandas nếu có
                        if PANDAS_AVAILABLE:
                            df = pd.DataFrame(report_data['data'])
                            df.to_csv(filename, index=False, encoding='utf-8')
                        else:
                            # Fallback manual CSV
                            writer = csv.writer(f)
                            for row in report_data['data']:
                                if isinstance(row, dict):
                                    writer.writerow(list(row.values()))
                                else:
                                    writer.writerow(row)
                
                messagebox.showinfo('Thành công', f'Đã xuất báo cáo ra CSV:\n{filename}')
                
        except Exception as e:
            messagebox.showerror('Lỗi', f'Không thể xuất CSV: {str(e)}')
    
    def get_current_report_data(self):
        """Lấy dữ liệu báo cáo hiện tại đang hiển thị"""
        try:
            # Lấy thông tin từ tab hiện tại
            current_tab = self.adv_report_nb.select()
            tab_text = self.adv_report_nb.tab(current_tab, 'text')
            
            start_date = self.adv_de_from.entry.get().strip()
            end_date = self.adv_de_to.entry.get().strip()
            
            if not start_date or not end_date:
                return None
            
            # Xác định loại báo cáo dựa trên tab hiện tại
            if 'Tóm tắt' in tab_text:
                # Báo cáo tóm tắt
                summary_data = self.report_manager.get_daily_sales_summary(start_date, end_date)
                summary = summary_data['summary']
                data_list = [
                    ['Tổng số phiếu xuất', summary.get('total_orders', 0)],
                    ['Tổng số lượng xuất', summary.get('total_revenue', 0)],
                    ['Số lượng TB/phiếu', round(summary.get('avg_order_value', 0), 1)],
                    ['Phiếu xuất lớn nhất', summary.get('max_order', 0)]
                ]
                return {
                    'title': f'Báo cáo tóm tắt từ {start_date} đến {end_date}',
                    'data': data_list,
                    'headers': ['Chỉ số', 'Giá trị']
                }
            elif 'Chi tiết' in tab_text:
                if self.current_report_type == 'dispatch':
                    data = self.report_manager.get_revenue_report(start_date, end_date, self.current_report_group_by)
                    data_list = []
                    for row in data:
                        data_list.append([
                            row.get('period', ''),
                            row.get('total_orders', 0),
                            row.get('total_revenue', 0),
                            row.get('total_paid', 0),
                            round(row.get('avg_order_value', 0), 1)
                        ])
                    return {
                        'title': f'Báo cáo cấp phát chi tiết từ {start_date} đến {end_date}',
                        'data': data_list,
                        'headers': ['Thời gian', 'Số phiếu xuất', 'Tổng số lượng xuất', 'Số loại sản phẩm', 'SL trung bình/phiếu']
                    }
                elif self.current_report_type == 'receiving_unit':
                    data = self.db.get_dispatch_stats_by_unit(start_date, end_date)
                    data_list = []
                    for row in data:
                        data_list.append([
                            row.get('receivingUnit', ''),
                            row.get('noteCount', 0),
                            row.get('totalQty', 0),
                            row.get('totalValue', 0.0)
                        ])
                    return {
                        'title': f'Thống kê cấp phát theo đơn vị nhận từ {start_date} đến {end_date}',
                        'data': data_list,
                        'headers': ['Đơn vị nhận', 'Số phiếu xuất', 'Tổng số lượng nhận', 'Tổng giá trị nhận (VNĐ)']
                    }
                elif self.current_report_type == 'top_products':
                    data = self.report_manager.get_top_products(start_date, end_date, 20)
                    data_list = []
                    for idx, row in enumerate(data):
                        data_list.append([
                            f"#{idx+1}",
                            row.get('product_name', ''),
                            row.get('total_qty', 0),
                            row.get('total_orders', 0)
                        ])
                    return {
                        'title': f'Top sản phẩm cấp phát nhiều nhất từ {start_date} đến {end_date}',
                        'data': data_list,
                        'headers': ['Hạng', 'Sản phẩm', 'Tổng SL cấp', 'Số phiếu xuất']
                    }
            else:
                # Mặc định là báo cáo cấp phát
                data = self.report_manager.get_revenue_report(start_date, end_date, 'day')
                data_list = []
                for row in data:
                    data_list.append([
                        row.get('period', ''),
                        row.get('total_orders', 0),
                        row.get('total_revenue', 0),
                        row.get('total_paid', 0),
                        round(row.get('avg_order_value', 0), 1)
                    ])
                return {
                    'title': f'Báo cáo cấp phát từ {start_date} đến {end_date}',
                    'data': data_list,
                    'headers': ['Thời gian', 'Số phiếu xuất', 'Tổng số lượng xuất', 'Số loại sản phẩm', 'SL trung bình/phiếu']
                }
                
        except Exception as e:
            print(f"Lỗi lấy dữ liệu báo cáo: {e}")
            return None

    # helpers chung
    def sort_tree(self, tree: tb.Treeview, col: str):
        data = [(tree.set(k, col), k) for k in tree.get_children('')]
        def to_num(x):
            try:
                if isinstance(x, str): x = x.replace(',', '')
                return float(x)
            except:
                return x.lower() if isinstance(x, str) else x
        data.sort(key=lambda t: to_num(t[0]))
        for i, (_, k) in enumerate(data): tree.move(k, '', i)

    def refresh_products(self):
        self._products = self.db.q('SELECT id, name FROM products ORDER BY name')
        opts = [f"{p['id']} — {p['name']}" for p in self._products]
        self.cmb_prod['values'] = opts; self.cmb_prod_pos['values'] = opts
        if opts:
            self.cmb_prod.current(0); self.cmb_prod_pos.current(0)
            self.update_purchase_unit_and_price()
            self.update_dispatch_unit_label()

    def _fill_tree(self, tree: tb.Treeview, rows):
        for i in tree.get_children(): tree.delete(i)
        keymap = {'product':'productId','productName':'productName','batch':'batchId','lot':'lotNo','exp':'expiryDate',
                  'qty':'qtyBase','cost':'costBase','value':'valueBase'}
        for idx, r in enumerate(rows):
            tree.insert('', 'end', values=[r.get(keymap.get(c, c), '') for c in tree['columns']],
                        tags=('odd',) if idx%2 else ())

    def refresh_stock(self):
        rows = self.db.stock_view()
        if hasattr(self, 'tree_stock') and self.tree_stock:
            self._fill_tree(self.tree_stock, rows)
        if hasattr(self, 'tree_stock2') and self.tree_stock2:
            self._fill_tree(self.tree_stock2, rows)

    def refresh_alerts(self):
        try: days = int(self.ent_warn_days.get())
        except: days = 180
        self._fill_tree(self.tree_alerts, self.db.expiring_view(days))

    def refresh_report(self):
        start_s = self.de_from.entry.get().strip() if hasattr(self, 'de_from') else ''
        end_s   = self.de_to.entry.get().strip() if hasattr(self, 'de_to') else ''

        # Clear
        for i in self.tree_report.get_children():
            self.tree_report.delete(i)

        if not start_s or not end_s:
            messagebox.showwarning('Thiếu ngày', 'Chọn đủ Từ ngày và Đến ngày'); return

        fund_source = self.cmb_report_fund.get().strip() if hasattr(self, 'cmb_report_fund') else 'Tất cả'
        rows = self.db.xnt_report(start_s, end_s, fund_source)

        tot_open = tot_in = tot_out = tot_close = 0.0
        for idx, r in enumerate(rows):
            tag = 'odd' if idx % 2 else ''
            tot_open  += float(r['opening'])
            tot_in    += float(r['inbound'])
            tot_out   += float(r['outbound'])
            tot_close += float(r['closing'])
            self.tree_report.insert(
                '',
                'end',
                values=(
                    r['productId'], 
                    r['productName'],
                    r['lotNo'] or '',
                    r['expiryDate'] or '',
                    r.get('fundSource', ''),
                    f"{r['opening']:g}", 
                    f"{r['inbound']:g}", 
                    f"{r['outbound']:g}", 
                    f"{r['closing']:g}"
                ),
                tags=(tag,)
            )

        # Dòng tổng
        if rows:
            self.tree_report.insert(
                '', 'end',
                values=('', 'TỔNG CỘNG', '', '', '', f"{tot_open:g}", f"{tot_in:g}", f"{tot_out:g}", f"{tot_close:g}"),
                tags=('total',)
            )

    def refresh_report_funds_combo(self):
        try:
            funds = self.db.get_fund_sources()
            vals = ['Tất cả']
            for f in funds:
                if f and f not in vals:
                    vals.append(f)
            # Add defaults to list if they aren't in DB yet to make it easier for user to select
            defaults = ['TCMR (Tiêm chủng mở rộng)', 'Ngân sách địa phương', 'Dự án viện trợ', 'Mua sắm đấu thầu', 'Nguồn khác']
            for d in defaults:
                if d not in vals:
                    vals.append(d)
            if hasattr(self, 'cmb_report_fund'):
                current_val = self.cmb_report_fund.get()
                self.cmb_report_fund['values'] = vals
                if current_val in vals:
                    self.cmb_report_fund.set(current_val)
                else:
                    self.cmb_report_fund.set('Tất cả')

            if hasattr(self, 'cmb_item_fund'):
                current_item_val = self.cmb_item_fund.get()
                item_vals = [v for v in vals if v != 'Tất cả']
                self.cmb_item_fund['values'] = item_vals
                if current_item_val:
                    self.cmb_item_fund.set(current_item_val)
                else:
                    self.cmb_item_fund.set('TCMR (Tiêm chủng mở rộng)')
        except Exception as e:
            print(f"Lỗi tải danh sách nguồn kinh phí: {e}")


    def on_ready(self):
        if hasattr(self, 'refresh_report_funds_combo'):
            self.refresh_report_funds_combo()
        self.refresh_products(); self.refresh_stock(); self.refresh_alerts(); self.refresh_report()
        self.refresh_backup_list()
        # Load advanced reports summary
        if hasattr(self, 'load_advanced_summary'):
            self.load_advanced_summary()
        # Update catalog info
        if hasattr(self, 'update_catalog_info'):
            self.update_catalog_info()
        # Tự động load thuoc.csv nếu có
        self.auto_load_medicine_catalog()
        # Cập nhật status database
        self.update_db_status()

    def auto_load_medicine_catalog(self):
        """Tự động load file thuoc.csv khi khởi động"""
        try:
            # Tìm file thuoc.csv trong thư mục hiện tại
            csv_path = os.path.join(os.getcwd(), 'thuoc.csv')
            
            if os.path.exists(csv_path):
                self.medicine_catalog.load_catalog_from_excel(csv_path)
                self.update_catalog_info()
                print(f"Đã tự động load danh mục thuốc: {os.path.basename(csv_path)}")
            else:
                print("Không tìm thấy file thuoc.csv để tự động load")
                
        except Exception as e:
            print(f"Lỗi khi tự động load danh mục thuốc: {e}")

    def export_import_template(self):
        """Xuất file mẫu Excel để nhập dữ liệu hàng loạt"""
        global pd, PANDAS_AVAILABLE
        if not PANDAS_AVAILABLE:
            response = messagebox.askyesno(
                "Thiếu thư viện", 
                "Hệ thống thiếu thư viện 'pandas' và 'openpyxl' để xử lý Excel.\nBạn có muốn tự động cài đặt không? (Quá trình này mất khoảng vài giây)"
            )
            if response:
                import subprocess, sys
                try:
                    self.toast("Đang cài đặt thư viện pandas và openpyxl...")
                    subprocess.run([sys.executable, "-m", "pip", "install", "pandas", "openpyxl"], check=True)
                    import pandas as pd
                    PANDAS_AVAILABLE = True
                    self.toast("Đã cài đặt thành công!")
                except Exception as ex:
                    messagebox.showerror("Lỗi cài đặt", f"Không thể tự động cài đặt: {str(ex)}\nHãy chạy lệnh 'pip install pandas openpyxl' trong terminal."); return
            else:
                return
        
        try:
            path = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[('Excel files', '*.xlsx'), ('CSV files', '*.csv'), ('All files', '*.*')],
                initialfile='mau_nhap_lieu_hang_loat.xlsx',
                title='Chọn nơi lưu file mẫu'
            )
            if not path:
                return

            headers = [
                'Tên sản phẩm',
                'Đơn vị cơ sở',
                'Mã vạch (Barcode)',
                'Loại sản phẩm (thuoc/vaccine/vtyt/khac)',
                'Số đăng ký',
                'Đơn vị quy đổi 1',
                'Tỷ lệ quy đổi 1',
                'Giá bán đơn vị quy đổi 1',
                'Đơn vị quy đổi 2',
                'Tỷ lệ quy đổi 2',
                'Giá bán đơn vị quy đổi 2',
                'Đơn vị quy đổi 3',
                'Tỷ lệ quy đổi 3',
                'Giá bán đơn vị quy đổi 3',
                'Số lô',
                'Hạn sử dụng (YYYY-MM-DD)',
                'Số lượng tồn (Đơn vị cơ sở)',
                'Giá nhập (Đơn vị cơ sở)'
            ]

            sample_data = [
                [
                    'Paracetamol 500mg',
                    'vien',
                    '8931234567890',
                    'thuoc',
                    'VD-12345-20',
                    'vi',
                    10,
                    15000,
                    'hop',
                    100,
                    140000,
                    '',
                    '',
                    '',
                    'LOT123',
                    '2027-12-31',
                    500,
                    1200
                ],
                [
                    'Vaccine Quinvaxem',
                    'lo',
                    '8930987654321',
                    'vaccine',
                    'QLSP-987-19',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    'B2209',
                    '2026-09-30',
                    50,
                    150000
                ]
            ]

            if path.lower().endswith('.csv'):
                df = pd.DataFrame(sample_data, columns=headers)
                df.to_csv(path, index=False, encoding='utf-8-sig')
            else:
                df = pd.DataFrame(sample_data, columns=headers)
                with pd.ExcelWriter(path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Sheet1', index=False)
                    worksheet = writer.sheets['Sheet1']
                    
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 3, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
                    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
                    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
                    thin_border = Border(
                        left=Side(style='thin', color='DDDDDD'),
                        right=Side(style='thin', color='DDDDDD'),
                        top=Side(style='thin', color='DDDDDD'),
                        bottom=Side(style='thin', color='DDDDDD')
                    )

                    worksheet.row_dimensions[1].height = 28
                    
                    for cell in worksheet[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_align
                        cell.border = thin_border
                    
                    data_font = Font(name='Segoe UI', size=10)
                    for row in range(2, worksheet.max_row + 1):
                        worksheet.row_dimensions[row].height = 20
                        for col in range(1, worksheet.max_column + 1):
                            cell = worksheet.cell(row=row, column=col)
                            cell.font = data_font
                            cell.border = thin_border
                            if col in [7, 8, 10, 11, 13, 14, 17, 18]:
                                cell.alignment = Alignment(horizontal='right')
                            elif col in [2, 4, 15, 16]:
                                cell.alignment = Alignment(horizontal='center')
                            else:
                                cell.alignment = Alignment(horizontal='left')

            self.toast('Đã tải file Excel mẫu thành công')
            
        except Exception as e:
            messagebox.showerror('Lỗi', f'Không thể xuất file mẫu: {str(e)}')

    def bulk_import_from_excel(self):
        """Nhập sản phẩm và tồn kho hàng loạt từ file Excel/CSV"""
        global pd, PANDAS_AVAILABLE
        if not PANDAS_AVAILABLE:
            response = messagebox.askyesno(
                "Thiếu thư viện", 
                "Hệ thống thiếu thư viện 'pandas' và 'openpyxl' để xử lý Excel.\nBạn có muốn tự động cài đặt không? (Quá trình này mất khoảng vài giây)"
            )
            if response:
                import subprocess, sys
                try:
                    self.toast("Đang cài đặt thư viện pandas và openpyxl...")
                    subprocess.run([sys.executable, "-m", "pip", "install", "pandas", "openpyxl"], check=True)
                    import pandas as pd
                    PANDAS_AVAILABLE = True
                    self.toast("Đã cài đặt thành công!")
                except Exception as ex:
                    messagebox.showerror("Lỗi cài đặt", f"Không thể tự động cài đặt: {str(ex)}\nHãy chạy lệnh 'pip install pandas openpyxl' trong terminal."); return
            else:
                return
        
        try:
            path = filedialog.askopenfilename(
                title="Chọn file Excel hoặc CSV để nhập hàng loạt",
                filetypes=[
                    ('Excel/CSV files', '*.xlsx;*.xls;*.csv'),
                    ('Excel files', '*.xlsx;*.xls'),
                    ('CSV files', '*.csv'),
                    ('All files', '*.*')
                ]
            )
            if not path:
                return
            
            if path.lower().endswith('.csv'):
                df = pd.read_csv(path, encoding='utf-8')
            else:
                df = pd.read_excel(path)
            
            df.columns = df.columns.str.strip()
            
            required_cols = ['Tên sản phẩm', 'Đơn vị cơ sở']
            missing = [c for c in required_cols if c not in df.columns]
            if missing:
                messagebox.showerror('Lỗi định dạng', f'File thiếu các cột bắt buộc: {", ".join(missing)}')
                return
            
            if df.empty:
                messagebox.showwarning('Cảnh báo', 'File Excel/CSV không có dữ liệu')
                return
            
            total_rows = len(df)
            imported_products = 0
            updated_products = 0
            imported_units = 0
            imported_stock = 0
            errors = []
            
            def parse_import_date(val):
                if pd.isna(val) or val is None:
                    return None
                if isinstance(val, (dt.datetime, dt.date)):
                    return val.strftime('%Y-%m-%d')
                val_str = str(val).strip()
                if not val_str or val_str.lower() in ('nan', 'none', 'null', ''):
                    return None
                for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d', '%d-%m-%Y'):
                    try:
                        return dt.datetime.strptime(val_str, fmt).strftime('%Y-%m-%d')
                    except ValueError:
                        continue
                try:
                    val_float = float(val_str)
                    d = dt.datetime(1899, 12, 30) + dt.timedelta(days=int(val_float))
                    return d.strftime('%Y-%m-%d')
                except:
                    pass
                return val_str

            self.db.conn.execute("BEGIN TRANSACTION")
            
            items_to_record = []
            
            for idx, row in df.iterrows():
                row_num = idx + 2
                
                name = str(row.get('Tên sản phẩm', '')).strip()
                if not name or name.lower() == 'nan':
                    errors.append(f"Dòng {row_num}: Tên sản phẩm trống")
                    continue
                
                default_unit = str(row.get('Đơn vị cơ sở', '')).strip()
                if not default_unit or default_unit.lower() == 'nan':
                    errors.append(f"Dòng {row_num}: Đơn vị cơ sở trống")
                    continue
                
                barcode = str(row.get('Mã vạch (Barcode)', '')).strip()
                if not barcode or barcode.lower() in ('nan', 'none', ''):
                    barcode = None
                
                product_type = str(row.get('Loại sản phẩm (thuoc/vaccine/vtyt/khac)', '')).strip().lower()
                if not product_type or product_type not in ('thuoc', 'vaccine', 'vtyt', 'khac'):
                    product_type = 'thuoc'
                
                reg_num = str(row.get('Số đăng ký', '')).strip()
                if not reg_num or reg_num.lower() in ('nan', 'none', ''):
                    reg_num = None
                
                prod_row = self.db.q("SELECT id, defaultUnit FROM products WHERE name = ?", (name,))
                is_new = False
                if prod_row:
                    product_id_db = prod_row[0]['id']
                    self.db.conn.execute(
                        "UPDATE products SET barcode=COALESCE(?, barcode), productType=?, registrationNumber=COALESCE(?, registrationNumber) WHERE id=?",
                        (barcode, product_type, reg_num, product_id_db)
                    )
                    updated_products += 1
                else:
                    if barcode:
                        prod_barcode = self.db.q("SELECT id, defaultUnit FROM products WHERE barcode = ?", (barcode,))
                        if prod_barcode:
                            product_id_db = prod_barcode[0]['id']
                            self.db.conn.execute(
                                "UPDATE products SET name=?, productType=?, registrationNumber=COALESCE(?, registrationNumber) WHERE id=?",
                                (name, product_type, reg_num, product_id_db)
                            )
                            updated_products += 1
                        else:
                            is_new = True
                    else:
                        is_new = True
                
                if is_new:
                    cur = self.db.conn.execute(
                        "INSERT INTO products (name, defaultUnit, barcode, productType, registrationNumber) VALUES (?, ?, ?, ?, ?)",
                        (name, default_unit, barcode, product_type, reg_num)
                    )
                    product_id_db = cur.lastrowid
                    imported_products += 1
                
                self.db.conn.execute(
                    "INSERT OR IGNORE INTO product_units(productId, unitCode, toBaseQty, price) VALUES(?,?,1,0)", 
                    (product_id_db, default_unit)
                )
                
                for i in range(1, 4):
                    unit_name = str(row.get(f'Đơn vị quy đổi {i}', '')).strip()
                    if not unit_name or unit_name.lower() in ('nan', 'none', ''):
                        continue
                    
                    try:
                        ratio_val = row.get(f'Tỷ lệ quy đổi {i}')
                        if pd.isna(ratio_val):
                            continue
                        ratio = float(ratio_val)
                        if ratio <= 0:
                            errors.append(f"Dòng {row_num}: Tỷ lệ quy đổi {i} của '{name}' phải > 0 (bỏ qua đơn vị này)")
                            continue
                    except ValueError:
                        errors.append(f"Dòng {row_num}: Tỷ lệ quy đổi {i} của '{name}' không phải là số (bỏ qua đơn vị này)")
                        continue
                    
                    try:
                        price_val = row.get(f'Giá bán đơn vị quy đổi {i}')
                        price = float(price_val) if not pd.isna(price_val) else 0.0
                        if price < 0:
                            price = 0.0
                    except ValueError:
                        price = 0.0
                    
                    self.db.conn.execute(
                        "INSERT OR REPLACE INTO product_units (productId, unitCode, toBaseQty, price) VALUES (?, ?, ?, ?)",
                        (product_id_db, unit_name, ratio, price)
                    )
                    imported_units += 1
                
                lot_no = str(row.get('Số lô', '')).strip()
                if lot_no and lot_no.lower() not in ('nan', 'none', ''):
                    expiry_val = row.get('Hạn sử dụng (YYYY-MM-DD)')
                    expiry_date = parse_import_date(expiry_val)
                    
                    if not expiry_date:
                        errors.append(f"Dòng {row_num}: Số lô '{lot_no}' cho '{name}' thiếu hoặc sai hạn sử dụng (Bỏ qua nhập lô)")
                        continue
                    
                    try:
                        dt.datetime.strptime(expiry_date, '%Y-%m-%d')
                    except ValueError:
                        errors.append(f"Dòng {row_num}: Hạn sử dụng '{expiry_date}' của lô '{lot_no}' không đúng định dạng YYYY-MM-DD (Bỏ qua nhập lô)")
                        continue
                    
                    try:
                        qty_val = row.get('Số lượng tồn (Đơn vị cơ sở)')
                        if pd.isna(qty_val):
                            errors.append(f"Dòng {row_num}: Thiếu số lượng tồn cho lô '{lot_no}' của '{name}' (Bỏ qua nhập lô)")
                            continue
                        qty = float(qty_val)
                        if qty <= 0:
                            errors.append(f"Dòng {row_num}: Số lượng tồn {qty} cho lô '{lot_no}' của '{name}' phải > 0 (Bỏ qua nhập lô)")
                            continue
                    except ValueError:
                        errors.append(f"Dòng {row_num}: Số lượng tồn cho lô '{lot_no}' của '{name}' không hợp lệ (Bỏ qua nhập lô)")
                        continue
                    
                    try:
                        cost_val = row.get('Giá nhập (Đơn vị cơ sở)')
                        cost = float(cost_val) if not pd.isna(cost_val) else 0.0
                        if cost < 0:
                            cost = 0.0
                    except ValueError:
                        cost = 0.0
                        
                    fund_source = str(row.get('Nguồn kinh phí', '')).strip()
                    if fund_source.lower() in ('nan', 'none', ''):
                        fund_source = ''
                    
                    # Gọi validate_batch để kiểm tra ngay tại đây mà không chèn bản ghi dở dang vào DB
                    try:
                        self.db.validate_batch(product_id_db, lot_no, expiry_date)
                        items_to_record.append({
                            'productId': product_id_db,
                            'lotNo': lot_no,
                            'expiryDate': expiry_date,
                            'unitCode': default_unit,
                            'qty': qty,
                            'cost': cost,
                            'fundSource': fund_source
                        })
                        imported_stock += 1
                    except Exception as ex_batch:
                        errors.append(f"Dòng {row_num}: Lỗi lô hàng: {str(ex_batch)}")
                        continue
            
            # Commit phần cập nhật sản phẩm & đơn vị quy đổi
            self.db.conn.commit()
            
            # Nhập kho tồn ban đầu sử dụng hàm chuẩn record_purchase
            note_number = ""
            if items_to_record:
                _, note_number, _ = self.db.record_purchase(
                    items=items_to_record,
                    supplier="Nhập kho ban đầu",
                    reason="Nhập kho ban đầu",
                    note="Nhập hàng loạt từ Excel"
                )
            
            self.refresh_products()
            self.refresh_stock()
            self.refresh_alerts()
            self.refresh_report()
            
            success_msg = f"Đã nhập dữ liệu thành công:\n"
            success_msg += f"- Thêm mới {imported_products} sản phẩm\n"
            success_msg += f"- Cập nhật {updated_products} sản phẩm\n"
            if imported_units > 0:
                success_msg += f"- Thêm {imported_units} đơn vị quy đổi\n"
            if imported_stock > 0:
                success_msg += f"- Nhập {imported_stock} lô tồn kho ban đầu (Số phiếu: {note_number})\n"
            
            if errors:
                self.show_import_log_dialog(success_msg, errors)
            else:
                messagebox.showinfo('Thành công', success_msg)
                self.toast('Đã nhập dữ liệu hàng loạt thành công')
                
        except Exception as e:
            try:
                self.db.conn.rollback()
            except:
                pass
            messagebox.showerror('Lỗi', f'Lỗi trong quá trình nhập dữ liệu: {str(e)}')

    def show_import_log_dialog(self, summary, errors):
        """Hiển thị thông báo kết quả nhập và danh sách lỗi/cảnh báo"""
        dialog = tb.Toplevel(self)
        dialog.title("Kết quả nhập hàng loạt")
        dialog.geometry("650x500")
        dialog.transient(self)
        dialog.grab_set()
        
        dialog.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() // 2) - (650 // 2)
        y = self.winfo_y() + (self.winfo_height() // 2) - (500 // 2)
        dialog.geometry(f"650x500+{x}+{y}")
        
        main_frame = tb.Frame(dialog)
        main_frame.pack(fill='both', expand=True, padx=15, pady=15)
        
        summary_title = tb.Label(main_frame, text="TÓM TẮT KẾT QUẢ", font=('Segoe UI', 11, 'bold'), bootstyle='success')
        summary_title.pack(anchor='w', pady=(0, 5))
        
        summary_box = tb.Label(main_frame, text=summary, font=('Segoe UI', 9), justify='left')
        summary_box.pack(anchor='w', pady=(0, 15))
        
        error_title = tb.Label(main_frame, text=f"DANH SÁCH CHI TIẾT BỎ QUA/CẢNH BÁO ({len(errors)} dòng bị lỗi)", 
                               font=('Segoe UI', 11, 'bold'), bootstyle='danger')
        error_title.pack(anchor='w', pady=(0, 5))
        
        txt_frame = tb.Frame(main_frame)
        txt_frame.pack(fill='both', expand=True)
        
        scrollbar = tb.Scrollbar(txt_frame)
        scrollbar.pack(side='right', fill='y')
        
        text_area = tb.Text(txt_frame, yscrollcommand=scrollbar.set, font=('Consolas', 9), wrap='word')
        text_area.pack(fill='both', expand=True, side='left')
        scrollbar.config(command=text_area.yview)
        
        for err in errors:
            text_area.insert('end', f"• {err}\n")
        
        text_area.config(state='disabled')
        
        tb.Button(main_frame, text="Đóng", bootstyle='secondary', command=dialog.destroy).pack(pady=(15, 0), side='right')

    def build_temp_log_tab(self):
        """Tạo giao diện Nhật ký nhiệt độ"""
        container = tb.Frame(self.tab_temp_log, padding=10)
        container.pack(fill=BOTH, expand=True)

        # Tiêu đề phân hệ
        title_frame = tb.Frame(container)
        title_frame.pack(fill='x', pady=(0, 10))
        tb.Label(title_frame, text="🌡️ NHẬT KÝ THEO DÕI NHIỆT ĐỘ - ĐỘ ẨM (GSP)", 
                 font=('Segoe UI', 14, 'bold'), bootstyle='primary').pack(side='left')

        # Split layout: Form (Trái) & Bảng lịch sử (Phải)
        main_layout = tb.Frame(container)
        main_layout.pack(fill=BOTH, expand=True)

        left_panel = tb.LabelFrame(main_layout, text=" Nhập chỉ số hàng ngày ", padding=15, width=320)
        left_panel.pack(side='left', fill='y', padx=(0, 10))
        left_panel.pack_propagate(False)

        right_panel = tb.LabelFrame(main_layout, text=" Lịch sử theo dõi ", padding=15)
        right_panel.pack(side='right', fill=BOTH, expand=True)

        # --- Form nhập liệu (left_panel) ---
        # 1. Ngày ghi
        tb.Label(left_panel, text="Ngày theo dõi:").pack(anchor='w', pady=(5, 2))
        self.temp_date = DateEntry(left_panel, dateformat='%Y-%m-%d', bootstyle='primary')
        self.temp_date.pack(fill='x', pady=(0, 10))

        # 2. Buổi
        tb.Label(left_panel, text="Buổi:").pack(anchor='w', pady=(0, 2))
        self.temp_session = tb.Combobox(left_panel, values=['Sáng', 'Chiều'], state='readonly')
        self.temp_session.pack(fill='x', pady=(0, 10))
        self.temp_session.current(0)

        # 3. Vị trí thiết bị
        tb.Label(left_panel, text="Tủ lạnh / Kho bảo quản:").pack(anchor='w', pady=(0, 2))
        self.temp_location = tb.Combobox(left_panel, values=[
            'Tủ vaccine 1 (2-8°C)', 
            'Tủ vaccine 2 (2-8°C)', 
            'Kho lạnh 1 (2-8°C)', 
            'Kho mát VTYT (15-25°C)'
        ])
        self.temp_location.pack(fill='x', pady=(0, 10))
        self.temp_location.current(0)

        # 4. Nhiệt độ
        tb.Label(left_panel, text="Nhiệt độ (°C):").pack(anchor='w', pady=(0, 2))
        self.temp_val = tb.Entry(left_panel)
        self.temp_val.pack(fill='x', pady=(0, 10))
        self.temp_val.insert(0, "5.0")

        # 5. Độ ẩm
        tb.Label(left_panel, text="Độ ẩm (% RH - không bắt buộc):").pack(anchor='w', pady=(0, 2))
        self.temp_humidity = tb.Entry(left_panel)
        self.temp_humidity.pack(fill='x', pady=(0, 10))

        # 6. Người ghi nhận
        tb.Label(left_panel, text="Người ghi nhận:").pack(anchor='w', pady=(0, 2))
        self.temp_recorded_by = tb.Entry(left_panel)
        self.temp_recorded_by.pack(fill='x', pady=(0, 15))
        
        # Load tên thủ kho mặc định nếu có
        self.temp_recorded_by.insert(0, "Thủ kho CDC")

        # Nút Lưu
        btn_save = tb.Button(left_panel, text="💾 Lưu chỉ số", bootstyle='success', command=self.save_temp_log)
        btn_save.pack(fill='x')

        # --- Bảng lịch sử & bộ lọc (right_panel) ---
        # Bộ lọc
        filter_frame = tb.Frame(right_panel)
        filter_frame.pack(fill='x', pady=(0, 10))

        tb.Label(filter_frame, text="Tháng lọc:").pack(side='left', padx=(0, 5))
        
        # List các tháng của năm hiện tại và trước đó
        now = dt.datetime.now()
        months_list = []
        for i in range(12):
            m = now - dt.timedelta(days=30 * i)
            months_list.append(m.strftime('%Y-%m'))
        months_list = sorted(list(set(months_list)), reverse=True)
        
        self.temp_filter_month = tb.Combobox(filter_frame, values=months_list, state='readonly', width=10)
        self.temp_filter_month.pack(side='left', padx=(0, 15))
        if months_list:
            self.temp_filter_month.current(0)
        self.temp_filter_month.bind('<<ComboboxSelected>>', lambda e: self.load_temp_logs_tree())

        tb.Label(filter_frame, text="Tủ/Kho:").pack(side='left', padx=(0, 5))
        self.temp_filter_loc = tb.Combobox(filter_frame, values=['Tất cả'], state='readonly', width=20)
        self.temp_filter_loc.pack(side='left', padx=(0, 10))
        self.temp_filter_loc.current(0)
        self.temp_filter_loc.bind('<<ComboboxSelected>>', lambda e: self.load_temp_logs_tree())

        btn_refresh = tb.Button(filter_frame, text="🔄 Tải lại", bootstyle='secondary-outline', 
                                command=self.refresh_temp_tab_data, padding=(10, 5))
        btn_refresh.pack(side='left')

        # Treeview bảng dữ liệu
        tree_frame = tb.Frame(right_panel)
        tree_frame.pack(fill=BOTH, expand=True)

        self.temp_tree = tb.Treeview(
            tree_frame, 
            columns=('id', 'date', 'session', 'location', 'temp', 'humidity', 'recorded_by', 'status'),
            show='headings',
            bootstyle='primary'
        )
        self.temp_tree.pack(side='left', fill=BOTH, expand=True)

        scrollbar = tb.Scrollbar(tree_frame, orient="vertical", command=self.temp_tree.yview)
        scrollbar.pack(side='right', fill='y')
        self.temp_tree.configure(yscrollcommand=scrollbar.set)

        self.temp_tree.heading('id', text='ID')
        self.temp_tree.heading('date', text='Ngày ghi')
        self.temp_tree.heading('session', text='Buổi')
        self.temp_tree.heading('location', text='Vị trí tủ/kho')
        self.temp_tree.heading('temp', text='Nhiệt độ (°C)')
        self.temp_tree.heading('humidity', text='Độ ẩm (%)')
        self.temp_tree.heading('recorded_by', text='Người ghi')
        self.temp_tree.heading('status', text='Trạng thái')

        self.temp_tree.column('id', width=40, anchor='center')
        self.temp_tree.column('date', width=90, anchor='center')
        self.temp_tree.column('session', width=70, anchor='center')
        self.temp_tree.column('location', width=180, anchor='w')
        self.temp_tree.column('temp', width=95, anchor='center')
        self.temp_tree.column('humidity', width=95, anchor='center')
        self.temp_tree.column('recorded_by', width=110, anchor='w')
        self.temp_tree.column('status', width=120, anchor='center')

        # Thiết lập tag cảnh báo màu sắc
        self.temp_tree.tag_configure('warning', foreground='orange', font=('Segoe UI', 10, 'bold'))
        self.temp_tree.tag_configure('danger', foreground='red', font=('Segoe UI', 10, 'bold'))

        # Khu vực các nút chức năng (phía dưới bảng)
        action_frame = tb.Frame(right_panel)
        action_frame.pack(fill='x', pady=(10, 0))

        btn_delete = tb.Button(action_frame, text="❌ Xóa dòng chọn", bootstyle='danger-outline', 
                               command=self.delete_selected_temp_log)
        btn_delete.pack(side='left', padx=(0, 10))

        self.btn_export_pdf = tb.Button(action_frame, text="📄 Xuất PDF Sổ nhật ký", bootstyle='info', 
                                        command=self.export_temp_log_pdf)
        self.btn_export_pdf.pack(side='right', padx=(10, 0))

        self.btn_plot = tb.Button(action_frame, text="📈 Vẽ biểu đồ xu hướng", bootstyle='primary-outline', 
                                  command=self.plot_temp_chart)
        self.btn_plot.pack(side='right')

        # Tải dữ liệu ban đầu
        self.refresh_temp_tab_data()

    def refresh_temp_tab_data(self):
        """Cập nhật bộ lọc vị trí tủ và tải danh sách nhật ký"""
        try:
            locations = self.db.get_temperature_locations()
            current_filter = self.temp_filter_loc.get()
            
            # Giữ nguyên danh sách cố định và thêm các vị trí trong DB
            fixed_locations = ['Tủ vaccine 1 (2-8°C)', 'Tủ vaccine 2 (2-8°C)', 'Kho lạnh 1 (2-8°C)', 'Kho mát VTYT (15-25°C)']
            all_locs = sorted(list(set(fixed_locations + locations)))
            
            self.temp_filter_loc.config(values=['Tất cả'] + all_locs)
            self.temp_location.config(values=all_locs)
            
            if current_filter in all_locs:
                self.temp_filter_loc.set(current_filter)
            else:
                self.temp_filter_loc.current(0)
                
            self.load_temp_logs_tree()
        except Exception as e:
            print(f"Lỗi refresh danh sách tủ: {e}")

    def load_temp_logs_tree(self):
        """Đọc và điền dữ liệu nhật ký nhiệt độ lên bảng"""
        # Xóa bảng cũ
        for item in self.temp_tree.get_children():
            self.temp_tree.delete(item)

        month = self.temp_filter_month.get()
        loc = self.temp_filter_loc.get()

        try:
            logs = self.db.get_temperature_logs(month, loc)
            for r in logs:
                t = float(r['temperature'])
                h = r['humidity']
                h_str = f"{h}%" if h is not None and h != "" else "-"
                
                # Xác định trạng thái vượt ngưỡng
                loc_lower = r['locationName'].lower()
                status = "Bình thường"
                tag = ""
                
                if "2-8" in loc_lower or "vaccine" in loc_lower or "lạnh" in loc_lower:
                    if t < 2.0 or t > 8.0:
                        status = "Vượt ngưỡng (2-8°C)"
                        tag = "danger"
                elif "15-25" in loc_lower or "mát" in loc_lower:
                    if t < 15.0 or t > 25.0:
                        status = "Vượt ngưỡng (15-25°C)"
                        tag = "warning"
                else:
                    # Mặc định kho thường
                    if t > 30.0:
                        status = "Nhiệt độ cao (>30°C)"
                        tag = "warning"
                    if h is not None and h != "" and float(h) > 75.0:
                        status = "Độ ẩm cao (>75%)"
                        tag = "warning"
                
                self.temp_tree.insert('', 'end', values=(
                    r['id'],
                    r['logDate'],
                    r['session'],
                    r['locationName'],
                    f"{t} °C",
                    h_str,
                    r['recordedBy'] or "-",
                    status
                ), tags=(tag,) if tag else ())
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể tải nhật ký nhiệt độ: {str(e)}")

    def save_temp_log(self):
        """Lưu chỉ số nhiệt độ/độ ẩm mới nhập vào DB"""
        log_date = self.temp_date.entry.get()
        session = self.temp_session.get()
        location = self.temp_location.get().strip()
        temp_s = self.temp_val.get().strip()
        humidity_s = self.temp_humidity.get().strip()
        recorded_by = self.temp_recorded_by.get().strip()

        if not location:
            messagebox.showwarning("Thiếu dữ liệu", "Vui lòng nhập tên tủ lạnh hoặc kho bảo quản."); return
        if not temp_s:
            messagebox.showwarning("Thiếu dữ liệu", "Vui lòng nhập chỉ số nhiệt độ."); return

        try:
            temp = float(temp_s)
        except ValueError:
            messagebox.showerror("Sai định dạng", "Chỉ số nhiệt độ phải là số thực (ví dụ: 5.5)."); return

        humidity = None
        if humidity_s:
            try:
                humidity = float(humidity_s)
            except ValueError:
                messagebox.showerror("Sai định dạng", "Chỉ số độ ẩm phải là số thực (ví dụ: 65)."); return

        # Kiểm tra nhanh ngưỡng để đưa cảnh báo ngay khi nhập
        loc_lower = location.lower()
        is_alert = False
        limit_desc = ""
        if "2-8" in loc_lower or "vaccine" in loc_lower or "lạnh" in loc_lower:
            if temp < 2.0 or temp > 8.0:
                is_alert = True
                limit_desc = "2-8°C"
        elif "15-25" in loc_lower or "mát" in loc_lower:
            if temp < 15.0 or temp > 25.0:
                is_alert = True
                limit_desc = "15-25°C"
        else:
            if temp > 30.0:
                is_alert = True
                limit_desc = "dưới 30°C"

        if is_alert:
            self.toast(f"⚠️ Cảnh báo: Nhiệt độ {temp}°C vượt ngưỡng an toàn ({limit_desc})!", ms=3000)

        try:
            self.db.add_temperature_log(log_date, session, location, temp, humidity, recorded_by)
            self.toast("Đã lưu chỉ số nhiệt độ thành công!")
            
            # Reset ô nhập nhiệt độ/độ ẩm để nhập tiếp buổi khác
            self.temp_val.delete(0, 'end')
            self.temp_val.insert(0, "5.0")
            self.temp_humidity.delete(0, 'end')
            
            self.refresh_temp_tab_data()
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể lưu dữ liệu: {str(e)}")

    def delete_selected_temp_log(self):
        """Xóa dòng nhật ký đang được chọn"""
        selected = self.temp_tree.selection()
        if not selected:
            messagebox.showwarning("Chưa chọn", "Vui lòng chọn dòng cần xóa trong bảng."); return

        values = self.temp_tree.item(selected[0])['values']
        log_id = values[0]
        date_str = values[1]
        session = values[2]
        loc = values[3]

        confirm = messagebox.askyesno("Xác nhận", f"Bạn có chắc muốn xóa bản ghi ngày {date_str} ({session}) của {loc} không?")
        if confirm:
            try:
                self.db.delete_temperature_log(log_id)
                self.toast("Đã xóa dòng nhật ký thành công!")
                self.refresh_temp_tab_data()
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể xóa bản ghi: {str(e)}")

    def plot_temp_chart(self):
        """Vẽ biểu đồ dao động nhiệt độ trong tháng sử dụng matplotlib"""
        # Kiểm tra thư viện matplotlib
        try:
            import matplotlib
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            from matplotlib.figure import Figure
        except ImportError:
            messagebox.showerror("Thiếu thư viện", "Bạn cần cài đặt matplotlib để sử dụng chức năng vẽ biểu đồ.\nHãy chạy lệnh 'pip install matplotlib' trong terminal."); return

        month = self.temp_filter_month.get()
        loc = self.temp_filter_loc.get()

        if loc == "Tất cả":
            messagebox.showwarning("Chọn vị trí", "Vui lòng chọn một tủ bảo quản hoặc kho lạnh cụ thể để vẽ biểu đồ."); return

        try:
            logs = self.db.get_temperature_logs(month, loc)
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể đọc dữ liệu: {str(e)}"); return

        if not logs:
            messagebox.showwarning("Không có dữ liệu", f"Không tìm thấy bản ghi nhiệt độ nào cho {loc} trong tháng {month} để vẽ biểu đồ."); return

        # Đảo ngược danh sách để vẽ từ ngày 1 đến cuối tháng
        logs = list(reversed(logs))

        # Chuẩn bị dữ liệu
        dates_session = []
        temps = []
        
        for r in logs:
            day_str = r['logDate'].split('-')[-1]
            dates_session.append(f"{day_str} ({r['session']})")
            temps.append(float(r['temperature']))

        # Tạo cửa sổ hiển thị biểu đồ
        chart_win = tb.Toplevel(self)
        chart_win.title(f"Biểu đồ nhiệt độ - {loc} - Tháng {month}")
        chart_win.geometry("900x550")
        chart_win.transient(self)
        
        # Center cửa sổ
        chart_win.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() // 2) - (900 // 2)
        y = self.winfo_y() + (self.winfo_height() // 2) - (550 // 2)
        chart_win.geometry(f"900x550+{x}+{y}")

        # Thêm tiêu đề text ở trên
        tb.Label(chart_win, text=f"BIỂU ĐỒ BIẾN THIÊN NHIỆT ĐỘ - THÁNG {month}", 
                 font=('Segoe UI', 12, 'bold'), bootstyle='primary').pack(pady=(10, 5))
        tb.Label(chart_win, text=f"Thiết bị: {loc}", font=('Segoe UI', 10)).pack(pady=(0, 10))

        # Thiết lập figure matplotlib
        fig = Figure(figsize=(8, 4), dpi=100)
        ax = fig.add_subplot(111)

        # Vẽ đường nhiệt độ
        ax.plot(dates_session, temps, marker='o', color='#0275d8', linewidth=2, label='Nhiệt độ thực tế (°C)')
        
        # Thêm nhãn giá trị nhiệt độ tại các điểm
        for idx, temp in enumerate(temps):
            ax.annotate(f"{temp}°", (dates_session[idx], temps[idx]), textcoords="offset points", 
                        xytext=(0,10), ha='center', fontsize=8, fontweight='bold', color='#111')

        # Thêm đường giới hạn cảnh báo dựa trên loại tủ
        loc_lower = loc.lower()
        if "2-8" in loc_lower or "vaccine" in loc_lower or "lạnh" in loc_lower:
            ax.axhline(y=2.0, color='red', linestyle='--', linewidth=1.5, label='Giới hạn dưới (2°C)')
            ax.axhline(y=8.0, color='red', linestyle='--', linewidth=1.5, label='Giới hạn trên (8°C)')
            ax.set_ylim(0, 12)
        elif "15-25" in loc_lower or "mát" in loc_lower:
            ax.axhline(y=15.0, color='orange', linestyle='--', linewidth=1.5, label='Giới hạn dưới (15°C)')
            ax.axhline(y=25.0, color='orange', linestyle='--', linewidth=1.5, label='Giới hạn trên (25°C)')
            ax.set_ylim(10, 30)
        else:
            ax.axhline(y=30.0, color='red', linestyle='--', linewidth=1.5, label='Giới hạn nhiệt độ thường (30°C)')
            ax.set_ylim(15, 35)

        ax.set_ylabel("Nhiệt độ (°C)", fontsize=10, fontweight='bold')
        ax.set_xlabel("Ngày & Buổi kiểm tra", fontsize=10, fontweight='bold')
        ax.grid(True, linestyle=':', alpha=0.6)
        ax.legend(loc='upper right')
        
        # Xoay label x để đỡ bị dính chữ
        fig.autofmt_xdate(rotation=45)

        canvas = FigureCanvasTkAgg(fig, master=chart_win)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=BOTH, expand=True, padx=15, pady=10)

        tb.Button(chart_win, text="Đóng biểu đồ", bootstyle='secondary', command=chart_win.destroy).pack(pady=(0, 15))

    def export_temp_log_pdf(self):
        """Xuất PDF Sổ nhật ký nhiệt độ tháng chuyên nghiệp theo quy định GSP"""
        # Kiểm tra reportlab
        try:
            import reportlab
        except ImportError:
            response = messagebox.askyesno(
                "Thiếu thư viện", 
                "Hệ thống thiếu thư viện 'reportlab' để xuất PDF.\nBạn có muốn tự động cài đặt không? (Quá trình này mất khoảng vài giây)"
            )
            if response:
                import subprocess
                import sys
                try:
                    self.toast("Đang cài đặt thư viện reportlab, vui lòng đợi...")
                    subprocess.run([sys.executable, "-m", "pip", "install", "reportlab"], check=True)
                    self.toast("Đã cài đặt reportlab thành công!")
                except Exception as ex:
                    messagebox.showerror("Lỗi cài đặt", f"Không thể tự động cài đặt reportlab: {str(ex)}\nHãy chạy lệnh 'pip install reportlab' trong terminal."); return
            else:
                return

        month = self.temp_filter_month.get()
        loc = self.temp_filter_loc.get()

        if loc == "Tất cả":
            messagebox.showwarning("Chọn thiết bị", "Vui lòng chọn cụ thể một tủ lạnh hoặc kho bảo quản cần xuất nhật ký."); return

        try:
            logs = self.db.get_temperature_logs(month, loc)
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể lấy dữ liệu: {str(e)}"); return

        if not logs:
            messagebox.showwarning("Không có dữ liệu", f"Không có dữ liệu ghi chép nào của {loc} trong tháng {month} để xuất báo cáo."); return

        # Cho phép chọn vị trí lưu file PDF
        month_file_str = month.replace('-', '_')
        loc_file_str = "".join([c if c.isalnum() else "_" for c in loc])
        initial_filename = f"Nhat_Ky_Nhiet_Do_{loc_file_str}_{month_file_str}.pdf"
        
        pdf_path = filedialog.asksaveasfilename(
            title="Chọn vị trí lưu Sổ nhật ký nhiệt độ PDF",
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            initialfile=initial_filename
        )
        if not pdf_path:
            return

        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            
            # Đăng ký font hỗ trợ tiếng Việt
            try:
                pdfmetrics.registerFont(TTFont('TimesNewRoman', "C:\\Windows\\Fonts\\times.ttf"))
                pdfmetrics.registerFont(TTFont('TimesNewRoman-Bold', "C:\\Windows\\Fonts\\timesbd.ttf"))
                pdfmetrics.registerFont(TTFont('TimesNewRoman-Italic', "C:\\Windows\\Fonts\\timesi.ttf"))
                font_normal = 'TimesNewRoman'
                font_bold = 'TimesNewRoman-Bold'
                font_italic = 'TimesNewRoman-Italic'
            except Exception:
                font_normal = 'Helvetica'
                font_bold = 'Helvetica-Bold'
                font_italic = 'Helvetica-Oblique'

            # Trang nằm ngang (A4 Landscape)
            from reportlab.lib.pagesizes import landscape
            doc = SimpleDocTemplate(pdf_path, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
            story = []

            styles = getSampleStyleSheet()

            style_header_left = ParagraphStyle(
                'HLeft', parent=styles['Normal'], fontName=font_bold, fontSize=10, leading=13, alignment=0
            )
            style_header_right = ParagraphStyle(
                'HRight', parent=styles['Normal'], fontName=font_normal, fontSize=9, leading=12, alignment=2
            )
            style_title = ParagraphStyle(
                'TitleTemp', parent=styles['Heading1'], fontName=font_bold, fontSize=15, leading=18, alignment=1, spaceAfter=5
            )
            style_subtitle = ParagraphStyle(
                'SubTemp', parent=styles['Normal'], fontName=font_italic, fontSize=10, leading=13, alignment=1, spaceAfter=15
            )
            style_tbl_hdr = ParagraphStyle(
                'TblHdr', parent=styles['Normal'], fontName=font_bold, fontSize=9, leading=11, alignment=1
            )
            style_cell = ParagraphStyle(
                'CVal', parent=styles['Normal'], fontName=font_normal, fontSize=9, leading=11, alignment=1
            )
            style_cell_left = ParagraphStyle(
                'CValL', parent=styles['Normal'], fontName=font_normal, fontSize=9, leading=11, alignment=0
            )

            # Header hai bên
            header_data = [
                [
                    Paragraph("SỞ Y TẾ THÀNH PHỐ CẦN THƠ<br/>TRUNG TÂM KIỂM SOÁT BỆNH TẬT (CDC)", style_header_left),
                    Paragraph("<b>BIỂU MẪU THEO DÕI GSP</b><br/><i>(Quy định theo Thông tư số 36/2018/TT-BYT)</i>", style_header_right)
                ]
            ]
            header_table = Table(header_data, colWidths=[380, 400])
            header_table.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
                ('BOTTOMPADDING', (0,0), (-1,-1), 10),
            ]))
            story.append(header_table)
            
            # Tiêu đề báo cáo
            story.append(Paragraph("NHẬT KÝ THEO DÕI NHIỆT ĐỘ, ĐỘ ẨM TỦ BẢO QUẢN / KHO LẠNH", style_title))
            
            # Lấy thông tin tháng
            parts = month.split('-')
            month_text = f"Tháng {parts[1]} năm {parts[0]}"
            story.append(Paragraph(f"Thiết bị/Vị trí theo dõi: {loc} • Thời gian: {month_text}", style_subtitle))

            # Xây dựng bảng nhật ký (bố cục ngang, chia cột rõ ràng)
            # Cột: STT | Ngày | Buổi | Chỉ số nhiệt độ | Chỉ số độ ẩm | Trạng thái | Người ký xác nhận | Ghi chú/Biện pháp khắc phục
            table_data = [
                [
                    Paragraph("<b>STT</b>", style_tbl_hdr),
                    Paragraph("<b>Ngày ghi</b>", style_tbl_hdr),
                    Paragraph("<b>Buổi</b>", style_tbl_hdr),
                    Paragraph("<b>Nhiệt độ (°C)</b>", style_tbl_hdr),
                    Paragraph("<b>Độ ẩm (% RH)</b>", style_tbl_hdr),
                    Paragraph("<b>Đánh giá GSP</b>", style_tbl_hdr),
                    Paragraph("<b>Người ký ghi tên</b>", style_tbl_hdr),
                    Paragraph("<b>Biện pháp khắc phục khi vượt ngưỡng</b>", style_tbl_hdr)
                ]
            ]

            # Sắp xếp từ ngày đầu đến cuối tháng
            logs = list(reversed(logs))
            
            for idx, r in enumerate(logs):
                t = float(r['temperature'])
                h = r['humidity']
                h_str = f"{h}%" if h is not None and h != "" else "-"
                
                # Trạng thái
                loc_lower = r['locationName'].lower()
                status = "Đạt tiêu chuẩn"
                if "2-8" in loc_lower or "vaccine" in loc_lower or "lạnh" in loc_lower:
                    if t < 2.0 or t > 8.0:
                        status = "⚠️ VƯỢT NGƯỠNG"
                elif "15-25" in loc_lower or "mát" in loc_lower:
                    if t < 15.0 or t > 25.0:
                        status = "⚠️ VƯỢT NGƯỠNG"
                else:
                    if t > 30.0 or (h is not None and h != "" and float(h) > 75.0):
                        status = "⚠️ VƯỢT NGƯỠNG"
                
                table_data.append([
                    Paragraph(str(idx + 1), style_cell),
                    Paragraph(r['logDate'], style_cell),
                    Paragraph(r['session'], style_cell),
                    Paragraph(f"<b>{t} °C</b>" if "VƯỢT" in status else f"{t} °C", style_cell),
                    Paragraph(h_str, style_cell),
                    Paragraph(f"<font color='red'><b>{status}</b></font>" if "VƯỢT" in status else status, style_cell),
                    Paragraph(r['recordedBy'] or "-", style_cell),
                    Paragraph("" if "Đạt" in status else "Đã chuyển vaccine sang tủ dự phòng / Báo cáo kỹ thuật", style_cell_left)
                ])

            # Rộng cột
            col_widths = [35, 75, 55, 90, 80, 110, 110, 225]
            t_style = TableStyle([
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ('BACKGROUND', (0,0), (-1,0), colors.whitesmoke),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('TOPPADDING', (0,0), (-1,-1), 5),
                ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ])
            
            # Tạo bảng
            temp_table = Table(table_data, colWidths=col_widths)
            temp_table.setStyle(t_style)
            story.append(temp_table)
            story.append(Spacer(1, 20))

            # Chữ ký phê duyệt
            now_dt = dt.datetime.now()
            date_sign_str = f"Ngày {now_dt.strftime('%d')} tháng {now_dt.strftime('%m')} năm {now_dt.strftime('%Y')}"
            
            sig_data = [
                [
                    "",
                    Paragraph(f"<i>Cần Thơ, {date_sign_str}</i>", style_cell)
                ],
                [
                    Paragraph("<b>THỦ KHO CDC</b><br/><i>(Ký và ghi rõ họ tên)</i>", style_cell),
                    Paragraph("<b>LÃNH ĐẠO TRUNG TÂM PHÊ DUYỆT</b><br/><i>(Ký tên và đóng dấu)</i>", style_cell)
                ],
                [
                    Spacer(1, 40),
                    Spacer(1, 40)
                ]
            ]
            sig_table = Table(sig_data, colWidths=[390, 390])
            sig_table.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ]))
            story.append(sig_table)

            # Build PDF
            doc.build(story)
            self.toast("Đã xuất bản Sổ nhật ký nhiệt độ thành công!")
            
            # Mở file PDF ngay sau khi lưu
            try:
                os.startfile(pdf_path)
            except:
                pass
        except Exception as ex:
            messagebox.showerror("Lỗi xuất PDF", f"Lỗi không thể tạo file PDF: {str(ex)}")




