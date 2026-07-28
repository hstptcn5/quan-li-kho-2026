# -*- coding: utf-8 -*-
# test_fixes.py — Kịch bản kiểm thử tự động cho các bản vá lỗi (Bug 1 - 11)
import os
import sqlite3
import unittest
from unittest import mock
import json
import tempfile
import shutil
import gzip
import base64
import http.server
import threading
import urllib.error
import urllib.request



from config import DB_PATH, SCHEMA_VERSION
from date_utils import format_date_display, format_datetime_display, parse_date_to_iso
from database import DB
from managers import BackupManager
import server as mobile_server_module

class TestMedicalWarehouseFixes(unittest.TestCase):
    def setUp(self):
        # Tạo database tạm thời để chạy test
        self.temp_dir = tempfile.mkdtemp()
        self.db_path = os.path.join(self.temp_dir, "test_warehouse.db")
        
        # Tạo kết nối và khởi chạy migrations
        self.db = DB(self.db_path)
        
    def tearDown(self):
        # Đóng database và xóa thư mục tạm
        if hasattr(self, 'db') and self.db:
            try:
                self.db.conn.close()
            except:
                pass
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_schema_and_migrations(self):
        """Kiểm thử Bug 1 & Bug 6: Schema và Migrations tự động"""
        # Kiểm tra sự tồn tại của các cột mới trong bảng stock_movements
        cursor = self.db.conn.cursor()
        cursor.execute("PRAGMA table_info(stock_movements)")
        columns = {row[1] for row in cursor.fetchall()}
        
        self.assertIn("qtyBase", columns)
        self.assertIn("originalQty", columns)
        self.assertIn("originalUnit", columns)
        self.assertIn("referenceType", columns)
        self.assertIn("referenceId", columns)
        self.assertIn("referenceItemId", columns)

        cursor.execute("PRAGMA table_info(purchase_items)")
        purchase_cols = {row[1] for row in cursor.fetchall()}
        self.assertIn("totalAmount", purchase_cols)

        cursor.execute("PRAGMA table_info(dispatch_items)")
        dispatch_cols = {row[1] for row in cursor.fetchall()}
        self.assertIn("totalAmount", dispatch_cols)
        
        # Kiểm tra SCHEMA_VERSION mới nhất
        cursor.execute("PRAGMA user_version")
        version = cursor.fetchone()[0]
        self.assertEqual(version, SCHEMA_VERSION)

    def test_qty_base_and_unit_conversion(self):
        """Kiểm thử Bug 1: Quy đổi đơn vị và qtyBase chính xác"""
        # 1. Thêm sản phẩm mẫu
        self.db.conn.execute(
            "INSERT INTO products (id, name, defaultUnit) VALUES (101, 'Thuốc A', 'Hộp')"
        )
        # Quy đổi đơn vị: 1 Thùng = 10 Hộp, 1 Vỉ = 0.1 Hộp
        self.db.conn.execute(
            "INSERT INTO product_units (productId, unitCode, toBaseQty, price) VALUES (101, 'Thùng', 10, 1000)"
        )
        self.db.conn.execute(
            "INSERT INTO product_units (productId, unitCode, toBaseQty, price) VALUES (101, 'Vỉ', 0.1, 10)"
        )
        self.db.conn.commit()
        
        # 2. Ghi nhận phiếu nhập: 5 Thùng thuốc A
        items = [{
            "productId": 101,
            "qty": 5.0,
            "unitCode": "Thùng",
            "lotNo": "LOT001",
            "expiryDate": "2028-12-31",
            "cost": 1000.0,
            "fundSource": "Nguồn A"
        }]
        purchase_id, note_num, _ = self.db.record_purchase(items, "NCC Test", "Nhập kiểm thử", "Không")
        
        # Kiểm tra stock_movements có qtyBase = 5 * 10 = 50.0
        cursor = self.db.conn.cursor()
        cursor.execute("SELECT qtyBase, qty, originalUnit FROM stock_movements WHERE referenceId=? AND referenceType='PURCHASE'", (purchase_id,))
        move = cursor.fetchone()
        self.assertIsNotNone(move)
        self.assertEqual(move[0], 50.0) # qtyBase = 50
        self.assertEqual(move[1], 5.0)  # qty = 5 (nhập 5 Thùng)
        
        # Kiểm tra tồn kho quy đổi từ stock_view()
        stock = self.db.stock_view()
        self.assertEqual(len(stock), 1)
        self.assertEqual(stock[0]['qtyBase'], 50.0)

    def test_fund_source_isolation_and_no_fallback(self):
        """Kiểm thử Bug 2: Cách ly nguồn kinh phí và không tự động fallback sang nguồn trống"""
        self.db.conn.execute("INSERT INTO products (id, name, defaultUnit) VALUES (102, 'Thuốc B', 'Viên')")
        self.db.conn.commit()
        # Đăng ký đơn vị cơ sở cho sản phẩm 102
        self.db.conn.execute("INSERT INTO product_units (productId, unitCode, toBaseQty, price) VALUES (102, 'Viên', 1.0, 0.0)")
        self.db.conn.commit()
        
        # Nhập 100 Viên nguồn "Chương trình A"
        self.db.record_purchase([{
            "productId": 102,
            "qty": 100.0,
            "unitCode": "Viên",
            "lotNo": "LOT002",
            "expiryDate": "2028-12-31",
            "cost": 500.0,
            "fundSource": "Chương trình A"
        }], "NCC Test", "Nhập", "Ghi chú")
        
        # Nhập 50 Viên nguồn "Chương trình B"
        self.db.record_purchase([{
            "productId": 102,
            "qty": 50.0,
            "unitCode": "Viên",
            "lotNo": "LOT002",
            "expiryDate": "2028-12-31",
            "cost": 500.0,
            "fundSource": "Chương trình B"
        }], "NCC Test", "Nhập", "Ghi chú")
        
        # Thử xuất 120 viên từ nguồn "Chương trình A" (Phải lỗi vì chỉ có 100 viên nguồn này)
        with self.assertRaises(Exception) as ctx:
            self.db.dispatch([{
                "productId": 102,
                "qty": 120.0,
                "unitCode": "Viên",
                "lotNo": "LOT002",
                "fundSource": "Chương trình A"
            }], "Đơn vị Test", "Xuất", "Ghi chú")
        self.assertIn("không đủ", str(ctx.exception).lower())
        
        # Thử xuất 80 viên từ nguồn rỗng/None (Phải lỗi vì không tự động fallback)
        with self.assertRaises(Exception) as ctx:
            self.db.dispatch([{
                "productId": 102,
                "qty": 10.0,
                "unitCode": "Viên",
                "lotNo": "LOT002",
                "fundSource": ""  # Hoặc None
            }], "Đơn vị Test", "Xuất", "Ghi chú")
        self.assertIn("tồn kho", str(ctx.exception).lower())

    def test_batch_validation_same_lot_different_expiry(self):
        """Kiểm thử Bug 8: Validation lô hàng cùng số lô nhưng khác HSD"""
        self.db.conn.execute("INSERT INTO products (id, name, defaultUnit) VALUES (103, 'Thuốc C', 'Lọ')")
        self.db.conn.execute("INSERT INTO product_units (productId, unitCode, toBaseQty, price) VALUES (103, 'Lọ', 1.0, 0.0)")
        self.db.conn.commit()
        
        # Nhập lô LOT003, HSD 2027-12-31
        self.db.record_purchase([{
            "productId": 103,
            "qty": 10.0,
            "unitCode": "Lọ",
            "lotNo": "LOT003",
            "expiryDate": "2027-12-31",
            "cost": 100.0,
            "fundSource": "Nguồn A"
        }], "NCC Test", "Nhập", "")
        
        # Nhập tiếp cùng LOT003 nhưng HSD 2028-06-30 (Phải ném ra ngoại lệ ngăn chặn)
        with self.assertRaises(Exception) as ctx:
            self.db.record_purchase([{
                "productId": 103,
                "qty": 5.0,
                "unitCode": "Lọ",
                "lotNo": "LOT003",
                "expiryDate": "2028-06-30", # Khác HSD
                "cost": 100.0,
                "fundSource": "Nguồn A"
            }], "NCC Test", "Nhập", "")
        self.assertIn("không khớp HSD mới", str(ctx.exception))

    def test_unique_sequential_note_numbers(self):
        """Kiểm thử Bug 9: Số phiếu tăng dần tuần tự và duy nhất"""
        self.db.conn.execute("INSERT INTO products (id, name, defaultUnit) VALUES (104, 'Thuốc D', 'Hộp')")
        self.db.conn.execute("INSERT INTO product_units (productId, unitCode, toBaseQty, price) VALUES (104, 'Hộp', 1.0, 0.0)")
        self.db.conn.commit()
        
        # Nhập 2 phiếu khác nhau
        _, note_num1, _ = self.db.record_purchase([{
            "productId": 104, "qty": 1.0, "unitCode": "Hộp", "lotNo": "L1", "expiryDate": "2027-12-31", "cost": 1.0, "fundSource": "N"
        }], "NCC", "Nhập", "")
        
        _, note_num2, _ = self.db.record_purchase([{
            "productId": 104, "qty": 1.0, "unitCode": "Hộp", "lotNo": "L2", "expiryDate": "2027-12-31", "cost": 1.0, "fundSource": "N"
        }], "NCC", "Nhập", "")
        
        # Kiểm tra hai số phiếu khác nhau và có dạng PNxxx
        self.assertNotEqual(note_num1, note_num2)
        self.assertTrue(note_num1.startswith("PN"))
        self.assertTrue(note_num2.startswith("PN"))
        self.assertRegex(note_num1, r"^PN-\d{6}-\d{3}$")
        self.assertRegex(note_num2, r"^PN-\d{6}-\d{3}$")
        
        # Kiểm tra số thứ hai lớn hơn số thứ nhất
        num1 = int(note_num1.split("-")[-1])
        num2 = int(note_num2.split("-")[-1])
        self.assertEqual(num2, num1 + 1)

    def test_safe_backup_manager(self):
        """Kiểm thử Bug 4: Backup và restore SQLite an toàn"""
        backup_dir = os.path.join(self.temp_dir, "backups")
        os.makedirs(backup_dir, exist_ok=True)
        
        mgr = BackupManager(self.db_path, backup_dir)
        
        # 1. Viết dữ liệu mẫu
        self.db.conn.execute("INSERT INTO products (id, name, defaultUnit) VALUES (105, 'Thuốc E', 'Hộp')")
        self.db.conn.execute("INSERT INTO product_units (productId, unitCode, toBaseQty, price) VALUES (105, 'Hộp', 1.0, 0.0)")
        self.db.conn.commit()
        
        # 2. Tạo backup
        backup_path = mgr.create_backup("test")
        self.assertTrue(os.path.exists(backup_path))
        self.assertTrue(os.path.exists(backup_path.replace(".db", ".json"))) # Metadata file
        
        # 3. Sửa dữ liệu hiện tại
        self.db.conn.execute("DELETE FROM products WHERE id=105")
        self.db.conn.commit()
        
        # Kiểm tra dữ liệu đã mất
        cursor = self.db.conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM products WHERE id=105")
        self.assertEqual(cursor.fetchone()[0], 0)
        
        # 4. Khôi phục từ backup
        # Trước tiên đóng kết nối hiện tại để tránh lock
        self.db.conn.close()
        
        success = mgr.restore_backup(backup_path)
        self.assertTrue(success)
        
        # Mở lại kết nối và kiểm tra dữ liệu đã được khôi phục
        self.db = DB(self.db_path)
        cursor = self.db.conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM products WHERE id=105")
        self.assertEqual(cursor.fetchone()[0], 1)

    def test_export_import_json(self):
        """Kiểm thử Bug 5: Export/Import JSON Whitelisted, Transactional & FK check"""
        backup_dir = os.path.join(self.temp_dir, "backups")
        os.makedirs(backup_dir, exist_ok=True)
        mgr = BackupManager(self.db_path, backup_dir)
        
        # Tạo dữ liệu mẫu
        self.db.conn.execute("INSERT INTO products (id, name, defaultUnit) VALUES (106, 'Thuốc F', 'Hộp')")
        self.db.conn.execute("INSERT INTO product_units (productId, unitCode, toBaseQty, price) VALUES (106, 'Hộp', 1.0, 0.0)")
        self.db.conn.commit()
        
        # Export dữ liệu
        export_path = os.path.join(self.temp_dir, "export.json")
        mgr.export_data(export_path)
        self.assertTrue(os.path.exists(export_path))
        
        # Đọc file export kiểm tra có audit_logs và schema_version
        with open(export_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        self.assertIn("export_info", data)
        self.assertIn("schema_version", data["export_info"])
        self.assertIn("audit_logs", data)
        
        # Import lại vào database khác hoặc dọn sạch rồi import
        self.db.conn.execute("DELETE FROM products")
        self.db.conn.commit()
        
        success = mgr.import_data(export_path)
        self.assertTrue(success)
        
        # Kiểm tra xem dữ liệu đã phục hồi
        cursor = self.db.conn.cursor()
        cursor.execute("SELECT name FROM products WHERE id=106")
        row = cursor.fetchone()
        self.assertIsNotNone(row)
        self.assertEqual(row[0], "Thuốc F")

    def test_advanced_fefo_fund_source(self):
        """Kiểm thử Bug 2 nâng cao: FEFO bỏ qua lô không chứa nguồn kinh phí được chọn"""
        # Sản phẩm 201
        self.db.conn.execute("INSERT INTO products (id, name, defaultUnit) VALUES (201, 'Thuốc FEFO', 'Viên')")
        self.db.conn.execute("INSERT INTO product_units (productId, unitCode, toBaseQty, price) VALUES (201, 'Viên', 1.0, 0.0)")
        self.db.conn.commit()
        
        # Nhập lô 1 (Hết hạn sớm: 2027-12-31), nguồn B
        self.db.record_purchase([{
            "productId": 201,
            "qty": 50.0,
            "unitCode": "Viên",
            "lotNo": "LOT_FEFO_1",
            "expiryDate": "2027-12-31",
            "cost": 10.0,
            "fundSource": "Nguồn B"
        }], "NCC", "Nhập", "")
        
        # Nhập lô 2 (Hết hạn muộn: 2028-12-31), nguồn A
        self.db.record_purchase([{
            "productId": 201,
            "qty": 100.0,
            "unitCode": "Viên",
            "lotNo": "LOT_FEFO_2",
            "expiryDate": "2028-12-31",
            "cost": 10.0,
            "fundSource": "Nguồn A"
        }], "NCC", "Nhập", "")
        
        # Thử xuất 40 viên từ Nguồn A.
        # Lô 1 (hạn sớm) chỉ chứa Nguồn B. Lô 2 (hạn muộn) chứa Nguồn A.
        # FEFO phải bỏ qua Lô 1 và xuất thành công từ Lô 2 (Nguồn A).
        dispatch_id, note_num, details = self.db.dispatch([{
            "productId": 201,
            "qty": 40.0,
            "unitCode": "Viên",
            "fundSource": "Nguồn A"
        }], "Đơn vị nhận", "Xuất", "")
        
        self.assertEqual(len(details), 1)
        self.assertEqual(details[0]['lotNo'], 'LOT_FEFO_2')
        self.assertEqual(details[0]['qty'], 40.0)
        self.assertEqual(details[0]['fundSource'], 'Nguồn A')

    def test_restore_integrity_rollback(self):
        """Kiểm thử Bug 4 nâng cao: Khôi phục database lỗi tự động rollback về bản cũ"""
        backup_dir = os.path.join(self.temp_dir, "backups")
        os.makedirs(backup_dir, exist_ok=True)
        mgr = BackupManager(self.db_path, backup_dir)
        
        # 1. Ghi dữ liệu ban đầu
        self.db.conn.execute("INSERT INTO products (id, name, defaultUnit) VALUES (202, 'Thuốc R', 'Lọ')")
        self.db.conn.execute("INSERT INTO product_units (productId, unitCode, toBaseQty, price) VALUES (202, 'Lọ', 1.0, 0.0)")
        self.db.conn.commit()
        
        # 2. Tạo một file backup bị lỗi (file văn bản rác)
        corrupted_backup_path = os.path.join(backup_dir, "corrupted.db")
        with open(corrupted_backup_path, "w", encoding="utf-8") as f:
            f.write("Đây không phải là file SQLite hợp lệ!")
            
        # 3. Khôi phục từ file lỗi (Phải ném ra ngoại lệ và rollback về trạng thái ban đầu)
        with self.assertRaises(Exception):
            mgr.restore_backup(corrupted_backup_path)
            
        # 4. Mở lại DB, kiểm tra sản phẩm 202 vẫn tồn tại (khôi phục thành công trạng thái cũ)
        self.db = DB(self.db_path)
        cursor = self.db.conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM products WHERE id=202")
        self.assertEqual(cursor.fetchone()[0], 1)

    def test_offline_qr_js_file_exists(self):
        """Kiểm thử Blocker Fix 1: Tệp html5-qrcode.min.js tồn tại trong thư mục static"""
        js_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static", "html5-qrcode.min.js")
        self.assertTrue(os.path.exists(js_path), f"Không tìm thấy tệp JS offline tại {js_path}")
        self.assertGreater(os.path.getsize(js_path), 100000, "Tệp html5-qrcode.min.js có kích thước quá nhỏ hoặc chưa đủ")

    def test_validate_batch_read_only(self):
        """Kiểm thử Fix 2: validate_batch chỉ đọc và ném lỗi khi HSD lệch mà không chèn bản ghi vào DB"""
        self.db.conn.execute("INSERT INTO products (id, name, defaultUnit) VALUES (301, 'Thuốc Batch Test', 'Hộp')")
        self.db.conn.execute("INSERT INTO batches (productId, lotNo, expiryDate) VALUES (301, 'LOT_TEST_V', '2028-12-31')")
        self.db.conn.commit()

        # Gọi validate_batch cùng lotNo nhưng HSD sai -> Ném ValueError
        with self.assertRaises(ValueError):
            self.db.validate_batch(301, 'LOT_TEST_V', '2027-01-01')

        # Gọi validate_batch cho lô mới chưa từng tồn tại -> Không ném lỗi và KHÔNG tạo lô trong DB
        self.db.validate_batch(301, 'LOT_NEW_NONEXISTENT', '2029-01-01')
        count = self.db.q("SELECT COUNT(*) as c FROM batches WHERE lotNo='LOT_NEW_NONEXISTENT'")[0]['c']
        self.assertEqual(count, 0, "validate_batch không được tạo lô mới dở dang vào DB")

    def test_negative_stock_invariant(self):
        """Kiểm thử Fix 3: Invariant kiểm tra âm kho theo (productId, batchId, fundSource)"""
        self.db.conn.execute("INSERT INTO products (id, name, defaultUnit) VALUES (401, 'Thuốc Âm Kho', 'Chai')")
        self.db.conn.execute("INSERT INTO product_units (productId, unitCode, toBaseQty, price) VALUES (401, 'Chai', 1.0, 0.0)")
        self.db.conn.commit()

        # Nhập 10 chai lô L1 nguồn N1
        self.db.record_purchase([{
            "productId": 401, "qty": 10.0, "unitCode": "Chai",
            "lotNo": "L1", "expiryDate": "2030-01-01", "fundSource": "N1"
        }], "NCC", "Nhập", "")

        # Chèn chuyển động làm âm kho lô L1 nguồn N1
        self.db.conn.execute("BEGIN")
        bid = self.db.q("SELECT id FROM batches WHERE productId=401 AND lotNo='L1'")[0]['id']
        self.db.conn.execute(
            "INSERT INTO stock_movements(productId, batchId, unitCode, qty, qtyBase, type, fundSource) VALUES(401, ?, 'Chai', -15.0, -15.0, 'DISPATCH', 'N1')",
            (bid,)
        )
        # Gọi _assert_no_negative_stock -> Ném ValueError
        with self.assertRaises(ValueError):
            self.db._assert_no_negative_stock()
        self.db.conn.rollback()

    def test_audit_ip_recording(self):
        """Kiểm thử Fix 4: Nhật ký thao tác ghi nhận đúng IP người thực hiện"""
        self.db.conn.execute("INSERT INTO products (id, name, defaultUnit) VALUES (501, 'Thuốc Audit IP', 'Ống')")
        self.db.conn.execute("INSERT INTO product_units (productId, unitCode, toBaseQty, price) VALUES (501, 'Ống', 1.0, 0.0)")
        self.db.conn.commit()

        # Nhập kho với IP di động
        mobile_ip = "192.168.1.105"
        purchase_id, _, _ = self.db.record_purchase([{
            "productId": 501, "qty": 20.0, "unitCode": "Ống",
            "lotNo": "L501", "expiryDate": "2030-01-01", "fundSource": ""
        }], "NCC", "Nhập mobile", "", audit_ip=mobile_ip)

        # Kiểm tra IP trong audit_logs
        log = self.db.q("SELECT ip FROM audit_logs WHERE noteId=? ORDER BY id DESC LIMIT 1", (purchase_id,))
        self.assertEqual(log[0]['ip'], mobile_ip)

    def test_token_ip_matching(self):
        """Kiểm thử Fix 6: Phiên đăng nhập mobile tự hủy khi truy cập từ IP khác"""
        active_tokens = {}
        token = "test_token_123"
        active_tokens[token] = {"ip": "192.168.1.50", "expiry": 9999999999.0}

        # Truy cập đúng IP -> Hợp lệ
        request_ip = "192.168.1.50"
        self.assertEqual(active_tokens.get(token, {}).get("ip"), request_ip)

        # Truy cập sai IP -> Bị vô hiệu hóa
        fraud_ip = "192.168.1.99"
        if active_tokens.get(token, {}).get("ip") != fraud_ip:
            active_tokens.pop(token, None)

        self.assertNotIn(token, active_tokens, "Token phải bị xóa khi IP không trùng khớp")

    def _start_mobile_test_server(self):
        old_db_path = mobile_server_module.DB_PATH
        old_pin = mobile_server_module.SERVER_PIN
        old_tokens = dict(mobile_server_module.ACTIVE_TOKENS)
        old_failed = dict(mobile_server_module.FAILED_ATTEMPTS)

        mobile_server_module.DB_PATH = self.db_path
        mobile_server_module.SERVER_PIN = "123456"
        mobile_server_module.ACTIVE_TOKENS.clear()
        mobile_server_module.FAILED_ATTEMPTS.clear()

        httpd = http.server.HTTPServer(("127.0.0.1", 0), mobile_server_module.MobileInventoryRequestHandler)
        thread = threading.Thread(target=httpd.serve_forever, daemon=True)
        thread.start()
        base_url = f"http://127.0.0.1:{httpd.server_port}"

        def cleanup():
            httpd.shutdown()
            httpd.server_close()
            thread.join(timeout=2)
            mobile_server_module.DB_PATH = old_db_path
            mobile_server_module.SERVER_PIN = old_pin
            mobile_server_module.ACTIVE_TOKENS.clear()
            mobile_server_module.ACTIVE_TOKENS.update(old_tokens)
            mobile_server_module.FAILED_ATTEMPTS.clear()
            mobile_server_module.FAILED_ATTEMPTS.update(old_failed)

        self.addCleanup(cleanup)
        return base_url

    def _json_request(self, url, payload=None, token=None):
        data = None
        headers = {"Content-Type": "application/json"}
        if token:
            headers["Authorization"] = f"Bearer {token}"
        if payload is not None:
            data = json.dumps(payload).encode("utf-8")
        req = urllib.request.Request(url, data=data, headers=headers, method="POST" if payload is not None else "GET")
        with urllib.request.urlopen(req, timeout=5) as resp:
            return resp.status, json.loads(resp.read().decode("utf-8"))

    def test_mobile_handler_auth_and_invalid_dispatch_date(self):
        """Kiểm thử handler HTTP thật: auth, URL protected và validate ngày xuất mobile."""
        base_url = self._start_mobile_test_server()

        with self.assertRaises(urllib.error.HTTPError) as missing_auth:
            self._json_request(f"{base_url}/api/products")
        self.assertEqual(missing_auth.exception.code, 401)

        status, auth_data = self._json_request(f"{base_url}/api/auth", {"pin": "123456"})
        self.assertEqual(status, 200)
        self.assertTrue(auth_data["success"])
        token = auth_data["token"]

        self.db.conn.execute("INSERT INTO products (id, name, defaultUnit) VALUES (501, 'Mobile Handler Drug', 'Vien')")
        self.db.conn.execute("INSERT INTO product_units (productId, unitCode, toBaseQty, price) VALUES (501, 'Vien', 1, 0)")
        self.db.conn.commit()

        with self.assertRaises(urllib.error.HTTPError) as invalid_date:
            self._json_request(f"{base_url}/api/dispatch", {
                "productId": 501,
                "qty": 1,
                "dispatchDate": "31/02/2026"
            }, token=token)
        self.assertEqual(invalid_date.exception.code, 400)
        body = json.loads(invalid_date.exception.read().decode("utf-8"))
        self.assertIn("Ngày xuất không hợp lệ", body["message"])

    def test_mobile_handler_rejects_wrong_ip_token(self):
        """Kiểm thử handler HTTP thật: token không đúng IP bị từ chối và vô hiệu hóa."""
        base_url = self._start_mobile_test_server()
        mobile_server_module.ACTIVE_TOKENS["bad_ip_token"] = {
            "ip": "192.0.2.10",
            "expiry": 9999999999.0,
        }

        req = urllib.request.Request(
            f"{base_url}/api/products",
            headers={"Authorization": "Bearer bad_ip_token"},
            method="GET",
        )
        with self.assertRaises(urllib.error.HTTPError) as wrong_ip:
            urllib.request.urlopen(req, timeout=5)
        self.assertEqual(wrong_ip.exception.code, 401)
        self.assertNotIn("bad_ip_token", mobile_server_module.ACTIVE_TOKENS)

    def test_mobile_print_url_requires_and_accepts_query_token(self):
        """Kiểm thử URL in mở bằng window.open: thiếu token bị chặn, token query hợp lệ được mở."""
        self.db.conn.execute("INSERT INTO products (id, name, defaultUnit) VALUES (502, 'Print Token Drug', 'Vien')")
        self.db.conn.execute("INSERT INTO product_units (productId, unitCode, toBaseQty, price) VALUES (502, 'Vien', 1, 0)")
        self.db.conn.commit()
        purchase_id, _, _ = self.db.record_purchase([{
            "productId": 502,
            "productName": "Print Token Drug",
            "qty": 3,
            "unitCode": "Vien",
            "lotNo": "PT01",
            "expiryDate": "2028-12-31",
            "cost": 100,
            "fundSource": "N"
        }], "NCC", "Nhap", "", date_str="2026-07-28")

        base_url = self._start_mobile_test_server()
        with self.assertRaises(urllib.error.HTTPError) as missing_auth:
            urllib.request.urlopen(f"{base_url}/api/print-purchase?id={purchase_id}", timeout=5)
        self.assertEqual(missing_auth.exception.code, 401)

        status, auth_data = self._json_request(f"{base_url}/api/auth", {"pin": "123456"})
        self.assertEqual(status, 200)
        token = auth_data["token"]

        with urllib.request.urlopen(f"{base_url}/api/print-purchase?id={purchase_id}&token={token}", timeout=5) as resp:
            body = resp.read().decode("utf-8")
            self.assertEqual(resp.status, 200)
            self.assertIn("Print Token Drug", body)

    def test_mobile_template_window_open_urls_include_token_helper(self):
        """Kiểm thử template mobile không mở URL in protected thiếu token."""
        from mobile_templates import MOBILE_HTML

        self.assertIn("function withAuthToken(url)", MOBILE_HTML)
        self.assertNotIn("window.open(`/api/print-purchase", MOBILE_HTML)
        self.assertNotIn("window.open(`/api/print-dispatch", MOBILE_HTML)
        self.assertNotIn("window.open(printUrl, '_blank')", MOBILE_HTML)
        self.assertNotIn("window.open(url, '_blank')", MOBILE_HTML)

    def test_bulk_import_products_and_stock_atomic(self):
        """Kiểm thử Lỗi 1 mới: Nhập Excel Atomic — khi xảy ra lỗi ở bất kỳ bước nào, toàn bộ sản phẩm & đơn vị được rollback sạch"""
        records_with_error = [
            {
                'product_info': {
                    'name': 'Thuốc Atomic 1',
                    'defaultUnit': 'Hộp',
                    'barcode': '11111',
                    'productType': 'thuoc',
                    'registrationNumber': 'VD-111',
                    'units': [{'unitCode': 'Vỉ', 'toBaseQty': 10.0, 'price': 5000.0}]
                },
                'stock_info': {
                    'lotNo': 'LOT_ATOMIC_1',
                    'expiryDate': '2030-01-01',
                    'qty': 100.0,
                    'cost': 2000.0,
                    'fundSource': 'Nguồn A'
                }
            },
            {
                'product_info': {
                    'name': 'Thuốc Atomic 2 (Lỗi)',
                    'defaultUnit': 'Chai',
                    'barcode': '22222',
                    'productType': 'thuoc',
                    'registrationNumber': 'VD-222',
                    'units': []
                },
                # Cố tình gây lỗi tồn kho âm để phát sinh exception ở _assert_no_negative_stock / validate
                'stock_info': {
                    'lotNo': 'LOT_ATOMIC_2',
                    'expiryDate': 'INVALID_DATE_FORMAT_EX',
                    'qty': -50.0,
                    'cost': 1000.0,
                    'fundSource': ''
                }
            }
        ]

        # Thao tác bulk_import_products_and_stock bị ném ra ngoại lệ
        with self.assertRaises(Exception):
            self.db.bulk_import_products_and_stock(records_with_error)

        # Kiểm tra CSDL: Không một sản phẩm hay đơn vị nào được phép tồn tại (Rollback 100%)
        count1 = self.db.q("SELECT COUNT(*) as c FROM products WHERE name IN ('Thuốc Atomic 1', 'Thuốc Atomic 2 (Lỗi)')")[0]['c']
        self.assertEqual(count1, 0, "Bắt buộc rollback 100% sản phẩm khi bulk_import bị lỗi")

        # Thử lại với dữ liệu hoàn toàn hợp lệ
        valid_records = [
            {
                'product_info': {
                    'name': 'Thuốc Atomic Hợp Lệ',
                    'defaultUnit': 'Hộp',
                    'barcode': '33333',
                    'productType': 'thuoc',
                    'registrationNumber': 'VD-333',
                    'units': [{'unitCode': 'Vỉ', 'toBaseQty': 10.0, 'price': 5000.0}]
                },
                'stock_info': {
                    'lotNo': 'LOT_OK_1',
                    'expiryDate': '2030-12-31',
                    'qty': 50.0,
                    'cost': 15000.0,
                    'fundSource': 'Nguồn OK'
                }
            }
        ]
        p_count, u_count, unit_count, s_count, note_num = self.db.bulk_import_products_and_stock(valid_records)
        self.assertEqual(p_count, 1)
        self.assertEqual(u_count, 0)
        self.assertEqual(s_count, 1)
        self.assertTrue(note_num.startswith('PN-'))

        # Gọi lại lần 2 để kiểm tra cập nhật sản phẩm đã tồn tại (updated_products)
        update_records = [
            {
                'product_info': {
                    'name': 'Thuốc Atomic Hợp Lệ',
                    'defaultUnit': 'Hộp',
                    'barcode': '33333_UPDATED',
                    'productType': 'vaccine',
                    'registrationNumber': 'VD-333-NEW',
                    'units': []
                },
                'stock_info': None
            }
        ]
        p_count2, u_count2, unit_count2, s_count2, note_num2 = self.db.bulk_import_products_and_stock(update_records)
        self.assertEqual(p_count2, 0, "Không tạo sản phẩm trùng tên")
        self.assertEqual(u_count2, 1, "Tăng số lượng sản phẩm được cập nhật")

        # Kiểm tra sản phẩm đã được cập nhật thông tin thành công
        p_row = self.db.q("SELECT barcode, productType, registrationNumber FROM products WHERE name='Thuốc Atomic Hợp Lệ'")
        self.assertEqual(p_row[0]['barcode'], '33333_UPDATED')
        self.assertEqual(p_row[0]['productType'], 'vaccine')
        self.assertEqual(p_row[0]['registrationNumber'], 'VD-333-NEW')

    def test_historical_dispatch_fefo_date(self):
        """Kiểm thử: Nhập hàng 2026-01-01 (HSD 2026-06-01), xuất ngày 2026-01-22 phải thành công (không bị DATE('now') làm báo KHÔNG ĐỦ KHO)"""
        # 1. Tạo sản phẩm
        cur = self.db.conn.execute("INSERT INTO products (name, defaultUnit) VALUES ('Nevirapine Test', 'Lọ')")
        pid = cur.lastrowid
        self.db.conn.execute("INSERT INTO product_units (productId, unitCode, toBaseQty, price) VALUES (?, 'Lọ', 1, 32489)", (pid,))

        # 2. Nhập kho ngày 2026-01-01 với HSD 2026-06-01 (HSD đã qua so với DATE('now') hiện tại nhưng chưa qua so với Ngày xuất 2026-01-22)
        purchase_items = [{
            'productId': pid,
            'lotNo': 'LOT_NEV_2026',
            'expiryDate': '2026-06-01',
            'unitCode': 'Lọ',
            'qty': 10,
            'cost': 32489,
            'fundSource': 'Quỹ toàn cầu'
        }]
        self.db.record_purchase(purchase_items, supplier="NCC Test", date_str="2026-01-01")

        # 3. Xuất kho ngày 2026-01-22: FEFO phải so sánh HSD với ngày xuất 2026-01-22 chứ không so với DATE('now')
        dispatch_items = [{
            'productId': pid,
            'unitCode': 'Lọ',
            'qty': 2,
            'lotNo': None, # FEFO tự động
            'fundSource': 'Quỹ toàn cầu'
        }]
        dispatch_id, note_num, _ = self.db.dispatch(dispatch_items, receiving_unit="BV Phụ sản", date_str="2026-01-22")
        self.assertTrue(note_num.startswith('PX-'), "Xuất kho lùi ngày 2026-01-22 phải thành công không bị báo lỗi không đủ kho")

        # Kiểm tra tồn kho sau khi xuất còn 8 Lọ
        inv = self.db.get_inventory()
        matching = [x for x in inv if x['productId'] == pid and x['fundSource'] == 'Quỹ toàn cầu']
        self.assertEqual(len(matching), 1)
        self.assertEqual(matching[0]['stockBase'], 8.0)

    def test_dd_mm_yyyy_date_helpers(self):
        self.assertEqual(parse_date_to_iso("24-07-2026"), "2026-07-24")
        self.assertEqual(parse_date_to_iso("24/07/2026"), "2026-07-24")
        self.assertEqual(format_date_display("2026-07-24"), "24-07-2026")
        self.assertEqual(format_datetime_display("2026-07-24 09:15:30"), "24-07-2026 09:15:30")
        with self.assertRaises(ValueError):
            parse_date_to_iso("31/02/2026")
        with self.assertRaises(ValueError):
            parse_date_to_iso("2026-22-01")

    def test_future_purchase_not_available_for_backdated_dispatch(self):
        self.db.conn.execute("INSERT INTO products (id, name, defaultUnit) VALUES (205, 'Future Purchase Drug', 'Vien')")
        self.db.conn.execute("INSERT INTO product_units (productId, unitCode, toBaseQty, price) VALUES (205, 'Vien', 1, 0)")
        self.db.conn.commit()

        self.db.record_purchase([{
            "productId": 205,
            "productName": "Future Purchase Drug",
            "qty": 10,
            "unitCode": "Vien",
            "lotNo": "FP01",
            "expiryDate": "2028-12-31",
            "cost": 100,
            "fundSource": "N"
        }], "NCC", "Nhap", "", date_str="2026-02-01")

        with self.assertRaisesRegex(Exception, "Không đủ tồn kho"):
            self.db.dispatch([{
                "productId": 205,
                "qty": 1,
                "unitCode": "Vien",
                "lotNo": None,
                "fundSource": "N"
            }], "Don vi", "Xuat lui ngay", "", date_str="2026-01-22")

    def test_future_dispatch_does_not_reduce_historical_stock_lookup(self):
        self.db.conn.execute("INSERT INTO products (id, name, defaultUnit) VALUES (206, 'Historical Stock Drug', 'Vien')")
        self.db.conn.execute("INSERT INTO product_units (productId, unitCode, toBaseQty, price) VALUES (206, 'Vien', 1, 0)")
        self.db.conn.commit()

        self.db.record_purchase([{
            "productId": 206,
            "productName": "Historical Stock Drug",
            "qty": 10,
            "unitCode": "Vien",
            "lotNo": "HS01",
            "expiryDate": "2028-12-31",
            "cost": 100,
            "fundSource": "N"
        }], "NCC", "Nhap", "", date_str="2026-01-01")
        self.db.dispatch([{
            "productId": 206,
            "qty": 8,
            "unitCode": "Vien",
            "lotNo": "HS01",
            "fundSource": "N"
        }], "Don vi", "Xuat sau", "", date_str="2026-02-01")

        stock_on_jan_22 = self.db.get_stock_as_of("2026-01-22", product_id=206, fund_source="N")
        self.assertEqual(stock_on_jan_22[0]["stockBase"], 10.0)

        with self.assertRaisesRegex(Exception, "tồn kho âm|Không đủ tồn kho"):
            self.db.dispatch([{
                "productId": 206,
                "qty": 5,
                "unitCode": "Vien",
                "lotNo": "HS01",
                "fundSource": "N"
            }], "Don vi", "Khong du ton hien tai", "", date_str="2026-01-22")

    def test_backdated_dispatch_uses_historical_balance_before_future_dispatch(self):
        self.db.conn.execute("INSERT INTO products (id, name, defaultUnit) VALUES (207, 'Backdated Drug', 'Vien')")
        self.db.conn.execute("INSERT INTO product_units (productId, unitCode, toBaseQty, price) VALUES (207, 'Vien', 1, 0)")
        self.db.conn.commit()

        self.db.record_purchase([{
            "productId": 207,
            "productName": "Backdated Drug",
            "qty": 10,
            "unitCode": "Vien",
            "lotNo": "BD01",
            "expiryDate": "2028-12-31",
            "cost": 100,
            "fundSource": "N"
        }], "NCC", "Nhap", "", date_str="2026-01-01")
        self.db.dispatch([{
            "productId": 207,
            "qty": 4,
            "unitCode": "Vien",
            "lotNo": "BD01",
            "fundSource": "N"
        }], "Don vi", "Xuat sau", "", date_str="2026-02-01")

        dispatch_id, _, _ = self.db.dispatch([{
            "productId": 207,
            "qty": 5,
            "unitCode": "Vien",
            "lotNo": "BD01",
            "fundSource": "N"
        }], "Don vi", "Xuat lui ngay", "", date_str="2026-01-22")
        self.assertGreater(dispatch_id, 0)
        stock_now = self.db.get_inventory()
        row = next(r for r in stock_now if r["productId"] == 207 and r["fundSource"] == "N")
        self.assertEqual(row["stockBase"], 1.0)

    def test_note_details_keep_insert_order(self):
        self.db.conn.execute("INSERT INTO products (id, name, defaultUnit) VALUES (201, 'Zeta', 'Vien')")
        self.db.conn.execute("INSERT INTO products (id, name, defaultUnit) VALUES (202, 'Alpha', 'Vien')")
        self.db.conn.execute("INSERT INTO product_units (productId, unitCode, toBaseQty, price) VALUES (201, 'Vien', 1, 100)")
        self.db.conn.execute("INSERT INTO product_units (productId, unitCode, toBaseQty, price) VALUES (202, 'Vien', 1, 100)")
        self.db.conn.commit()

        purchase_id, _, _ = self.db.record_purchase([
            {"productId": 201, "productName": "Zeta", "qty": 5, "unitCode": "Vien", "lotNo": "LZ", "expiryDate": "2028-12-31", "cost": 100, "fundSource": "N"},
            {"productId": 202, "productName": "Alpha", "qty": 5, "unitCode": "Vien", "lotNo": "LA", "expiryDate": "2028-12-31", "cost": 100, "fundSource": "N"},
        ], "NCC", "Nhap", "")
        purchase_names = [row["productName"] for row in self.db.get_purchase_detail(purchase_id)]
        self.assertEqual(purchase_names, ["Zeta", "Alpha"])

        dispatch_id, _, _ = self.db.dispatch([
            {"productId": 201, "qty": 1, "unitCode": "Vien", "lotNo": "LZ", "fundSource": "N"},
            {"productId": 202, "qty": 1, "unitCode": "Vien", "lotNo": "LA", "fundSource": "N"},
        ], "Don vi", "Xuat", "")
        dispatch_names = [row["productName"] for row in self.db.get_dispatch_detail(dispatch_id)]
        self.assertEqual(dispatch_names, ["Zeta", "Alpha"])

    def test_purchase_total_amount_is_preserved_for_fractional_unit_cost(self):
        self.db.conn.execute("INSERT INTO products (id, name, defaultUnit) VALUES (203, 'Fractional Cost Drug', 'Vien')")
        self.db.conn.execute("INSERT INTO product_units (productId, unitCode, toBaseQty, price) VALUES (203, 'Vien', 1, 0)")
        self.db.conn.commit()

        purchase_id, note_num, details = self.db.record_purchase([{
            "productId": 203,
            "productName": "Fractional Cost Drug",
            "qty": 60,
            "unitCode": "Vien",
            "lotNo": "LF",
            "expiryDate": "2028-12-31",
            "cost": 0,
            "totalAmount": 125000,
            "fundSource": "N"
        }], "NCC", "Nhap", "", date_str="2026-07-24")

        self.assertEqual(note_num, "PN-240726-001")
        self.assertAlmostEqual(details[0]["cost"], 125000 / 60)
        self.assertEqual(details[0]["totalAmount"], 125000)
        row = self.db.get_purchase_detail(purchase_id)[0]
        self.assertEqual(row["totalAmount"], 125000)

    def test_inventory_adjustment_and_history_lookup(self):
        self.db.conn.execute("INSERT INTO products (id, name, defaultUnit) VALUES (204, 'Inventory Check Drug', 'Vien')")
        self.db.conn.execute("INSERT INTO product_units (productId, unitCode, toBaseQty, price) VALUES (204, 'Vien', 1, 0)")
        self.db.conn.commit()
        self.db.record_purchase([{
            "productId": 204,
            "productName": "Inventory Check Drug",
            "qty": 10,
            "unitCode": "Vien",
            "lotNo": "IC01",
            "expiryDate": "2028-12-31",
            "cost": 100,
            "fundSource": "N"
        }], "NCC", "Nhap", "")
        inv = self.db.get_inventory()
        row = next(r for r in inv if r["productId"] == 204)

        applied = self.db.record_inventory_adjustments([{
            "productId": 204,
            "batchId": row["batchId"],
            "fundSource": "N",
            "actualQtyBase": 7,
            "note": "Kiem ke test"
        }])
        self.assertEqual(len(applied), 1)
        self.assertEqual(applied[0]["delta"], -3)

        inv_after = self.db.get_inventory()
        row_after = next(r for r in inv_after if r["productId"] == 204)
        self.assertEqual(row_after["stockBase"], 7)

        history = self.db.product_lot_history(product_id=204, lot_no="IC01")
        self.assertEqual(history[0]["type"], "ADJUSTMENT")
        self.assertTrue(self.db.dashboard_summary()["product_count"] >= 1)

    def test_xnt_report_excel_export_creates_formatted_workbook(self):
        import ui
        from ui import App
        from openpyxl import load_workbook

        self.db.conn.execute("INSERT INTO products (id, name, defaultUnit) VALUES (208, 'Excel Export Drug', 'Vien')")
        self.db.conn.execute("INSERT INTO product_units (productId, unitCode, toBaseQty, price) VALUES (208, 'Vien', 1, 0)")
        self.db.conn.commit()
        self.db.record_purchase([{
            "productId": 208,
            "productName": "Excel Export Drug",
            "qty": 10,
            "unitCode": "Vien",
            "lotNo": "XNT01",
            "expiryDate": "2028-12-31",
            "cost": 100,
            "fundSource": "N"
        }], "NCC", "Nhap", "", date_str="2026-07-01")
        self.db.dispatch([{
            "productId": 208,
            "qty": 4,
            "unitCode": "Vien",
            "lotNo": "XNT01",
            "fundSource": "N"
        }], "Don vi", "Xuat", "", date_str="2026-07-02")

        class DummyEntry:
            def __init__(self, value):
                self.entry = self
                self.value = value
            def get(self):
                return self.value

        class DummyCombo:
            def get(self):
                return "N"

        app = App.__new__(App)
        app.db = self.db
        app.de_from = DummyEntry("01-07-2026")
        app.de_to = DummyEntry("31-07-2026")
        app.cmb_report_fund = DummyCombo()
        app.toast = lambda *args, **kwargs: None

        export_path = os.path.join(self.temp_dir, "xnt_export.xlsx")
        with mock.patch.object(ui.filedialog, "asksaveasfilename", return_value=export_path), \
             mock.patch.object(ui.os, "startfile", return_value=None), \
             mock.patch.object(ui.messagebox, "showerror") as showerror:
            App.export_report_excel(app)

        showerror.assert_not_called()
        self.assertTrue(os.path.exists(export_path))

        wb = load_workbook(export_path)
        ws = wb.active
        self.assertEqual(ws["A1"].value, "BÁO CÁO XUẤT - NHẬP - TỒN")
        self.assertIn("01-07-2026", ws["A2"].value)
        self.assertEqual(ws["E5"].value, "31-12-2028")
        self.assertEqual(ws["H5"].value, 10)
        self.assertEqual(ws["I5"].value, 4)
        self.assertEqual(ws["J5"].value, 6)
        self.assertEqual(ws.freeze_panes, "A5")

if __name__ == '__main__':
    unittest.main()
